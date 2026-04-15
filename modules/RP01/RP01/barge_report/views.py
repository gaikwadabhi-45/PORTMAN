from flask import render_template, request, jsonify, session, redirect, url_for, Response
from functools import wraps
import traceback
import io

from .. import bp
from database import get_db, get_cursor

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XL_AVAILABLE = True
except ImportError:
    _XL_AVAILABLE = False


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def _fmt(val):
    """Return string or empty string for None."""
    if val is None:
        return ''
    return str(val)


# ── routes ──────────────────────────────────────────────────────────────────────

@bp.route('/module/RP01/barge-report/')
@login_required
def barge_report_index():
    return render_template('barge_report/barge_report.html', username=session.get('username'))


@bp.route('/api/module/RP01/barge-report/vessels')
@login_required
def barge_report_vessels():
    """Return ldud records for a given operation type (import / export)."""
    op_type = request.args.get('op_type', 'import').strip()
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT h.id, h.doc_num, h.vessel_name, h.operation_type,
                   h.nor_tendered, h.doc_status
            FROM ldud_header h
            WHERE LOWER(h.operation_type) = LOWER(%s)
            ORDER BY h.id DESC
        """, (op_type,))
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'doc_num': r['doc_num'] or '',
                'vessel_name': r['vessel_name'] or '',
                'operation_type': r['operation_type'] or '',
                'nor_tendered': _fmt(r['nor_tendered']),
                'doc_status': r['doc_status'] or '',
                'label': f"{r['doc_num'] or ''} — {r['vessel_name'] or ''}",
            })
        return jsonify(result)
    except Exception:
        return jsonify({'error': traceback.format_exc()}), 500
    finally:
        conn.close()


@bp.route('/api/module/RP01/barge-report/data')
@login_required
def barge_report_data():
    """Return barge lines for selected ldud_ids, with vessel header repeated per line."""
    ids_param = request.args.get('ldud_ids', '').strip()
    if not ids_param:
        return jsonify([])

    try:
        ldud_ids = [int(x) for x in ids_param.split(',') if x.strip()]
    except ValueError:
        return jsonify({'error': 'Invalid ldud_ids'}), 400

    if not ldud_ids:
        return jsonify([])

    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT
                h.id                          AS ldud_id,
                h.doc_num,
                h.vessel_name,
                h.operation_type,
                h.nor_tendered,
                h.discharge_commenced,
                h.discharge_completed,
                h.doc_status,
                h.initial_draft_survey_quantity,
                v.vessel_agent_name,
                v.importer_exporter_name,
                bl.id                             AS line_id,
                bl.trip_number,
                bl.hold_name,
                bl.barge_name,
                bl.contractor_name,
                bl.cargo_name,
                bl.bpt_bfl,
                bl.discharge_quantity,
                bl.crane_loaded_from,
                bl.port_crane,
                bl.trip_start,
                bl.anchored_gull_island,
                bl.aweigh_gull_island,
                bl.amf_at_port,
                bl.along_side_vessel,
                bl.along_side_berth,
                bl.commenced_loading,
                bl.completed_loading,
                bl.cast_off_mv,
                bl.cast_off_berth,
                bl.cast_off_berth_nt,
                bl.cast_off_port,
                bl.commence_discharge_berth,
                bl.completed_discharge_berth,
                bl.anchored_gull_island_empty,
                bl.aweigh_gull_island_empty
            FROM ldud_header h
            LEFT JOIN vcn_header v ON v.id = h.vcn_id
            LEFT JOIN ldud_barge_lines bl ON bl.ldud_id = h.id
            WHERE h.id = ANY(%s)
            ORDER BY h.id, bl.trip_number, bl.barge_name, bl.id
        """, (ldud_ids,))
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'ldud_id':                        r['ldud_id'],
                'doc_num':                        _fmt(r['doc_num']),
                'vessel_name':                    _fmt(r['vessel_name']),
                'operation_type':                 _fmt(r['operation_type']),
                'nor_tendered':                   _fmt(r['nor_tendered']),
                'discharge_commenced':            _fmt(r['discharge_commenced']),
                'discharge_completed':            _fmt(r['discharge_completed']),
                'doc_status':                     _fmt(r['doc_status']),
                'initial_draft_survey_quantity':  _fmt(r['initial_draft_survey_quantity']),
                'vessel_agent_name':              _fmt(r['vessel_agent_name']),
                'importer_exporter_name':         _fmt(r['importer_exporter_name']),
                'line_id':                        r['line_id'],
                'trip_number':                    r['trip_number'],
                'hold_name':                      _fmt(r['hold_name']),
                'barge_name':                     _fmt(r['barge_name']),
                'contractor_name':                _fmt(r['contractor_name']),
                'cargo_name':                     _fmt(r['cargo_name']),
                'bpt_bfl':                        _fmt(r['bpt_bfl']),
                'discharge_quantity':             float(r['discharge_quantity']) if r['discharge_quantity'] is not None else None,
                'crane_loaded_from':              _fmt(r['crane_loaded_from']),
                'port_crane':                     _fmt(r['port_crane']),
                'trip_start':                     _fmt(r['trip_start']),
                'anchored_gull_island':           _fmt(r['anchored_gull_island']),
                'aweigh_gull_island':             _fmt(r['aweigh_gull_island']),
                'amf_at_port':                    _fmt(r['amf_at_port']),
                'along_side_vessel':              _fmt(r['along_side_vessel']),
                'along_side_berth':               _fmt(r['along_side_berth']),
                'commenced_loading':              _fmt(r['commenced_loading']),
                'completed_loading':              _fmt(r['completed_loading']),
                'cast_off_mv':                    _fmt(r['cast_off_mv']),
                'cast_off_berth':                 _fmt(r['cast_off_berth']),
                'cast_off_berth_nt':              _fmt(r['cast_off_berth_nt']),
                'cast_off_port':                  _fmt(r['cast_off_port']),
                'commence_discharge_berth':       _fmt(r['commence_discharge_berth']),
                'completed_discharge_berth':      _fmt(r['completed_discharge_berth']),
                'anchored_gull_island_empty':     _fmt(r['anchored_gull_island_empty']),
                'aweigh_gull_island_empty':       _fmt(r['aweigh_gull_island_empty']),
            })
        return jsonify(result)
    except Exception:
        return jsonify({'error': traceback.format_exc()}), 500
    finally:
        conn.close()


# ── Excel export ─────────────────────────────────────────────────────────────

_BARGE_COLUMNS = [
    ('Doc No',                      'doc_num'),
    ('Vessel Name',                 'vessel_name'),
    ('Operation',                   'operation_type'),
    ('NOR Tendered',                'nor_tendered'),
    ('Discharge Commenced',         'discharge_commenced'),
    ('Discharge Completed',         'discharge_completed'),
    ('Status',                      'doc_status'),
    ('Init. Draft Survey Qty',      'initial_draft_survey_quantity'),
    ('Agent',                       'vessel_agent_name'),
    ('Importer / Exporter',         'importer_exporter_name'),
    ('Trip #',                      'trip_number'),
    ('Hold',                        'hold_name'),
    ('Barge',                       'barge_name'),
    ('Stevedore / Contractor',      'contractor_name'),
    ('Cargo',                       'cargo_name'),
    ('MBPT/PLA',                    'bpt_bfl'),
    ('Qty (MT)',                    'discharge_quantity'),
    ('Crane Loaded From',           'crane_loaded_from'),
    ('Port Crane',                  'port_crane'),
    ('Trip Start',                  'trip_start'),
    ('Anch. Gull Island',           'anchored_gull_island'),
    ('Aweigh Gull Island',          'aweigh_gull_island'),
    ('AMF at Port',                 'amf_at_port'),
    ('Alongside Vessel (MV)',       'along_side_vessel'),
    ('Alongside Berth',             'along_side_berth'),
    ('Loading Start',               'commenced_loading'),
    ('Loading End',                 'completed_loading'),
    ('Cast Off MV',                 'cast_off_mv'),
    ('Cast Off Berth',              'cast_off_berth'),
    ('Cast Off Berth NT',           'cast_off_berth_nt'),
    ('Cast Off Port',               'cast_off_port'),
    ('Discharge Start (Berth)',     'commence_discharge_berth'),
    ('Discharge End (Berth)',       'completed_discharge_berth'),
    ('Anch. Gull Island (Loaded)',  'anchored_gull_island_empty'),
    ('Aweigh Gull Island (Loaded)', 'aweigh_gull_island_empty'),
]


@bp.route('/api/module/RP01/barge-report/download')
@login_required
def barge_report_download():
    """Export selected barge lines to Excel."""
    ids_param = request.args.get('ldud_ids', '').strip()
    if not ids_param:
        return Response('No ldud_ids provided', status=400)
    try:
        ldud_ids = [int(x) for x in ids_param.split(',') if x.strip()]
    except ValueError:
        return Response('Invalid ldud_ids', status=400)

    conn = get_db()
    cur  = get_cursor(conn)
    try:
        cur.execute("""
            SELECT
                h.doc_num, h.vessel_name, h.operation_type,
                h.nor_tendered, h.discharge_commenced, h.discharge_completed,
                h.doc_status, h.initial_draft_survey_quantity,
                v.vessel_agent_name, v.importer_exporter_name,
                bl.trip_number, bl.hold_name, bl.barge_name,
                bl.contractor_name, bl.cargo_name, bl.bpt_bfl, bl.discharge_quantity,
                bl.crane_loaded_from, bl.port_crane, bl.trip_start,
                bl.anchored_gull_island, bl.aweigh_gull_island, bl.amf_at_port,
                bl.along_side_vessel, bl.along_side_berth,
                bl.commenced_loading, bl.completed_loading,
                bl.cast_off_mv, bl.cast_off_berth, bl.cast_off_berth_nt,
                bl.cast_off_port, bl.commence_discharge_berth, bl.completed_discharge_berth,
                bl.anchored_gull_island_empty, bl.aweigh_gull_island_empty
            FROM ldud_header h
            LEFT JOIN vcn_header v ON v.id = h.vcn_id
            LEFT JOIN ldud_barge_lines bl ON bl.ldud_id = h.id
            WHERE h.id = ANY(%s)
            ORDER BY h.id, bl.trip_number, bl.barge_name, bl.id
        """, (ldud_ids,))
        xl_rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin      = Side(style='thin', color='000000')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill('solid', fgColor='C0C0C0')
    hdr_font  = Font(name='Calibri', bold=True,  size=10)
    cell_font = Font(name='Calibri', bold=False, size=10)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rgt = Alignment(horizontal='right',  vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Barge Lines'
    ws.freeze_panes = 'A2'

    ws.row_dimensions[1].height = 24
    for col_idx, (label, _) in enumerate(_BARGE_COLUMNS, 1):
        c = ws.cell(1, col_idx, label)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bdr; c.alignment = ctr
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(label) + 2)

    for row_idx, row in enumerate(xl_rows, 2):
        for col_idx, (_, key) in enumerate(_BARGE_COLUMNS, 1):
            val = row.get(key)
            if val is None:
                val = ''
            elif hasattr(val, 'isoformat'):
                val = str(val)
            c = ws.cell(row_idx, col_idx, val)
            c.font = cell_font; c.border = bdr
            c.alignment = rgt if key == 'discharge_quantity' else lft

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="BargeLines.xlsx"'},
    )
