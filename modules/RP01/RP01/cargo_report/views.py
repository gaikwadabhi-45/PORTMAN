from flask import (
    render_template, request, session,
    redirect, url_for, Response,
    jsonify, send_file
)
from functools import wraps
from io import BytesIO
from datetime import datetime
import io
import json
import re

from .. import bp
from database import get_db, get_cursor

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_thin = Side(style='thin', color='C7CDD4')
_bdr = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

_TITLE_FILL = 'F6F8FB'
_HEADER_FILL = 'EEF4FB'
_BODY_FILL = 'FFFFFF'
_GROUP_FILL = 'F8FAFC'
_SUBTOTAL_FILL = 'F6F7F9'
_TOTAL_FILL = 'E9EEF5'
_TYPE_TOTAL_FILL = 'F1F6FB'
_TEXT = '2C3E50'


# =========================================================
# LOGIN REQUIRED
# =========================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# =========================================================
# INDEX PAGE
# =========================================================

@bp.route('/module/RP01/cargo-report/')
@login_required
def cargo_report_index():
    return render_template(
        'cargo_report/cargo_report.html',
        username=session.get('username')
    )


# =========================================================
# FETCH REPORT DATA
# =========================================================

@bp.route('/api/cargo-report')
@login_required
def get_cargo_report():

    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')

    conn = get_db()
    cur  = get_cursor(conn)

    try:
        query = """
SELECT
    h.id,

    COALESCE(h.doc_num, '-') || ' / ' ||
    COALESCE(h.mbc_name, '-') AS vessel_name,

    'MBC' AS vessel_type,

    h.cargo_type,

    h.cargo_name,

    COALESCE(cd.quantity, 0) AS bl_qty_mt,

    COALESCE(cd.quantity, 0) AS actual_discharge,

    h.load_port,

    cd.material_po AS material_po,

    cd.customer_name AS consignee,

    dp.unloading_commenced::text AS discharge_commenced,

    CASE
        WHEN dp.unloading_commenced IS NOT NULL
             AND dp.unloading_completed IS NULL
        THEN 'InProgress'
        ELSE dp.unloading_completed::text
    END AS discharge_completed,

    COALESCE(v.nationality, '-') AS flag

FROM mbc_header h

LEFT JOIN vessels v
    ON LOWER(TRIM(v.vessel_name)) =
       LOWER(TRIM(h.mbc_name))

LEFT JOIN mbc_discharge_port_lines dp
    ON dp.mbc_id = h.id

LEFT JOIN mbc_customer_details cd
    ON cd.mbc_id = h.id

WHERE NULLIF(TRIM(dp.unloading_commenced), '') IS NOT NULL

  AND NULLIF(TRIM(dp.unloading_commenced), '')::date
      BETWEEN %s::date AND %s::date

GROUP BY
    h.id,
    h.doc_num,
    h.mbc_name,
    h.cargo_type,
    h.cargo_name,
    cd.quantity,
    h.load_port,
    cd.material_po,
    cd.customer_name,
    dp.unloading_commenced,
    dp.unloading_completed,
    v.nationality

UNION ALL

SELECT
    lh.id,

    COALESCE(vh.vcn_doc_num, '-') || ' / ' ||
    COALESCE(lh.vessel_name, '-') AS vessel_name,

    'MV' AS vessel_type,

    vh.cargo_type AS cargo_type,

    vc.cargo_name AS cargo_name,

    COALESCE(vc.bl_quantity, 0) AS bl_qty_mt,

    COALESCE(mv.actual_discharge, 0) AS actual_discharge,

    vh.load_port AS load_port,

    lh.material_po_number AS material_po,

    vc.customer_name AS consignee,

    la.first_discharge_started::text AS discharge_commenced,

    la.last_discharge_completed::text AS discharge_completed,

    COALESCE(v.nationality, '-') AS flag

FROM ldud_header lh

LEFT JOIN (
    SELECT
        ldud_id,

        MIN(discharge_started) AS first_discharge_started,

        CASE
            WHEN SUM(
                CASE
                    WHEN discharge_started IS NOT NULL
                     AND discharge_commenced IS NULL
                    THEN 1
                    ELSE 0
                END
            ) > 0
            THEN NULL
            ELSE MAX(discharge_commenced)
        END AS last_discharge_completed

    FROM ldud_anchorage
    GROUP BY ldud_id
) la ON la.ldud_id = lh.id

LEFT JOIN vessels v
    ON LOWER(TRIM(v.vessel_name)) =
       LOWER(TRIM(lh.vessel_name))

LEFT JOIN vcn_header vh
    ON vh.id = lh.vcn_id

LEFT JOIN vcn_cargo_declaration vc
    ON vc.vcn_id = lh.vcn_id

LEFT JOIN (
    SELECT
        LOWER(
            REGEXP_REPLACE(
                TRIM(source_display),
                '\s+',
                ' ',
                'g'
            )
        ) AS match_display,

        LOWER(TRIM(cargo_name)) AS match_cargo,

        SUM(
            CASE
                WHEN TRIM(COALESCE(quantity::text, '')) != ''
                THEN quantity::numeric
                ELSE 0
            END
        ) AS actual_discharge

    FROM lueu_lines

    WHERE source_type = 'VCN'
      AND is_deleted = false

    GROUP BY
        LOWER(
            REGEXP_REPLACE(
                TRIM(source_display),
                '\s+',
                ' ',
                'g'
            )
        ),
        LOWER(TRIM(cargo_name))

) mv
ON mv.match_display =
   LOWER(
       REGEXP_REPLACE(
           TRIM(
               COALESCE(vh.vcn_doc_num, '') || ' / ' ||
               COALESCE(lh.vessel_name, '')
           ),
           '\s+',
           ' ',
           'g'
       )
   )

AND mv.match_cargo =
    LOWER(TRIM(vc.cargo_name))

WHERE la.first_discharge_started IS NOT NULL

  AND la.first_discharge_started::date
      BETWEEN %s::date AND %s::date

GROUP BY
    lh.id,
    vh.vcn_doc_num,
    lh.vessel_name,
    vh.cargo_type,
    vc.cargo_name,
    vc.bl_quantity,
    vh.load_port,
    lh.material_po_number,
    vc.customer_name,
    la.first_discharge_started,
    la.last_discharge_completed,
    v.nationality,
    mv.actual_discharge

ORDER BY discharge_commenced DESC;
"""

        cur.execute(query, (from_date, to_date, from_date, to_date))
        rows = cur.fetchall()
        data = []
        for row in rows:
            data.append({
    'id': row['id'],
    'vessel_name': row['vessel_name'] or '-',
    'vessel_type': row['vessel_type'] or '-',
    'material_po': row['material_po'] or '-',
    'cargo_type': row['cargo_type'] or '-',
    'cargo_name': row['cargo_name'] or '-',
    'bl_qty_mt': float(row['bl_qty_mt']) if row['bl_qty_mt'] else 0,
    'actual_discharge': float(row['actual_discharge']) if row['actual_discharge'] else 0,
    'load_port': row['load_port'] or '-',

    'discharge_commenced':
        datetime.fromisoformat(
            str(row['discharge_commenced'])
        ).strftime('%d/%m/%Y %H:%M')
        if row['discharge_commenced'] else '-',

    'discharge_completed':
        'InProgress'
        if row['discharge_completed'] == 'InProgress'
        else datetime.fromisoformat(
            str(row['discharge_completed'])
        ).strftime('%d/%m/%Y %H:%M')
        if row['discharge_completed']
        else 'InProgress',

    'consignee': row['consignee'] or '-',
    'flag': row['flag'] or '-',
})

        return jsonify({'success': True, 'data': data})

    except Exception as e:
        print('CARGO REPORT ERROR:', e)
        return jsonify({'success': False, 'message': str(e)}), 500

    finally:
        cur.close()
        conn.close()

@bp.route('/api/update-material-po', methods=['POST'])
@login_required
def update_material_po():

    conn = get_db()
    cur = get_cursor(conn)

    try:

        data = request.get_json()

        print("UPDATE REQUEST:", data)

        record_id = data.get('id')
        vessel_type = data.get('vessel_type')
        material_po = data.get('material_po')

        if not record_id:
            return jsonify({
                'success': False,
                'message': 'Record ID missing'
            }), 400

        if vessel_type == 'MV':

            cur.execute("""
                UPDATE ldud_header
                SET material_po_number = %s
                WHERE id = %s
            """, (material_po, record_id))

        elif vessel_type == 'MBC':

            cur.execute("""
                UPDATE mbc_customer_details
                SET material_po = %s
                WHERE mbc_id = %s
            """, (material_po, record_id))

        else:

            return jsonify({
                'success': False,
                'message': f'Invalid vessel type: {vessel_type}'
            }), 400

        conn.commit()

        return jsonify({
            'success': True
        })

    except Exception as e:

        conn.rollback()
        print("UPDATE ERROR:", str(e))

        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

    finally:
        cur.close()
        conn.close()


# =========================================================
# DOWNLOAD FULL REPORT
# =========================================================

@bp.route('/api/module/RP01/cargo-report/download/')
@login_required
def download_cargo_handling_report():

    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')

    if not from_date or not to_date:
        return jsonify({'success': False, 'message': 'from_date and to_date are required'}), 400

    conn = get_db()
    cur  = get_cursor(conn)

    try:
        query = """
SELECT
    h.id,

    COALESCE(h.doc_num, '-') || ' / ' ||
    COALESCE(h.mbc_name, '-') AS vessel_name,

    'MBC' AS vessel_type,

    h.cargo_type,

    h.cargo_name,

    COALESCE(cd.quantity, 0) AS bl_qty_mt,

    COALESCE(cd.quantity, 0) AS actual_discharge,

    h.load_port,

    cd.material_po AS material_po,

    cd.customer_name AS consignee,

    dp.unloading_commenced::text AS discharge_commenced,

    dp.unloading_completed::text  AS discharge_completed,

    COALESCE(v.nationality, '-') AS flag

FROM mbc_header h

LEFT JOIN vessels v
    ON LOWER(TRIM(v.vessel_name)) =
       LOWER(TRIM(h.mbc_name))

LEFT JOIN mbc_discharge_port_lines dp
    ON dp.mbc_id = h.id

LEFT JOIN mbc_customer_details cd
    ON cd.mbc_id = h.id

WHERE NULLIF(TRIM(dp.unloading_completed), '') IS NOT NULL

  AND NULLIF(TRIM(dp.unloading_completed), '')::date
      BETWEEN %s::date AND %s::date

GROUP BY
    h.id,
    h.doc_num,
    h.mbc_name,
    h.cargo_type,
    h.cargo_name,
    cd.quantity,
    h.load_port,
    cd.material_po,
    cd.customer_name,
    dp.unloading_commenced,
    dp.unloading_completed,
    v.nationality

UNION ALL

SELECT
    lh.id,

    COALESCE(vh.vcn_doc_num, '-') || ' / ' ||
    COALESCE(lh.vessel_name, '-') AS vessel_name,

    'MV' AS vessel_type,

    vh.cargo_type AS cargo_type,

    vc.cargo_name AS cargo_name,

    COALESCE(vc.bl_quantity, 0) AS bl_qty_mt,

    COALESCE(mv.actual_discharge, 0) AS actual_discharge,

    vh.load_port AS load_port,

    lh.material_po_number AS material_po,

    vc.customer_name AS consignee,

    la.first_discharge_started::text  AS discharge_commenced,

    la.last_discharge_completed::text AS discharge_completed,

    COALESCE(v.nationality, '-') AS flag

FROM ldud_header lh

LEFT JOIN (
    SELECT
        ldud_id,

        MIN(discharge_started) AS first_discharge_started,

        CASE
            WHEN COUNT(
                CASE
                    WHEN discharge_started IS NOT NULL
                     AND discharge_commenced IS NULL
                    THEN 1
                END
            ) > 0
            THEN NULL
            ELSE MAX(discharge_commenced)
        END AS last_discharge_completed

    FROM ldud_anchorage
    GROUP BY ldud_id
) la ON la.ldud_id = lh.id

LEFT JOIN vessels v
    ON LOWER(TRIM(v.vessel_name)) =
       LOWER(TRIM(lh.vessel_name))

LEFT JOIN vcn_header vh
    ON vh.id = lh.vcn_id

LEFT JOIN vcn_cargo_declaration vc
    ON vc.vcn_id = lh.vcn_id

LEFT JOIN (
    SELECT
        LOWER(
            REGEXP_REPLACE(
                TRIM(source_display),
                '\s+',
                ' ',
                'g'
            )
        ) AS match_display,

        LOWER(TRIM(cargo_name)) AS match_cargo,

        SUM(
            CASE
                WHEN TRIM(COALESCE(quantity::text, '')) != ''
                THEN quantity::numeric
                ELSE 0
            END
        ) AS actual_discharge

    FROM lueu_lines

    WHERE source_type = 'VCN'
      AND is_deleted = false

    GROUP BY
        LOWER(
            REGEXP_REPLACE(
                TRIM(source_display),
                '\s+',
                ' ',
                'g'
            )
        ),
        LOWER(TRIM(cargo_name))

) mv
ON mv.match_display =
   LOWER(
       REGEXP_REPLACE(
           TRIM(
               COALESCE(vh.vcn_doc_num, '') || ' / ' ||
               COALESCE(lh.vessel_name, '')
           ),
           '\s+',
           ' ',
           'g'
       )
   )

AND mv.match_cargo =
    LOWER(TRIM(vc.cargo_name))

WHERE la.last_discharge_completed IS NOT NULL

  AND la.last_discharge_completed::date
      BETWEEN %s::date AND %s::date

GROUP BY
    lh.id,
    vh.vcn_doc_num,
    lh.vessel_name,
    vh.cargo_type,
    vc.cargo_name,
    vc.bl_quantity,
    vh.load_port,
    lh.material_po_number,
    vc.customer_name,
    la.first_discharge_started,
    la.last_discharge_completed,
    v.nationality,
    mv.actual_discharge

ORDER BY discharge_commenced DESC;
"""

        cur.execute(query, (from_date, to_date, from_date, to_date))
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Cargo Handling Report'

        # ── Styles ────────────────────────────────────────────
        thin      = Side(style='thin',   color='B8860B')
        thick     = Side(style='medium', color='B8860B')
        bdr_thin  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
        bdr_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

        fill_title  = PatternFill('solid', fgColor='F4B942')
        fill_header = PatternFill('solid', fgColor='E8E7B5')
        fill_even   = PatternFill('solid', fgColor='EBF3FB')
        fill_odd    = PatternFill('solid', fgColor='FFFFFF')
        fill_total  = PatternFill('solid', fgColor='FFFF00')

        font_title  = Font(name='Arial', bold=True, size=13, color='000000')
        font_header = Font(name='Arial', bold=True, size=11, color='000000')
        font_data   = Font(name='Arial', size=11,   color='344054')
        font_total  = Font(name='Arial', bold=True, size=11, color='000000')

        align_ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        headers = [
            'Sr No', 'M.Vessel Name', 'MV/MBC', 'Material PO',
            'Type', 'Cargo', 'B/L Qty.\n(MT)', 'Actual Discharge\n(MT)',
            'Load Port', 'Discharge\nCommence', 'Discharge\nCompleted',
            'Consignee /\nCustomer', 'Flag'
        ]
        col_widths = [7, 40, 9, 16, 9, 16, 12, 24, 20, 22, 22, 22, 18]
        num_cols   = len(headers)

        # ── Row 1: Title ──────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell           = ws.cell(row=1, column=1)
        title_cell.value     = 'Cargo Handling details_JSW Dharamtar Port'
        title_cell.font      = font_title
        title_cell.fill      = fill_title
        title_cell.alignment = align_ctr
        title_cell.border    = bdr_thick
        ws.row_dimensions[1].height = 28

        # ── Row 2: Column headers ─────────────────────────────
        for col, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell           = ws.cell(row=2, column=col)
            cell.value     = hdr
            cell.font      = font_header
            cell.fill      = fill_header
            cell.alignment = align_ctr
            cell.border    = bdr_thin
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 32

        # ── Data rows ─────────────────────────────────────────
        total_qty    = 0
        total_actual = 0
        data_start   = 3

        for idx, row in enumerate(rows, start=1):
            r      = data_start + idx - 1
            qty    = float(row['bl_qty_mt'])        if row['bl_qty_mt']        else 0
            actual = float(row['actual_discharge'])  if row['actual_discharge']  else 0
            total_qty    += qty
            total_actual += actual
            fill = fill_even if idx % 2 == 0 else fill_odd

            values = [
    idx,
    row['vessel_name'] or '',
    row['vessel_type'] or '',
    row['material_po'] or '',
    row['cargo_type'] or '',
    row['cargo_name'] or '',
    qty,
    actual,
    row['load_port'] or '',

    datetime.fromisoformat(
        str(row['discharge_commenced'])
    ).strftime('%d/%m/%Y %H:%M')
    if row['discharge_commenced'] else '',

    datetime.fromisoformat(
        str(row['discharge_completed'])
    ).strftime('%d/%m/%Y %H:%M')
    if row['discharge_completed'] else '',

    row['consignee'] or '',
    row['flag'] or '',
]

            for col, val in enumerate(values, start=1):
                cell           = ws.cell(row=r, column=col)
                cell.value     = val
                cell.font      = font_data
                cell.fill      = fill
                cell.border    = bdr_thin
                cell.alignment = align_ctr if col in [1, 3, 7, 8] else align_left
            ws.row_dimensions[r].height = 20

        # ── Total row ─────────────────────────────────────────
        total_row = data_start + len(rows)

        ws.merge_cells(start_row=total_row, start_column=5,
                       end_row=total_row,   end_column=6)

        lbl           = ws.cell(row=total_row, column=5)
        lbl.value     = 'Total'
        lbl.font      = font_total
        lbl.fill      = fill_header
        lbl.alignment = align_ctr
        lbl.border    = bdr_thin

        bl_cell               = ws.cell(row=total_row, column=7)
        bl_cell.value         = total_qty
        bl_cell.font          = font_total
        bl_cell.fill          = fill_total
        bl_cell.alignment     = align_ctr
        bl_cell.border        = bdr_thin
        bl_cell.number_format = '#,##0'

        act_cell               = ws.cell(row=total_row, column=8)
        act_cell.value         = total_actual
        act_cell.font          = font_total
        act_cell.fill          = fill_total
        act_cell.alignment     = align_ctr
        act_cell.border        = bdr_thin
        act_cell.number_format = '#,##0'

        for col in [1, 2, 3, 4, 9, 10, 11, 12, 13]:
            c        = ws.cell(row=total_row, column=col)
            c.fill   = fill_header
            c.border = bdr_thin

        ws.row_dimensions[total_row].height = 20

        # ── Freeze panes ──────────────────────────────────────
        ws.freeze_panes = 'A3'

        # ── Export ────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='Cargo_Handling_Report.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print('ERROR DOWNLOADING REPORT:', str(e))
        return jsonify({'success': False, 'message': str(e)}), 500

    finally:
        cur.close()
        conn.close()