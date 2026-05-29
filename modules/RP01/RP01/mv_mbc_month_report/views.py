from flask import render_template, request, jsonify, session, redirect, url_for, Response
from functools import wraps
from datetime import date, datetime, timedelta
from collections import defaultdict
import io
import re

from .. import bp
from database import get_db, get_cursor

# ── Excel colour / style constants ─────────────────────────────────────────
XL_GREY     = 'C0C0C0'
XL_LAVEND   = 'CCCCFF'
XL_CYAN     = 'CCFFFF'
XL_WHITE    = 'FFFFFF'
XL_NAVY     = '1F4E78'
XL_YELLOW   = 'FFFF00'
XL_TITLE_SZ = 14
XL_NORM_SZ  = 10
XL_SMALL_SZ = 9

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_thin   = Side(style='thin',   color='000000')
_med    = Side(style='medium', color='000000')
_bdr    = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
_bdr_ml = Border(left=_med,   right=_thin,  top=_thin,  bottom=_thin)
_ctr    = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_right  = Alignment(horizontal='right',  vertical='center', wrap_text=True)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=XL_NORM_SZ, color='000000'):
    return Font(name='Calibri', bold=bold, size=size, color=color)


def _safe_float(val):
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return 0.0


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── Data fetch ──────────────────────────────────────────────────────────────

def _fetch_mv_monthly_data(
    from_date=None,
    to_date=None,
    operation_type=None
):
    """
    Pivot daily quantity data by Mother Vessel OR MBC.

    Previous Month Qty logic:
    Only shows last previous-month carry-forward qty
    for the same MV/MBC.
    """

    conn = get_db()
    cur = get_cursor(conn)

    sql = """
    SELECT

        l.entry_date,
        l.quantity,
        l.cargo_name AS lueu_cargo,
        l.operation_type,
        l.source_type,
        l.source_id,

        COALESCE(
            CASE
                WHEN l.source_type = 'VCN'
                    THEN CONCAT(v.vcn_doc_num, ' / ', v.vessel_name)

                WHEN l.source_type = 'MBC'
                    THEN CONCAT(m.doc_num, ' / ', m.mbc_name)

                ELSE COALESCE(l.barge_name, 'Unknown')
            END,
            'Unknown'
        ) AS header_name,

        CASE
            WHEN l.source_type = 'VCN'
                THEN vc.cargo_name

            WHEN l.source_type = 'MBC'
                THEN COALESCE(m.cargo_name, l.cargo_name)

            ELSE l.cargo_name
        END AS cargo_name,

        CASE
            WHEN l.source_type = 'VCN'
                THEN COALESCE(
                    vcargo.cargo_type,
                    vc.cargo_name
                )

            WHEN l.source_type = 'MBC'
                THEN COALESCE(
                    m.cargo_type,
                    m.cargo_name,
                    l.cargo_name
                )

            ELSE l.cargo_name
        END AS cargo_type,

        CASE
            WHEN l.source_type = 'VCN'
                THEN COALESCE(vc_total.bl_quantity, 0)

            WHEN l.source_type = 'MBC'
                THEN COALESCE(mc_total.quantity, 0)

            ELSE 0
        END AS bl_qty,

        CASE
            WHEN l.source_type = 'VCN'
                THEN 'Mother Vessel'

            WHEN l.source_type = 'MBC'
                 AND UPPER(COALESCE(mm.mbc_owner_name, '')) LIKE '%%SHIPPING%%'
                THEN 'JSW SHIPPING'

            WHEN l.source_type = 'MBC'
                 AND UPPER(COALESCE(mm.mbc_owner_name, '')) LIKE '%%INFRA%%'
                THEN 'JSW INFRA'

            WHEN l.source_type = 'MBC'
                THEN 'OTHERS'

            ELSE 'Unknown'
        END AS company

    FROM lueu_lines l

    LEFT JOIN vcn_header v
        ON l.source_type = 'VCN'
       AND l.source_id = v.id

    LEFT JOIN vcn_cargo_declaration vc
        ON v.id = vc.vcn_id
       AND vc.cargo_name = l.cargo_name

    LEFT JOIN LATERAL (
        SELECT cargo_type
        FROM vessel_cargo
        WHERE cargo_name = vc.cargo_name
        LIMIT 1
    ) vcargo ON TRUE

    LEFT JOIN (
        SELECT
            vcn_id,
            SUM(bl_quantity) AS bl_quantity
        FROM vcn_cargo_declaration
        GROUP BY vcn_id
    ) vc_total
        ON v.id = vc_total.vcn_id

    LEFT JOIN mbc_header m
        ON l.source_type = 'MBC'
       AND l.source_id = m.id

    LEFT JOIN mbc_master mm
        ON UPPER(TRIM(m.mbc_name)) = UPPER(TRIM(mm.mbc_name))

    LEFT JOIN (
        SELECT
            mbc_id,
            cargo_name,
            SUM(quantity) AS quantity
        FROM mbc_customer_details
        GROUP BY mbc_id, cargo_name
    ) mc
        ON m.id = mc.mbc_id
       AND mc.cargo_name = l.cargo_name

    LEFT JOIN (
        SELECT
            mbc_id,
            SUM(quantity) AS quantity
        FROM mbc_customer_details
        GROUP BY mbc_id
    ) mc_total
        ON m.id = mc_total.mbc_id

    WHERE l.is_deleted IS NOT TRUE
      AND l.quantity IS NOT NULL
      AND l.quantity > 0
      AND l.source_type IN ('VCN', 'MBC')
    """

    params = []

    if from_date:
        sql += " AND l.entry_date >= %s"
        params.append(from_date)

    if to_date:
        sql += " AND l.entry_date <= %s"
        params.append(to_date)

    if operation_type:
        sql += " AND l.operation_type = %s"
        params.append(operation_type)

    sql += """
        ORDER BY
            l.entry_date DESC,
            header_name
    """

    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    vessel_meta = {}
    date_vessel_qty = defaultdict(lambda: defaultdict(float))
    all_dates = set()

    for r in rows:

        v_name = re.sub(
            r'\s+', ' ',
            (r['header_name'] or '').strip()
        ).upper()

        dt = str(r['entry_date'])
        qty = _safe_float(r['quantity'])

        all_dates.add(dt)
        date_vessel_qty[dt][v_name] += qty

        if v_name not in vessel_meta:
            vessel_meta[v_name] = {
                'vessel_name':        v_name,
                'cargo_name':         r['cargo_name'] or r['lueu_cargo'] or '-',
                'cargo_type':         r['cargo_type'] or r['lueu_cargo'] or '-',
                'bl_qty':             _safe_float(r['bl_qty']) if r['source_type'] == 'VCN' else 0,
                'company':            r['company'] or 'Mother Vessel',
                'source_type':        r['source_type'],
                'previous_month_qty': 0
            }

    # ─────────────────────────────────────────────
    # Previous Month Continuing Qty
    # ─────────────────────────────────────────────

    from datetime import date as _date

    if from_date:
        try:
            fd               = datetime.strptime(from_date, '%Y-%m-%d').date()
            prev_month_end   = fd - timedelta(days=1)
            prev_month_start = prev_month_end.replace(day=1)
        except Exception:
            prev_month_start = None
            prev_month_end   = None
    else:
        today            = _date.today()
        prev_month_end   = today.replace(day=1) - timedelta(days=1)
        prev_month_start = prev_month_end.replace(day=1)

    if prev_month_start and prev_month_end:

        current_month_vessels = set(vessel_meta.keys())

        conn2 = get_db()
        cur2  = get_cursor(conn2)

        prev_sql = """
    SELECT
        COALESCE(
            CASE
                WHEN l.source_type = 'VCN'
                    THEN CONCAT(v.vcn_doc_num, ' / ', v.vessel_name)

                WHEN l.source_type = 'MBC'
                    THEN CONCAT(m.doc_num, ' / ', m.mbc_name)

                ELSE COALESCE(l.barge_name, 'Unknown')
            END,
            'Unknown'
        ) AS header_name,

        COALESCE(SUM(l.quantity), 0) AS prev_qty

    FROM lueu_lines l

    LEFT JOIN vcn_header v
        ON l.source_type = 'VCN'
       AND l.source_id = v.id

    LEFT JOIN mbc_header m
        ON l.source_type = 'MBC'
       AND l.source_id = m.id

    LEFT JOIN mbc_master mm
        ON UPPER(TRIM(m.mbc_name)) = UPPER(TRIM(mm.mbc_name))

    WHERE l.is_deleted IS NOT TRUE
      AND l.quantity IS NOT NULL
      AND l.quantity > 0
      AND l.source_type IN ('VCN', 'MBC')
      AND l.entry_date >= %s
      AND l.entry_date <= %s

    GROUP BY header_name
        """

        cur2.execute(
            prev_sql,
            [
                prev_month_start.strftime('%Y-%m-%d'),
                prev_month_end.strftime('%Y-%m-%d')
            ]
        )

        prev_rows = cur2.fetchall()
        conn2.close()

        prev_qty_map = {}

        for pr in prev_rows:
            vn = re.sub(
                r'\s+', ' ',
                (pr['header_name'] or '').strip()
            ).upper()

            # ONLY CONTINUING VESSELS
            if vn in current_month_vessels:
                prev_qty_map[vn] = _safe_float(pr['prev_qty'])

        for v_name in vessel_meta:
            vessel_meta[v_name]['previous_month_qty'] = prev_qty_map.get(v_name, 0)

    # ─────────────────────────────────────────────
    # Sorting & Final Output
    # ─────────────────────────────────────────────

    sorted_dates = sorted(all_dates, reverse=False)

    vessel_totals = {}

    for v_name in vessel_meta:
        total = 0
        for dt in sorted_dates:
            total += date_vessel_qty[dt].get(v_name, 0)
        vessel_totals[v_name] = total

    for v_name in vessel_meta:
        if vessel_meta[v_name]['source_type'] == 'MBC':
            current_total = vessel_totals.get(v_name, 0)
            prev_total    = vessel_meta[v_name].get('previous_month_qty', 0)
            vessel_meta[v_name]['bl_qty'] = current_total + prev_total

    vessels_with_data = [
        v for v in vessel_meta.values()
        if vessel_totals.get(v['vessel_name'], 0) > 0
    ]

    def _vessel_sort_key(v):
        name = v['vessel_name'].upper()
        if v['source_type'] == 'VCN':
            return (0, name)
        else:
            m = re.search(r'MBC(\d+)', name)
            num = int(m.group(1)) if m else 999999999
            return (1, num)

    sorted_vessels = sorted(vessels_with_data, key=_vessel_sort_key)

    return {
        'vessels': sorted_vessels,
        'dates':   sorted_dates,
        'data': {
            dt: dict(date_vessel_qty[dt])
            for dt in sorted_dates
        }
    }


# ── Cargo Summary fetch ─────────────────────────────────────────────────────

def _fetch_cargo_summary(from_date=None, to_date=None):

    conn = get_db()
    cur  = get_cursor(conn)

    sql = """
        SELECT
            l.quantity,
            l.source_type,

            CASE
                WHEN l.source_type = 'VCN'
                    THEN COALESCE(
                        vcargo.cargo_type,
                        vc.cargo_name,
                        l.cargo_name
                    )

                WHEN l.source_type = 'MBC'
                    THEN COALESCE(
                        m.cargo_type,
                        m.cargo_name,
                        l.cargo_name
                    )

                ELSE COALESCE(l.cargo_name, 'Unknown')
            END AS cargo_type,

            CASE
                WHEN l.source_type = 'VCN'
                    THEN 'MV'

                WHEN l.source_type = 'MBC'
                     AND UPPER(COALESCE(mm.mbc_owner_name, '')) LIKE '%%SHIPPING%%'
                    THEN 'MBC-Shipping'

                WHEN l.source_type = 'MBC'
                     AND UPPER(COALESCE(mm.mbc_owner_name, '')) LIKE '%%INFRA%%'
                    THEN 'MBC-Infra'

                WHEN l.source_type = 'MBC'
                    THEN 'Other MBC'

                ELSE 'Double Handling'
            END AS op_category

        FROM lueu_lines l

        LEFT JOIN vcn_header v
            ON l.source_type = 'VCN'
           AND l.source_id = v.id

        LEFT JOIN vcn_cargo_declaration vc
            ON v.id = vc.vcn_id
           AND vc.cargo_name = l.cargo_name

        LEFT JOIN LATERAL (
            SELECT cargo_type
            FROM vessel_cargo
            WHERE cargo_name = vc.cargo_name
            LIMIT 1
        ) vcargo ON TRUE

        LEFT JOIN mbc_header m
            ON l.source_type = 'MBC'
           AND l.source_id = m.id

        LEFT JOIN mbc_master mm
            ON UPPER(TRIM(m.mbc_name)) = UPPER(TRIM(mm.mbc_name))

        WHERE l.is_deleted IS NOT TRUE
          AND l.quantity IS NOT NULL
          AND l.quantity > 0
          AND l.source_type IN ('VCN', 'MBC')
    """

    params = []

    if from_date:
        sql += " AND l.entry_date >= %s"
        params.append(from_date)

    if to_date:
        sql += " AND l.entry_date <= %s"
        params.append(to_date)

    sql += " ORDER BY cargo_type, op_category"

    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    cargo_op_qty = defaultdict(lambda: defaultdict(float))
    all_cargos   = set()

    # Fixed ordered list — no 'Total' here
    all_ops = [
        'MV',
        'MBC-Shipping',
        'MBC-Infra',
        'Other MBC',
        'Double Handling'
    ]

    for r in rows:
        cargo = (r['cargo_type'] or 'Unknown').strip()
        op    = (r['op_category'] or 'Other MBC')
        qty   = _safe_float(r['quantity'])
        all_cargos.add(cargo)
        cargo_op_qty[cargo][op] += qty

    result_data = {}

    # ✅ Loop is complete BEFORE return
    for cargo in all_cargos:
        row = dict(cargo_op_qty[cargo])
        row['Total'] = (
            row.get('MV', 0)
            + row.get('MBC-Shipping', 0)
            + row.get('MBC-Infra', 0)
            + row.get('Other MBC', 0)
            + row.get('Double Handling', 0)
        )
        result_data[cargo] = row

    # ✅ 'Total' added ONCE, outside the loop
    all_ops_with_total = all_ops + ['Total']

    return {
        'cargos':     sorted(all_cargos),
        'operations': all_ops_with_total,
        'data':       result_data,
    }


# ── Excel helpers ───────────────────────────────────────────────────────────

def _mbdr(ws, row, c1, c2, fill=XL_WHITE):
    """Apply perimeter thin borders + fill to every cell in a merged range."""
    for ci in range(c1, c2 + 1):
        b = Border(
            left   = _thin if ci == c1 else None,
            right  = _thin if ci == c2 else None,
            top    = _thin,
            bottom = _thin,
        )
        try:
            ws.cell(row, ci).border = b
            ws.cell(row, ci).fill   = _fill(fill)
        except AttributeError:
            pass


# ── Excel sheet writer ──────────────────────────────────────────────────────

def _write_mv_monthly_sheet(ws, report_data):

    vessels = report_data.get('vessels', [])
    dates   = report_data.get('dates', [])
    data    = report_data.get('data', {})

    if not vessels:
        ws['A1'] = "No data available"
        return

    # ----------------------------------------------------
    # COLUMN WIDTHS
    # ----------------------------------------------------

    ws.column_dimensions['A'].width = 18

    total_cols = len(vessels) + 4

    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ----------------------------------------------------
    # HEADER ROWS
    # ----------------------------------------------------

    header_rows = [
        ("Cargo Name", "cargo_name"),
        ("Cargo Type", "cargo_type"),
        ("BL Qty",     "bl_qty"),
        ("Company",    "company"),
    ]

    current_row = 1

    for title, key in header_rows:

        cell = ws.cell(current_row, 1, title)
        cell.font      = _font(bold=True)
        cell.fill      = _fill(XL_GREY)
        cell.alignment = _ctr
        cell.border    = _bdr

        for idx, v in enumerate(vessels, start=2):
            value = v.get(key, '-')
            c = ws.cell(current_row, idx, value)
            c.font      = _font(bold=False)
            c.alignment = _ctr
            c.border    = _bdr

        current_row += 1

    # ----------------------------------------------------
    # VESSEL HEADER ROW
    # ----------------------------------------------------

    c = ws.cell(current_row, 1, "Date")
    c.font      = _font(bold=True)
    c.fill      = _fill(XL_GREY)
    c.alignment = _ctr
    c.border    = _bdr

    for idx, v in enumerate(vessels, start=2):
        cell = ws.cell(current_row, idx, v['vessel_name'])
        cell.fill      = _fill(XL_YELLOW)
        cell.font      = _font(bold=True, color='C00000')
        cell.alignment = _ctr
        cell.border    = _bdr

    mv_col    = len(vessels) + 2
    mbc_col   = len(vessels) + 3
    grand_col = len(vessels) + 4

    for col, text, color in [
        (mv_col,    "MV Total",    "1E40AF"),
        (mbc_col,   "MBC Total",   "166534"),
        (grand_col, "Grand Total", "B45309"),
    ]:
        cell = ws.cell(current_row, col, text)
        cell.fill      = _fill(color)
        cell.font      = _font(bold=True, color='FFFFFF')
        cell.alignment = _ctr
        cell.border    = _bdr

    current_row += 1

    # ----------------------------------------------------
    # DATE ROWS
    # ----------------------------------------------------

    for dt in dates:

        row_total_mv  = 0
        row_total_mbc = 0
        row_total_all = 0

        # ✅ Convert string to date object so Excel formats it correctly
        try:
            dt_obj = datetime.strptime(dt, '%Y-%m-%d').date()
        except Exception:
            dt_obj = dt   # fallback to string if parse fails

        dcell = ws.cell(current_row, 1, dt_obj)   # ✅ pass date object, not string
        dcell.number_format = 'DD-MM-YYYY'         # ✅ formats as 01-05-2026
        dcell.font          = _font(bold=True)
        dcell.alignment     = _ctr
        dcell.border        = _bdr

        for idx, v in enumerate(vessels, start=2):
            qty  = data.get(dt, {}).get(v['vessel_name'], 0)
            cell = ws.cell(current_row, idx, qty if qty > 0 else "")
            cell.number_format = '#,##0.00'
            cell.alignment     = _ctr
            cell.border        = _bdr

            if v['source_type'] == 'VCN':
                row_total_mv += qty
            else:
                row_total_mbc += qty
            row_total_all += qty

        for col, val, clr in [
            (mv_col,    row_total_mv,  'DBEAFE'),
            (mbc_col,   row_total_mbc, 'DCFCE7'),
            (grand_col, row_total_all, 'FEF9C3'),
        ]:
            c = ws.cell(current_row, col, val)
            c.fill         = _fill(clr)
            c.font         = _font(bold=True)
            c.number_format = '#,##0.00'
            c.alignment    = _ctr
            c.border       = _bdr

        current_row += 1

    # ----------------------------------------------------
    # PREVIOUS MONTH QTY ROW
    # ----------------------------------------------------

    prev_row = current_row

    c = ws.cell(prev_row, 1, "Previous Month Qty")
    c.fill      = _fill('D9EAD3')
    c.font      = _font(bold=True)
    c.alignment = _ctr
    c.border    = _bdr

    mv_prev = mbc_prev = grand_prev = 0

    for idx, v in enumerate(vessels, start=2):
        qty  = v.get('previous_month_qty', 0)
        cell = ws.cell(prev_row, idx, qty)
        cell.fill          = _fill('D9EAD3')
        cell.font          = _font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment     = _ctr
        cell.border        = _bdr

        if v['source_type'] == 'VCN':
            mv_prev += qty
        else:
            mbc_prev += qty
        grand_prev += qty

    for col, val in [
        (mv_col,    mv_prev),
        (mbc_col,   mbc_prev),
        (grand_col, grand_prev),
    ]:
        cc = ws.cell(prev_row, col, val)
        cc.fill          = _fill('B6D7A8')
        cc.font          = _font(bold=True)
        cc.number_format = '#,##0.00'
        cc.alignment     = _ctr
        cc.border        = _bdr

    current_row += 1

    # ----------------------------------------------------
    # TOTAL QTY ROW
    # ----------------------------------------------------

    total_row = current_row

    c = ws.cell(total_row, 1, "Total Qty")
    c.fill      = _fill(XL_YELLOW)
    c.font      = _font(bold=True)
    c.alignment = _ctr
    c.border    = _bdr

    mv_total = mbc_total = all_total = 0
    vessel_totals = {}

    for idx, v in enumerate(vessels, start=2):
        total = sum(data.get(dt, {}).get(v['vessel_name'], 0) for dt in dates)
        vessel_totals[v['vessel_name']] = total

        cell = ws.cell(total_row, idx, total)
        cell.fill          = _fill(XL_YELLOW)
        cell.font          = _font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment     = _ctr
        cell.border        = _bdr

        if v['source_type'] == 'VCN':
            mv_total += total
        else:
            mbc_total += total
        all_total += total

    for col, val, clr in [
        (mv_col,    mv_total,  '1E40AF'),
        (mbc_col,   mbc_total, '166534'),
        (grand_col, all_total, 'B45309'),
    ]:
        cc = ws.cell(total_row, col, val)
        cc.fill          = _fill(clr)
        cc.font          = _font(bold=True, color='FFFFFF')
        cc.number_format = '#,##0.00'
        cc.alignment     = _ctr
        cc.border        = _bdr

    current_row += 1

    # ----------------------------------------------------
    # BL QTY ROW
    # ----------------------------------------------------

    bl_row = current_row

    c = ws.cell(bl_row, 1, "BL Qty")
    c.font      = _font(bold=True)
    c.alignment = _ctr
    c.border    = _bdr

    mv_bl = mbc_bl = all_bl = 0

    for idx, v in enumerate(vessels, start=2):
        qty  = v.get('bl_qty', 0)
        cell = ws.cell(bl_row, idx, qty)
        cell.number_format = '#,##0.00'
        cell.alignment     = _ctr
        cell.border        = _bdr

        if v['source_type'] == 'VCN':
            mv_bl += qty
        else:
            mbc_bl += qty
        all_bl += qty

    for col, val, clr in [
        (mv_col,    mv_bl,  'DBEAFE'),
        (mbc_col,   mbc_bl, 'DCFCE7'),
        (grand_col, all_bl, 'FEF9C3'),
    ]:
        cc = ws.cell(bl_row, col, val)
        cc.fill          = _fill(clr)
        cc.font          = _font(bold=True)
        cc.number_format = '#,##0.00'
        cc.alignment     = _ctr
        cc.border        = _bdr

    current_row += 1

    # ----------------------------------------------------
    # DIFFERENCE ROW
    # ----------------------------------------------------

    diff_row = current_row

    c = ws.cell(diff_row, 1, "Difference")
    c.font      = _font(bold=True)
    c.alignment = _ctr
    c.border    = _bdr

    for idx, v in enumerate(vessels, start=2):
        diff = (
            v.get('bl_qty', 0)
            - v.get('previous_month_qty', 0)
            - vessel_totals[v['vessel_name']]
        )
        cell = ws.cell(diff_row, idx, diff)
        cell.number_format = '#,##0.00'
        cell.alignment     = _ctr
        cell.border        = _bdr
        if diff < 0:
            cell.font = _font(color='FF0000')

    # ----------------------------------------------------
    # FREEZE PANES
    # ----------------------------------------------------

    ws.freeze_panes = 'B6'


# ── Cargo Summary Excel sheet ───────────────────────────────────────────────

def _write_cargo_summary_sheet(ws, report_data):
    """Write Cargo-wise summary sheet — ONE Total column only."""

    cargos     = report_data.get('cargos', [])
    operations = report_data.get('operations', [])
    data       = report_data.get('data', {})

    if not cargos:
        ws['A1'] = "No data available"
        return

    # ✅ Strip 'Total' from operations — we write it manually as the last column
    ops = [o for o in operations if o != 'Total']

    # Columns: Cargo | op1 | op2 | ... | Total
    NC = len(ops) + 2   # +1 for Cargo label col, +1 for Total col

    # Column widths
    ws.column_dimensions['A'].width = 22
    for i in range(2, NC + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ----------------------------------------------------
    # Title row
    # ----------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    c = ws['A1']
    c.value     = 'Cargo Summary'
    c.font      = _font(bold=True, size=14)
    c.alignment = _ctr
    c.fill      = _fill(XL_WHITE)
    c.border    = _bdr

    # ----------------------------------------------------
    # Header row  (yellow)
    # ----------------------------------------------------
    ws.row_dimensions[2].height = 24
    headers = ['Cargo'] + ops + ['Total']   # ✅ exactly ONE Total

    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.fill      = _fill(XL_YELLOW)
        cell.font      = _font(bold=True, size=11)
        cell.alignment = _ctr
        cell.border    = _bdr

    # ----------------------------------------------------
    # Data rows
    # ----------------------------------------------------
    row_idx = 3

    for cargo in cargos:
        ws.row_dimensions[row_idx].height = 20

        # Cargo label
        cell = ws.cell(row=row_idx, column=1, value=cargo)
        cell.font      = _font(bold=True, size=10)
        cell.alignment = _left
        cell.border    = _bdr

        # Operation columns
        row_total = 0
        for c_idx, op in enumerate(ops, start=2):
            qty = data.get(cargo, {}).get(op, 0)
            row_total += qty
            cell = ws.cell(row=row_idx, column=c_idx, value=qty if qty > 0 else 0)
            cell.alignment     = _ctr
            cell.border        = _bdr
            cell.number_format = '#,##0.00'
            cell.font          = _font(size=10)

        # ✅ Single Total column at NC
        cell = ws.cell(row=row_idx, column=NC, value=row_total)
        cell.alignment     = _ctr
        cell.border        = _bdr
        cell.number_format = '#,##0.00'
        cell.font          = _font(bold=True, size=10)

        row_idx += 1

    # ----------------------------------------------------
    # Grand Total row  (yellow)
    # ----------------------------------------------------
    ws.row_dimensions[row_idx].height = 24

    cell = ws.cell(row=row_idx, column=1, value='Grand Total')
    cell.fill      = _fill(XL_YELLOW)
    cell.font      = _font(bold=True, size=10)
    cell.alignment = _left
    cell.border    = _bdr

    grand_total = 0
    for c_idx, op in enumerate(ops, start=2):
        col_total = sum(data.get(cargo, {}).get(op, 0) for cargo in cargos)
        grand_total += col_total
        cell = ws.cell(row=row_idx, column=c_idx, value=col_total)
        cell.fill          = _fill(XL_YELLOW)
        cell.font          = _font(bold=True, size=10)
        cell.alignment     = _ctr
        cell.border        = _bdr
        cell.number_format = '#,##0.00'

    # ✅ Grand total at NC
    cell = ws.cell(row=row_idx, column=NC, value=grand_total)
    cell.fill          = _fill(XL_YELLOW)
    cell.font          = _font(bold=True, size=10)
    cell.alignment     = _ctr
    cell.border        = _bdr
    cell.number_format = '#,##0.00'

    # Freeze: keep Cargo col + title/header rows fixed
    ws.freeze_panes = 'B3'


# ── Excel builder ───────────────────────────────────────────────────────────

def _build_mv_monthly_excel(report_data, cargo_data=None):
    from openpyxl import Workbook
    wb = Workbook()

    # Sheet 1: Vessel Matrix
    ws1 = wb.active
    ws1.title = 'Vessel Report'
    _write_mv_monthly_sheet(ws1, report_data)

    # Sheet 2: Cargo Summary
    if cargo_data:
        ws2 = wb.create_sheet('Cargo Summary')
        _write_cargo_summary_sheet(ws2, cargo_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ───────────────────────────────────────────────────────────────────

@bp.route('/module/RP01/mv-monthly-report/')
@login_required
def mv_monthly_report_index():
    return render_template(
        'mv_mbc_month_report/mv_mbc_report.html',
        username=session.get('username')
    )


@bp.route('/api/module/RP01/mv-monthly-report/data')
@login_required
def mv_monthly_report_data():
    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')
    op_type   = request.args.get('operation_type')
    report    = _fetch_mv_monthly_data(from_date, to_date, op_type)
    return jsonify(report)


@bp.route('/api/module/RP01/mv-monthly-report/download')
@login_required
def mv_monthly_report_download():
    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')
    op_type   = request.args.get('operation_type')

    vessel_report = _fetch_mv_monthly_data(from_date, to_date, op_type)
    cargo_report  = _fetch_cargo_summary(from_date, to_date)

    buf   = _build_mv_monthly_excel(vessel_report, cargo_report)
    fname = f'MV_Monthly_Loading_Report_{date.today().strftime("%Y%m%d")}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@bp.route('/api/module/RP01/mv-monthly-report/cargo-summary')
@login_required
def mv_monthly_cargo_summary():
    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')
    report    = _fetch_cargo_summary(from_date, to_date)
    return jsonify(report)