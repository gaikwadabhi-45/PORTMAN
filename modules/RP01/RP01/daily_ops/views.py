from flask import render_template, request, session, redirect, url_for, Response, jsonify
from functools import wraps
from datetime import date, datetime, timedelta
import io
import json
from copy import copy
from openpyxl.styles import Alignment
from datetime import timedelta

from .. import bp
from database import get_db, get_cursor
from .model import build_fy_throughput

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants (same as rest of RP01) ──────────────────────────────────
XL_NORM_SZ = 11
_thin  = Side(style='thin',   color='000000')
_bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr   = Alignment(horizontal='center', vertical='center', wrap_text=False)
_left  = Alignment(horizontal='left',   vertical='center', wrap_text=False)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


# def _font(bold=False, size=XL_NORM_SZ):
#     return Font(name='Calibri', bold=bold, size=size)
def _font(bold=False):
    return Font(
        name="Calibri",
        size=21,
        bold=bold
    )

def _parse_dt(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val))
    except Exception:
        return None


def _fmt_dt(val, strfmt='%d-%m-%Y %H:%M'):
    dt = _parse_dt(val)
    return dt.strftime(strfmt) if dt else ''


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


_MBC_OWNERS      = ['JSW INFRA', 'JSW SHIPPING', 'OTHERS']
_MBC_CARGO_TYPES = ['Break Bulk', 'Container', 'Liquid', 'Bulk']


# ── Routes ──────────────────────────────────────────────────────────────────

@bp.route('/module/RP01/daily-ops/')
@login_required
def daily_ops_index():
    return render_template(
        'daily_ops/daily_ops.html',
        username=session.get('username'),
        is_admin=session.get('is_admin'),
    )


# ── Routes API (from CRM01 conveyor_routes) ─────────────────────────────────

@bp.route('/api/module/RP01/daily-ops/routes', methods=['GET'])
@login_required
def daily_ops_routes():
    """Return all active route names from conveyor_routes."""
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("SELECT route_name FROM conveyor_routes WHERE is_active = 1 ORDER BY route_name")
    routes = [r['route_name'] for r in cur.fetchall()]
    conn.close()
    return jsonify(routes)

@bp.route("/daily-ops/save-cargo-table", methods=["POST"])
def save_cargo_table():

    data = request.get_json()

    report_date = data["report_date"]
    table_data = data["table_data"]

    # save to database
    save_to_db(report_date, json.dumps(table_data))

    return jsonify({"success": True})


# ── Cutoff API ──────────────────────────────────────────────────────────────

@bp.route('/api/module/RP01/daily-ops/cutoff', methods=['GET'])
@login_required
def daily_ops_cutoff_get():
    """Return the latest cutoff record (or empty defaults)."""
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("""
        SELECT id, cutoff_date, cutoff_values
        FROM daily_ops_cutoff
        ORDER BY id DESC LIMIT 1
    """)
    row = cur.fetchone()
    conn.close()
    if row:
        return jsonify({
            'id':            row['id'],
            'cutoff_date':   row['cutoff_date'],
            'cutoff_values': json.loads(row['cutoff_values']),
        })
    return jsonify({
        'id':            None,
        'cutoff_date':   '',
        'cutoff_values': {'fy_throughput': {}},
    })


@bp.route('/api/module/RP01/daily-ops/cutoff', methods=['POST'])
@login_required
def daily_ops_cutoff_save():
    """Admin-only: set the cutoff date and store the computed FY snapshot."""
    if not session.get('is_admin'):
        return Response('Admin access required', status=403)

    data               = request.get_json(force=True)
    cutoff_date        = data.get('cutoff_date', '')
    editable_fy_values = data.get('editable_fy_values') or {}

    if not cutoff_date:
        return Response('cutoff_date is required', status=400)

    fy_throughput = _compute_fy_throughput(cutoff_date, editable_fy_values)
    values_json   = json.dumps({'fy_throughput': fy_throughput})
    user          = session.get('username', '')

    conn = get_db()
    cur  = get_cursor(conn)
    # Single-row table: clear then insert.
    cur.execute("DELETE FROM daily_ops_cutoff")
    cur.execute("""
        INSERT INTO daily_ops_cutoff (cutoff_date, cutoff_values, created_by)
        VALUES (%s, %s, %s)
    """, (cutoff_date, values_json, user))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'fy_throughput': fy_throughput})


@bp.route('/api/module/RP01/daily-ops/cutoff', methods=['DELETE'])
@login_required
def daily_ops_cutoff_clear():
    """Admin-only: remove any stored cutoff so FY values revert to live queries."""
    if not session.get('is_admin'):
        return Response('Admin access required', status=403)

    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("DELETE FROM daily_ops_cutoff")
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ── FY throughput snapshot ───────────────────────────────────────────────────

def _compute_fy_throughput(cutoff_date, editable_fy_values=None):
    """Aggregate quantity by (financial year, cargo type) up to cutoff_date.

    Unions historical (rp01_historical_lueu) and live (lueu_lines) rows, maps
    cargo_name -> cargo_type via the VCG01 vessel_cargo master, buckets by
    April-start financial year, and returns {fy_label: {cargo_type: qty}}.
    The cutoff FY is naturally partial (entry_date <= cutoff_date).
    
    Ensures all fiscal years from 2012 to the current fiscal year are displayed,
    with zero values for years that have no data.
    """
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("""
        WITH throughput AS (
            SELECT
                COALESCE(vc.cargo_type, 'OTHERS') AS cargo_type,
                (EXTRACT(YEAR FROM TO_DATE(l.entry_date, 'YYYY-MM-DD'))::int
                    - CASE WHEN EXTRACT(MONTH FROM TO_DATE(l.entry_date, 'YYYY-MM-DD')) < 4
                           THEN 1 ELSE 0 END) AS fy_start,
                COALESCE(l.quantity, 0) AS quantity
            FROM lueu_lines l
            LEFT JOIN vessel_cargo vc
                ON UPPER(TRIM(vc.cargo_name)) = UPPER(TRIM(l.cargo_name))
            WHERE l.is_deleted = false
              AND l.cargo_name IS NOT NULL
              AND TO_DATE(l.entry_date, 'YYYY-MM-DD') <= %s::date

            UNION ALL

            SELECT
                COALESCE(vc.cargo_type, 'OTHERS') AS cargo_type,
                (EXTRACT(YEAR FROM h.entry_date)::int
                    - CASE WHEN EXTRACT(MONTH FROM h.entry_date) < 4
                           THEN 1 ELSE 0 END) AS fy_start,
                COALESCE(h.quantity, 0) AS quantity
            FROM rp01_historical_lueu h
            LEFT JOIN vessel_cargo vc
                ON UPPER(TRIM(vc.cargo_name)) = UPPER(TRIM(h.cargo_name))
            WHERE h.entry_date <= %s::date
        )
        SELECT
            fy_start,
            cargo_type,
            SUM(quantity) AS qty
        FROM throughput
        GROUP BY fy_start, cargo_type
        ORDER BY fy_start, cargo_type
    """, (cutoff_date, cutoff_date))

    rows = cur.fetchall()
    conn.close()
    fy_data = build_fy_throughput(rows)
    
    # Ensure all years from 2012 through current FY are present
    current_year = date.today().year
    current_month = date.today().month
    current_fy_start = current_year if current_month >= 4 else current_year - 1

    # Get all cargo types from the computed data and any editable values
    all_cargo_types = set()
    for fy_dict in fy_data.values():
        all_cargo_types.update(fy_dict.keys())
    for fy_dict in (editable_fy_values or {}).values():
        if isinstance(fy_dict, dict):
            all_cargo_types.update(fy_dict.keys())

    for fy_start in range(2012, current_fy_start + 1):
        fy_label = f"{fy_start}-{fy_start + 1}"
        if fy_label not in fy_data:
            fy_data[fy_label] = {cargo_type: 0.0 for cargo_type in all_cargo_types}
        else:
            for cargo_type in all_cargo_types:
                fy_data[fy_label].setdefault(cargo_type, 0.0)

    if editable_fy_values:
        for fy_label, cargo_map in editable_fy_values.items():
            if not isinstance(cargo_map, dict):
                continue
            row = fy_data.setdefault(fy_label, {})
            for cargo_type, qty in cargo_map.items():
                try:
                    cleaned_qty = float(qty)
                except (TypeError, ValueError):
                    continue
                row[cargo_type] = cleaned_qty

        all_cargo_types = set()
        for fy_dict in fy_data.values():
            all_cargo_types.update(fy_dict.keys())
        for fy_dict in fy_data.values():
            for cargo_type in all_cargo_types:
                fy_dict.setdefault(cargo_type, 0.0)

    return fy_data


# ── Data fetchers ───────────────────────────────────────────────────────────

from datetime import datetime, timedelta

def _fetch_data(report_date):
    window_end = datetime(
        report_date.year,
        report_date.month,
        report_date.day,
        8, 0, 0
    )

    # Previous day 08:00 AM
    window_start = window_end - timedelta(hours=24)

    ws_str = window_start.strftime('%Y-%m-%d %H:%M:%S')
    we_str = window_end.strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cur = get_cursor(conn)

    # Fetch vessels that were active during the reporting window
    cur.execute("""
    SELECT DISTINCT

        h.id,
        h.vcn_id,
        h.vessel_name,
        h.operation_type,
        h.nor_tendered,

        first_anchor.discharge_started AS discharge_commenced,

        last_anchor.discharge_completed AS discharge_completed,

        CASE
            WHEN last_anchor.discharge_completed IS NULL THEN 1
            ELSE 0
        END AS sort_order,

        h.doc_status

    FROM ldud_header h

    LEFT JOIN LATERAL (
        SELECT
            MIN(a1.discharge_started) AS discharge_started
        FROM ldud_anchorage a1
        WHERE a1.ldud_id = h.id
        AND a1.discharge_started IS NOT NULL
    ) first_anchor ON TRUE

    LEFT JOIN LATERAL (
        SELECT
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM ldud_anchorage x
                    WHERE x.ldud_id = h.id
                    AND x.discharge_started IS NOT NULL
                    AND x.discharge_commenced IS NULL
                )
                THEN NULL
                ELSE MAX(a2.discharge_commenced)
            END AS discharge_completed
        FROM ldud_anchorage a2
        WHERE a2.ldud_id = h.id
    ) last_anchor ON TRUE

    WHERE

        first_anchor.discharge_started IS NOT NULL

        -- Vessel started before report end
        AND first_anchor.discharge_started < %s

        -- Vessel active OR barges still active
        AND (
            last_anchor.discharge_completed IS NULL
            OR last_anchor.discharge_completed >= %s

            OR EXISTS (
            SELECT 1
            FROM ldud_barge_lines b
            WHERE b.ldud_id = h.id
            AND (
                    b.completed_discharge_berth IS NULL
                    OR b.cast_off_berth IS NULL
                )
        )
        )

    ORDER BY
        sort_order,
        first_anchor.discharge_started,
        h.id

    """,  (
    window_end,
    window_start
))

    vessels  = [dict(r) for r in cur.fetchall()]
    ldud_to_vcn = {
    v['id']: v['vcn_id']
    for v in vessels
    if v.get('vcn_id')
    }
    ldud_ids = [v['id'] for v in vessels]
    vcn_ids  = [v['vcn_id'] for v in vessels if v.get('vcn_id')]

    bl_import, bl_export, vcn_meta, cargo_map = {}, {}, {}, {}

    if vcn_ids:

        # Import BL Qty
        cur.execute("""
            SELECT
                vcn_id,
                COALESCE(SUM(bl_quantity), 0) AS total
            FROM vcn_cargo_declaration
            WHERE vcn_id = ANY(%s)
            GROUP BY vcn_id
        """, (vcn_ids,))

        for r in cur.fetchall():
            bl_import[r['vcn_id']] = float(r['total'])

        # Export BL Qty
        cur.execute("""
            SELECT
                vcn_id,
                COALESCE(SUM(bl_quantity), 0) AS total
            FROM vcn_export_cargo_declaration
            WHERE vcn_id = ANY(%s)
            GROUP BY vcn_id
        """, (vcn_ids,))

        for r in cur.fetchall():
            bl_export[r['vcn_id']] = float(r['total'])

        # Stevedore / Customer
        cur.execute("""
            SELECT
                id,
                importer_exporter_name
            FROM vcn_header
            WHERE id = ANY(%s)
        """, (vcn_ids,))

        vcn_meta = {
            r['id']: r['importer_exporter_name'] or ''
            for r in cur.fetchall()
        }

        # Cargo Names
        cur.execute("""
            SELECT
                vcn_id,
                STRING_AGG(
                    DISTINCT cargo_name,
                    ', '
                    ORDER BY cargo_name
                ) AS cargo_names
            FROM vcn_cargo_declaration
            WHERE vcn_id = ANY(%s)
            GROUP BY vcn_id
        """, (vcn_ids,))

        for r in cur.fetchall():
            cargo_map[r['vcn_id']] = r['cargo_names'] or ''

    # ops_till — total discharged from lueu_lines up to selected date
    lueu_total = {}
    if vcn_ids:
        cur.execute("""
            SELECT source_id, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE source_type = 'VCN'
              AND source_id = ANY(%s)
              AND entry_date::date < %s::date
            GROUP BY source_id
        """, (vcn_ids, we_str))
        for r in cur.fetchall():
            lueu_total[r['source_id']] = float(r['qty'])

    # ops_24h — discharged in the 24h window
   # 24 Hr Discharge (Till Previous Day)

    ops_24h = {}

    prev_date = report_date - timedelta(days=1)

    if ldud_ids:
        cur.execute("""
            SELECT
                ldud_id,
                COALESCE(SUM(quantity), 0) AS qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
            AND TO_DATE(start_time, 'YYYY-MM-DD') = %s
            GROUP BY ldud_id
        """, (ldud_ids, prev_date))

        for r in cur.fetchall():
            ops_24h[r['ldud_id']] = float(r['qty'])


    #LOADED TILL DATE (TILL PREVIOUS DAY)

    ops_till = {}

    cutoff_date = report_date - timedelta(days=1)

    if ldud_ids:
        cur.execute("""
            SELECT
                ldud_id,
                COALESCE(SUM(quantity), 0) AS qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
            AND TO_DATE(start_time, 'YYYY-MM-DD') <= %s
            GROUP BY ldud_id
        """, (ldud_ids, cutoff_date))

        for r in cur.fetchall():
            ops_till[r['ldud_id']] = float(r['qty'])

    # Barges
    
    # Barges

    # Fetch actual discharged quantity for barges from lueu_lines

    # Actual discharged quantity from LUEU by VCN + Base Barge

    barge_actual = {}

    if vcn_ids:

        cur.execute("""
            SELECT
                source_id,
                UPPER(TRIM(SPLIT_PART(barge_name,'/',1))) AS base_barge,
                TRIM(SPLIT_PART(barge_name,'/',2)) AS trip_no,

                MAX(route_name) AS route_name,
                MAX(equipment_name) AS crane,

                SUM(COALESCE(quantity,0)) AS actual_qty

            FROM lueu_lines

            WHERE source_type = 'VCN'
            AND is_deleted = false
            AND source_id = ANY(%s)
            AND barge_name IS NOT NULL
            AND quantity IS NOT NULL
            AND TO_DATE(entry_date,'YYYY-MM-DD') <= %s

            GROUP BY
                source_id,
                UPPER(TRIM(SPLIT_PART(barge_name,'/',1))),
                TRIM(SPLIT_PART(barge_name,'/',2))
        """, (
            vcn_ids,
            report_date - timedelta(days=1)
        ))

        for r in cur.fetchall():

            barge_actual[
                (
                    r['source_id'],
                    r['base_barge'],
                    r['trip_no']
                )
            ] = {
                'actual_qty': float(r['actual_qty']),
                'route_name': r['route_name'] or '',
                'crane': r['crane'] or ''
            }


    barge_stats = {}

    _STATUS_KEYS = (
        'at_jetty', 'waiting_discharge', 'waiting_empty_jetty',
        'at_gull_loaded', 'under_loading', 'waiting_loading',
        'in_transit_jetty_to_mv', 'Non-Operational',
    )

    if ldud_ids:
        cur.execute("""
    SELECT
        h.vcn_id,

        b.ldud_id,
        b.barge_name,
        b.discharge_quantity,
        b.port_crane,

        b.along_side_vessel,
        b.commenced_loading,
        b.completed_loading,
        b.cast_off_mv,

        b.anchored_gull_island,
        b.aweigh_gull_island,
        b.amf_at_port,

        b.along_side_berth,
        b.commence_discharge_berth,
        b.completed_discharge_berth,
        b.cast_off_berth,
        b.cast_off_port,

        ROW_NUMBER() OVER (
            PARTITION BY
                h.vcn_id,
                UPPER(TRIM(b.barge_name))
            ORDER BY
                b.commence_discharge_berth
        ) AS trip_no

        FROM ldud_barge_lines b

        JOIN ldud_header h
            ON h.id = b.ldud_id

        WHERE b.ldud_id = ANY(%s)
        AND (b.cast_off_port IS NULL OR b.cast_off_port > %s)

        ORDER BY
            h.vcn_id,
            b.barge_name,
            trip_no
    """, (ldud_ids, ws_str))

        for r in cur.fetchall():

            lid = r['ldud_id']

            bn = (r['barge_name'] or '').strip()
            bn_key = bn.upper()

            bl_qty = float(r['discharge_quantity'] or 0)

            vcn_id = ldud_to_vcn.get(lid)

            trip_no = str(r['trip_no'])

            barge_info = barge_actual.get(
                (
                    vcn_id,
                    bn_key,
                    trip_no
                ),
                {}
            )

            actual_qty = float(barge_info.get('actual_qty', 0))
            route_name = barge_info.get('route_name', '')
            crane = barge_info.get('crane', '') or (r['port_crane'] or '').strip()

            balance_qty = max(0, bl_qty - actual_qty)

            if lid not in barge_stats:
                barge_stats[lid] = {
                    'all': set(),
                    **{k: [] for k in _STATUS_KEYS}
                }

            if bn:
                barge_stats[lid]['all'].add(bn)

            if r['cast_off_port']:
                status = 'Non-Operational'
            elif r['completed_discharge_berth'] and not r['cast_off_berth']:
                status = 'waiting_empty_jetty'
            elif r['commence_discharge_berth'] and not r['cast_off_berth']:
                status = 'at_jetty'
            elif r['along_side_berth'] and not r['commence_discharge_berth']:
                status = 'waiting_discharge'
            elif r['cast_off_mv'] and not r['along_side_berth']:
                status = 'at_gull_loaded'
            elif r['commenced_loading'] and not r['completed_loading']:
                status = 'under_loading'
            elif r['along_side_vessel'] and not r['commenced_loading']:
                status = 'waiting_loading'
            else:
                status = None

            if status and bn:

                if status == 'at_jetty':
                    entry = (
                        f"{bn} / {trip_no}"
                        f"{' - ' + route_name if route_name else ''}"
                        f"{' - ' + crane if crane else ''} "
                        f"(Bal:{int(round(balance_qty))} MT)"
                    )
                elif status == 'waiting_discharge' and bl_qty:
                    entry = f"{bn} ({int(round(bl_qty))} MT)"

                else:
                    entry = bn

                barge_stats[lid][status].append(entry)
    conn.close()

    def _make_names(bs_dict, key):
        return '\n'.join(bs_dict.get(key, []))

    for v in vessels:
        lid    = v['id']
        vid    = v.get('vcn_id')
        op     = v.get('operation_type', '')

        bl_qty = (
            bl_export.get(vid, 0)
            if op == 'Export'
            else bl_import.get(vid, 0)
        ) if vid else 0

        actual = lueu_total.get(vid, 0)
        bs     = barge_stats.get(lid, {})

        v['stevedore_group'] = vcn_meta.get(vid, '') if vid else ''
        v['cargo_name']      = cargo_map.get(vid, '') if vid else ''

        print(
            "VESSEL:",
            v['vessel_name'],
            "VCN:",
            vid,
            "CARGO:",
            v['cargo_name']
        )

        v['bl_qty'] = bl_qty
        v['ops_24h'] = ops_24h.get(lid, 0)
        v['ops_till'] = ops_till.get(lid, 0)
        v['balance'] = round(
            bl_qty - ops_till.get(lid, 0),
            2
        )

        active_statuses = (
            'at_jetty',
            'at_gull_loaded',
            'waiting_discharge',
            'under_loading'
        )

        v['num_barges'] = len({
            name.split('/')[0].strip()
            for key in active_statuses
            for name in bs.get(key, [])
        }) or ''

        v['at_jetty'] = _make_names(bs, 'at_jetty')
        v['waiting_discharge'] = _make_names(bs, 'waiting_discharge')
        v['waiting_empty_jetty'] = _make_names(bs, 'waiting_empty_jetty')
        v['at_gull_loaded'] = _make_names(bs, 'at_gull_loaded')
        v['under_loading'] = _make_names(bs, 'under_loading')
        v['waiting_loading'] = _make_names(bs, 'waiting_loading')
        v['in_transit_jetty_to_mv'] = _make_names(bs, 'in_transit_jetty_to_mv')
        v['breakdown'] = _make_names(bs, 'breakdown')

    return vessels

def _fetch_upcoming_vessels(report_date):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
    SELECT
        vh.vessel_name,

        vc.cargo_name,

        COALESCE(vc.bl_quantity, 0) AS bl_quantity,

        vh.vessel_agent_name,

        CASE
            WHEN lh.nor_tendered IS NULL
                THEN 'ETA : ' ||
                     TO_CHAR(vn.eta::timestamp, 'DD-MM-YYYY HH24:MI')

            WHEN lh.nor_tendered IS NOT NULL
                 AND fa.discharge_started IS NULL
                THEN 'ARRIVED AT : ' ||
                     TO_CHAR(lh.nor_tendered::timestamp, 'DD-MM-YYYY HH24:MI')
        END AS eta,

        CASE
            WHEN lh.nor_tendered IS NULL
                THEN 'ETA'

            WHEN lh.nor_tendered IS NOT NULL
                 AND fa.discharge_started IS NULL
                THEN 'ARRIVED'
        END AS vessel_status,

        COALESCE(
            lh.nor_tendered::timestamp,
            vn.eta::timestamp
        ) AS status_time

    FROM vcn_header vh

    JOIN vcn_nominations vn
        ON vn.vcn_id = vh.id

    LEFT JOIN ldud_header lh
        ON lh.vcn_id = vh.id

    LEFT JOIN vcn_cargo_declaration vc
        ON vc.vcn_id = vh.id

    LEFT JOIN LATERAL (
        SELECT MIN(a.discharge_started) AS discharge_started
        FROM ldud_anchorage a
        WHERE a.ldud_id = lh.id
    ) fa ON TRUE

    WHERE
        fa.discharge_started IS NULL

    ORDER BY status_time
    """)

    rows = cur.fetchall()

    conn.close()

    return rows


def _fetch_discharging_mbcs(report_date):

    day_start = datetime(
        report_date.year,
        report_date.month,
        report_date.day,
        0, 0, 0
    ) - timedelta(days=1)

    day_end = datetime(
        report_date.year,
        report_date.month,
        report_date.day,
        23, 59, 59
    )

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
SELECT
    h.id,
    h.mbc_name,
    p.vessel_unloaded_by AS equipment,
    h.cargo_name,

    h.bl_quantity AS bl_quantity,

    COALESCE(l.actual_qty, 0) AS actual_quantity,

    (h.bl_quantity - COALESCE(l.actual_qty, 0)) AS discharge_quantity,

    p.vessel_arrival_port,
    p.unloading_commenced,
    p.unloading_completed,

    CASE
        WHEN
            NULLIF(TRIM(p.vessel_arrival_port), '') IS NOT NULL
            AND NULLIF(TRIM(p.vessel_arrival_port), '')::timestamp >= %s
            AND NULLIF(TRIM(p.vessel_arrival_port), '')::timestamp <= %s
            AND (
                p.unloading_commenced IS NULL
                OR TRIM(COALESCE(p.unloading_commenced, '')) = ''
            )
        THEN 'ARRIVED'

        WHEN
            p.unloading_commenced IS NOT NULL
            AND TRIM(COALESCE(p.unloading_commenced, '')) <> ''
            AND (
                p.unloading_completed IS NULL
                OR TRIM(COALESCE(p.unloading_completed, '')) = ''
            )
            AND NULLIF(TRIM(p.unloading_commenced), '')::timestamp >= %s
            AND NULLIF(TRIM(p.unloading_commenced), '')::timestamp <= %s
        THEN 'DISCHARGING'

        WHEN
            p.unloading_completed IS NOT NULL
            AND TRIM(COALESCE(p.unloading_completed, '')) <> ''
            AND NULLIF(TRIM(p.unloading_completed), '')::timestamp >= %s
            AND NULLIF(TRIM(p.unloading_completed), '')::timestamp <= %s
        THEN 'COMPLETED'
    END AS status

FROM mbc_header h

JOIN mbc_discharge_port_lines p
    ON p.mbc_id = h.id

LEFT JOIN (
    SELECT
        source_id AS mbc_id,
        SUM(COALESCE(quantity, 0)) AS actual_qty
    FROM lueu_lines
    WHERE source_type = 'MBC'
      AND is_deleted = false
    GROUP BY source_id
) l
    ON l.mbc_id = h.id

WHERE

(
    NULLIF(TRIM(p.vessel_arrival_port), '') IS NOT NULL
    AND NULLIF(TRIM(p.vessel_arrival_port), '')::timestamp >= %s
    AND NULLIF(TRIM(p.vessel_arrival_port), '')::timestamp <= %s
    AND (
        p.unloading_commenced IS NULL
        OR TRIM(COALESCE(p.unloading_commenced, '')) = ''
    )
)

OR

(
    p.unloading_commenced IS NOT NULL
    AND TRIM(COALESCE(p.unloading_commenced, '')) <> ''
    AND (
        p.unloading_completed IS NULL
        OR TRIM(COALESCE(p.unloading_completed, '')) = ''
    )
    AND NULLIF(TRIM(p.unloading_commenced), '')::timestamp >= %s
    AND NULLIF(TRIM(p.unloading_commenced), '')::timestamp <= %s
)

OR

(
    p.unloading_completed IS NOT NULL
    AND TRIM(COALESCE(p.unloading_completed, '')) <> ''
    AND NULLIF(TRIM(p.unloading_completed), '')::timestamp >= %s
    AND NULLIF(TRIM(p.unloading_completed), '')::timestamp <= %s
)

ORDER BY
    CASE
        WHEN
            NULLIF(TRIM(p.vessel_arrival_port), '') IS NOT NULL
            AND (
                p.unloading_commenced IS NULL
                OR TRIM(COALESCE(p.unloading_commenced, '')) = ''
            )
        THEN 1

        WHEN
            p.unloading_commenced IS NOT NULL
            AND (
                p.unloading_completed IS NULL
                OR TRIM(COALESCE(p.unloading_completed, '')) = ''
            )
        THEN 2

        WHEN
            p.unloading_completed IS NOT NULL
        THEN 3
    END,

    COALESCE(
        NULLIF(TRIM(p.unloading_completed), '')::timestamp,
        NULLIF(TRIM(p.unloading_commenced), '')::timestamp,
        NULLIF(TRIM(p.vessel_arrival_port), '')::timestamp
    ) DESC

""", (
    day_start, day_end,   # CASE ARRIVED
    day_start, day_end,   # CASE DISCHARGING
    day_start, day_end,   # CASE COMPLETED

    day_start, day_end,   # WHERE ARRIVED
    day_start, day_end,   # WHERE DISCHARGING
    day_start, day_end    # WHERE COMPLETED
))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

def _fetch_upcoming_mbcs(report_date):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT
            h.id,
            h.mbc_name,
            m.mbc_owner_name AS owner,
            h.cargo_name,
            h.bl_quantity,

            l.fwd_draft,
            l.mid_draft,
            l.aft_draft,

            CASE
                WHEN NULLIF(TRIM(l.eta), '') IS NOT NULL
                     AND NULLIF(TRIM(d.arrival_gull_island), '') IS NULL
                THEN l.eta

                WHEN NULLIF(TRIM(d.arrival_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.departure_gull_island), '') IS NULL
                THEN d.arrival_gull_island

                WHEN NULLIF(TRIM(d.departure_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.vessel_arrival_port), '') IS NULL
                THEN d.departure_gull_island
            END AS event_time,

            CASE
                WHEN NULLIF(TRIM(l.eta), '') IS NOT NULL
                     AND NULLIF(TRIM(d.arrival_gull_island), '') IS NULL
                THEN l.eta

                WHEN NULLIF(TRIM(d.arrival_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.departure_gull_island), '') IS NULL
                THEN d.arrival_gull_island

                WHEN NULLIF(TRIM(d.departure_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.vessel_arrival_port), '') IS NULL
                THEN d.departure_gull_island
            END AS event_date,

            CASE
                WHEN NULLIF(TRIM(l.eta), '') IS NOT NULL
                     AND NULLIF(TRIM(d.arrival_gull_island), '') IS NULL
                THEN 'ETA GULL'

                WHEN NULLIF(TRIM(d.arrival_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.departure_gull_island), '') IS NULL
                THEN 'WAITING AT GULL'

                WHEN NULLIF(TRIM(d.departure_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.vessel_arrival_port), '') IS NULL
                THEN 'ETA DHARAMTAR'
            END AS status

        FROM mbc_header h

        JOIN mbc_load_port_lines l
            ON l.mbc_id = h.id

        LEFT JOIN mbc_discharge_port_lines d
            ON d.mbc_id = h.id

        LEFT JOIN mbc_master m
            ON TRIM(m.mbc_name) = TRIM(h.mbc_name)

        WHERE
            (
                NULLIF(TRIM(l.eta), '') IS NOT NULL
                AND NULLIF(TRIM(d.arrival_gull_island), '') IS NULL
            )

            OR

            (
                NULLIF(TRIM(d.arrival_gull_island), '') IS NOT NULL
                AND NULLIF(TRIM(d.departure_gull_island), '') IS NULL
            )

            OR

            (
                NULLIF(TRIM(d.departure_gull_island), '') IS NOT NULL
                AND NULLIF(TRIM(d.vessel_arrival_port), '') IS NULL
            )

        ORDER BY
            CASE
                WHEN NULLIF(TRIM(l.eta), '') IS NOT NULL
                     AND NULLIF(TRIM(d.arrival_gull_island), '') IS NULL
                THEN 1

                WHEN NULLIF(TRIM(d.arrival_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.departure_gull_island), '') IS NULL
                THEN 2

                WHEN NULLIF(TRIM(d.departure_gull_island), '') IS NOT NULL
                     AND NULLIF(TRIM(d.vessel_arrival_port), '') IS NULL
                THEN 3
            END,
            event_time
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

# def _fetch_cargo_availability(report_date):

#     conn = get_db()
#     cur = get_cursor(conn)

#     balance_date = report_date - timedelta(days=1)

#     cur.execute("""
#         SELECT
#             c.cargo_name,
#             ROUND(COALESCE(SUM(x.balance_qty),0)::numeric,0) AS at_jetty_qty

#         FROM
#         (
#             SELECT DISTINCT cargo_name
#             FROM mbc_header
#             WHERE cargo_name IS NOT NULL

#             UNION

#             SELECT DISTINCT cargo_name
#             FROM ldud_barge_lines
#             WHERE cargo_name IS NOT NULL
#         ) c

#         LEFT JOIN
#         (

#             /* MBC Balance */

#             SELECT
#                 h.cargo_name,

#                 GREATEST(
#                     h.bl_quantity - COALESCE(l.qty,0),
#                     0
#                 ) AS balance_qty

#             FROM mbc_header h

#             JOIN mbc_discharge_port_lines p
#                 ON p.mbc_id = h.id

#             LEFT JOIN
#             (
#                 SELECT
#                     source_id,
#                     SUM(COALESCE(quantity,0)) AS qty

#                 FROM lueu_lines

#                 WHERE source_type = 'MBC'
#                   AND is_deleted = false
#                   AND TO_DATE(entry_date,'YYYY-MM-DD') <= %s

#                 GROUP BY source_id

#             ) l
#                 ON l.source_id = h.id

#             WHERE
#                 p.unloading_commenced IS NOT NULL
#                 AND TRIM(COALESCE(p.unloading_commenced,'')) <> ''
#                 AND (
#                     p.unloading_completed IS NULL
#                     OR TRIM(COALESCE(p.unloading_completed,'')) = ''
#                 )

#             UNION ALL

#             /* Barge Balance */

#             SELECT
#                 b.cargo_name,

#                 GREATEST(
#                     COALESCE(b.discharge_quantity,0)
#                     - COALESCE(lb.actual_qty,0),
#                     0
#                 ) AS balance_qty

#             FROM ldud_barge_lines b

#             LEFT JOIN
#             (
#                 SELECT
#                     UPPER(TRIM(barge_name)) AS barge_name,
#                     SUM(COALESCE(quantity,0)) AS actual_qty

#                 FROM lueu_lines

#                 WHERE is_deleted = false
#                   AND barge_name IS NOT NULL
#                   AND TO_DATE(entry_date,'YYYY-MM-DD') <= %s

#                 GROUP BY UPPER(TRIM(barge_name))

#             ) lb
#                 ON lb.barge_name = UPPER(TRIM(b.barge_name))

#             WHERE
#                 b.commence_discharge_berth IS NOT NULL
#                 AND (
#                     b.cast_off_berth IS NULL
#                     OR TRIM(COALESCE(b.cast_off_berth,'')) = ''
#                 )

#         ) x
#             ON x.cargo_name = c.cargo_name

#         GROUP BY c.cargo_name
#         ORDER BY c.cargo_name

#     """, (
#         balance_date,
#         balance_date
#     ))

#     rows = cur.fetchall()

#     cur.close()
#     conn.close()

#     return rows

def _fetch_tide_data(report_date):

    start_datetime = report_date.strftime('%Y-%m-%d 00:00:00')

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT
            tide_datetime,
            tide_meters
        FROM tide_master
        WHERE tide_datetime >= %s
        ORDER BY tide_datetime
        LIMIT 6
    """, (start_datetime,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows


def _fetch_cargo_handled(report_date):
    """Fetch cargo handled by route (day + month)."""

    # Previous day window
    previous_date = report_date - timedelta(days=1)

    window_start = datetime(
        previous_date.year,
        previous_date.month,
        previous_date.day,
        0, 0, 0
    )

    window_end = datetime(
        previous_date.year,
        previous_date.month,
        previous_date.day,
        23, 59, 59
    )

    month_start = datetime(
        report_date.year,
        report_date.month,
        1,
        0, 0, 0
    )

    ws_str = window_start.strftime('%Y-%m-%d %H:%M:%S')
    we_str = window_end.strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cur = get_cursor(conn)

    def _period(start, end):

        cur.execute("""
            SELECT
                route_name,
                COALESCE(SUM(quantity),0) AS qty
            FROM lueu_lines
            WHERE route_name IS NOT NULL
              AND route_name <> ''
              AND entry_date IS NOT NULL
              AND (entry_date || ' ' || COALESCE(from_time,'00:00')) >= %s
              AND (entry_date || ' ' || COALESCE(from_time,'00:00')) <= %s
            GROUP BY route_name
            ORDER BY route_name
        """, (start, end))

        return {
            r['route_name']: float(r['qty'])
            for r in cur.fetchall()
        }

    def _group_routes(data):

        grouped = {
            'DIRECT PLANT': 0,
            'STACKER / SHED': 0,
            'CEMENT SILO': 0,
            'BY ROAD': 0,
            'OTHERS': 0
        }

        for route, qty in data.items():

            route_upper = (route or '').strip().upper()

            if route_upper in (
                'C-131',
                'C-131 A',
                'C-131 C',
                'C-131 D',
                'C-131 E',
                'C-131 F'
            ):
                grouped['DIRECT PLANT'] += qty

            elif route_upper in (
                'COAL STACKER',
                'LS-01',
                'LS-02',
                'LS-03'
            ):
                grouped['STACKER / SHED'] += qty

            elif route_upper == 'LS-05':
                grouped['CEMENT SILO'] += qty

            elif route_upper == 'BY ROAD':
                grouped['BY ROAD'] += qty

            else:
                grouped['OTHERS'] += qty

        return {
            k: v
            for k, v in grouped.items()
            if v > 0
        }

    # Day Data (Previous Date)
    day_dict = _group_routes(
        _period(ws_str, we_str)
    )

    # Month Data (live, 1st of month -> report date)
    month_dict = _group_routes(
        _period(
            month_start.strftime('%Y-%m-%d %H:%M:%S'),
            report_date.strftime('%Y-%m-%d 23:59:59')
        )
    )

    conn.close()

    day_rows = sorted(day_dict.items())
    month_rows = sorted(month_dict.items())

    return day_rows, month_rows


def _fetch_cargo_statistics(report_date):

    # Show previous day's report
    report_date = report_date - timedelta(days=1)

    month_start = report_date.replace(day=1)

    conn = get_db()
    cur = get_cursor(conn)

    def _period(start_date, end_date):

        cur.execute("""
            SELECT
                CASE
                    WHEN source_type = 'VCN'
                        THEN 'Mumbai Anchorage'
                    WHEN source_type = 'MBC'
                        THEN 'MBC (Jaigad/Other)'
                    ELSE source_type
                END AS cargo_source,

                COALESCE(SUM(quantity), 0) AS qty

            FROM lueu_lines

            WHERE is_deleted = false
            AND entry_date IS NOT NULL
            AND TO_DATE(entry_date,'YYYY-MM-DD') >= %s
            AND TO_DATE(entry_date,'YYYY-MM-DD') <= %s

            GROUP BY 1
            ORDER BY 1
        """, (
            start_date,
            end_date
        ))

        return [
            (r['cargo_source'], float(r['qty']))
            for r in cur.fetchall()
        ]

    # Day = only previous day
    day_rows = _period(
        report_date,
        report_date
    )

    # MTD = 1st of month to previous day
    month_rows = _period(
        month_start,
        report_date
    )

    conn.close()

    return day_rows, month_rows



def _fetch_mbc_cargo_handling(report_date):


    target_date = report_date - timedelta(days=1)

    month_start = date(
        target_date.year,
        target_date.month,
        1
    )

    # Financial Year Start (April)
    if target_date.month >= 4:
        fy_start = date(target_date.year, 4, 1)
    else:
        fy_start = date(target_date.year - 1, 4, 1)

    conn = get_db()
    cur = get_cursor(conn)

    def _period(start_date, end_date):

        cur.execute("""
            SELECT

                h.cargo_type,

                COALESCE(
                    m.mbc_owner_name,
                    'OTHERS'
                ) AS owner,

                COALESCE(
                    SUM(l.quantity),
                    0
                ) AS qty

            FROM lueu_lines l

            JOIN mbc_header h
                ON h.id = l.source_id

            LEFT JOIN mbc_master m
                ON TRIM(m.mbc_name) = TRIM(h.mbc_name)

            WHERE l.is_deleted = false
            AND l.source_type = 'MBC'
            AND l.entry_date::date BETWEEN %s AND %s

            GROUP BY
                h.cargo_type,
                m.mbc_owner_name

            ORDER BY
                m.mbc_owner_name,
                h.cargo_type

        """, (
            start_date,
            end_date
        ))

        return cur.fetchall()

    # Previous Day
    day_rows = _period(
        target_date,
        target_date
    )

    # Month To Date
    month_rows = _period(
        month_start,
        target_date
    )

    # Financial Year To Date
    year_rows = _period(
        fy_start,
        target_date
    )

    cur.close()
    conn.close()

    return day_rows, month_rows, year_rows


def _fetch_mbc_status(report_date):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT
            m.mbc_name,

            CASE

                /* Empty : Waiting at Jaigad */
                WHEN h.id IS NULL
                THEN 'EMPTY : WAITING AT JAIGAD'

                /* Empty : Waiting at Load Port */
                WHEN
                    NULLIF(TRIM(l.arrived_load_port), '') IS NOT NULL
                    AND NULLIF(TRIM(l.loading_commenced), '') IS NULL
                THEN
                    'EMPTY : WAITING AT LOAD PORT'

                /* Under Loading */
                WHEN
                    NULLIF(TRIM(l.loading_commenced), '') IS NOT NULL
                    AND NULLIF(TRIM(l.loading_completed), '') IS NULL
                THEN
                    'UNDER LOADING'

                /* Loaded : Waiting at Load Port */
                WHEN
                    NULLIF(TRIM(l.loading_completed), '') IS NOT NULL
                    AND NULLIF(TRIM(l.cast_off_load_port), '') IS NULL
                THEN
                    'LOADED : WAITING AT LOAD PORT'

                /* Loaded : On the way to Gull */
                WHEN
                    NULLIF(TRIM(l.cast_off_load_port), '') IS NOT NULL
                    AND NULLIF(TRIM(d.arrival_gull_island), '') IS NULL
                THEN
                    'LOADED : ON THE WAY TO GULL'

                /* Loaded : Waiting at Gull */
                WHEN
                    NULLIF(TRIM(d.arrival_gull_island), '') IS NOT NULL
                    AND NULLIF(TRIM(d.departure_gull_island), '') IS NULL
                THEN
                    'LOADED : WAITING AT GULL'

                /* Loaded : On the way to Dharamtar */
                WHEN
                    NULLIF(TRIM(d.departure_gull_island), '') IS NOT NULL
                    AND NULLIF(TRIM(d.vessel_arrival_port), '') IS NULL
                THEN
                    'LOADED : ON THE WAY TO DHARAMTAR'

                /* Loaded : Waiting at Dharamtar */
                WHEN
                    NULLIF(TRIM(d.vessel_arrival_port), '') IS NOT NULL
                    AND NULLIF(TRIM(d.unloading_commenced), '') IS NULL
                THEN
                    'LOADED : WAITING AT DHARAMTAR'

                /* Under Discharge */
                WHEN
                    NULLIF(TRIM(d.unloading_commenced), '') IS NOT NULL
                    AND NULLIF(TRIM(d.unloading_completed), '') IS NULL
                THEN
                    'UNDER DISCHARGE'

                /* Empty : On the way to Jaigad */
                WHEN
                    NULLIF(TRIM(d.unloading_completed), '') IS NOT NULL
                    AND d.sailed_out_load_port IS NOT NULL
                THEN
                    'EMPTY : ON THE WAY TO JAIGAD'

                /* Empty : Waiting at Dharamtar */
                WHEN
                    NULLIF(TRIM(d.unloading_completed), '') IS NOT NULL
                THEN
                    'EMPTY : WAITING AT DHARAMTAR'

                ELSE
                    'NA'

            END AS mbc_status

        FROM mbc_master m

        LEFT JOIN LATERAL (
            SELECT h.*
            FROM mbc_header h
            WHERE TRIM(h.mbc_name) = TRIM(m.mbc_name)
            ORDER BY h.id DESC
            LIMIT 1
        ) h ON TRUE

        LEFT JOIN mbc_load_port_lines l
            ON l.mbc_id = h.id

        LEFT JOIN mbc_discharge_port_lines d
            ON d.mbc_id = h.id

        WHERE
            UPPER(TRIM(COALESCE(m.mbc_owner_name, '')))
            IN ('JSW INFRA', 'JSW SHIPPING')

        ORDER BY m.mbc_name
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

def _fetch_cargo_type_throughput(report_date):

    target_date = report_date - timedelta(days=1)

    month_start = date(
        target_date.year,
        target_date.month,
        1
    )

    if target_date.month >= 4:
        fy_start = date(target_date.year, 4, 1)
    else:
        fy_start = date(target_date.year - 1, 4, 1)

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        WITH hist AS (

            SELECT
                COALESCE(vc.cargo_type, 'OTHERS') AS cargo_type,
                h.entry_date AS txn_date,
                SUM(COALESCE(h.quantity, 0)) AS quantity

            FROM rp01_historical_lueu h

            LEFT JOIN vessel_cargo vc
                ON UPPER(TRIM(vc.cargo_name))
                = UPPER(TRIM(h.cargo_name))

            WHERE h.cargo_name IS NOT NULL

            GROUP BY
                COALESCE(vc.cargo_type, 'OTHERS'),
                h.entry_date
        ),

        live AS (

            SELECT
                COALESCE(vc.cargo_type, 'OTHERS') AS cargo_type,
                TO_DATE(l.entry_date,'YYYY-MM-DD') AS txn_date,
                SUM(COALESCE(l.quantity,0)) AS quantity

            FROM lueu_lines l

            LEFT JOIN vessel_cargo vc
                ON UPPER(TRIM(vc.cargo_name))
                = UPPER(TRIM(l.cargo_name))

            WHERE
                l.is_deleted = false
                AND l.cargo_name IS NOT NULL

            GROUP BY
                COALESCE(vc.cargo_type,'OTHERS'),
                TO_DATE(l.entry_date,'YYYY-MM-DD')
        ),

        throughput AS (

            SELECT
                cargo_type,
                txn_date,
                quantity
            FROM hist

            UNION ALL

            SELECT
                l.cargo_type,
                l.txn_date,
                l.quantity
            FROM live l

            WHERE NOT EXISTS (
                SELECT 1
                FROM hist h
                WHERE h.txn_date = l.txn_date
            )
        ),

        cargo_summary AS (

            SELECT

                cargo_type,

                COALESCE(
                    SUM(
                        CASE
                            WHEN txn_date = %s
                            THEN quantity
                            ELSE 0
                        END
                    ),
                    0
                ) AS day_qty,

                COALESCE(
                    SUM(
                        CASE
                            WHEN txn_date BETWEEN %s AND %s
                            THEN quantity
                            ELSE 0
                        END
                    ),
                    0
                ) AS month_qty,

                COALESCE(
                    SUM(
                        CASE
                            WHEN txn_date BETWEEN %s AND %s
                            THEN quantity
                            ELSE 0
                        END
                    ),
                    0
                ) AS year_qty

            FROM throughput

            GROUP BY cargo_type
        )

        SELECT
            cargo_type,
            day_qty,
            month_qty,
            year_qty

        FROM (

            SELECT
                cargo_type,
                day_qty,
                month_qty,
                year_qty,
                1 AS sort_order

            FROM cargo_summary

            UNION ALL

            SELECT
                'STEEL PLANT CARGO',

                SUM(day_qty)
                - SUM(
                    CASE
                        WHEN UPPER(cargo_type) = 'CLINKER'
                        THEN day_qty
                        ELSE 0
                    END
                )
                - SUM(
                    CASE
                        WHEN UPPER(cargo_type) = 'SLAG'
                        THEN day_qty
                        ELSE 0
                    END
                ),

                SUM(month_qty)
                - SUM(
                    CASE
                        WHEN UPPER(cargo_type) = 'CLINKER'
                        THEN month_qty
                        ELSE 0
                    END
                )
                - SUM(
                    CASE
                        WHEN UPPER(cargo_type) = 'SLAG'
                        THEN month_qty
                        ELSE 0
                    END
                ),

                SUM(year_qty)
                - SUM(
                    CASE
                        WHEN UPPER(cargo_type) = 'CLINKER'
                        THEN year_qty
                        ELSE 0
                    END
                )
                - SUM(
                    CASE
                        WHEN UPPER(cargo_type) = 'SLAG'
                        THEN year_qty
                        ELSE 0
                    END
                ),

                2

            FROM cargo_summary

            UNION ALL

            SELECT
                'AVG (STEEL PLANT CARGO)',

                ROUND(
                    (
                        SUM(day_qty)
                        - SUM(CASE WHEN UPPER(cargo_type)='CLINKER' THEN day_qty ELSE 0 END)
                        - SUM(CASE WHEN UPPER(cargo_type)='SLAG' THEN day_qty ELSE 0 END)
                    )::numeric,
                    2
                ),

                ROUND(
                    (
                        SUM(month_qty)
                        - SUM(CASE WHEN UPPER(cargo_type)='CLINKER' THEN month_qty ELSE 0 END)
                        - SUM(CASE WHEN UPPER(cargo_type)='SLAG' THEN month_qty ELSE 0 END)
                    )::numeric
                    /
                    EXTRACT(DAY FROM %s::date),
                    2
                ),

                ROUND(
                    (
                        SUM(year_qty)
                        - SUM(CASE WHEN UPPER(cargo_type)='CLINKER' THEN year_qty ELSE 0 END)
                        - SUM(CASE WHEN UPPER(cargo_type)='SLAG' THEN year_qty ELSE 0 END)
                    )::numeric
                    /
                    (%s::date - %s::date + 1),
                    2
                ),

                3

            FROM cargo_summary

        ) x

        ORDER BY
            sort_order,
            cargo_type

    """, (
        target_date,
        month_start,
        target_date,
        fy_start,
        target_date,
        target_date,
        target_date,
        fy_start
    ))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows




def _fetch_port_throughput(report_date):

    target_date = report_date - timedelta(days=1)

    month_start = date(
        target_date.year,
        target_date.month,
        1
    )

    # Financial Year Start (April)
    if target_date.month >= 4:
        fy_start = date(target_date.year, 4, 1)
    else:
        fy_start = date(target_date.year - 1, 4, 1)

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        WITH hist AS (

            SELECT
                entry_date,
                SUM(quantity) qty
            FROM rp01_historical_lueu
            WHERE cargo_name IS NOT NULL
            GROUP BY entry_date

        ),

        live AS (

            SELECT
                TO_DATE(entry_date,'YYYY-MM-DD') AS entry_date,
                SUM(quantity) qty
            FROM lueu_lines
            WHERE is_deleted = false
              AND cargo_name IS NOT NULL
            GROUP BY TO_DATE(entry_date,'YYYY-MM-DD')

        ),

        throughput AS (

            SELECT
                h.entry_date,
                h.qty
            FROM hist h

            UNION ALL

            SELECT
                l.entry_date,
                l.qty
            FROM live l
            WHERE NOT EXISTS (
                SELECT 1
                FROM hist h
                WHERE h.entry_date = l.entry_date
            )
        )

        SELECT

            COALESCE(
                SUM(
                    CASE
                        WHEN entry_date = %s
                        THEN qty
                        ELSE 0
                    END
                ),
                0
            ) AS day_qty,

            COALESCE(
                SUM(
                    CASE
                        WHEN entry_date BETWEEN %s AND %s
                        THEN qty
                        ELSE 0
                    END
                ),
                0
            ) AS month_qty,

            COALESCE(
                SUM(
                    CASE
                        WHEN entry_date BETWEEN %s AND %s
                        THEN qty
                        ELSE 0
                    END
                ),
                0
            ) AS year_qty

        FROM throughput
    """, (
        target_date,
        month_start,
        target_date,
        fy_start,
        target_date
    ))

    row = cur.fetchone()

    # =====================================================
    # MONTH TPD
    # =====================================================

    month_days_elapsed = (
        target_date - month_start
    ).days + 1

    month_tpd = round(
        float(row["month_qty"] or 0) /
        month_days_elapsed,
        2
    ) if month_days_elapsed else 0

    # =====================================================
    # YEAR TPD (FINANCIAL YEAR)
    # =====================================================

    fy_days_elapsed = (
        target_date - fy_start
    ).days + 1

    year_tpd = round(
        float(row["year_qty"] or 0) /
        fy_days_elapsed,
        2
    ) if fy_days_elapsed else 0

    # =====================================================
    # CUMULATIVE SINCE OCT 2012
    # =====================================================

    cur.execute("""
        WITH cutoff_total AS (
            SELECT
                COALESCE(
                    SUM(cargo.value::numeric),
                    0
                ) AS qty
            FROM (
                SELECT cutoff_values::jsonb AS j
                FROM daily_ops_cutoff
                ORDER BY cutoff_date DESC
                LIMIT 1
            ) d,
            LATERAL jsonb_each(
                d.j->'fy_throughput'
            ) fy,
            LATERAL jsonb_each_text(
                fy.value
            ) cargo
        )
        SELECT
            cutoff_total.qty AS cumulative_qty
        FROM cutoff_total
    """)

    cumulative_row = cur.fetchone()

    cumulative_qty = (
        float(cumulative_row["cumulative_qty"] or 0)
        + float(row["year_qty"] or 0)
    )

    cur.close()
    conn.close()

    return {
        "day_qty": int(row["day_qty"] or 0),
        "mtd_qty": int(row["month_qty"] or 0),
        "ytd_qty": int(row["year_qty"] or 0),
        "cumulative_qty": int(cumulative_qty),
        "month_tpd": float(month_tpd),
        "year_tpd": float(year_tpd)
    }
# =====================================================================
# MODIFIED _build_excel_a4
# Changes:
#   1. ALL borders are BOLD (thick/medium) — no thin borders anywhere
#   2. NO bold text anywhere (bold=False throughout)
#   3. Title rows use a larger font size (14pt instead of default)
#   4. Vessel detail column widths are kept CONSTANT (not overridden)
# =====================================================================

# ── shared style helpers (assumed to exist in the calling module) ────
# _font(bold, size)  – returns Font(...)
# _fill(hex)         – returns PatternFill(...)
# _bdr               – replaced below with thick_bdr
# _ctr / _left       – Alignment objects
# _fmt_dt            – date formatter
# ─────────────────────────────────────────────────────────────────────

def _build_excel_a4(
    vessels,
    report_date,
    day_rows=None,
    month_rows=None,
    tide_rows=None,
    mbc_day=None,
    mbc_month=None,
    mbc_year=None,
    upcoming_vessels=None,
    upcoming_mbcs=None,
    discharging_mbcs=None,
    mbc_status_rows=None,
    cargo_availability=None,
    mbc_day_rows=None,
    mbc_month_rows=None,
    mbc_year_rows=None,
    cargo_type_throughput=None,
    cargo_stats_day=None,
    cargo_stats_month=None,
    port_throughput=None,
    editable_table=None,
    rainfall_table=None,
    bf_table=None,
    rm_table=None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from datetime import datetime
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Ops"

    # ── page setup ────────────────────────────────────────────────────
    # ── page setup ────────────────────────────────────────────────────
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Fit entire report on ONE page
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # Remove fixed scaling
    ws.page_setup.scale = None

    ws.page_margins = PageMargins(
        left=0.05,
        right=0.05,
        top=0.05,
        bottom=0.05,
        header=0,
        footer=0
    )

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # Excel view only (does not affect printing)
    ws.sheet_view.zoomScale = 100

    # Set print area after sheet is fully built
    ws.print_area = (
        f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    )

    # ── style constants ───────────────────────────────────────────────
    TITLE_SIZE   = 34   # font size for section titles
    NORMAL_SIZE  = 32   # font size for everything else

    _thick = Side(style="medium")          # BOLD border side
    _thin  = Side(style="medium")          # use medium everywhere → all bold
    thick_bdr = Border(
        left=_thick, right=_thick, top=_thick, bottom=_thick
    )

    _ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _left = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _font(bold=False, size=NORMAL_SIZE):
        # CHANGE 2: ignore bold parameter — never bold
        return Font(name="Calibri", size=size, bold=False)

    def _title_font():
        return Font(name="Calibri", size=TITLE_SIZE, bold=False)

    def _fill(hex_color="FFFFFF"):
        return PatternFill("solid", fgColor=hex_color)

    def _fmt_dt(val):
        if not val:
            return ""
        try:
            if isinstance(val, str):
                val = datetime.fromisoformat(val)
            return val.strftime("%d-%m-%Y %H:%M")
        except Exception:
            return str(val)

    def _fmt_num(v):
        if v is None:
            return ""
        try:
            return int(round(float(v)))
        except Exception:
            return v

    # ── column layout ─────────────────────────────────────────────────
    LABEL_START  = 2
    LABEL_END    = 4
    V_START      = 5
    COLS_PER_V   = 4
    # CHANGE 4: constant vessel column width — never reassigned later
    VESSEL_COL_WIDTH = 22

    vessel_count = len(vessels)

    def v_start(idx):  return V_START + idx * COLS_PER_V
    def v_end(idx):    return v_start(idx) + COLS_PER_V - 1

    last_vessel_col = (
        v_end(vessel_count - 1) if vessel_count else LABEL_END
    )

    # ── column widths ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # CHANGE 4: set vessel column widths ONCE here; never changed later
    for i in range(vessel_count):
        for dc in range(COLS_PER_V):
            col_letter = get_column_letter(v_start(i) + dc)
            ws.column_dimensions[col_letter].width = VESSEL_COL_WIDTH

    # ── cell helpers ──────────────────────────────────────────────────
    def _cell(r, c, val="", bold=False, fill="FFFFFF", align=_ctr, title=False):
        cell = ws.cell(r, c)
        cell.value     = val
        cell.font      = _title_font() if title else _font()
        cell.fill      = _fill(fill)
        cell.alignment = align
        cell.border    = thick_bdr
        return cell

    def _merge_write(r1, c1, r2, c2, val="", bold=False,
                     fill="FFFFFF", align=_ctr, title=False):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        anchor            = ws.cell(r1, c1)
        anchor.value      = val
        anchor.font       = _title_font() if title else _font()
        anchor.fill       = _fill(fill)
        anchor.alignment  = align
        anchor.border     = thick_bdr
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(rr, cc).border = thick_bdr

    def safe_cell(ws, row, col, value=None):
        cell_coord = f"{get_column_letter(col)}{row}"
        for merge in list(ws.merged_cells.ranges):
            if cell_coord in merge:
                ws.unmerge_cells(str(merge))
                break
        c = ws.cell(row, col)
        if value is not None:
            c.value = value
        return c

    def safe_merge(ws, start_row, start_col, end_row, end_col):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                coord = f"{get_column_letter(c)}{r}"
                for merge in list(ws.merged_cells.ranges):
                    if coord in merge:
                        ws.unmerge_cells(str(merge))
                        break
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row,     end_column=end_col)

    # ── report header ─────────────────────────────────────────────────
    report_title = (
        f"Daily Report of JSW Dharamtar Port Operation : "
        f"{report_date.day}.{report_date.month}.{report_date.year}"
    )
    date_str = report_date.strftime("%d-%m-%Y")
    ws.row_dimensions[2].height = 30

    _merge_write(2, LABEL_START, 2, LABEL_START + 2,
                 date_str, align=_left, title=True)
    _merge_write(2, LABEL_START + 3, 2, last_vessel_col - 10,
                 report_title, align=_ctr, title=True)
    _merge_write(2, last_vessel_col - 9, 2, last_vessel_col - 5,
                 "Doc No.OPE/0100/F/11| REV.02 | Issue no. 02",
                 align=_left, title=True)
    _merge_write(2, last_vessel_col - 4, 2, last_vessel_col,
                 "Issue Date : 01/04/2024", align=_left, title=True)

    # ── vessel header ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 80
    _merge_write(3, LABEL_START, 3, LABEL_END, "")

    for idx, vessel in enumerate(vessels):

        print(
            "HEADER CARGO =",
            vessel.get("cargo_name")
        )

        _merge_write(
            3,
            v_start(idx),
            3,
            v_end(idx),
            f"Vessel {idx + 1}: {vessel['vessel_name']} | {vessel.get('cargo_name','')}",
            align=_ctr,
            title=True
        )
    ws.print_title_rows = "2:3"
    ws.print_options.horizontalCentered = True

    current_row = 4

    # ── vessel status section ─────────────────────────────────────────
    STATUS_ROWS = [
        ("Stevedore / Barge Group",       "stevedore_group",        lambda x: x or "",  _ctr),
        ("BL Qty",                         "bl_qty",                 _fmt_num,            _ctr),
        ("24 Hrs Discharge",               "ops_24h",                _fmt_num,            _ctr),
        ("Unloaded Till Date",             "ops_till",               _fmt_num,            _ctr),
        ("Balance",                        "balance",                _fmt_num,            _ctr),
        ("Vsl Arrived / NOR",              "nor_tendered",           _fmt_dt,             _ctr),
        ("Discharge Commenced",            "discharge_commenced",    _fmt_dt,             _ctr),
        ("Discharge Completed",            "discharge_completed",    _fmt_dt,             _ctr),
        (None, None, None, None),
        ("No Of Barges",                   "num_barges",             lambda x: x or "",   _ctr),
        ("Discharge At Jetty",             "at_jetty",               lambda x: x or "",   _left),
        ("Waiting For Discharge At Jetty", "waiting_discharge",      lambda x: x or "",   _left),
        ("Waiting Empty At Jetty",         "waiting_empty_jetty",    lambda x: x or "",   _left),
        ("At Gull - Waiting (Loaded)",     "at_gull_loaded",         lambda x: x or "",   _left),
        ("Under Loading",                  "under_loading",          lambda x: x or "",   _left),
        ("Waiting For Loading",            "waiting_loading",        lambda x: x or "",   _left),
        ("In Transit Jetty To MV",         "in_transit_jetty_to_mv", lambda x: x or "",   _left),
        ("Breakdown/Offhire/Costal/DD",    "breakdown",              lambda x: x or "",   _left),
    ]

    for label, field, formatter, align in STATUS_ROWS:

        if label is None:
            current_row += 1
            continue

        ws.row_dimensions[current_row].height = 30

        if field in (
            "at_jetty",
            "waiting_discharge",
            "at_gull_loaded",
            "under_loading"
        ):
            ws.row_dimensions[current_row].height = 400

        _merge_write(
            current_row,
            LABEL_START,
            current_row,
            LABEL_END,
            label,
            align=_left
        )

        row_total = 0

        for idx, vessel in enumerate(vessels):

            raw = vessel.get(field)
            value = formatter(raw) if formatter else raw

            _merge_write(
                current_row,
                v_start(idx),
                current_row,
                v_end(idx),
                value,
                align=align
            )

            if field in ("ops_24h", "balance", "num_barges"):
                try:
                    row_total += float(raw or 0)
                except:
                    pass

        # =====================================================
        # SHOW TOTALS ON RIGHT SIDE
        # =====================================================
        if field in ("ops_24h", "balance", "num_barges"):

            total_start_col = last_vessel_col + 1
            total_end_col = total_start_col

            _merge_write(
                current_row,
                total_start_col,
                current_row,
                total_end_col,
                f"{int(round(row_total)):,}",
                align=_ctr
            )

            total_cell = ws.cell(current_row, total_start_col)
            total_cell.font = Font(
                name="Calibri",
                size=NORMAL_SIZE,
                bold=False,
                color="FF0000"
            )

        if label == "Discharge At Jetty":

            mbc_col = last_vessel_col + 2

            discharging_at_jetty = [
                m for m in (discharging_mbcs or [])
                if (m.get("status") or "") == "DISCHARGING"
            ]

            val = "\n".join([
                f"{(m.get('mbc_name','') or '').replace('JSW ','').strip()}({m.get('cargo_name','')})-Bal:{int(round(float(m.get('discharge_quantity') or 0)))}MT"
                for m in discharging_at_jetty
            ])

            c = ws.cell(current_row, mbc_col, val)
            c.font = _font()
            c.fill = _fill("FFFFFF")
            c.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True
            )
            c.border = thick_bdr

            ws.merge_cells(
                start_row=current_row,
                start_column=mbc_col,
                end_row=current_row,
                end_column=mbc_col + 3
            )

            ws.row_dimensions[current_row].height = 400

        elif label == "Waiting For Discharge At Jetty":

            mbc_col = last_vessel_col + 2

            discharging_waiting = [
                m for m in (discharging_mbcs or [])
                if (m.get("status") or "") == "ARRIVED"
            ]

            val = "\n".join([
                f"{(m.get('mbc_name','') or '').replace('JSW ','').strip()} "
                f"({m.get('cargo_name','')}) "
                f"{int(round(float(m.get('bl_quantity') or 0)))}MT"
                for m in discharging_waiting
            ])

            c = ws.cell(current_row, mbc_col, val)
            c.font = _font()
            c.fill = _fill("FFFFFF")
            c.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True
            )
            c.border = thick_bdr

            ws.merge_cells(
                start_row=current_row,
                start_column=mbc_col,
                end_row=current_row,
                end_column=mbc_col + 3
            )

            ws.row_dimensions[current_row].height = 400

        current_row += 1

        # =====================================================
        # Draw complete Total Column Border
        # =====================================================

        TOTAL_COL = last_vessel_col + 1

        status_start_row = 4
        status_end_row = current_row - 1

        for r in range(status_start_row, status_end_row + 1):

            c = ws.cell(r, TOTAL_COL)

            if c.value is None:
                c.value = ""

            c.border = thick_bdr

    # ── upcoming vessels ──────────────────────────────────────────────
    # Move Upcoming Vessels section down by 1 row
    current_row += 1

    upcoming_vessels = upcoming_vessels or []
    uv_section_start_row = current_row  # ← save for Cargo Availability anchor

    # Explicit column widths for the UV/MBC table area (cols B..J)
    UV_COL_WIDTHS = [30, 45, 20, 20, 22, 22, 30, 30, 35, 35, 35]
    for i, w in enumerate(UV_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(LABEL_START + i)].width = w

    # spans: Vessel Name=3, Cargo=2, Qty=1, Agent=2, ETA/Status=2 → total=10
    uv_headers = ["Vessel Name", "Cargo", "Qty (MT)", "Agent", "ETA / Status"]
    uv_spans   = [3, 2, 2, 2, 2]          # total = 10 cols
    UV_TOTAL   = sum(uv_spans)             # 10

    ws.merge_cells(start_row=current_row, start_column=LABEL_START,
                   end_row=current_row,   end_column=LABEL_START + UV_TOTAL - 1)
    c = ws.cell(current_row, LABEL_START, "Upcoming Vessels")
    c.font = _title_font(); c.fill = _fill("D9EAF7")
    c.alignment = _ctr; c.border = thick_bdr
    for cc in range(LABEL_START, LABEL_START + UV_TOTAL):
        ws.cell(current_row, cc).border = thick_bdr
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    col = LABEL_START
    for h, span in zip(uv_headers, uv_spans):
        if span > 1:
            ws.merge_cells(start_row=current_row, start_column=col,
                           end_row=current_row,   end_column=col + span - 1)
        c = ws.cell(current_row, col, h)
        c.font = _font(); c.fill = _fill("D9EAF7")
        c.alignment = _ctr; c.border = thick_bdr
        for cc in range(col, col + span):
            ws.cell(current_row, cc).border = thick_bdr
        col += span
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    for v in upcoming_vessels:
        col = LABEL_START
        for val, span in zip(
            [
                v.get("vessel_name", ""),
                v.get("cargo_name", "") or "-",
                int(round(float(v.get("bl_quantity") or 0))) or "-",
                v.get("vessel_agent_name", "") or "-",
                v.get("eta", "") or "-",
            ],
            uv_spans,
        ):
            if span > 1:
                ws.merge_cells(start_row=current_row, start_column=col,
                               end_row=current_row,   end_column=col + span - 1)
            c = ws.cell(current_row, col, val)
            c.font = _font(); c.fill = _fill("FFFFFF")
            c.alignment = _left; c.border = thick_bdr
            for cc in range(col, col + span):
                ws.cell(current_row, cc).border = thick_bdr
            col += span
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # ── cargo availability ────────────────────────────────────────────
    cargo_availability = cargo_availability or []
    data_row = current_row

    if cargo_availability:
        CA_START_COL   = LABEL_START + UV_TOTAL + 1
        HDR_ROW_HEIGHT = 75

        cargo_names = (
            editable_table[0][1:-1]
            if editable_table and len(editable_table) > 0
            else [row["cargo_name"] for row in cargo_availability]
        )

        grand_total  = sum(float(row["at_jetty_qty"] or 0) for row in cargo_availability)
        uv_start_row = uv_section_start_row

        # ── Auto-fit helper: measures all cells in a column and sets width ─────
        def _autofit_col(col_idx):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows():
                cell = ws.cell(row=row[0].row, column=col_idx)
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # ── title ──────────────────────────────────────────────────────────────
        last_ca_col = CA_START_COL + len(cargo_names) + 1
        ws.merge_cells(
            start_row=uv_start_row, start_column=CA_START_COL,
            end_row=uv_start_row,   end_column=last_ca_col
        )
        c = ws.cell(uv_start_row, CA_START_COL, "Cargo Availability For The Day")
        c.font      = _title_font()
        c.fill      = _fill("D9EAF7")
        c.alignment = _ctr
        c.border    = thick_bdr
        for cc in range(CA_START_COL, last_ca_col + 1):
            ws.cell(uv_start_row, cc).border = thick_bdr

        # ── header row ─────────────────────────────────────────────────────────
        hdr_row = uv_start_row + 1
        col     = CA_START_COL

        # Empty label column
        c = ws.cell(hdr_row, col, "")
        c.font      = _font()
        c.fill      = _fill("D9EAF7")
        c.alignment = _ctr
        c.border    = thick_bdr
        col += 1

        # Cargo name columns
        for cargo in cargo_names:
            c = ws.cell(hdr_row, col, cargo)
            c.font      = _font()
            c.fill      = _fill("D9EAF7")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = thick_bdr
            col += 1

        # Total column
        c = ws.cell(hdr_row, col, "Total")
        c.font      = _font()
        c.fill      = _fill("D9EAF7")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thick_bdr
        total_col = col

        ws.row_dimensions[hdr_row].height = HDR_ROW_HEIGHT

        # ── data rows ──────────────────────────────────────────────────────────
        data_row   = hdr_row + 1
        row_labels = ["At Jetty", "", "", "", "Total"]
        for r, lbl in enumerate(row_labels):
            c = ws.cell(data_row + r, CA_START_COL, lbl)
            c.font      = _font()
            c.border    = thick_bdr
            c.alignment = _left

        if editable_table:
            print("EXCEL TABLE =", editable_table)
            for r_idx, row_data in enumerate(editable_table[1:]):
                excel_row = data_row + r_idx
                for c_idx, value in enumerate(row_data):
                    excel_col = CA_START_COL + c_idx
                    c = ws.cell(excel_row, excel_col, value)
                    c.border    = thick_bdr
                    c.alignment = _ctr
        else:
            col = CA_START_COL + 1
            for row in cargo_availability:
                c = ws.cell(data_row, col, row.get("at_jetty_qty", ""))
                c.alignment = _ctr
                c.border    = thick_bdr
                col += 1

        # ── total column & borders ─────────────────────────────────────────────
        c = ws.cell(data_row + 4, total_col, int(round(grand_total)))
        c.font      = _font()
        c.alignment = _ctr
        c.border    = thick_bdr

        for cc in range(CA_START_COL, total_col + 1):
            ws.cell(data_row + 4, cc).border = thick_bdr
        for rr in range(hdr_row, data_row + 5):
            ws.cell(rr, total_col).border = thick_bdr
        for rr in range(uv_start_row, data_row + 5):
            for cc in range(CA_START_COL, total_col + 1):
                ws.cell(rr, cc).border = thick_bdr

        # ── grand total row ────────────────────────────────────────────────────
        grand_total_row = data_row + 5
        c = ws.cell(grand_total_row, CA_START_COL, "")
        c.border = thick_bdr

        ws.merge_cells(start_row=grand_total_row, start_column=CA_START_COL + 1,
                    end_row=grand_total_row,   end_column=CA_START_COL + 5)
        ws.merge_cells(start_row=grand_total_row, start_column=CA_START_COL + 6,
                    end_row=grand_total_row,   end_column=CA_START_COL + 15)
        ws.merge_cells(start_row=grand_total_row, start_column=CA_START_COL + 16,
                    end_row=grand_total_row,   end_column=CA_START_COL + 18)

        grand_ibrm = grand_cbrm = grand_fluxes = ""
        grand_slag = grand_clinker = grand_total_val = ""
        try:
            if editable_table and len(editable_table) > 6:
                grand_row       = editable_table[6]
                grand_ibrm      = grand_row[1] if len(grand_row) > 1 else ""
                grand_cbrm      = grand_row[2] if len(grand_row) > 2 else ""
                grand_fluxes    = grand_row[3] if len(grand_row) > 3 else ""
                grand_slag      = grand_row[4] if len(grand_row) > 4 else ""
                grand_clinker   = grand_row[5] if len(grand_row) > 5 else ""
                grand_total_val = grand_row[6] if len(grand_row) > 6 else ""
        except Exception as e:
            print("GRAND TOTAL ERROR =", str(e))

        for (col_offset, val) in [
            (1,  grand_ibrm),
            (6,  grand_cbrm),
            (16, grand_fluxes),
            (19, grand_slag),
            (20, grand_clinker),
        ]:
            c = ws.cell(grand_total_row, CA_START_COL + col_offset, val)
            c.font      = _font()
            c.alignment = _ctr
            c.border    = thick_bdr

        c = ws.cell(grand_total_row, total_col, grand_total_val)
        c.font      = _font()
        c.alignment = _ctr
        c.border    = thick_bdr

        for col in range(CA_START_COL, total_col + 1):
            cell = ws.cell(grand_total_row, col)
            cell.border    = thick_bdr
            cell.alignment = _ctr
            cell.font      = _font()

        # ── Auto-fit ALL cargo columns based on cargo name length ──────────────
        # Runs at the very end so it always wins over any earlier width setting
        ws.column_dimensions[get_column_letter(CA_START_COL)].width = 15  # label col

        col = CA_START_COL + 1
        for cargo in cargo_names:
            col_letter = get_column_letter(col)
            # width = full cargo name length + padding, minimum 12
            cargo_width = max(len(str(cargo)) + 4, 12)
            ws.column_dimensions[col_letter].width = cargo_width
            col += 1

        # Total col
        ws.column_dimensions[get_column_letter(col)].width = 12

        data_row = grand_total_row + 1
    # ── tide ──────────────────────────────────────────────────────────

    dashboard_row = current_row

    # Place Tide table BELOW Cargo Availability table
    if cargo_availability:
        tide_start_row = data_row + 2
    else:
        tide_start_row = current_row + 2

    # Move table 2 columns right
    TIDE_COL = 14

    ws.column_dimensions[get_column_letter(TIDE_COL)].width     = 30
    ws.column_dimensions[get_column_letter(TIDE_COL + 1)].width = 30

    safe_merge(ws, tide_start_row, TIDE_COL, tide_start_row, TIDE_COL + 1)
    for cc in range(TIDE_COL, TIDE_COL + 2):
        ws.cell(tide_start_row, cc).fill   = _fill("D9EAF7")
        ws.cell(tide_start_row, cc).border = thick_bdr
        ws.cell(tide_start_row, cc).font   = _title_font()
    ws.cell(tide_start_row, TIDE_COL).value     = "Tide - Dharamtar Port"
    ws.cell(tide_start_row, TIDE_COL).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[tide_start_row].height    = 18

    header_row = tide_start_row + 1
    ws.row_dimensions[header_row].height = 18
    for col_off, lbl in [(0, "Time"), (1, "Tide")]:
        c = safe_cell(ws, header_row, TIDE_COL + col_off, lbl)
        c.font      = _font()
        c.fill      = _fill("D9EAF7")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thick_bdr

    row_no = header_row + 1
    for tide in (tide_rows or []):
        tide_dt = ""
        if tide.get("tide_datetime"):
            try:
                dt = tide["tide_datetime"]
                if isinstance(dt, str):
                    dt = datetime.fromisoformat(dt)
                tide_dt = dt.strftime("%d/%H:%M")
            except Exception:
                tide_dt = str(tide["tide_datetime"])

        c = safe_cell(ws, row_no, TIDE_COL, tide_dt)
        c.font      = _font()
        c.border    = thick_bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_no].height = 18

        c = safe_cell(ws, row_no, TIDE_COL + 1, tide.get("tide_meters", ""))
        c.font      = _font()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thick_bdr
        row_no += 1

    # ── port throughput ───────────────────────────────────────────────
    PT_COL       = TIDE_COL + 3
    pt_start_row = tide_start_row - 1

    ws.column_dimensions[get_column_letter(PT_COL)].width     = 30
    ws.column_dimensions[get_column_letter(PT_COL + 1)].width = 30

    safe_merge(ws, pt_start_row, PT_COL, pt_start_row, PT_COL + 1)
    for cc in range(PT_COL, PT_COL + 2):
        ws.cell(pt_start_row, cc).fill   = _fill("D9EAF7")
        ws.cell(pt_start_row, cc).border = thick_bdr
        ws.cell(pt_start_row, cc).font   = _title_font()
    ws.cell(pt_start_row, PT_COL).value     = "Port Throughput"
    ws.cell(pt_start_row, PT_COL).alignment = Alignment(horizontal="center", vertical="center")

    port_throughput = port_throughput or {}
    pt_rows_data = [
        ("Jetty Throughput (Day)", port_throughput.get("day_qty", "")),
        ("Month",                  port_throughput.get("mtd_qty", "")),
        ("Year",                   port_throughput.get("ytd_qty", "")),
        ("Cumulative Since Oct 2012", port_throughput.get("cumulative_qty", "")),
        ("Month TPD",              f"{port_throughput.get('month_tpd', 0):,.2f}"),
        ("Year TPD",               f"{port_throughput.get('year_tpd', 0):,.2f}"),
    ]

    pt_row = pt_start_row + 1
    for label, value in pt_rows_data:
        c = safe_cell(ws, pt_row, PT_COL, label)
        c.font      = _font()
        c.border    = thick_bdr
        c.alignment = Alignment(horizontal="left", vertical="center")

        c = safe_cell(ws, pt_row, PT_COL + 1, value)
        c.font      = _font()  
        c.border    = thick_bdr
        c.alignment = Alignment(horizontal="right", vertical="center")
        pt_row += 1

    # ── MBC cargo handling ────────────────────────────────────────────
    MBC_COL       = PT_COL + 3
    mbc_start_row = tide_start_row - 1

    day_lookup = {
        (r["owner"], r["cargo_type"]): float(r["qty"] or 0)
        for r in (mbc_day_rows or [])
    }
    month_lookup = {
        (r["owner"], r["cargo_type"]): float(r["qty"] or 0)
        for r in (mbc_month_rows or [])
    }
    year_lookup = {
        (r["owner"], r["cargo_type"]): float(r["qty"] or 0)
        for r in (mbc_year_rows or [])
    }

    all_rows_mbc = (
        list(mbc_day_rows or []) +
        list(mbc_month_rows or []) +
        list(mbc_year_rows or [])
    )
    cargo_types = list(dict.fromkeys(r["cargo_type"] for r in all_rows_mbc))
    owners      = list(dict.fromkeys(r["owner"]      for r in all_rows_mbc))
    n_cargo     = len(cargo_types)

    DAY_COLS   = n_cargo + 1
    MTD_COLS   = n_cargo + 1
    YTD_COLS   = 1
    total_cols = 1 + DAY_COLS + MTD_COLS + YTD_COLS

    ws.column_dimensions[get_column_letter(MBC_COL)].width = 16
    for i in range(n_cargo):
        ws.column_dimensions[get_column_letter(MBC_COL + 1 + i)].width = 16

    day_start_col = MBC_COL + 1
    mtd_start_col = day_start_col + n_cargo + 1
    ytd_start_col = mtd_start_col + n_cargo + 1

    for i in range(n_cargo):
        ws.column_dimensions[get_column_letter(day_start_col + n_cargo)].width = 25
        ws.column_dimensions[get_column_letter(mtd_start_col + i)].width       = 16
    ws.column_dimensions[get_column_letter(mtd_start_col + n_cargo)].width     = 30
    ws.column_dimensions[get_column_letter(ytd_start_col)].width               = 25

    # title
    safe_merge(ws, mbc_start_row, MBC_COL, mbc_start_row, MBC_COL + total_cols - 1)
    for cc in range(MBC_COL, MBC_COL + total_cols):
        ws.cell(mbc_start_row, cc).fill   = _fill("D9EAF7")
        ws.cell(mbc_start_row, cc).border = thick_bdr
        ws.cell(mbc_start_row, cc).font   = _title_font()
    ws.cell(mbc_start_row, MBC_COL).value     = "MBC Cargo Handling"
    ws.cell(mbc_start_row, MBC_COL).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[mbc_start_row].height   = 18

    cargo_hdr_row = mbc_start_row + 1
    col_hdr_row   = cargo_hdr_row + 1

    # owner merged header
    safe_merge(ws, cargo_hdr_row, MBC_COL, col_hdr_row, MBC_COL)
    c = safe_cell(ws, cargo_hdr_row, MBC_COL, "Owner")
    c.font = _font(); c.fill = _fill("D9EAF7"); c.border = thick_bdr
    c.alignment = Alignment(horizontal="center", vertical="center")

    for label, start, end in [
        ("Day", day_start_col, day_start_col + n_cargo),
        ("MTD", mtd_start_col, mtd_start_col + n_cargo),
        ("YTD", ytd_start_col, ytd_start_col),
    ]:
        safe_merge(ws, cargo_hdr_row, start, cargo_hdr_row, end)
        c = safe_cell(ws, cargo_hdr_row, start, label)
        c.font = _font(); c.fill = _fill("D9EAF7"); c.border = thick_bdr
        c.alignment = Alignment(horizontal="center")

    for i, ct in enumerate(cargo_types):
        for base in [day_start_col, mtd_start_col]:
            c = safe_cell(ws, col_hdr_row, base + i, ct)
            c.font = _font(); c.fill = _fill("D9EAF7"); c.border = thick_bdr
            c.alignment = Alignment(horizontal="center")

    for base in [day_start_col + n_cargo, mtd_start_col + n_cargo, ytd_start_col]:
        c = safe_cell(ws, col_hdr_row, base, "Total")
        c.font = _font(); c.fill = _fill("D9EAF7"); c.border = thick_bdr
        c.alignment = Alignment(horizontal="center")

    for cc in range(MBC_COL, MBC_COL + total_cols):
        ws.cell(cargo_hdr_row, cc).border = thick_bdr
        ws.cell(col_hdr_row,   cc).border = thick_bdr

    data_start = col_hdr_row + 1
    for owner in owners:
        c = safe_cell(ws, data_start, MBC_COL, owner)
        c.border = thick_bdr; c.font = _font()

        day_total = month_total = 0
        for i, ct in enumerate(cargo_types):
            qty = day_lookup.get((owner, ct), 0)
            day_total += qty
            c = safe_cell(ws, data_start, day_start_col + i, qty if qty else "")
            c.border = thick_bdr; c.font = _font()

        c = safe_cell(ws, data_start, day_start_col + n_cargo, day_total if day_total else "")
        c.border = thick_bdr; c.font = _font()

        for i, ct in enumerate(cargo_types):
            qty = month_lookup.get((owner, ct), 0)
            month_total += qty
            c = safe_cell(ws, data_start, mtd_start_col + i, qty if qty else "")
            c.border = thick_bdr; c.font = _font()

        c = safe_cell(ws, data_start, mtd_start_col + n_cargo, month_total if month_total else "")
        c.border = thick_bdr; c.font = _font()

        year_total = sum(year_lookup.get((owner, ct), 0) for ct in cargo_types)
        c = safe_cell(ws, data_start, ytd_start_col, year_total if year_total else "")
        c.border = thick_bdr; c.font = _font()
        data_start += 1

    # total row
    for i, ct in enumerate(cargo_types):
        total_day = sum(day_lookup.get((o, ct), 0) for o in owners)
        c = safe_cell(ws, data_start, day_start_col + i, total_day if total_day else "")
        c.font = _font(); c.fill = _fill("F2F2F2"); c.border = thick_bdr

        total_mtd = sum(month_lookup.get((o, ct), 0) for o in owners)
        c = safe_cell(ws, data_start, mtd_start_col + i, total_mtd if total_mtd else "")
        c.font = _font(); c.fill = _fill("F2F2F2"); c.border = thick_bdr

    c = safe_cell(ws, data_start, day_start_col + n_cargo, sum(day_lookup.values()))
    c.font = _font(); c.fill = _fill("F2F2F2"); c.border = thick_bdr
    c = safe_cell(ws, data_start, mtd_start_col + n_cargo, sum(month_lookup.values()))
    c.font = _font(); c.fill = _fill("F2F2F2"); c.border = thick_bdr
    c = safe_cell(ws, data_start, ytd_start_col, sum(year_lookup.values()))
    c.font = _font(); c.fill = _fill("F2F2F2"); c.border = thick_bdr

    for cc in range(MBC_COL, MBC_COL + total_cols):
        ws.cell(data_start, cc).font   = _font()
        ws.cell(data_start, cc).fill   = _fill("F2F2F2")
        ws.cell(data_start, cc).border = thick_bdr
    ws.cell(data_start, MBC_COL).value = "Total"
    data_start += 1

    # ── cargo handled ─────────────────────────────────────────────────
    cargo_start_row = row_no + 5
    CARGO_COL = TIDE_COL

    day_dict_ch   = dict(day_rows   or [])
    month_dict_ch = dict(month_rows or [])

    all_routes = []
    for route, _ in (day_rows or []):
        if route not in all_routes: all_routes.append(route)
    for route, _ in (month_rows or []):
        if route not in all_routes: all_routes.append(route)

    safe_merge(ws, cargo_start_row, CARGO_COL, cargo_start_row, CARGO_COL + 2)
    for cc in range(CARGO_COL, CARGO_COL + 3):
        ws.cell(cargo_start_row, cc).fill   = _fill("D9EAF7")
        ws.cell(cargo_start_row, cc).border = thick_bdr
        ws.cell(cargo_start_row, cc).font   = _title_font()
    ws.cell(cargo_start_row, CARGO_COL).value     = "Cargo Handled"
    ws.cell(cargo_start_row, CARGO_COL).alignment = Alignment(horizontal="center", vertical="center")

    r = cargo_start_row + 1
    day_s = r
    day_e = day_s + len(all_routes)
    safe_merge(ws, day_s, CARGO_COL, day_e, CARGO_COL)
    c = safe_cell(ws, day_s, CARGO_COL, "For The Day")
    c.font = _font(); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thick_bdr

    for route in all_routes:
        qty = int(round(day_dict_ch.get(route, 0)))
        c = safe_cell(ws, r, CARGO_COL + 1, route)
        c.font = _font(); c.alignment = Alignment(horizontal="left", vertical="center"); c.border = thick_bdr
        c = safe_cell(ws, r, CARGO_COL + 2, qty)
        c.font = _font(); c.alignment = Alignment(horizontal="right", vertical="center"); c.border = thick_bdr
        r += 1

    c = safe_cell(ws, r, CARGO_COL + 1, "Total")
    c.font = _font(); c.border = thick_bdr
    c = safe_cell(ws, r, CARGO_COL + 2, int(round(sum(day_dict_ch.values()))))
    c.font = _font(); c.alignment = Alignment(horizontal="right", vertical="center"); c.border = thick_bdr
    r += 1

    month_s = r
    month_e = month_s + len(all_routes)
    safe_merge(ws, month_s, CARGO_COL, month_e, CARGO_COL)
    c = safe_cell(ws, month_s, CARGO_COL, "For The Month")
    c.font = _font(); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thick_bdr

    for route in all_routes:
        qty = int(round(month_dict_ch.get(route, 0)))
        c = safe_cell(ws, r, CARGO_COL + 1, route)
        c.font = _font(); c.alignment = Alignment(horizontal="left", vertical="center"); c.border = thick_bdr
        c = safe_cell(ws, r, CARGO_COL + 2, qty)
        c.font = _font(); c.alignment = Alignment(horizontal="right", vertical="center"); c.border = thick_bdr
        r += 1

    c = safe_cell(ws, r, CARGO_COL + 1, "Total")
    c.font = _font(); c.border = thick_bdr
    c = safe_cell(ws, r, CARGO_COL + 2, int(round(sum(month_dict_ch.values()))))
    c.font = _font(); c.alignment = Alignment(horizontal="right", vertical="center"); c.border = thick_bdr

    for rr in range(cargo_start_row, r + 1):
        for cc in range(CARGO_COL, CARGO_COL + 3):
            ws.cell(rr, cc).border = thick_bdr
    for rr in range(day_s, day_e + 1):
        ws.cell(rr, CARGO_COL).border = thick_bdr
    for rr in range(month_s, month_e + 1):
        ws.cell(rr, CARGO_COL).border = thick_bdr

    ws.column_dimensions[get_column_letter(CARGO_COL)].width     = 30
    ws.column_dimensions[get_column_letter(CARGO_COL + 1)].width = 30
    ws.column_dimensions[get_column_letter(CARGO_COL + 2)].width = 30

    cargo_end_row = r

    # ── cargo wise throughput ─────────────────────────────────────────
    throughput_rows = cargo_type_throughput or []

    THR_COL = CARGO_COL + 4
    THR_ROW = cargo_start_row

    # ================= COLUMN WIDTHS =================
    ws.column_dimensions[get_column_letter(THR_COL)].width = 45      # Cargo
    ws.column_dimensions[get_column_letter(THR_COL + 1)].width = 45  # Day
    ws.column_dimensions[get_column_letter(THR_COL + 2)].width = 45  # Month
    ws.column_dimensions[get_column_letter(THR_COL + 3)].width = 45  # YTD

    safe_merge(ws, THR_ROW, THR_COL, THR_ROW, THR_COL + 3)

    for cc in range(THR_COL, THR_COL + 4):
        ws.cell(THR_ROW, cc).fill = _fill("D9EAF7")
        ws.cell(THR_ROW, cc).border = thick_bdr
        ws.cell(THR_ROW, cc).font = _title_font()

    ws.cell(THR_ROW, THR_COL).value = "Cargo Wise Throughput"
    ws.cell(THR_ROW, THR_COL).alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    hdr_row = THR_ROW + 1

    for i, hdr in enumerate(["Cargo", "Day", "Month", "YTD"]):
        c = safe_cell(ws, hdr_row, THR_COL + i, hdr)
        c.font = _font()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thick_bdr

    r = hdr_row + 1

    total_day_t = 0
    total_month_t = 0
    total_year_t = 0

    steel_row = None
    avg_row = None

    for row in throughput_rows:

        cargo_type = (row["cargo_type"] or "").upper()

        if cargo_type == "STEEL PLANT CARGO":
            steel_row = row
            continue

        if cargo_type == "AVG (STEEL PLANT CARGO)":
            avg_row = row
            continue

        day_qty = int(float(row["day_qty"] or 0))
        month_qty = int(float(row["month_qty"] or 0))
        year_qty = int(float(row["year_qty"] or 0))

        total_day_t += day_qty
        total_month_t += month_qty
        total_year_t += year_qty

        c = safe_cell(ws, r, THR_COL, row["cargo_type"])
        c.font = _font()
        c.border = thick_bdr
        c.alignment = Alignment(horizontal="left", vertical="center")

        for off, val in [
            (1, day_qty),
            (2, month_qty),
            (3, year_qty)
        ]:
            c = safe_cell(
                ws,
                r,
                THR_COL + off,
                int(round(val)) if val else "-"
            )
            c.font = _font()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thick_bdr

        r += 1

    # ================= TOTAL =================
    c = safe_cell(ws, r, THR_COL, "Total")
    c.font = _font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thick_bdr

    for off, val in [
        (1, total_day_t),
        (2, total_month_t),
        (3, total_year_t)
    ]:
        c = safe_cell(ws, r, THR_COL + off, int(round(val)))
        c.font = _font(bold=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = thick_bdr

    r += 1

    # ================= STEEL PLANT CARGO =================
    if steel_row:

        c = safe_cell(ws, r, THR_COL, "Steel Plant Cargo")
        c.font = _font()
        c.border = thick_bdr
        c.alignment = Alignment(horizontal="left", vertical="center")

        for off, key in [
            (1, "day_qty"),
            (2, "month_qty"),
            (3, "year_qty")
        ]:
            val = int(float(steel_row[key] or 0))

            c = safe_cell(ws, r, THR_COL + off, val)
            c.font = _font()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thick_bdr

        r += 1

    # ================= AVG =================
    if avg_row:

        c = safe_cell(ws, r, THR_COL, "Avg (Steel Plant Cargo)")
        c.font = _font()
        c.border = thick_bdr
        c.alignment = Alignment(horizontal="left", vertical="center")

        for off, key in [
            (1, "day_qty"),
            (2, "month_qty"),
            (3, "year_qty")
        ]:
            val = int(float(avg_row[key] or 0))

            c = safe_cell(ws, r, THR_COL + off, val)
            c.font = _font()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thick_bdr

        r += 1
    # ── rainfall ──────────────────────────────────────────────────────
    RAINFALL_COL = THR_COL + 6
    RAINFALL_ROW = THR_ROW
    current_year = report_date.year - 1
    prev_year    = current_year - 1

    safe_merge(ws, RAINFALL_ROW, RAINFALL_COL, RAINFALL_ROW, RAINFALL_COL + 3)
    for cc in range(RAINFALL_COL, RAINFALL_COL + 4):
        ws.cell(RAINFALL_ROW, cc).fill   = _fill("D9EAF7")
        ws.cell(RAINFALL_ROW, cc).border = thick_bdr
        ws.cell(RAINFALL_ROW, cc).font   = _title_font()
    ws.cell(RAINFALL_ROW, RAINFALL_COL).value     = "Rainfall Details"
    ws.cell(RAINFALL_ROW, RAINFALL_COL).alignment = Alignment(horizontal="center", vertical="center")

    hdr_row = RAINFALL_ROW + 1
    for i, hdr in enumerate(["Year", "Period", "Rainfall", "Max."]):
        c = safe_cell(ws, hdr_row, RAINFALL_COL + i, hdr)
        c.font = _font(); c.fill = _fill("D9EAF7")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thick_bdr

    r = hdr_row + 1
    for yr_idx, (yr_key, periods_list, base_idx) in enumerate([
        (current_year, ["For The Day", "MTD",   "YTD"],   2),
        (prev_year,    ["For The Day", "Month", "Year"],  5),
    ]):
        safe_merge(ws, r, RAINFALL_COL, r + 2, RAINFALL_COL)
        yr_val = yr_key
        if rainfall_table:
            try: yr_val = rainfall_table[base_idx][0]
            except: pass
        c = safe_cell(ws, r, RAINFALL_COL, yr_val)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thick_bdr

        safe_merge(ws, r, RAINFALL_COL + 3, r + 2, RAINFALL_COL + 3)
        max_val = ""
        if rainfall_table:
            try: max_val = rainfall_table[base_idx][3]
            except: pass
        c = safe_cell(ws, r, RAINFALL_COL + 3, max_val)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); c.border = thick_bdr

        for i, period in enumerate(periods_list):
            c = safe_cell(ws, r + i, RAINFALL_COL + 1, period)
            c.border = thick_bdr
            rf_val = ""
            if rainfall_table:
                try:
                    rd = rainfall_table[base_idx + i]
                    rf_val = rd[2] if len(rd) >= 3 else (rd[1] if len(rd) >= 2 else "")
                except: pass
            c = safe_cell(ws, r + i, RAINFALL_COL + 2, rf_val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = thick_bdr
        r += 3

        last_row = r - 1

        for rr in range(RAINFALL_ROW, last_row + 1):
            for cc in range(RAINFALL_COL, RAINFALL_COL + 4):
                ws.cell(rr, cc).font   = _font()
                ws.cell(rr, cc).border = thick_bdr

    for i in range(4):
        ws.column_dimensions[get_column_letter(RAINFALL_COL + i)].width = 45

    # ── cargo statistics ──────────────────────────────────────────────
    STAT_COL = RAINFALL_COL + 6
    STAT_ROW = THR_ROW

    cargo_stats_day   = cargo_stats_day   or []
    cargo_stats_month = cargo_stats_month or []
    day_dict_s   = dict(cargo_stats_day)
    month_dict_s = dict(cargo_stats_month)

    sources = []
    for src, _ in cargo_stats_day:
        if src not in sources: sources.append(src)
    for src, _ in cargo_stats_month:
        if src not in sources: sources.append(src)

    safe_merge(ws, STAT_ROW, STAT_COL, STAT_ROW, STAT_COL + 2)
    for cc in range(STAT_COL, STAT_COL + 3):
        ws.cell(STAT_ROW, cc).fill   = _fill("D9EAF7")
        ws.cell(STAT_ROW, cc).border = thick_bdr
        ws.cell(STAT_ROW, cc).font   = _title_font()
    ws.cell(STAT_ROW, STAT_COL).value     = "Cargo Statistics"
    ws.cell(STAT_ROW, STAT_COL).alignment = Alignment(horizontal="center", vertical="center")

    hdr_row = STAT_ROW + 1
    for idx, hdr in enumerate(["Source", "Day", "MTD"]):
        c = safe_cell(ws, hdr_row, STAT_COL + idx, hdr)
        c.font = _font(); c.fill = _fill("D9EAF7")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thick_bdr

    r = hdr_row + 1
    total_day_s = total_month_s = 0
    for src in sources:
        day_qty   = float(day_dict_s.get(src, 0)   or 0)
        month_qty = float(month_dict_s.get(src, 0) or 0)
        total_day_s += day_qty; total_month_s += month_qty

        c = safe_cell(ws, r, STAT_COL, src)
        c.font = _font(); c.border = thick_bdr
        for off, qty in [(1, day_qty), (2, month_qty)]:
            c = safe_cell(ws, r, STAT_COL + off, int(round(qty)) if qty else "")
            c.font = _font(); c.alignment = Alignment(horizontal="right"); c.border = thick_bdr
        r += 1

    c = safe_cell(ws, r, STAT_COL, "Total")
    c.font = _font(); c.fill = _fill("F2F2F2"); c.border = thick_bdr
    for off, tot in [(1, total_day_s), (2, total_month_s)]:
        c = safe_cell(ws, r, STAT_COL + off, int(round(tot)))
        c.font = _font(); c.fill = _fill("F2F2F2"); c.alignment = Alignment(horizontal="right"); c.border = thick_bdr

    for i in range(3):
        ws.column_dimensions[get_column_letter(STAT_COL + i)].width = 30
    for rr in range(STAT_ROW, r + 1):
        for cc in range(STAT_COL, STAT_COL + 3):
            ws.cell(rr, cc).border = thick_bdr
            ws.cell(rr, cc).font   = _font()

    # ── BF production details ─────────────────────────────────────────
    BF_COL = STAT_COL
    BF_ROW = r + 3

    ws.column_dimensions[get_column_letter(BF_COL)].width     = 20
    ws.column_dimensions[get_column_letter(BF_COL + 1)].width = 70
    ws.column_dimensions[get_column_letter(BF_COL + 2)].width = 70

    safe_merge(ws, BF_ROW, BF_COL, BF_ROW, BF_COL + 2)
    for cc in range(BF_COL, BF_COL + 3):
        ws.cell(BF_ROW, cc).fill   = _fill("D9EAF7")
        ws.cell(BF_ROW, cc).border = thick_bdr
        ws.cell(BF_ROW, cc).font   = _title_font()
    ws.cell(BF_ROW, BF_COL).value     = "BF Production Details"
    ws.cell(BF_ROW, BF_COL).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[BF_ROW].height  = 24

    hdr_row = BF_ROW + 1
    for i, hdr in enumerate(["Plant", "Target Production (TPD)", "Actual Production (TPD)"]):
        c = safe_cell(ws, hdr_row, BF_COL + i, hdr)
        c.font = _font(); c.fill = _fill("D9EAF7"); c.border = thick_bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 35

    for row_off, plant_name, tbl_idx in [(1, "BF1", 1), (2, "BF2", 2)]:
        rr = hdr_row + row_off
        c = safe_cell(ws, rr, BF_COL, plant_name)
        c.font = _font(); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thick_bdr

        target = actual = ""
        if bf_table:
            try: target = bf_table[tbl_idx][1]
            except: pass
            try: actual = bf_table[tbl_idx][2]
            except: pass

        for off, val in [(1, target), (2, actual)]:
            c = safe_cell(ws, rr, BF_COL + off, val)
            c.font = _font(); c.border = thick_bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[rr].height = 22

    for rr in range(BF_ROW, hdr_row + 3):
        for cc in range(BF_COL, BF_COL + 3):
            ws.cell(rr, cc).border = thick_bdr
            ws.cell(rr, cc).font   = _font()

    # ── RM stock details ──────────────────────────────────────────────
    RM_COL = BF_COL + 6
    RM_ROW = BF_ROW - 5   # adjust based on the gap between Cargo Statistics and BF Production Details

    ws.column_dimensions[get_column_letter(RM_COL)].width     = 22
    ws.column_dimensions[get_column_letter(RM_COL + 1)].width = 40

    safe_merge(ws, RM_ROW, RM_COL, RM_ROW, RM_COL + 1)
    for cc in range(RM_COL, RM_COL + 2):
        ws.cell(RM_ROW, cc).fill   = _fill("D9EAF7")
        ws.cell(RM_ROW, cc).border = thick_bdr
        ws.cell(RM_ROW, cc).font   = _title_font()
    ws.cell(RM_ROW, RM_COL).value     = "RM Stock Details"
    ws.cell(RM_ROW, RM_COL).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[RM_ROW].height  = 24

    hdr_row = RM_ROW + 1
    for i, hdr in enumerate(["Material", "Qty (LMT)"]):
        c = safe_cell(ws, hdr_row, RM_COL + i, hdr)
        c.font = _font(); c.fill = _fill("D9EAF7"); c.border = thick_bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 30

    rm_items = [
        ("IBRM",  1), ("CBRM", 2), ("FLUXES", 3), ("TOTAL", 4)
    ]
    for off, (material, tbl_idx) in enumerate(rm_items):
        rr  = hdr_row + 1 + off
        qty = ""
        if rm_table:
            try: qty = rm_table[tbl_idx][1]
            except: pass

        c = safe_cell(ws, rr, RM_COL, material)
        c.font = _font(); c.border = thick_bdr
        c.alignment = Alignment(horizontal="left", vertical="center")
        if material == "TOTAL":
            c.fill = _fill("F2F2F2")

        c = safe_cell(ws, rr, RM_COL + 1, qty)
        c.font = _font(); c.border = thick_bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        if material == "TOTAL":
            c.fill = _fill("F2F2F2")
        ws.row_dimensions[rr].height = 22

    for rr in range(RM_ROW, hdr_row + 5):
        for cc in range(RM_COL, RM_COL + 2):
            ws.cell(rr, cc).border = thick_bdr
            ws.cell(rr, cc).font   = _font()

    # ── upcoming MBCs ─────────────────────────────────────────────────
    current_row = dashboard_row
    upcoming_mbcs = upcoming_mbcs or []
    

    # MBC table column widths — 1 col each, wider for name/owner/cargo/date/status
    MBC_COL_WIDTHS = [58, 41, 41, 23, 20, 20, 20, 49, 52]
    # MBC Name, Owner, Cargo, Qty, FWD, MID, AFT, Date, Status  → 9 cols
    MBC_HEADERS = ["MBC Name", "Owner", "Cargo", "Qty (MT)", "FWD", "MID", "AFT", "Date", "Status"]
    MBC_TOTAL   = len(MBC_HEADERS)  # 9

    for i, w in enumerate(MBC_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(LABEL_START + i)].width = w

    # Title row
    ws.merge_cells(start_row=current_row, start_column=LABEL_START,
                   end_row=current_row,   end_column=LABEL_START + MBC_TOTAL - 1)
    c = ws.cell(current_row, LABEL_START, "Upcoming MBC's")
    c.font = _title_font(); c.fill = _fill("D9EAF7")
    c.alignment = _ctr; c.border = thick_bdr
    for cc in range(LABEL_START, LABEL_START + MBC_TOTAL):
        ws.cell(current_row, cc).border = thick_bdr
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Header row — one col per header, no merging
    for col_i, h in enumerate(MBC_HEADERS):
        c = ws.cell(current_row, LABEL_START + col_i, h)
        c.font = _font(); c.fill = _fill("D9EAF7")
        c.alignment = _ctr; c.border = thick_bdr
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Data rows
    # =====================================================
    # Move Owner='Other' records to bottom
    # =====================================================
    # Custom status order
    status_order = {
        "ETA DHARAMTAR": 1,
        "WAITING AT GULL": 2,
        "ETA GULL": 3
    }

    upcoming_mbcs = sorted(
        upcoming_mbcs,
        key=lambda m: (
            str(m.get("owner", "")).strip().upper() == "OTHERS",  # OTHERS always last
            status_order.get(
                str(m.get("status", "")).strip().upper(),
                99
            ),
            str(m.get("event_date", ""))
        )
    )

# =====================================================
# Data rows
# =====================================================
    for m in upcoming_mbcs:

        event_date_fmt = ""

        if m.get("event_date"):
            try:
                event_date_fmt = datetime.fromisoformat(
                    str(m["event_date"])
                ).strftime("%d-%m-%Y %H:%M")
            except:
                event_date_fmt = str(m["event_date"])

        row_vals = [
            (m.get("mbc_name", "") or "").replace("JSW ", "").strip(),
            m.get("owner", "") or "-",
            m.get("cargo_name", "") or "-",
            int(round(float(m.get("bl_quantity") or 0))) or "-",
            m.get("fwd_draft", "") or "-",
            m.get("mid_draft", "") or "-",
            m.get("aft_draft", "") or "-",
            event_date_fmt or "-",
            m.get("status", "") or "-",
        ]

        for col_i, val in enumerate(row_vals):
            c = ws.cell(current_row, LABEL_START + col_i, val)
            c.font = _font()
            c.fill = _fill("FFFFFF")
            c.alignment = _left
            c.border = thick_bdr

        ws.row_dimensions[current_row].height = 30
        current_row += 1
    current_row += 2

    # ── MBC status ────────────────────────────────────────────────────
    # =====================================================
    # MBC Status
    # =====================================================
    mbc_status_rows = mbc_status_rows or []

    # Bring these MBCs to the top
    priority_mbcs = [
        "JSW RAIGAD",
        "JSW PRATAPGAD",
        "JSW LOHGAD",
        "JSW SINHGAD"
    ]

    mbc_status_rows = sorted(
        mbc_status_rows,
        key=lambda r: (
            0 if (r.get("mbc_name", "").strip().upper() in priority_mbcs)
            else 1,
            priority_mbcs.index(r.get("mbc_name", "").strip().upper())
            if r.get("mbc_name", "").strip().upper() in priority_mbcs
            else 999,
            r.get("mbc_name", "").strip().upper()
        )
    )

    ws.merge_cells(
        start_row=current_row,
        start_column=LABEL_START,
        end_row=current_row,
        end_column=LABEL_START + MBC_TOTAL - 1
    )

    c = ws.cell(current_row, LABEL_START, "MBC Status")
    c.font = _title_font()
    c.fill = _fill("D9EAF7")
    c.alignment = _ctr
    c.border = thick_bdr

    for cc in range(LABEL_START, LABEL_START + MBC_TOTAL):
        ws.cell(current_row, cc).border = thick_bdr

    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # MBC Name = 3 cols, Status = remaining cols
    ms_spans = [3, MBC_TOTAL - 3]
    ms_headers = ["MBC Name", "Status"]

    col = LABEL_START

    for h, span in zip(ms_headers, ms_spans):

        ws.merge_cells(
            start_row=current_row,
            start_column=col,
            end_row=current_row,
            end_column=col + span - 1
        )

        c = ws.cell(current_row, col, h)
        c.font = _font()
        c.fill = _fill("D9EAF7")
        c.alignment = _ctr
        c.border = thick_bdr

        for cc in range(col, col + span):
            ws.cell(current_row, cc).border = thick_bdr

        col += span

    ws.row_dimensions[current_row].height = 30
    current_row += 1

    for row in mbc_status_rows:

        col = LABEL_START

        row_values = [
            (row.get("mbc_name", "") or "").replace("JSW ", "").strip(),
            row.get("mbc_status", "") or ""
        ]

        for val, span in zip(row_values, ms_spans):

            ws.merge_cells(
                start_row=current_row,
                start_column=col,
                end_row=current_row,
                end_column=col + span - 1
            )

            c = ws.cell(current_row, col, val)
            c.font = _font()
            c.fill = _fill("FFFFFF")
            c.alignment = _left
            c.border = thick_bdr

            for cc in range(col, col + span):
                ws.cell(current_row, cc).border = thick_bdr

            col += span

        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2
    # ── uniform row height ────────────────────────────────────────────
    for row_num in range(1, ws.max_row + 1):

        label = ws.cell(row_num, LABEL_START).value

        if label in (
            "Discharge At Jetty",
            "Waiting For Discharge At Jetty",
            "At Gull - Waiting (Loaded)",
            "Under Loading"
        ):
            ws.row_dimensions[row_num].height = 400
        else:
            ws.row_dimensions[row_num].height = 50

    # ── page setup ────────────────────────────────────────────────────
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Allow Excel to use 2 pages horizontally instead of shrinking everything
    # ws.sheet_properties.pageSetUpPr.fitToPage = True
    # ws.page_setup.fitToWidth = 2
    # ws.page_setup.fitToHeight = 1

    # Small margins
    ws.page_margins = PageMargins(
        left=0.05,
        right=0.05,
        top=0.05,
        bottom=0.05,
        header=0,
        footer=0
    )

    # FORCE HEIGHT FOR THESE TWO ROWS
    for row in range(1, ws.max_row + 1):

        label = ws.cell(row, LABEL_START).value

        if label in (
            "Discharge At Jetty",
            "Waiting For Discharge At Jetty",
            "At Gull - Waiting (Loaded)",
            "Under Loading"
        ):
            ws.row_dimensions[row].height = 400

        else:
            ws.row_dimensions[row].height = 40
    # Center horizontally only
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # Excel view zoom
    ws.sheet_view.zoomScale = 100

    # Recalculate print area AFTER entire sheet is built
    max_row = ws.max_row
    max_col = ws.max_column

    print("MAX ROW =", max_row)
    print("MAX COL =", max_col)

    ws.print_area = (
        f"A1:{get_column_letter(max_col)}{max_row}"
    )

    print("PRINT AREA =", ws.print_area)
    print("MAX COL =", ws.max_column)
    print("LAST COL =", get_column_letter(ws.max_column))
    print("MAX ROW =", ws.max_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
# def _fetch_mbc_cargo(report_date):
#     """Return (day_data, month_data) as dicts { owner: { cargo_type: qty } }.
#     Month values incorporate cutoff if the cutoff date falls within the report month.
#     """
#     from datetime import date as date_type
#     prev_date = report_date - timedelta(days=1)
#     day_str   = prev_date.strftime('%Y-%m-%d')
#     month_start_date = date_type(prev_date.year, prev_date.month, 1)

#     # ── Load cutoff ─────────────────────────────────────────────────────
#     cutoff_date_str, cutoff_vals = _load_cutoff()
#     mbc_cutoff = cutoff_vals.get('mbc_cargo', {})

#     cutoff_date_obj = None
#     if cutoff_date_str and mbc_cutoff:
#         try:
#             cutoff_date_obj = datetime.strptime(cutoff_date_str, '%Y-%m-%d').date()
#         except ValueError:
#             pass

#     use_cutoff = (cutoff_date_obj is not None
#                   and month_start_date < cutoff_date_obj
#                   and cutoff_date_obj <= prev_date)

#     conn = get_db()
#     cur  = get_cursor(conn)

#     def _period(date_from, date_to):
#         cur.execute("""
#             SELECT COALESCE(m.mbc_owner_name, 'OTHERS') AS owner,
#                    h.cargo_type,
#                    SUM(h.bl_quantity) AS qty
#             FROM mbc_header h
#             LEFT JOIN mbc_master m ON m.mbc_name = h.mbc_name
#             WHERE h.doc_date IS NOT NULL
#               AND h.doc_date >= %s
#               AND h.doc_date <= %s
#             GROUP BY owner, h.cargo_type
#         """, (date_from, date_to))
#         data = {o: {ct: 0.0 for ct in _MBC_CARGO_TYPES} for o in _MBC_OWNERS}
#         for r in cur.fetchall():
#             owner = r['owner'] if r['owner'] in _MBC_OWNERS else 'OTHERS'
#             ct    = r['cargo_type']
#             if ct in _MBC_CARGO_TYPES:
#                 data[owner][ct] += float(r['qty'] or 0)
#         return data

#     day_data = _period(day_str, day_str)

#     if use_cutoff:
#         # Query only from the day after cutoff onwards
#         live_from = (cutoff_date_obj + timedelta(days=1)).strftime('%Y-%m-%d')
#         live_data = _period(live_from, day_str)
#         # Merge cutoff + live
#         month_data = {o: {ct: 0.0 for ct in _MBC_CARGO_TYPES} for o in _MBC_OWNERS}
#         for owner in _MBC_OWNERS:
#             for ct in _MBC_CARGO_TYPES:
#                 co_key = f'{owner}|{ct}'
#                 co_val = float(mbc_cutoff.get(co_key, 0))
#                 live_val = live_data[owner][ct]
#                 month_data[owner][ct] = co_val + live_val
#     else:
#         mth_str = month_start_date.strftime('%Y-%m-%d')
#         month_data = _period(mth_str, day_str)

#     conn.close()
#     return day_data, month_data




# def _fmt_tide_dt(dt_str):
#     """'2026-01-27T16:00' -> '27/16:00'"""
#     try:
#         dt = datetime.fromisoformat(dt_str)
#         return dt.strftime('%d/%H:%M')
#     except Exception:
#         return dt_str


# ── Excel builder ───────────────────────────────────────────────────────────

# def _build_excel(vessels, report_date,
#                  day_rows=None, month_rows=None, tide_rows=None,
#                  mbc_day=None, mbc_month=None):
#     from openpyxl import Workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Daily Ops'
#     day_rows   = day_rows   or []
#     month_rows = month_rows or []
#     tide_rows  = tide_rows  or []
#     _empty_mbc = lambda: {o: {ct: 0.0 for ct in _MBC_CARGO_TYPES} for o in _MBC_OWNERS}
#     mbc_day    = mbc_day   or _empty_mbc()
#     mbc_month  = mbc_month or _empty_mbc()

#         # Dynamic column calculation
#     vessel_end_col = 1 + len(vessels)
#     doc_col = vessel_end_col + 1
#     issue_col = vessel_end_col + 2

#     # Dynamic widths
#     col_widths = {1: 30}

#     # Vessel columns
#     for i in range(len(vessels)):
#         col_widths[2 + i] = 35

#     # Extra columns
#     col_widths[doc_col] = 32
#     col_widths[issue_col] = 22

#     # Apply widths
#     for ci, w in col_widths.items():
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     def _cell(r, c, val='', bold=False, fill='FFFFFF', align=_ctr):
#         cell = ws.cell(r, c, val)
#         cell.font      = _font(bold=bold)
#         cell.fill      = _fill(fill)
#         cell.alignment = align
#         cell.border    = _bdr
#         return cell

#     def _merge_row(r, c1, c2, val='', bold=False, fill='FFFFFF', align=_ctr):
#         ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
#         for ci in range(c1, c2 + 1):
#             b = Border(
#                 left   = _thin if ci == c1 else None,
#                 right  = _thin if ci == c2 else None,
#                 top    = _thin,
#                 bottom = _thin,
#             )
#             try:
#                 cell        = ws.cell(r, ci)
#                 cell.fill   = _fill(fill)
#                 cell.border = b
#             except AttributeError:
#                 pass
#         anchor           = ws.cell(r, c1)
#         anchor.value     = val
#         anchor.font      = _font(bold=bold)
#         anchor.alignment = align

#     def _merge_col(r1, r2, c, val='', bold=False, fill='FFFFFF', align=_ctr):
#         ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
#         for ri in range(r1, r2 + 1):
#             b = Border(
#                 left   = _thin,
#                 right  = _thin,
#                 top    = _thin if ri == r1 else None,
#                 bottom = _thin if ri == r2 else None,
#             )
#             try:
#                 cell        = ws.cell(ri, c)
#                 cell.fill   = _fill(fill)
#                 cell.border = b
#             except AttributeError:
#                 pass
#         anchor           = ws.cell(r1, c)
#         anchor.value     = val
#         anchor.font      = _font(bold=bold)
#         anchor.alignment = align

#     date_str  = f"{report_date.day}.{report_date.month}.{report_date.year}"
#     title_str = f'Daily Report of JSW Dharamtar Port Operation : {date_str}'

#     # Row 1
#     # Row 1
#     ws.row_dimensions[1].height = 20

#     vessel_end_col = 1 + len(vessels)
#     doc_col = vessel_end_col + 1
#     issue_col = vessel_end_col + 2

#     _cell(1, 1, report_date.strftime('%d-%m-%Y'), align=_left)

#     _merge_row(1, 2, vessel_end_col, title_str, align=_ctr)

#     _cell(doc_col, 1)
#     _cell(1, doc_col, 'Doc No. | REV.02 | Issue no. 02', align=_left)

#     _cell(issue_col, 1)
#     _cell(1, issue_col, f'Issue Date: {report_date.strftime("%d-%m-%Y")}', align=_left)

#     # Row 2: vessel name headers
# # Row 2: vessel name headers
#     ws.row_dimensions[2].height = 20

#     _cell(2, 1, '')

#     for i, v in enumerate(vessels):
#         _cell(
#             2,
#             2 + i,
#             f'Vessel {i + 1}: {v["vessel_name"]}',
#             bold=True,
#             align=_ctr
#         )

#     # Empty cells after vessels
#     _cell(2, doc_col, '')
#     _cell(2, issue_col, '')

#     label_discharge = 'Unloaded till Date (LUEU)'
#     label_balance   = 'Balance'
#     label_commenced = 'Disch Commenced'
#     label_completed = 'Disch Completed'

#     _q = lambda x: int(round(x)) if x else ''
#     _n = lambda x: x if x else ''
#     ROWS = [
#         ('Stevedore/ Barge Group',          'stevedore_group',          None,       _left),
#         ('BL Qty',                          'bl_qty',                   _q,         _ctr),
#         ('24 hrs Discharge',                'ops_24h',                  _q,         _ctr),
#         (label_discharge,                   'ops_till',                 _q,         _ctr),
#         (label_balance,                     'balance',                  _q,         _ctr),
#         ('Vsl Arrived/NOR',                 'nor_tendered',             _fmt_dt,    _ctr),
#         (label_commenced,                   'discharge_commenced',      _fmt_dt,    _ctr),
#         (label_completed,                   'discharge_completed',      _fmt_dt,    _ctr),
#         (None, None, None, None),
#         ('No of Barges',                    'num_barges',               _n,         _ctr),
#         ('At Jetty',                        'at_jetty',                 _n,         _left),
#         ('Waiting for Discharge',           'waiting_discharge',        _n,         _left),
#         ('Waiting Empty at Jetty',          'waiting_empty_jetty',      _n,         _left),
#         ('In transit- MV/Gull to Jetty',    None,                       None,       _left),
#         ('At Gull- waiting (Loaded)',        'at_gull_loaded',           _n,         _left),
#         ('Under Loading at MV',             'under_loading',            _n,         _left),
#         ('Waiting for loading',             'waiting_loading',          _n,         _left),
#         ('In transit- from Jetty to MV',    'in_transit_jetty_to_mv',   _n,         _left),
#     ]

#     for idx, (label, field, formatter, align) in enumerate(ROWS):
#         r = 3 + idx
#         ws.row_dimensions[r].height = 18

#         if label is None:
#             for ci in range(1, 10):
#                 _cell(r, ci, '')
#             continue

#         _cell(r, 1, label, bold=True, align=_left)
#         for i, v in enumerate(vessels):
#             raw = v.get(field)
#             val = formatter(raw) if (formatter and raw is not None) else (raw or '')
#             _cell(r, 2 + i, val, align=align)
#             _cell(r, doc_col, '')
#             _cell(r, issue_col, '')

#     # ── Cargo Handled section ────────────────────────────────────────────────
#     cargo_start = 3 + len(ROWS)

#     def _cargo_section(row_start, period_rows, period_label):
#         r = row_start
#         n = len(period_rows) + 1
#         _merge_col(r, r + n - 1, 1, period_label, bold=True, align=_ctr)
#         for route_name, qty in period_rows:
#             _cell(r, 2, route_name, align=_left)
#             _cell(r, 3, int(round(qty)) if qty else '', align=_ctr)
#             for ci in range(4, 10):
#                 _cell(r, ci, '')
#             ws.row_dimensions[r].height = 18
#             r += 1
#         total = sum(q for _, q in period_rows)
#         _cell(r, 2, 'Total:', bold=True, align=_left)
#         _cell(r, 3, int(round(total)) if total else '', bold=True, align=_ctr)
#         for ci in range(4, 10):
#             _cell(r, ci, '')
#         ws.row_dimensions[r].height = 18
#         r += 1
#         return r

#     r = cargo_start
#     for ci in range(1, 10):
#         _cell(r, ci, '')
#     ws.row_dimensions[r].height = 18
#     r += 1
#     _merge_row(r, 1, 3, 'Cargo Handled', bold=True, align=_left)
#     for ci in range(4, 10):
#         _cell(r, ci, '')
#     ws.row_dimensions[r].height = 18
#     r += 1
#     r = _cargo_section(r, day_rows, 'For the Day')
#     r = _cargo_section(r, month_rows, 'For the Month')

#     # ── Tide — Dharamtar Port section ────────────────────────────────────────
#     for ci in range(1, 10):
#         _cell(r, ci, '')
#     ws.row_dimensions[r].height = 18
#     r += 1
#     _merge_row(r, 1, 2, 'Tide- Dharamtar Port', bold=False, align=_ctr)
#     for ci in range(3, 10):
#         ws.cell(r, ci).value = None
#     ws.row_dimensions[r].height = 18
#     r += 1
#     _cell(r, 1, 'Time', align=_ctr)
#     _cell(r, 2, 'Tide', align=_ctr)
#     ws.row_dimensions[r].height = 18
#     r += 1
#     for td_str, td_m in tide_rows:
#         _cell(r, 1, _fmt_tide_dt(td_str), align=_ctr)
#         _cell(r, 2, td_m, align=_ctr)
#         ws.row_dimensions[r].height = 18
#         r += 1

#     # ── MBC's Cargo Handling section ─────────────────────────────────────────
#     # Layout: col1=Owner | cols2-6=Day(BB,Container,Liquid,Bulk,Total)
#     #                     | cols7-11=MTD(BB,Container,Liquid,Bulk,Total)
#     MBC_TOTAL_COLS = 11

#     for ci in range(1, MBC_TOTAL_COLS + 1):
#         ws.cell(r, ci).value = None
#     ws.row_dimensions[r].height = 18
#     r += 1

#     _merge_row(r, 1, MBC_TOTAL_COLS, "MBC's Cargo Handling", bold=False, align=_ctr)
#     ws.row_dimensions[r].height = 18
#     r += 1

#     _merge_col(r, r + 1, 1, '', bold=False, align=_ctr)
#     _merge_row(r, 2, 6,              'Day', bold=False, align=_ctr)
#     _merge_row(r, 7, MBC_TOTAL_COLS, 'MTD', bold=False, align=_ctr)
#     ws.row_dimensions[r].height = 18
#     r += 1

#     # Sub-header row 2: cargo type labels + Total for both Day and MTD
#     for ci in range(2, 7):
#         label = (_MBC_CARGO_TYPES + ['Total'])[ci - 2]
#         _cell(r, ci, label, align=_ctr)
#     for ci in range(7, 12):
#         label = (_MBC_CARGO_TYPES + ['Total'])[ci - 7]
#         _cell(r, ci, label, align=_ctr)
#     ws.row_dimensions[r].height = 18
#     r += 1

#     # Widen col 11 for MTD Total
#     ws.column_dimensions[get_column_letter(11)].width = 12

#     totals_day   = {ct: 0.0 for ct in _MBC_CARGO_TYPES}
#     totals_month = {ct: 0.0 for ct in _MBC_CARGO_TYPES}
#     for owner in _MBC_OWNERS:
#         _cell(r, 1, owner, align=_ctr)
#         day_row   = mbc_day.get(owner,   {})
#         month_row = mbc_month.get(owner, {})
#         day_total = 0.0
#         mtd_total = 0.0
#         for idx, ct in enumerate(_MBC_CARGO_TYPES):
#             dv = day_row.get(ct, 0.0)
#             mv = month_row.get(ct, 0.0)
#             _cell(r, 2 + idx, int(round(dv)) if dv else '', align=_ctr)
#             _cell(r, 7 + idx, int(round(mv)) if mv else '', align=_ctr)
#             day_total        += dv
#             mtd_total        += mv
#             totals_day[ct]   += dv
#             totals_month[ct] += mv
#         _cell(r, 6, int(round(day_total)) if day_total else '', align=_ctr)
#         _cell(r, 11, int(round(mtd_total)) if mtd_total else '', align=_ctr)
#         ws.row_dimensions[r].height = 18
#         r += 1

#     # Grand total row
#     _cell(r, 1, 'Total', align=_ctr)
#     grand_day = 0.0
#     grand_mtd = 0.0
#     for idx, ct in enumerate(_MBC_CARGO_TYPES):
#         td = totals_day[ct]
#         tm = totals_month[ct]
#         _cell(r, 2 + idx, int(round(td)) if td else '', align=_ctr)
#         _cell(r, 7 + idx, int(round(tm)) if tm else '', align=_ctr)
#         grand_day += td
#         grand_mtd += tm
#     _cell(r, 6, int(round(grand_day)) if grand_day else '', align=_ctr)
#     _cell(r, 11, int(round(grand_mtd)) if grand_mtd else '', align=_ctr)
#     ws.row_dimensions[r].height = 18
#     r += 1

#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return buf


@bp.route('/api/module/RP01/daily-ops/preview')
@login_required
def daily_ops_preview():

    date_str = request.args.get(
        'report_date',
        date.today().strftime('%Y-%m-%d')
    )

    try:
        report_date = datetime.strptime(
            date_str,
            '%Y-%m-%d'
        ).date()
    except ValueError:
        return Response('Invalid date', status=400)

    vessels = _fetch_data(report_date)

    html = """
<div style="
    width:100%;
    overflow-x:auto;
    overflow-y:hidden;
">
<table style="
    border-collapse:collapse;
    font-family:Arial;
    table-layout:fixed;
    min-width:max-content;
">
    <tr style='background:#4a90d9;color:white'>
        <th style="
            border:1px solid #ccc;
            padding:8px;
            min-width:220px;
            width:220px;
            position:sticky;
            left:0;
            background:#4a90d9;
            z-index:10;
        ">
            Parameter
        </th>
"""

    for i, v in enumerate(vessels):

        cargo_name = v.get('cargo_name', '')

        html += f"""
            <th style="
                border:1px solid #ccc;
                padding:8px;
                min-width:280px;
                width:280px;
                max-width:280px;
                word-wrap:break-word;
                white-space:normal;
            ">
                Vessel {i+1}<br>
                {v['vessel_name']}<br>
                <span style="font-size:12px;font-weight:normal;color:#f5f5f5;">
                    {cargo_name}
                </span>
            </th>
        """

    html += "</tr>"

    rows = [
    ("Stevedore Group", "stevedore_group"),
    ("BL Qty", "bl_qty"),
    ("24 Hrs Discharge", "ops_24h"),
    ("Unloaded Till Date", "ops_till"),
    ("Balance", "balance"),
    ("Vsl Arrived/NOR", "nor_tendered"),
    ("Disch Commenced", "discharge_commenced"),
    ("Disch Completed", "discharge_completed"),
    ("No Of Barges", "num_barges"),
    ("At Jetty", "at_jetty"),
    ("At Jetty Waiting Discharge", "waiting_discharge"),
    ("Waiting Empty At Jetty", "waiting_empty_jetty"),
    ("At Gull-waiting(Loaded)", "at_gull_loaded"),
    ("Under Loading", "under_loading"),
    ("Waiting Loading", "waiting_loading"),
    ("In Transit Jetty To MV", "in_transit_jetty_to_mv"),
    ("Breakdown/Offhire/Costal/DD", "breakdown")
    ]

    for label, field in rows:

        html += f"""
        <tr>
            <td style="
                border:1px solid #ccc;
                padding:8px;
                font-weight:bold;
                min-width:220px;
                width:220px;
                position:sticky;
                left:0;
                background:white;
                z-index:5;
            ">
                {label}
            </td>
        """

        for v in vessels:

            value = v.get(field, '')

            if field in (
                'nor_tendered',
                'discharge_commenced',
                'discharge_completed'
            ):
                value = _fmt_dt(value)

            html += f"""
            <td style="
                border:1px solid #ccc;
                padding:8px;
                min-width:280px;
                width:280px;
                max-width:280px;
                vertical-align:top;
                white-space:normal;
                word-break:break-word;
            ">
                {value}
            </td>
            """

        html += "</tr>"

    html += """
    </table>
    </div>
    """

    upcoming_vessels = _fetch_upcoming_vessels(report_date)

    html += """
    <br><br>
    <h3>Upcoming Vessels</h3>
    <p>


    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
        <tr style='background:#4a90d9;color:white'>
            <th style='border:1px solid #ccc;padding:8px'>Vessel Name</th>
            <th>Cargo</th>
            <th>Qty (MT)</th>
            <th style='border:1px solid #ccc;padding:8px'>Vessel Agent</th>
            <th style='border:1px solid #ccc;padding:8px'>ETA</th>
        </tr>
    """

    for v in upcoming_vessels:
        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>{v['vessel_name']}</td>
            <td style='border:1px solid #ccc;padding:8px'>{v['cargo_name'] or '-'}</td>
            <td style='border:1px solid #ccc;padding:8px'>{v['bl_quantity'] or '-'}</td>
            <td style='border:1px solid #ccc;padding:8px'>{v['vessel_agent_name']}</td>
            <td style='border:1px solid #ccc;padding:8px'>{v['eta']}</td>
        </tr>
        """

    html += "</table>"

    discharging_mbcs = _fetch_discharging_mbcs(report_date)

    html += """
    <br><br>
    <h3>MBCs Discharging</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
        <tr style='background:#4a90d9;color:white'>
            <th style='border:1px solid #ccc;padding:8px'>MBC Name</th>
            <th style='border:1px solid #ccc;padding:8px'>Equipment</th>
            <th style='border:1px solid #ccc;padding:8px'>Cargo Name</th>
            <th style='border:1px solid #ccc;padding:8px'>Discharge Quantity (MT)</th>
        </tr>
    """

    for m in discharging_mbcs:

        equipment = m['equipment'] or '-'

        if m['status'] == 'DISCHARGING':
            row_style = "background-color:#fff3cd;"   # Yellow

        elif m['status'] == 'ARRIVED':
            row_style = "background-color:#f8d7da;"   # Light Red

        else:
            row_style = ""

        html += f"""
        <tr style="{row_style}">
            <td style='border:1px solid #ccc;padding:8px'>
                {m['mbc_name']}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:center'>
                {equipment}
            </td>

            <td style='border:1px solid #ccc;padding:8px'>
                {m['cargo_name']}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(m['discharge_quantity']):,.2f}
            </td>
        </tr>
        """

    html += "</table>"

    upcoming_mbcs = _fetch_upcoming_mbcs(report_date)

    html += """
    <br><br>
    <h3>Upcoming MBCs</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
        <tr style='background:#4a90d9;color:white'>
            <th style='border:1px solid #ccc;padding:8px'>MBC Name</th>
            <th style='border:1px solid #ccc;padding:8px'>Owner</th>
            <th style='border:1px solid #ccc;padding:8px'>Cargo Name</th>
            <th style='border:1px solid #ccc;padding:8px'>Quantity (MT)</th>
            <th style='border:1px solid #ccc;padding:8px'>FWD</th>
            <th style='border:1px solid #ccc;padding:8px'>MID</th>
            <th style='border:1px solid #ccc;padding:8px'>AFT</th>
            <th style='border:1px solid #ccc;padding:8px'>Date</th>
            <th style='border:1px solid #ccc;padding:8px'>Status</th>
        </tr>
    """

    for m in upcoming_mbcs:

        row_color = "#d1ecf1" if m["status"] == "AT GULL" else "#fff3cd"

        html += f"""
        <tr style="background-color:{row_color};">
            <td style='border:1px solid #ccc;padding:8px'>
                {m['mbc_name']}
            </td>

            <td style='border:1px solid #ccc;padding:8px'>
                {m.get('owner', '-')}
            </td>

            <td style='border:1px solid #ccc;padding:8px'>
                {m['cargo_name']}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(m['bl_quantity']):,.2f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:center'>
                {m['fwd_draft'] if m['fwd_draft'] else '-'}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:center'>
                {m['mid_draft'] if m['mid_draft'] else '-'}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:center'>
                {m['aft_draft'] if m['aft_draft'] else '-'}
            </td>

            <td style='border:1px solid #ccc;padding:8px'>
                {datetime.fromisoformat(m['event_date']).strftime('%d-%m-%Y %H:%M')
                if m['event_date']
                    else '-'
                }
            </td>

            <td style='border:1px solid #ccc;padding:8px'>
                {m['status']}
            </td>
        </tr>
        """

    html += """
    </table>
    """

    html += """
<br><br>
    <h3>Cargo Availability for the Day</h3>
    <div style="overflow-x:auto;width:100%;">
    <table id="cargo-availability-table"
       style="border-collapse:collapse;font-family:Arial;font-size:12px;white-space:nowrap;">
    <tr style="background:#4a90d9;color:white;">
        <th style="border:1px solid #ccc;padding:8px;height:38px;"></th>

<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">BRBF</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Orissa Fines</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Goa Fines</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">HBI</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">KDL CLO</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Jimblebar Fines</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Bacheli Fines</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Goa Clo</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Mabu</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Illavara</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Uval + Kestrel</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">MLV</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">PCI</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Antracite</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Limestone</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Bentonite</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Oliflux</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Dolomite</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Slag Loading/Unloading</th>
<th contenteditable="true" style="border:1px solid #ccc;padding:8px;height:38px;">Clinker</th>

<th style="border:1px solid #ccc;padding:8px;height:38px;">Total</th>
    </tr>
    <tr>
        <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">At Jetty</td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_BRBF" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Orissa Fines" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Goa Fines" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_HBI" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_KDL CLO" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Jimblebar Fines" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Bacheli Fines" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Goa Clo" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Mabu" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Illavara" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Uval + Kestrel" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_MLV" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_PCI" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Antracite" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Limestone" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Bentonite" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Oliflux" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Dolomite" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Slag Loading/Unloading" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="at_jetty_Clinker" style="border:1px solid #ccc;padding:8px;text-align:right;min-width:80px;"></td>
        <td id="cargo-avail-total" style="border:1px solid #ccc;padding:8px;text-align:right;font-weight:bold;min-width:80px;"></td>
    </tr>
    <tr>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_label" style="border:1px solid #ccc;padding:8px;min-width:100px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_BRBF" style="border:1px solid #ccc;padding:8px;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Orissa Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Goa Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_HBI" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_KDL CLO" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Jimblebar Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Bacheli Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Goa Clo" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Mabu" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Illavara" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Uval + Kestrel" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_MLV" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_PCI" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Antracite" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Limestone" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Bentonite" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Oliflux" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Dolomite" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Slag Loading/Unloading" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row2_Clinker" style="border:1px solid #ccc;padding:8px;"></td>
        <td style="border:1px solid #ccc;padding:8px;"></td>
    </tr>
    <tr>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_label" style="border:1px solid #ccc;padding:8px;min-width:100px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_BRBF" style="border:1px solid #ccc;padding:8px;min-width:80px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Orissa Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Goa Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_HBI" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_KDL CLO" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Jimblebar Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Bacheli Fines" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Goa Clo" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Mabu" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Illavara" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Uval + Kestrel" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_MLV" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_PCI" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Antracite" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Limestone" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Bentonite" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Oliflux" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Dolomite" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Slag Loading/Unloading" style="border:1px solid #ccc;padding:8px;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="row3_Clinker" style="border:1px solid #ccc;padding:8px;"></td>
        <td style="border:1px solid #ccc;padding:8px;"></td>
    </tr>
    <tr style="background:#f2f2f2;font-weight:bold;">
        <td style="border:1px solid #ccc;padding:8px;">Total</td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_BRBF" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Orissa Fines" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Goa Fines" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_HBI" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_KDL CLO" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Jimblebar Fines" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Bacheli Fines" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Goa Clo" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Mabu" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Illavara" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Uval + Kestrel" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_MLV" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_PCI" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Antracite" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Limestone" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Bentonite" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Oliflux" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Dolomite" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Slag Loading/Unloading" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_Clinker" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
        <td contenteditable="true" data-section="cargo_avail" data-key="total_grand" style="border:1px solid #ccc;padding:8px;text-align:right;"></td>
    </tr>

    <tr style="background:#f2f2f2;font-weight:bold;">

    <tr style="background:#f2f2f2;font-weight:bold;">

    <td style="border:1px solid #ccc;padding:8px;"></td>

    <td colspan="5"
        id="grand_ibrm"
        contenteditable="true"
        data-section="cargo_avail"
        data-key="grand_ibrm"
        style="border:1px solid #ccc;padding:8px;text-align:center;">
    </td>

    <td colspan="10"
        id="grand_cbrm"
        contenteditable="true"
        data-section="cargo_avail"
        data-key="grand_cbrm"
        style="border:1px solid #ccc;padding:8px;text-align:center;">
    </td>

    <td colspan="3"
        id="grand_fluxes"
        contenteditable="true"
        data-section="cargo_avail"
        data-key="grand_fluxes"
        style="border:1px solid #ccc;padding:8px;text-align:center;">
    </td>

    <td id="grand_slag"
        contenteditable="true"
        data-section="cargo_avail"
        data-key="grand_slag"
        style="border:1px solid #ccc;padding:8px;text-align:center;">
    </td>

    <td id="grand_clinker"
        contenteditable="true"
        data-section="cargo_avail"
        data-key="grand_clinker"
        style="border:1px solid #ccc;padding:8px;text-align:center;">
    </td>

    <td id="grand_total"
        contenteditable="true"
        data-section="cargo_avail"
        data-key="grand_total"
        style="border:1px solid #ccc;padding:8px;text-align:center;">
    </td>

</tr>

    </table>
    </div>
    """

    tide_rows = _fetch_tide_data(report_date)

    html += """
    <br><br>
    <h3>Tide - Dharamtar Port</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    <tr style='background:#4a90d9;color:white'>
        <th style='border:1px solid #ccc;padding:8px'>Time</th>
        <th style='border:1px solid #ccc;padding:8px'>Tide (Meters)</th>
    </tr>
    """

    for row in tide_rows:

        tide_dt = datetime.fromisoformat(
            str(row['tide_datetime'])
        ).strftime('%d-%m-%Y %H:%M')

        tide_meters = f"{float(row.get('tide_meters') or 0):05.2f}"

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                {tide_dt}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {tide_meters}
            </td>
        </tr>
        """



    html += "</table>"

    day_rows, month_rows = _fetch_cargo_handled(report_date)

    html += """
    <br><br>
    <h3>Cargo Handled - For the Day</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    <tr style='background:#4a90d9;color:white'>
        <th style='border:1px solid #ccc;padding:8px'>Route</th>
        <th style='border:1px solid #ccc;padding:8px'>Quantity (MT)</th>
    </tr>
    """

    day_total = 0

    for route_name, qty in day_rows:

        day_total += qty

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                {route_name}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {qty:,.0f}
            </td>
        </tr>
        """

    html += f"""
    <tr style='font-weight:bold;background:#f2f2f2'>
        <td style='border:1px solid #ccc;padding:8px'>
            TOTAL
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {day_total:,.0f}
        </td>
    </tr>
    """

    html += "</table>"

    html += """
    <br><br>
    <h3>Cargo Handled - MTD</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    <tr style='background:#4a90d9;color:white'>
        <th style='border:1px solid #ccc;padding:8px'>Route</th>
        <th style='border:1px solid #ccc;padding:8px'>Quantity (MT)</th>
    </tr>
    """

    month_total = 0

    for route_name, qty in month_rows:

        month_total += qty

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                {route_name}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {qty:,.0f}
            </td>
        </tr>
        """

    html += f"""
    <tr style='font-weight:bold;background:#f2f2f2'>
        <td style='border:1px solid #ccc;padding:8px'>
            TOTAL
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {month_total:,.0f}
        </td>
    </tr>
    """

    html += "</table>"

    cargo_stat_day, cargo_stat_month = _fetch_cargo_statistics(report_date)

    html += """
    <br><br>
    <h3>Cargo Statistics - For the Day</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    <tr style='background:#4a90d9;color:white'>
        <th style='border:1px solid #ccc;padding:8px'>Location</th>
        <th style='border:1px solid #ccc;padding:8px'>Quantity (MT)</th>
    </tr>
    """

    total_qty = 0

    for source_name, qty in cargo_stat_day:

        if not source_name:
            continue

        total_qty += qty

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                {source_name}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {qty:,.0f}
            </td>
        </tr>
        """

    html += f"""
    <tr style='font-weight:bold'>
        <td style='border:1px solid #ccc;padding:8px'>
            Total
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {total_qty:,.0f}
        </td>
    </tr>
    """

    html += "</table>"

    mbc_day_rows, mbc_month_rows, mbc_year_rows = _fetch_mbc_cargo_handling(report_date)

    day_dict = {
        (r['owner'], r['cargo_type']): float(r['qty'])
        for r in mbc_day_rows
    }

    month_dict = {
        (r['owner'], r['cargo_type']): float(r['qty'])
        for r in mbc_month_rows
    }

    year_dict = {
        (r['owner'], r['cargo_type']): float(r['qty'])
        for r in mbc_year_rows
    }

    owners = sorted(
        set(owner for owner, cargo in day_dict.keys()) |
        set(owner for owner, cargo in month_dict.keys()) |
        set(owner for owner, cargo in year_dict.keys())
    )

    cargo_types = sorted(
        set(cargo for owner, cargo in day_dict.keys()) |
        set(cargo for owner, cargo in month_dict.keys())|
        set(cargo for owner, cargo in year_dict.keys())
    )

    html += """
    <br><br>
    <h3>MBC's - Cargo Handling</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    """

    # Header row 1
    html += f"""
    <tr style='background:#4a90d9;color:white'>
        <th rowspan='2' style='border:1px solid #ccc;padding:8px'>Owner</th>

        <th colspan='{len(cargo_types) + 1}'
            style='border:1px solid #ccc;padding:8px'>
            Day
        </th>

        <th colspan='{len(cargo_types) + 1}'
            style='border:1px solid #ccc;padding:8px'>
            MTD
        </th>
        <th colspan='{len(cargo_types) + 1}'
            style='border:1px solid #ccc;padding:8px'>
            YTD
        </th>
    </tr>
    """

    # Header row 2
    html += "<tr style='background:#4a90d9;color:white'>"

    for cargo in cargo_types:
        html += f"""
        <th style='border:1px solid #ccc;padding:8px'>
            {cargo}
        </th>
        """

    html += """
    <th style='border:1px solid #ccc;padding:8px'>Total</th>
    """

    for cargo in cargo_types:
        html += f"""
        <th style='border:1px solid #ccc;padding:8px'>
            {cargo}
        </th>
        """

    html += """
    <th style='border:1px solid #ccc;padding:8px'>Total</th>
    <th style='border:1px solid #ccc;padding:8px'>Total</th>
    </tr>
    """

    # Owner rows
    # Owner rows
    for owner in owners:

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                {owner}
            </td>
        """

        day_total = 0

        for cargo in cargo_types:

            qty = day_dict.get((owner, cargo), 0)
            day_total += qty

            display_qty = "-" if qty == 0 else format(qty, ",.0f")

            html += f"""
            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {display_qty}
            </td>
            """

        html += f"""
        <td style='border:1px solid #ccc;padding:8px;text-align:right;font-weight:bold'>
            {format(day_total, ",.0f")}
        </td>
        """

        month_total = 0

        for cargo in cargo_types:

            qty = month_dict.get((owner, cargo), 0)
            month_total += qty

            display_qty = "-" if qty == 0 else format(qty, ",.0f")

            html += f"""
            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {display_qty}
            </td>
            """

        year_total = sum(
        year_dict.get((owner, cargo), 0)
        for cargo in cargo_types
        )

        html += f"""
        <td style='border:1px solid #ccc;padding:8px;text-align:right;font-weight:bold'>
            {format(month_total, ",.0f")}
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right;font-weight:bold'>
            {format(year_total, ",.0f")}
        </td>
        </tr>
        """

   
    # Grand total row
    html += """
    <tr style='font-weight:bold;background:#f2f2f2'>
        <td style='border:1px solid #ccc;padding:8px'>
            Total
        </td>
    """

    for cargo in cargo_types:

        total = sum(
            day_dict.get((owner, cargo), 0)
            for owner in owners
        )

        html += f"""
        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {format(total, ",.0f")}
        </td>
        """

    html += f"""
    <td style='border:1px solid #ccc;padding:8px;text-align:right'>
        {format(sum(day_dict.values()), ",.0f")}
    </td>
    """

    for cargo in cargo_types:

        total = sum(
            month_dict.get((owner, cargo), 0)
            for owner in owners
        )

        html += f"""
        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {format(total, ",.0f")}
        </td>
        """

    html += f"""
    <td style='border:1px solid #ccc;padding:8px;text-align:right'>
        {format(sum(month_dict.values()), ",.0f")}
    </td>

    <td style='border:1px solid #ccc;padding:8px;text-align:right'>
        {format(sum(year_dict.values()), ",.0f")}
    </td>
    </tr>
    """

    html += "</table>"

    mbc_status_rows = _fetch_mbc_status(report_date)

    html += """
    <br><br>
    <h3>MBC Status</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    <tr style='background:#4a90d9;color:white'>
        <th style='border:1px solid #ccc;padding:8px'>MBC Name</th>
        <th style='border:1px solid #ccc;padding:8px'>Status</th>
    </tr>
    """

    for row in mbc_status_rows:

        html += f"""
    <tr>
        <td style='border:1px solid #ccc;padding:8px'>
            {row['mbc_name']}
        </td>

        <td style='border:1px solid #ccc;padding:8px'>
            {row['mbc_status']}
        </td>
    </tr>
    """
    html += "</table>"

    # Cargo Type Throughput

    cargo_type_rows = _fetch_cargo_type_throughput(report_date)

    html += """
    <br><br>
    <h3>Cargo Type Throughput</h3>

    <table style='width:100%;border-collapse:collapse;font-family:Arial'>
    <tr style='background:#4a90d9;color:white'>
        <th style='border:1px solid #ccc;padding:8px'>Cargo Type</th>
        <th style='border:1px solid #ccc;padding:8px'>Day Qty (MT)</th>
        <th style='border:1px solid #ccc;padding:8px'>MTD Qty (MT)</th>
        <th style='border:1px solid #ccc;padding:8px'>YTD Qty (MT)</th>
    </tr>
    """

    day_total = 0
    month_total = 0
    year_total = 0

    steel_row = None
    avg_row = None

    for row in cargo_type_rows:

        cargo_type = (row['cargo_type'] or '').upper()

        if cargo_type == "STEEL PLANT CARGO":
            steel_row = row
            continue

        if cargo_type == "AVG (STEEL PLANT CARGO)":
            avg_row = row
            continue

        day_qty = float(row['day_qty'] or 0)
        month_qty = float(row['month_qty'] or 0)
        year_qty = float(row['year_qty'] or 0)

        day_total += day_qty
        month_total += month_qty
        year_total += year_qty

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                {row['cargo_type']}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {day_qty:,.0f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {month_qty:,.0f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {year_qty:,.0f}
            </td>
        </tr>
        """

    # TOTAL ROW
    html += f"""
    <tr style='background:#f2f2f2;font-weight:bold'>
        <td style='border:1px solid #ccc;padding:8px;text-align:center'>
            Total
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {day_total:,.0f}
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {month_total:,.0f}
        </td>

        <td style='border:1px solid #ccc;padding:8px;text-align:right'>
            {year_total:,.0f}
        </td>
    </tr>
    """

    # STEEL PLANT CARGO
    if steel_row:

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                Steel Plant Cargo
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(steel_row['day_qty'] or 0):,.0f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(steel_row['month_qty'] or 0):,.0f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(steel_row['year_qty'] or 0):,.0f}
            </td>
        </tr>
        """

    # AVG (STEEL PLANT CARGO)
    if avg_row:

        html += f"""
        <tr>
            <td style='border:1px solid #ccc;padding:8px'>
                Avg (Steel Plant Cargo)
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(avg_row['day_qty'] or 0):,.0f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(avg_row['month_qty'] or 0):,.0f}
            </td>

            <td style='border:1px solid #ccc;padding:8px;text-align:right'>
                {float(avg_row['year_qty'] or 0):,.0f}
            </td>
        </tr>
        """

    html += "</table>"

    html += """
    <br><br>

    <div style="
        display:flex;
        align-items:flex-start;
        justify-content:flex-start;
        gap:20px;
        width:100%;
    ">

        <!-- RM STOCK DETAILS -->
        <div>

    <h3 style="margin-top:0;">RM Stock Details</h3>

    <table id="rm-stock-table"
           style="
           border-collapse:collapse;
           font-family:Arial;
           font-size:12px;
           width:250px;
    ">

        <tr style="background:#4a90d9;color:white;">

            <th style="
                border:1px solid #ccc;
                padding:8px;
                text-align:left;
            ">
                Material
            </th>

            <th style="
                border:1px solid #ccc;
                padding:8px;
                text-align:right;
            ">
                Qty (LMT)
            </th>

        </tr>

        <tr>

            <td style="
                border:1px solid #ccc;
                padding:8px;
            ">
                IBRM
            </td>

            <td contenteditable="true"
                style="
                    border:1px solid #ccc;
                    padding:8px;
                    text-align:right;
                    min-width:120px;
                ">
            </td>

        </tr>

        <tr>

            <td style="
                border:1px solid #ccc;
                padding:8px;
            ">
                CBRM
            </td>

            <td contenteditable="true"
                style="
                    border:1px solid #ccc;
                    padding:8px;
                    text-align:right;
                    min-width:120px;
                ">
            </td>

        </tr>

        <tr>

            <td style="
                border:1px solid #ccc;
                padding:8px;
            ">
                FLUXES
            </td>

            <td contenteditable="true"
                style="
                    border:1px solid #ccc;
                    padding:8px;
                    text-align:right;
                    min-width:120px;
                ">
            </td>

        </tr>

        <tr style="
            font-weight:bold;
            background:#f2f2f2;
        ">

            <td style="
                border:1px solid #ccc;
                padding:8px;
            ">
                TOTAL
            </td>

            <td contenteditable="true"
                style="
                    border:1px solid #ccc;
                    padding:8px;
                    text-align:right;
                    min-width:120px;
                ">
            </td>

        </tr>

    </table>

</div>

        <!-- BF PRODUCTION DETAILS -->
        <div>

            <h3 style="margin-top:0;">BF Production Details</h3>

            <table id="bf-production-table"
                style="
                border-collapse:collapse;
                font-family:Arial;
                font-size:12px;
                width:500px;
            ">

                <tr style="background:#4a90d9;color:white;">
                    <th style="border:1px solid #ccc;padding:8px;text-align:left;">
                        Plant
                    </th>

                    <th style="border:1px solid #ccc;padding:8px;text-align:right;">
                        Target Production (TPD)
                    </th>

                    <th style="border:1px solid #ccc;padding:8px;text-align:right;">
                        Actual Production (TPD)
                    </th>
                </tr>

                <tr>

                    <td style="border:1px solid #ccc;padding:8px;">
                        BF1
                    </td>

                    <td id="bf1-target"
                        contenteditable="true"
                        style="
                            border:1px solid #ccc;
                            padding:8px;
                            text-align:right;
                            min-width:120px;
                        ">
                    </td>

                    <td id="bf1-actual"
                        contenteditable="true"
                        style="
                            border:1px solid #ccc;
                            padding:8px;
                            text-align:right;
                            min-width:120px;
                        ">
                    </td>

                </tr>

                <tr>

                    <td style="border:1px solid #ccc;padding:8px;">
                        BF2
                    </td>

                    <td id="bf2-target"
                        contenteditable="true"
                        style="
                            border:1px solid #ccc;
                            padding:8px;
                            text-align:right;
                            min-width:120px;
                        ">
                    </td>

                    <td id="bf2-actual"
                        contenteditable="true"
                        style="
                            border:1px solid #ccc;
                            padding:8px;
                            text-align:right;
                            min-width:120px;
                        ">
                    </td>

                </tr>

            </table>

        </div>

        </div>


        <!-- RAINFALL DETAILS -->
        <div>

            <h3 style="margin-top:0;">Rainfall Details</h3>

            <table id="rainfall-table"
       style="
           border-collapse:collapse;
           font-family:Arial;
           font-size:12px;
           width:330px;
       ">

    <tr style="background:#4a90d9;color:white;">
        <th colspan="4"
            style="border:1px solid #ccc;padding:8px;text-align:center;">
            Rainfall Details
        </th>
    </tr>

    <tr style="background:#f2f2f2;">
        <th style="border:1px solid #ccc;padding:6px;">Year</th>
        <th style="border:1px solid #ccc;padding:6px;">Period</th>
        <th style="border:1px solid #ccc;padding:6px;">Rainfall</th>
        <th style="border:1px solid #ccc;padding:6px;">Max.</th>
    </tr>

    <!-- 2025 -->

    <tr>
        <td rowspan="3"
            contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:center;">
            2025
        </td>

        <td style="border:1px solid #ccc;padding:6px;">
            For the Day
        </td>

        <td contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:right;">
        </td>

        <td rowspan="3"
            contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:center;">
        </td>
    </tr>

    <tr>
        <td style="border:1px solid #ccc;padding:6px;">
            MTD
        </td>

        <td contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:right;">
        </td>
    </tr>

    <tr>
        <td style="border:1px solid #ccc;padding:6px;font-weight:bold;">
            YTD
        </td>

        <td contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:right;">
        </td>
    </tr>

    <!-- 2024 -->

    <tr>
        <td rowspan="3"
            contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:center;">
            2024
        </td>

        <td style="border:1px solid #ccc;padding:6px;">
            For the Day
        </td>

        <td contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:right;">
        </td>

        <td rowspan="3"
            contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:center;">
        </td>
    </tr>

    <tr>
        <td style="border:1px solid #ccc;padding:6px;">
            Month
        </td>

        <td contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:right;">
        </td>
    </tr>

    <tr>
        <td style="border:1px solid #ccc;padding:6px;font-weight:bold;">
            Year
        </td>

        <td contenteditable="true"
            style="border:1px solid #ccc;padding:6px;text-align:right;">
        </td>
    </tr>

</table>

        </div>

    </div>
    """
    port_throughput = _fetch_port_throughput(report_date)

    html += f"""
            </div>

            <!-- PORT THROUGHPUT -->
            <div>

                <h3 style="margin-top:0;">Port Throughput</h3>

                <table style="
                    border-collapse:collapse;
                    font-family:Arial;
                    font-size:12px;
                    width:260px;
                ">

                    <tr style="background:#4a90d9;color:white;">
                        <th colspan="2"
                            style="border:1px solid #ccc;padding:8px;text-align:center;">
                            Port Throughput
                        </th>
                    </tr>

                    <tr>
                        <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">
                            Jetty Throughput (Day)
                        </td>
                        <td style="border:1px solid #ccc;padding:8px;text-align:right;">
                            {port_throughput['day_qty']:,}
                        </td>
                    </tr>

                    <tr>
                        <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">
                            Month
                        </td>
                        <td style="border:1px solid #ccc;padding:8px;text-align:right;">
                            {port_throughput['mtd_qty']:,}
                        </td>
                    </tr>

                    <tr>
                        <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">
                            Year
                        </td>
                        <td style="border:1px solid #ccc;padding:8px;text-align:right;">
                            {port_throughput['ytd_qty']:,}
                        </td>
                    </tr>

                    <tr>
                        <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">
                            Cumulative Since Oct 2012
                        </td>
                        <td style="border:1px solid #ccc;padding:8px;text-align:right;">
                            {port_throughput['cumulative_qty']:,}
                        </td>
                    </tr>

                    <tr>
                        <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">
                            Month TPD
                        </td>
                        <td style="border:1px solid #ccc;padding:8px;text-align:right;">
                            {port_throughput['month_tpd']:,.2f}
                        </td>
                    </tr>

                    <tr>
                    <td style="border:1px solid #ccc;padding:8px;font-weight:bold;">
                        Year TPD
                    </td>
                    <td style="border:1px solid #ccc;padding:8px;text-align:right;">
                        {port_throughput['year_tpd']:,.2f}
                    </td>
                </tr>
                </tr>

                </table>

            </div>

        </div>
    """

    


        
    return html

# ── Download endpoint ───────────────────────────────────────────────────────

@bp.route('/api/module/RP01/daily-ops/download')
@login_required
def daily_ops_download():

    date_str = request.args.get(
        'report_date',
        date.today().strftime('%Y-%m-%d')
    )

    editable_table = []

    if request.args.get("editable_table"):
        try:
            editable_table = json.loads(
                request.args.get("editable_table")
            )
        except Exception:
            editable_table = []

    try:
        report_date = datetime.strptime(
            date_str,
            '%Y-%m-%d'
        ).date()

    except ValueError:
        return Response(
            'Invalid date',
            status=400
        )

    # =====================================
    # FETCH DATA
    # =====================================

    vessels = _fetch_data(report_date)

    if not vessels:
        return Response(
            'No active vessels found',
            status=404
        )

    day_rows, month_rows = _fetch_cargo_handled(
        report_date
    )

    tide_rows = _fetch_tide_data(
        report_date
    )

    mbc_day_rows, mbc_month_rows, mbc_year_rows = \
        _fetch_mbc_cargo_handling(
            report_date
        )

    upcoming_vessels = _fetch_upcoming_vessels(
        report_date
    )

    discharging_mbcs = _fetch_discharging_mbcs(
        report_date
    )

    upcoming_mbcs = _fetch_upcoming_mbcs(
        report_date
    )

    mbc_status_rows = _fetch_mbc_status(
        report_date
    )

    cargo_availability = [
        {"cargo_name": "BRBF", "at_jetty_qty": ""},
        {"cargo_name": "Orissa Fines", "at_jetty_qty": ""},
        {"cargo_name": "Goa Fines", "at_jetty_qty": ""},
        {"cargo_name": "HBI", "at_jetty_qty": ""},
        {"cargo_name": "KDL CLO", "at_jetty_qty": ""},
        {"cargo_name": "Jimblebar Fines", "at_jetty_qty": ""},
        {"cargo_name": "Bacheli Fines", "at_jetty_qty": ""},
        {"cargo_name": "Goa Clo", "at_jetty_qty": ""},
        {"cargo_name": "Mabu", "at_jetty_qty": ""},
        {"cargo_name": "Illavara", "at_jetty_qty": ""},
        {"cargo_name": "Uval + Kestrel", "at_jetty_qty": ""},
        {"cargo_name": "MLV", "at_jetty_qty": ""},
        {"cargo_name": "PCI", "at_jetty_qty": ""},
        {"cargo_name": "Antracite", "at_jetty_qty": ""},
        {"cargo_name": "Limestone", "at_jetty_qty": ""},
        {"cargo_name": "Bentonite", "at_jetty_qty": ""},
        {"cargo_name": "Oliflux", "at_jetty_qty": ""},
        {"cargo_name": "Dolomite", "at_jetty_qty": ""},
        {"cargo_name": "Slag Loading/Unloading", "at_jetty_qty": ""},
        {"cargo_name": "Clinker", "at_jetty_qty": ""}
    ]

    cargo_type_throughput = _fetch_cargo_type_throughput(
        report_date
    )

    cargo_stats_day, cargo_stats_month = \
        _fetch_cargo_statistics(
            report_date
        )

    port_throughput = _fetch_port_throughput(
        report_date
    )

    def _mbc_rows_to_dict(rows):

        data = {
            o: {
                ct: 0.0
                for ct in _MBC_CARGO_TYPES
            }
            for o in _MBC_OWNERS
        }

        for row in rows:

            owner = (
                row['owner']
                if row['owner'] in _MBC_OWNERS
                else 'OTHERS'
            )

            cargo_type = row['cargo_type']

            if cargo_type in _MBC_CARGO_TYPES:
                data[owner][cargo_type] += float(
                    row['qty'] or 0
                )

        return data

    mbc_day = _mbc_rows_to_dict(
        mbc_day_rows
    )

    mbc_month = _mbc_rows_to_dict(
        mbc_month_rows
    )

    mbc_year = _mbc_rows_to_dict(
        mbc_year_rows
    )

    print("EDITABLE TABLE")
    print(editable_table)

    # =====================================
    # RAINFALL TABLE
    # =====================================

    rainfall_table = []

    if request.args.get("rainfall_table"):

        try:

            rainfall_table = json.loads(
                request.args.get("rainfall_table")
            )

            print("RAINFALL TABLE")
            print(rainfall_table)

        except Exception as e:

            print("RAINFALL PARSE ERROR")
            print(e)

            rainfall_table = []

    # =====================================
    # BF TABLE
    # =====================================

    bf_table = []

    if request.args.get("bf_table"):
        try:
            bf_table = json.loads(
                request.args.get("bf_table")
            )

            print("BF TABLE =")
            print(bf_table)

        except Exception:
            bf_table = []
    
     # =====================================
    # RM TABLE
    # =====================================

    rm_table = []

    if request.args.get("rm_table"):
        try:
            rm_table = json.loads(
                request.args.get("rm_table")
            )

            print("RM TABLE =")
            print(rm_table)

        except Exception:
            rm_table = []
    # =====================================
    # BUILD EXCEL
    # =====================================

    buf = _build_excel_a4(
        vessels,
        report_date,
        day_rows=day_rows,
        month_rows=month_rows,
        tide_rows=tide_rows,
        mbc_day=mbc_day,
        mbc_month=mbc_month,
        mbc_year=mbc_year,
        upcoming_vessels=upcoming_vessels,
        discharging_mbcs=discharging_mbcs,
        upcoming_mbcs=upcoming_mbcs,
        mbc_status_rows=mbc_status_rows,
        cargo_availability=cargo_availability,
        mbc_day_rows=mbc_day_rows,
        mbc_month_rows=mbc_month_rows,
        mbc_year_rows=mbc_year_rows,
        cargo_type_throughput=cargo_type_throughput,
        cargo_stats_day=cargo_stats_day,
        cargo_stats_month=cargo_stats_month,
        port_throughput=port_throughput,
        editable_table=editable_table,
        rainfall_table=rainfall_table,
        bf_table=bf_table,
        rm_table=rm_table
    )

    fname = f'DailyOps_{date_str}.xlsx'

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition':
            f'attachment; filename="{fname}"'
        }
    )