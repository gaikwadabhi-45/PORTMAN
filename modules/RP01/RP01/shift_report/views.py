from flask import render_template, request, session, redirect, url_for, Response, jsonify
from functools import wraps
from datetime import datetime
import io
import json
import re

from .. import bp
from database import get_db, get_cursor

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_thin  = Side(style='thin', color='C7CDD4')

_FULL_BDR = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_LR_BDR   = Border(left=_thin, right=_thin)
_LRT_BDR  = Border(left=_thin, right=_thin, top=_thin)
_LRB_BDR  = Border(left=_thin, right=_thin, bottom=_thin)

_bdr  = _FULL_BDR   # alias so _build_cargo_sheet still works

_ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_lctr = Alignment(horizontal='left',   vertical='center', wrap_text=False)

# Delay sheet palette
_TITLE_BG      = 'C0392B'
_HEADER_BG     = '2C3E50'
_TYPE_BG       = 'D6EAF8'
_ACTIVITY_BG   = 'D6E4F0'
_SUBTOTAL_BG   = 'BDD7EE'
_TYPE_TOTAL_BG = 'AED6F1'
_GRAND_BG      = '2980B9'

# Cargo sheet palette (keep existing names so _build_cargo_sheet works)
_TITLE_FILL    = 'F6F8FB'
_HEADER_FILL   = 'EEF4FB'
_BODY_FILL     = 'FFFFFF'
_GROUP_FILL    = 'F8FAFC'
_SUBTOTAL_FILL = 'F6F7F9'
_TOTAL_FILL    = 'E9EEF5'
_TYPE_TOTAL_FILL = 'F1F6FB'
_TEXT  = '2C3E50'
_WHITE = 'FFFFFF'

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=11, color='000000'):
    return Font(name='Calibri', bold=bold, size=size, color=color)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated


# ---------------------------------------------------------------------------
# Page route
# ---------------------------------------------------------------------------

@bp.route('/module/RP01/shift-report/')
@login_required
def shift_report_index():
    return render_template('shift_report/shift_report.html',
                           username=session.get('username'))


# ---------------------------------------------------------------------------
# Report API routes
# ---------------------------------------------------------------------------

@bp.route('/api/module/RP01/shift-report/preview')
@login_required
def shift_report_preview():
    entry_date = request.args.get('entry_date', '')
    shift = request.args.get('shift', '')
    if not entry_date or not shift:
        return jsonify({'error': 'entry_date and shift are required'}), 400

    cargo_pivot = _fetch_cargo_pivot(entry_date, shift)
    delays = _fetch_delays(entry_date, shift)
    delay_keys = _parse_delay_keys(request.args.get('delay_keys'))
    cargo_tables = _build_cargo_tables(entry_date, shift, cargo_pivot)

    if _is_all_shifts(shift):
        shift_pivot = _fetch_shift_pivot(entry_date, shift)
        shift_wise = _build_shift_wise_table(entry_date, shift, shift_pivot)
        if shift_wise:
            cargo_tables.insert(0, shift_wise)

    delay_view = _build_delay_view(entry_date, shift, delays, delay_keys)
    return jsonify({
        'cargo_tables': cargo_tables,
        'delay_view': delay_view,
    })


@bp.route('/api/module/RP01/shift-report/delay-options')
@login_required
def shift_report_delay_options():
    entry_date = request.args.get('entry_date', '')
    shift = request.args.get('shift', '')
    if not entry_date or not shift:
        return jsonify([])
    return jsonify(_fetch_delay_options(entry_date, shift))


@bp.route('/api/module/RP01/shift-report/download')
@login_required
def shift_report_download():
    entry_date = request.args.get('entry_date', '')
    shift = request.args.get('shift', '')
    if not entry_date or not shift:
        return Response('entry_date and shift are required', status=400)

    cargo_pivot = _fetch_cargo_pivot(entry_date, shift)
    delays = _fetch_delays(entry_date, shift)
    delay_keys = _parse_delay_keys(request.args.get('delay_keys'))
    cargo_tables = _build_cargo_tables(entry_date, shift, cargo_pivot)

    if _is_all_shifts(shift):
        shift_pivot = _fetch_shift_pivot(entry_date, shift)
        shift_wise = _build_shift_wise_table(entry_date, shift, shift_pivot)
        if shift_wise:
            cargo_tables.insert(0, shift_wise)

    delay_view = _build_delay_view(entry_date, shift, delays, delay_keys)
    buf = _build_excel(cargo_tables, delay_view)
    fname = f'ShiftReport_{entry_date}_Shift{shift}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Saved report (user_delay_preferences) API routes
# ---------------------------------------------------------------------------

@bp.route('/api/module/RP01/save-report', methods=['POST'])
@login_required
def save_shift_report():
    conn = get_db()
    cur = get_cursor(conn)
    try:
        data = request.get_json()
        preference_name = (data.get('preference_name') or '').strip()
        delay_keys = data.get('delay_keys') or []

        if not preference_name:
            return jsonify({'success': False, 'message': 'Report name is required'}), 400
        if not delay_keys:
            return jsonify({'success': False, 'message': 'No delays selected'}), 400

        # Upsert: update existing by name, otherwise insert
        cur.execute("""
            SELECT id FROM user_delay_preferences
            WHERE LOWER(preference_name) = LOWER(%s)
        """, (preference_name,))
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE user_delay_preferences
                SET delay_keys = %s, updated_at = NOW()
                WHERE id = %s
            """, (json.dumps(delay_keys), existing['id']))
            report_id = existing['id']
        else:
            cur.execute("""
                INSERT INTO user_delay_preferences (preference_name, delay_keys, updated_at)
                VALUES (%s, %s, NOW())
                RETURNING id
            """, (preference_name, json.dumps(delay_keys)))
            report_id = cur.fetchone()['id']

        conn.commit()
        return jsonify({'success': True, 'id': report_id, 'message': 'Saved successfully'})

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/module/RP01/saved-reports')
@login_required
def get_saved_reports():
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT id, preference_name, delay_keys, updated_at
            FROM user_delay_preferences
            ORDER BY updated_at DESC
        """)
        rows = cur.fetchall()
        result = []
        for row in rows:
            delay_keys = row['delay_keys'] or []
            result.append({
                'id': row['id'],
                'preference_name': row['preference_name'],
                'delay_keys': delay_keys,
                'total_delays': len(delay_keys),
                'created_at': row['updated_at'].isoformat() if row['updated_at'] else None,
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/module/RP01/saved-reports/<int:report_id>')
@login_required
def get_saved_report_by_id(report_id):
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT id, preference_name, delay_keys, updated_at
            FROM user_delay_preferences
            WHERE id = %s
        """, (report_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        return jsonify({
            'id': row['id'],
            'preference_name': row['preference_name'],
            'delay_keys': row['delay_keys'] or [],
            'created_at': row['updated_at'].isoformat() if row['updated_at'] else None,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/module/RP01/delete-saved-report/<int:report_id>', methods=['DELETE'])
@login_required
def delete_saved_report(report_id):
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("DELETE FROM user_delay_preferences WHERE id = %s", (report_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Report deleted successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# DB fetch helpers
# ---------------------------------------------------------------------------

def _fetch_cargo_pivot(entry_date, shift):
    conn = get_db()
    cur = get_cursor(conn)
    query = """
        SELECT l.cargo_name, l.equipment_name, l.route_name,
               COALESCE(vc.cargo_type, l.cargo_name) AS cargo_type,
               COALESCE(SUM(l.quantity), 0) AS qty
        FROM lueu_lines l
        LEFT JOIN vessel_cargo vc ON vc.cargo_name = l.cargo_name
        WHERE l.entry_date = %s
          AND l.quantity > 0
          AND l.cargo_name IS NOT NULL AND l.cargo_name != ''
    """
    params = [entry_date]
    if not _is_all_shifts(shift):
        query += " AND l.shift = %s"
        params.append(shift)
    query += """
        GROUP BY l.cargo_name, l.equipment_name, l.route_name, COALESCE(vc.cargo_type, l.cargo_name)
        ORDER BY l.cargo_name, l.equipment_name, l.route_name
    """
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()

    equipment_set = set()
    route_set = set()
    pivot = {}
    cargo_type_map = {}

    for row in rows:
        cargo = row['cargo_name'] or 'Unknown'
        equip = row['equipment_name'] or 'Unknown'
        route = row['route_name'] or 'Unknown'
        qty = float(row['qty'] or 0)

        cargo_type_map[cargo] = row['cargo_type'] or cargo
        equipment_set.add(equip)
        route_set.add(route)
        pivot.setdefault(cargo, {}).setdefault(equip, {})
        pivot[cargo][equip][route] = pivot[cargo][equip].get(route, 0) + qty

    return {
        'data': pivot,
        'cargo_type_map': cargo_type_map,
        'equipments': sorted(equipment_set, key=_equipment_sort_key),
        'routes': sorted(route_set, key=_route_sort_key),
    }


def _fetch_shift_pivot(entry_date, shift):
    conn = get_db()
    cur = get_cursor(conn)
    query = """
        SELECT l.cargo_name, l.shift,
               COALESCE(SUM(l.quantity), 0) AS qty
        FROM lueu_lines l
        WHERE l.entry_date = %s
          AND l.quantity > 0
          AND l.cargo_name IS NOT NULL AND l.cargo_name != ''
          AND l.shift IS NOT NULL AND l.shift != ''
    """
    params = [entry_date]
    if not _is_all_shifts(shift):
        query += " AND l.shift = %s"
        params.append(shift)
    query += """
        GROUP BY l.cargo_name, l.shift
        ORDER BY l.cargo_name, l.shift
    """
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()

    shift_set = set()
    pivot = {}

    for row in rows:
        cargo = row['cargo_name'] or 'Unknown'
        s = row['shift'] or 'Unknown'
        qty = float(row['qty'] or 0)
        shift_set.add(s)
        pivot.setdefault(cargo, {})
        pivot[cargo][s] = pivot[cargo].get(s, 0) + qty

    return {
        'data': pivot,
        'shifts': sorted(shift_set),
    }


def _fetch_delays(entry_date, shift):
    conn = get_db()
    cur = get_cursor(conn)
    query = """
        SELECT l.delay_name, l.equipment_name, l.system_name, l.route_name,
               l.from_time, l.to_time,
               COALESCE(d.type, 'Other') AS delay_type
        FROM lueu_lines l
        LEFT JOIN port_delay_types d ON d.name = l.delay_name
        WHERE l.entry_date = %s
          AND l.delay_name IS NOT NULL AND l.delay_name != ''
    """
    params = [entry_date]
    if not _is_all_shifts(shift):
        query += " AND l.shift = %s"
        params.append(shift)
    query += """
        ORDER BY d.type, l.delay_name, l.equipment_name, l.system_name, l.route_name
    """
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()

    delays = []
    for row in rows:
        from_t = (row['from_time'] or '').strip()
        to_t = (row['to_time'] or '').strip()
        delays.append({
            'delay_type': row['delay_type'] or 'Other',
            'delay_name': row['delay_name'] or '',
            'equipment_name': row['equipment_name'] or '',
            'system_name': row['system_name'] or '',
            'route_name': row['route_name'] or '',
            'from_time': from_t,
            'to_time': to_t,
            'total_minutes': _calc_minutes(from_t, to_t),
        })

    return delays


def _fetch_delay_options(entry_date, shift):
    conn = get_db()
    cur = get_cursor(conn)
    query = """
        SELECT COALESCE(d.type, 'Other') AS delay_type,
               l.delay_name
        FROM lueu_lines l
        LEFT JOIN port_delay_types d ON d.name = l.delay_name
        WHERE l.entry_date = %s
          AND l.delay_name IS NOT NULL AND l.delay_name != ''
    """
    params = [entry_date]
    if not _is_all_shifts(shift):
        query += " AND l.shift = %s"
        params.append(shift)
    query += """
        GROUP BY COALESCE(d.type, 'Other'), l.delay_name
        ORDER BY COALESCE(d.type, 'Other'), l.delay_name
    """
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()

    options = []
    for row in rows:
        delay_type = _blank_label(row['delay_type'])
        delay_name = _blank_label(row['delay_name'])
        options.append({
            'key': _delay_key(delay_type, delay_name),
            'label': f'{delay_name} [{delay_type}]',
            'delay_name': delay_name,
            'delay_type': delay_type,
        })

    return sorted(
        options,
        key=lambda item: (_delay_type_sort_key(item['delay_type']), _natural_sort_key(item['delay_name'])),
    )


# ---------------------------------------------------------------------------
# Utility / formatting helpers
# ---------------------------------------------------------------------------

def _calc_minutes(from_t, to_t):
    try:
        fmt = '%H:%M'
        start = datetime.strptime(from_t.strip(), fmt)
        end = datetime.strptime(to_t.strip(), fmt)
        diff = int((end - start).total_seconds() / 60)
        if diff < 0:
            diff += 24 * 60
        return diff
    except Exception:
        return 0


def _fmt_minutes(minutes, blank_zero=False):
    if minutes in (None, ''):
        return ''
    minutes = int(minutes)
    if blank_zero and minutes == 0:
        return ''
    return f'{minutes // 60}:{minutes % 60:02d}'


def _fmt_qty(value):
    if not value:
        return ''
    return int(round(value))


def _is_all_shifts(shift):
    return str(shift or '').strip().upper() in {'ALL', 'ALL SHIFTS'}


def _shift_scope_label(shift):
    shift = str(shift or '').strip().upper()
    if _is_all_shifts(shift):
        return 'All Shifts'
    return f'{shift} Shift'


def _report_date(entry_date):
    try:
        return datetime.strptime(entry_date, '%Y-%m-%d').strftime('%d.%m.%Y')
    except Exception:
        return entry_date


def _blank_label(value):
    value = (value or '').strip()
    return value or '(blank)'


def _total_label(value):
    return f'{_blank_label(value)} Total'


def _natural_sort_key(value):
    text = str(value or '').strip()
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r'(\d+)', text)]


def _equipment_sort_key(value):
    upper = str(value or '').upper()
    if 'SENNEBOGEN' in upper:
        group = 0
    elif upper.startswith('BUL'):
        group = 1
    else:
        group = 2
    return (group, _natural_sort_key(value))


def _route_sort_key(value):
    upper = str(value or '').upper().strip()
    if upper == 'BY ROAD':
        group = 0
    elif upper.startswith('C-'):
        group = 1
    elif 'JETTY YARD' in upper:
        group = 2
    elif 'LS-' in upper:
        group = 3
    elif 'STACKER' in upper:
        group = 4
    else:
        group = 5
    return (group, _natural_sort_key(value))

def _shift_time_sort_key(from_time):
    """Sort times starting from 06:00 (Shift A), then 14:00 (B), then 22:00 (C)."""
    try:
        h, m = map(int, from_time.strip().split(':'))
        minutes = h * 60 + m
        # Rotate: 06:00 (360 mins) becomes 0
        return (minutes - 360) % (24 * 60)
    except Exception:
        return 9999


def _location_group(route_name):
    upper = str(route_name or '').upper().strip()
    if not upper:
        return '(blank)'
    if upper == 'BY ROAD' or 'ROAD' in upper:
        return 'By Road'
    if 'CEMENT SILO' in upper or upper == 'SILO' or upper.endswith(' SILO'):
        return 'Cement Silo'
    if 'STACKER' in upper or 'JETTY YARD' in upper or 'LS-' in upper or 'SHED' in upper or 'STOCKYARD' in upper:
        return 'Stacker/Shed'
    if 'DIRECT PLANT' in upper or 'C-131' in upper or 'PLANT' in upper:
        return 'Direct Plant'
    return route_name


def _location_sort_key(value):
    order = {
        'Direct Plant': 0,
        'Stacker/Shed': 1,
        'By Road': 2,
        'Cement Silo': 3,
    }
    return (order.get(value, 99), _natural_sort_key(value))


def _delay_type_sort_key(value):
    order = {
        'RMHS Delays': 0,
        'Jetty Delays': 1,
        'Process Delays': 2,
        'ProcessDelays': 2,
        'Process Requirement': 3,
        'ProcessRequirement': 3,
        'Maintenance Delays': 3,
        'MaintenanceDelays': 4,
        'Other': 5,
    }
    return (order.get(value, 99), _natural_sort_key(value))


def _delay_key(delay_type, delay_name):
    return f'{_blank_label(delay_type)}||{_blank_label(delay_name)}'


def _parse_delay_keys(raw_value):
    if raw_value is None:
        return None
    try:
        values = json.loads(raw_value)
    except Exception:
        return []
    if not isinstance(values, list):
        return []

    cleaned = []
    seen = set()
    for value in values:
        if not isinstance(value, str):
            continue
        value = value.strip()
        if not value or value in seen:
            continue
        seen.add(value)
        cleaned.append(value)
    return cleaned


# ---------------------------------------------------------------------------
# Table builders
# ---------------------------------------------------------------------------

def _make_matrix_table(title, row_header, column_headers, row_names, values):
    rows = []
    column_totals = [0] * len(column_headers)
    grand_total = 0

    for row_name in row_names:
        row_total = 0
        row_values = []
        for idx, column_name in enumerate(column_headers):
            value = float(values.get(row_name, {}).get(column_name, 0) or 0)
            row_total += value
            column_totals[idx] += value
            row_values.append(_fmt_qty(value))
        grand_total += row_total
        rows.append({
            'label': row_name,
            'values': row_values,
            'total': _fmt_qty(row_total),
        })

    return {
        'title': title,
        'row_header': row_header,
        'columns': column_headers,
        'rows': rows,
        'totals': {
            'label': 'Grand Total',
            'values': [_fmt_qty(value) for value in column_totals],
            'total': _fmt_qty(grand_total),
        },
    }


def _build_cargo_tables(entry_date, shift, cargo_pivot):
    data = cargo_pivot['data']
    cargo_type_map = cargo_pivot['cargo_type_map']
    equipments = cargo_pivot['equipments']
    routes = cargo_pivot['routes']
    cargo_names = sorted(data.keys(), key=_natural_sort_key)

    if not cargo_names:
        return []

    equipment_values = {}
    route_values = {}
    location_values = {}
    cargo_type_location_values = {}

    location_columns = sorted({_location_group(route) for route in routes}, key=_location_sort_key)

    for cargo in cargo_names:
        equipment_values[cargo] = {}
        route_values[cargo] = {}
        location_values[cargo] = {location: 0 for location in location_columns}

        for equipment in equipments:
            total = sum(data.get(cargo, {}).get(equipment, {}).get(route, 0) for route in routes)
            equipment_values[cargo][equipment] = total

        for route in routes:
            route_total = sum(data.get(cargo, {}).get(equipment, {}).get(route, 0) for equipment in equipments)
            route_values[cargo][route] = route_total
            location = _location_group(route)
            location_values[cargo][location] = location_values[cargo].get(location, 0) + route_total

        ctype = cargo_type_map.get(cargo, cargo)
        if ctype not in cargo_type_location_values:
            cargo_type_location_values[ctype] = {loc: 0 for loc in location_columns}
        for loc in location_columns:
            cargo_type_location_values[ctype][loc] = (
                cargo_type_location_values[ctype].get(loc, 0) + location_values[cargo].get(loc, 0)
            )

    cargo_type_rows = sorted(cargo_type_location_values.keys(), key=_natural_sort_key)

    route_rows = sorted(routes, key=_route_sort_key)
    route_wise_values = {
        route: {cargo: route_values[cargo].get(route, 0) for cargo in cargo_names}
        for route in route_rows
    }

    date_label = _report_date(entry_date)
    shift_label = _shift_scope_label(shift)
    return [
        _make_matrix_table(
            f'{shift_label} Jetty Discharge: {date_label}',
            'Cargo Name',
            equipments,
            cargo_names,
            equipment_values,
        ),
        _make_matrix_table(
            f'{shift_label} Location Wise Discharge: {date_label}',
            'Cargo Type',
            location_columns,
            cargo_type_rows,
            cargo_type_location_values,
        ),
        _make_matrix_table(
            f'{shift_label} Receiving Route Wise Discharge: {date_label}',
            'Row Labels',
            cargo_names,
            route_rows,
            route_wise_values,
        ),
    ]


def _build_shift_wise_table(entry_date, shift, shift_pivot):
    data = shift_pivot['data']
    shifts = shift_pivot['shifts']
    cargo_names = sorted(data.keys(), key=_natural_sort_key)

    if not cargo_names or not shifts:
        return None

    shift_wise_values = {}
    for s in shifts:
        shift_wise_values[s] = {}
        for cargo in cargo_names:
            shift_wise_values[s][cargo] = data.get(cargo, {}).get(s, 0)

    date_label = _report_date(entry_date)
    shift_label = _shift_scope_label(shift)
    return _make_matrix_table(
        f'{shift_label} Wise Discharge: {date_label}',
        'Shift',
        cargo_names,
        shifts,
        shift_wise_values,
    )


def _build_delay_view(entry_date, shift, delays, delay_keys=None):

    shift_label = _shift_scope_label(shift)

    view = {
        'title': f'{shift_label} Jetty & RMHS Delays: {_report_date(entry_date)}',
        'rows': [],
        'grand_total': '',
    }

    if delay_keys is None:
        filtered_delays = list(delays)
        delay_order = {}

    else:

        delay_order = {
            key: idx
            for idx, key in enumerate(delay_keys)
        }

        filtered_delays = [
            delay for delay in delays
            if _delay_key(
                delay['delay_type'],
                delay['delay_name']
            ) in delay_order
        ]

    if not filtered_delays:
        return view

    grouped = {}

    # =========================================================
    # GROUP DATA
    # =========================================================
    for delay in filtered_delays:

        delay_type = _blank_label(delay['delay_type'])

        activity = _blank_label(delay['delay_name'])

        equipment = _blank_label(delay['equipment_name'])

        system = _blank_label(delay['system_name'])

        grouped\
            .setdefault(delay_type, {})\
            .setdefault(activity, {})\
            .setdefault(equipment, {})\
            .setdefault(system, [])\
            .append(delay)

    grand_total = 0

    # =========================================================
    # DELAY TYPE LOOP
    # =========================================================
    for delay_type in sorted(
        grouped,
        key=_delay_type_sort_key
    ):

        type_rows = []

        type_total = 0

        # =====================================================
        # ACTIVITY LOOP
        # =====================================================
        for activity in sorted(
            grouped[delay_type],
            key=_natural_sort_key
        ):

            activity_rows = []

            activity_total = 0

            # =================================================
            # EQUIPMENT LOOP
            # =================================================
            for equipment in sorted(
                grouped[delay_type][activity],
                key=_equipment_sort_key
            ):

                equipment_rows = []

                systems = grouped[delay_type][activity][equipment]

                # =============================================
                # SYSTEM LOOP
                # =============================================
                for system in sorted(
                    systems,
                    key=_natural_sort_key
                ):

                    system_items = sorted(
                    systems[system],
                    key=lambda item: (
                        _blank_label(item['route_name']).lower(),
                        _shift_time_sort_key(item['from_time']),
                        item['to_time'],
                    ),
                )

                    system_rows = []

                    system_total = 0

                    # =========================================
                    # DETAIL ROWS
                    # =========================================
                    for item in system_items:

                        system_total += int(
                            item['total_minutes'] or 0
                        )

                        system_rows.append({

                            'kind': 'detail',

                            'route': _blank_label(
                                item['route_name']
                            ),

                            'from_time': item['from_time'],

                            'to_time': item['to_time'],

                            'total': _fmt_minutes(
                                item['total_minutes']
                            ),

                            'system_name': system,
                        })

                    # =========================================
                    # SYSTEM ROWSPAN
                    # =========================================
                    detail_rows = [
                        r for r in system_rows
                        if r['kind'] == 'detail'
                    ]

                    if detail_rows:

                        detail_rows[0]['show_system'] = True

                        detail_rows[0]['system_name'] = system

                        detail_rows[0]['system_rowspan'] = len(detail_rows)

                        for r in detail_rows[1:]:

                            r['show_system'] = False

                    equipment_rows.extend(system_rows)

                    # =========================================
                    # SYSTEM TOTAL
                    # =========================================
                    equipment_rows.append({

                        'kind': 'system_total',

                        'label': _total_label(system),

                        'total': _fmt_minutes(system_total),
                    })

                    activity_total += system_total

                # =============================================
                # EQUIPMENT ROWSPAN
                # =============================================
                detail_rows = [
                    r for r in equipment_rows
                    if r['kind'] == 'detail'
                ]

                if detail_rows:

                    detail_rows[0]['show_equipment'] = True

                    detail_rows[0]['equipment_name'] = equipment

                    detail_rows[0]['equipment_rowspan'] = len(detail_rows)

                    for r in detail_rows[1:]:

                        r['show_equipment'] = False

                activity_rows.extend(equipment_rows)

            # =================================================
            # ACTIVITY ROWSPAN
            # =================================================
            detail_rows = [
                r for r in activity_rows
                if r['kind'] == 'detail'
            ]

            if detail_rows:

                detail_rows[0]['show_activity'] = True

                detail_rows[0]['activity_name'] = activity

                detail_rows[0]['activity_rowspan'] = len(activity_rows) + 1

                for r in detail_rows[1:]:

                    r['show_activity'] = False

            # =================================================
            # ADD ACTIVITY ROWS
            # =================================================
            type_rows.extend(activity_rows)

            # =================================================
            # ACTIVITY TOTAL
            # =================================================
            type_rows.append({

                'kind': 'activity_total',

                'label': _total_label(activity),

                'total': _fmt_minutes(activity_total),
            })

            type_total += activity_total

        # =====================================================
        # TYPE ROWSPAN
        # =====================================================
        detail_rows = [
            r for r in type_rows
            if r['kind'] == 'detail'
        ]

        if detail_rows:

            detail_rows[0]['show_type'] = True

            detail_rows[0]['type_name'] = delay_type

            detail_rows[0]['type_rowspan'] = len(detail_rows)

            for r in detail_rows[1:]:

                r['show_type'] = False

        # =====================================================
        # ADD TYPE ROWS TO VIEW
        # =====================================================
        view['rows'].extend(type_rows)

        # =====================================================
        # TYPE TOTAL
        # =====================================================
        view['rows'].append({

            'kind': 'type_total',

            'label': _total_label(delay_type),

            'total': _fmt_minutes(type_total),
        })

        grand_total += type_total

    # =========================================================
    # GRAND TOTAL
    # =========================================================
    view['rows'].append({

        'kind': 'grand_total',

        'label': 'Grand Total',

        'total': _fmt_minutes(grand_total),
    })

    view['grand_total'] = _fmt_minutes(grand_total)

    return view

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _build_excel(cargo_tables, delay_view):
    wb = Workbook()
    _build_cargo_sheet(wb, cargo_tables)
    _build_delay_sheet(wb, delay_view)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _cell(ws, row, col, value='', bold=False, fill_color='FFFFFF',
          align=None, font_color='000000', size=11, border=None):
    align = align or _ctr
    c = ws.cell(row, col, value)
    c.font      = Font(name='Calibri', bold=bold, size=size, color=font_color)
    c.fill      = _fill(fill_color)
    c.alignment = align
    c.border    = border or _FULL_BDR
    return c


def _merge_title(ws, row, total_cols, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    _cell(ws, row, 1, title, bold=True, fill_color=_TITLE_FILL, font_color=_TEXT)
    ws.cell(row, 1).font = _font(bold=True, size=12, color=_TEXT)
    for col in range(2, total_cols + 1):
        ws.cell(row, col).fill = _fill(_TITLE_FILL)
        ws.cell(row, col).border = _bdr


def _build_cargo_sheet(wb, cargo_tables):
    ws = wb.active
    ws.title = 'Cargo Handled'

    if not cargo_tables:
        _cell(ws, 1, 1, 'No cargo data found for this shift.', align=_left)
        return

    row = 1
    max_cols = 0

    for table in cargo_tables:
        total_cols = 1 + len(table['columns']) + 1
        max_cols = max(max_cols, total_cols)

        _merge_title(ws, row, total_cols, table['title'])
        row += 1

        _cell(ws, row, 1, table['row_header'], bold=True, fill_color=_HEADER_FILL, font_color=_TEXT)
        for idx, column in enumerate(table['columns'], start=2):
            _cell(ws, row, idx, column, bold=True, fill_color=_HEADER_FILL, font_color=_TEXT)
        _cell(ws, row, total_cols, 'Grand Total', bold=True, fill_color=_HEADER_FILL, font_color=_TEXT)
        row += 1

        for data_row in table['rows']:
            _cell(ws, row, 1, data_row['label'], fill_color=_BODY_FILL, align=_left, font_color=_TEXT)
            for idx, value in enumerate(data_row['values'], start=2):
                _cell(ws, row, idx, value, fill_color=_BODY_FILL, font_color=_TEXT)
            _cell(ws, row, total_cols, data_row['total'], bold=True, fill_color=_BODY_FILL, font_color=_TEXT)
            row += 1

        _cell(ws, row, 1, table['totals']['label'], bold=True, fill_color=_TOTAL_FILL, align=_left, font_color=_TEXT)
        for idx, value in enumerate(table['totals']['values'], start=2):
            _cell(ws, row, idx, value, bold=True, fill_color=_TOTAL_FILL, font_color=_TEXT)
        _cell(ws, row, total_cols, table['totals']['total'], bold=True, fill_color=_TOTAL_FILL, font_color=_TEXT)
        row += 2

    ws.column_dimensions['A'].width = 28
    for col in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _set_merged_value(ws, row, start_col, end_col, value, fill_color, bold=True, font_color=_TEXT, align=_ctr):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    _cell(ws, row, start_col, value, bold=bold, fill_color=fill_color, align=align, font_color=font_color)
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row, col).fill = _fill(fill_color)
        ws.cell(row, col).border = _bdr


def _build_delay_sheet(wb, delay_view):
    ws = wb.create_sheet('Delay Report')

    headers = ['Delays Type', 'Activity', 'Equipment', 'System',
               'Route', 'From', 'To', 'Total']
    total_cols = len(headers)

    if not delay_view['rows']:
        _cell(ws, 1, 1, 'No delay data found for this shift.', align=_left)
        return

    # Title
    _merge_title(ws, 1, total_cols, delay_view['title'])

    # Headers
    for idx, header in enumerate(headers, start=1):
        _cell(ws, 2, idx, header, bold=True,
              fill_color=_HEADER_FILL, font_color=_TEXT)

    row = 3
    merge_queue   = []   # cols 2-4 (activity, equipment, system)
    detail_row_index = 0

    # Use a dict so nested function can mutate without UnboundLocalError
    type_group = {'start': None, 'name': None}

    def _flush_type_merge(end_row):
        if type_group['start'] is None:
            return
        start_r = type_group['start']
        if start_r != end_row:
            ws.merge_cells(start_row=start_r, start_column=1,
                        end_row=end_row,   end_column=1)
        tc = ws.cell(start_r, 1)
        tc.value     = type_group['name']   # e.g. "RMHS Delays"
        tc.fill      = _fill(_TYPE_BG)
        tc.font = _font(bold=True, size=11, color='C0392B')
        tc.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        for r in range(start_r, end_row + 1):
            c = ws.cell(r, 1)
            c.fill = _fill(_TYPE_BG)
            if start_r == end_row:
                c.border = _FULL_BDR
            elif r == start_r:
                c.border = _LRT_BDR
            elif r == end_row:
                c.border = _LRB_BDR
            else:
                c.border = _LR_BDR

                detail_rows = []
                other_rows = []

                for item in delay_view['rows']:

                    if item.get('kind') == 'detail':

                        detail_rows.append(item)

                    else:

                        other_rows.append(item)


                detail_rows = sorted(
                    detail_rows,
                    key=lambda x: (
                        x.get('from_time', '')
                    )
                )

                delay_view['rows'] = detail_rows + other_rows
    

    for item in delay_view['rows']:
        kind = item['kind']

        if kind == 'detail':
            bg = 'FFFFFF' if detail_row_index % 2 == 0 else 'F4F9FD'
            detail_row_index += 1

            if item.get('show_type'):
                type_group['start'] = row
                type_group['name']  = item.get('type_name', '')

            # Col 1 — blue fill, value set later by _flush_type_merge
            _cell(ws, row, 1, '', bold=True,
                  fill_color=_TYPE_BG, font_color=_WHITE, align=_ctr)

            # Col 2 — Activity
            _cell(ws, row, 2,
                  item.get('activity_name', '') if item.get('show_activity') else '',
                  bold=True, fill_color=_ACTIVITY_BG, font_color=_TEXT, align=_lctr)
            if item.get('show_activity') and item.get('activity_rowspan', 0) > 1:
                merge_queue.append((row, 2,
                                    row + item['activity_rowspan'] - 1,
                                    _ACTIVITY_BG,
                                    item.get('activity_name', '')))

            # Col 3 — Equipment
            _cell(ws, row, 3,
                  item.get('equipment_name', '') if item.get('show_equipment') else '',
                  bold=True, fill_color=bg, font_color=_TEXT, align=_ctr)
            if item.get('show_equipment') and item.get('equipment_rowspan', 0) > 1:
                merge_queue.append((row, 3,
                                    row + item['equipment_rowspan'] - 1,
                                    bg,
                                    item.get('equipment_name', '')))

            # Col 4 — System
            _cell(ws, row, 4,
                  item.get('system_name', '') if item.get('show_system') else '',
                  bold=False, fill_color=bg, font_color=_TEXT, align=_ctr)
            if item.get('show_system') and item.get('system_rowspan', 0) > 1:
                merge_queue.append((row, 4,
                                    row + item['system_rowspan'] - 1,
                                    bg,
                                    item.get('system_name', '')))

            _cell(ws, row, 5, item.get('route', ''),
                  fill_color=bg, font_color=_TEXT, align=_ctr)
            _cell(ws, row, 6, item.get('from_time', ''),
                  fill_color=bg, font_color=_TEXT, align=_ctr)
            _cell(ws, row, 7, item.get('to_time', ''),
                  fill_color=bg, font_color=_TEXT, align=_ctr)
            _cell(ws, row, 8, item.get('total', ''),
                  bold=True, fill_color=bg, font_color=_TEXT, align=_ctr)

        elif kind == 'system_total':
            for col in range(1, 9):
                _cell(ws, row, col, '', fill_color=_SUBTOTAL_BG, font_color=_TEXT)
            _cell(ws, row, 1, '', bold=True, fill_color=_TYPE_BG, font_color=_WHITE)
            _cell(ws, row, 2, '', fill_color=_ACTIVITY_BG, font_color=_TEXT)   # ← ADD THIS
            _cell(ws, row, 4, item.get('label', ''), bold=True, fill_color=_SUBTOTAL_BG, font_color=_TEXT, align=_lctr)

        elif kind == 'activity_total':
            for col in range(1, 9):

                ws.cell(row, col).value = None

                ws.cell(row, col).fill = _fill(_SUBTOTAL_BG)

                ws.cell(row, col).border = _FULL_BDR

            # =========================================
            # KEEP TYPE COLUMN
            # =========================================
            ws.cell(row, 1).fill = _fill(_TYPE_BG)

            # =========================================
            # ACTIVITY COLUMN EMPTY
            # =========================================
            ws.cell(row, 2).fill = _fill(_ACTIVITY_BG)

            # =========================================
            # MERGE C:G
            # =========================================
            ws.merge_cells(
                start_row=row,
                start_column=3,
                end_row=row,
                end_column=7
            )

            # =========================================
            # LABEL
            # =========================================
            c = ws.cell(row, 3)

            c.value = item.get('label', '')

            c.font = _font(
                bold=True,
                size=11,
                color=_TEXT
            )

            c.fill = _fill(_SUBTOTAL_BG)

            c.alignment = Alignment(
                horizontal='left',
                vertical='center'
            )

            c.border = _FULL_BDR

            # =========================================
            # STYLE MERGED CELLS
            # =========================================
            for col in range(4, 8):

                ws.cell(row, col).fill = _fill(_SUBTOTAL_BG)

                ws.cell(row, col).border = _FULL_BDR

            # =========================================
            # TOTAL VALUE
            # =========================================
            total_cell = ws.cell(row, 8)

            total_cell.value = item.get('total', '')

            total_cell.font = _font(
                bold=True,
                size=11,
                color=_TEXT
            )

            total_cell.fill = _fill(_SUBTOTAL_BG)

            total_cell.alignment = _ctr

            total_cell.border = _FULL_BDR

        elif kind == 'type_total':
            for col in range(1, 9):
                _cell(ws, row, col, '', fill_color=_TYPE_TOTAL_BG, font_color=_TEXT)
            # Col 1 stays blue (no text)
            _cell(ws, row, 1, '', bold=True,
      fill_color=_TYPE_BG, font_color='C0392B', align=_ctr)
            # Merge cols 2-7 and put the label there
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            _cell(ws, row, 2, item.get('label', ''), bold=True,
                fill_color=_TYPE_TOTAL_BG, font_color=_TEXT, align=_ctr)
            for col in range(3, 8):
                ws.cell(row, col).fill   = _fill(_TYPE_TOTAL_BG)
                ws.cell(row, col).border = _FULL_BDR
            # Col 8 — total value
            _cell(ws, row, 8, item.get('total', ''), bold=True,
                fill_color=_TYPE_TOTAL_BG, font_color=_TEXT, align=_ctr)
            # Flush col-1 merge (no label needed, col 1 just stays blue)
            _flush_type_merge(row)
            type_group['start'] = None
            type_group['name']  = None

        elif kind == 'grand_total':
            if type_group['start'] is not None:
                _flush_type_merge(row - 1)
                type_group['start'] = None
            for col in range(1, 9):
                _cell(ws, row, col, '', fill_color=_GRAND_BG,
                      font_color=_WHITE, bold=True)
            _cell(ws, row, 1, item.get('label', ''), bold=True,
                  fill_color=_GRAND_BG, font_color=_WHITE, align=_left)
            _cell(ws, row, 8, item.get('total', ''), bold=True,
                  fill_color=_GRAND_BG, font_color=_WHITE, align=_ctr)

        row += 1

    # Flush any unclosed type group
    if type_group['start'] is not None:
        _flush_type_merge(row - 1)

    # Apply cols 2-4 vertical merges
    for start_r, col, end_r, bg, label in merge_queue:
        ws.merge_cells(start_row=start_r, start_column=col,
                       end_row=end_r,     end_column=col)
        tc = ws.cell(start_r, col)
        tc.value     = label
        tc.fill      = _fill(bg)
        tc.font      = _font(bold=True, size=11, color=_TEXT)
        tc.alignment = Alignment(
            horizontal='left' if col == 2 else 'center',
            vertical='center', wrap_text=True)
        for r in range(start_r, end_r + 1):
            c = ws.cell(r, col)
            if start_r == end_r:
                c.border = _FULL_BDR
            elif r == start_r:
                c.border = _LRT_BDR
            elif r == end_r:
                c.border = _LRB_BDR
            else:
                c.border = _LR_BDR

    # Column widths
    for letter, width in [('A', 22), ('B', 32), ('C', 20), ('D', 20),
                           ('E', 18), ('F', 10), ('G', 10), ('H', 12)]:
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = 'A3'

    