from flask import render_template, session, redirect, url_for, jsonify, request
from datetime import date, datetime, timedelta
from functools import wraps
from database import get_db, get_cursor
from .. import bp
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def _classify_tide_types(rows):
    if not rows:
        return []
    heights = [float(r.get('tide_meters') or 0) for r in rows]
    current = 'HW' if len(heights) >= 2 and heights[0] > heights[1] else 'LW'
    types = []
    for _ in heights:
        types.append(current)
        current = 'LW' if current == 'HW' else 'HW'
    return types

def _parse_dt(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val).strip())
    except Exception:
        try:
            return datetime.strptime(str(val).strip(), '%Y-%m-%d %H:%M:%S')
        except Exception:
            return None

def _fmt_dt(val, strfmt='%d-%m-%Y %H:%M'):
    dt = _parse_dt(val)
    return dt.strftime(strfmt) if dt else ''

def _fetch_mother_vessels(from_datetime, to_datetime):

    conn = get_db()
    cur  = get_cursor(conn)

    cur.execute("""
        SELECT
            h.id,
            h.vcn_id,
            h.vessel_name,
            h.operation_type,
            h.nor_tendered,

            (
                SELECT MIN(a.discharge_started)
                FROM ldud_anchorage a
                WHERE a.ldud_id = h.id
            ) AS discharge_commenced,

            (
                SELECT MAX(a.discharge_commenced)
                FROM ldud_anchorage a
                WHERE a.ldud_id = h.id
            ) AS discharge_completed

        FROM ldud_header h
        WHERE h.nor_tendered IS NOT NULL
        ORDER BY h.nor_tendered
    """)

    all_vessels = [dict(r) for r in cur.fetchall()]

    vessels = []
    for v in all_vessels:
        commenced = _parse_dt(v.get("discharge_commenced"))
        completed = _parse_dt(v.get("discharge_completed"))

        if not commenced:
            continue

        if commenced > to_datetime:
            continue

        if completed and completed < from_datetime:
            continue

        vessels.append(v)

    ldud_ids = [v['id'] for v in vessels]
    vcn_ids  = [v['vcn_id'] for v in vessels if v.get('vcn_id')]

    bl_import        = {}
    bl_export        = {}
    vcn_meta         = {}
    ops_24h          = {}
    ops_till         = {}
    under_loading    = {}
    at_gull_loaded   = {}
    eta_to_dharamtar = {}
    mbc_eta_list     = []

    report_date = to_datetime.date()
    prev_date   = report_date - timedelta(days=1)

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id, COALESCE(SUM(quantity),0) qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
              AND TO_DATE(start_time,'YYYY-MM-DD') = %s
            GROUP BY ldud_id
        """, (ldud_ids, prev_date))
        for r in cur.fetchall():
            ops_24h[r['ldud_id']] = float(r['qty'])

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id, COALESCE(SUM(quantity),0) qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
              AND TO_DATE(start_time,'YYYY-MM-DD') <= %s
            GROUP BY ldud_id
        """, (ldud_ids, prev_date))
        for r in cur.fetchall():
            ops_till[r['ldud_id']] = float(r['qty'])

    if vcn_ids:
        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity),0) total
            FROM vcn_cargo_declaration
            WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        bl_import = {r['vcn_id']: float(r['total']) for r in cur.fetchall()}

        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity),0) total
            FROM vcn_export_cargo_declaration
            WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        bl_export = {r['vcn_id']: float(r['total']) for r in cur.fetchall()}

        cur.execute("""
            SELECT id, importer_exporter_name
            FROM vcn_header WHERE id = ANY(%s)
        """, (vcn_ids,))
        vcn_meta = {r['id']: r['importer_exporter_name'] or '' for r in cur.fetchall()}

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id,
                   STRING_AGG(TRIM(barge_name), ', ' ORDER BY barge_name) AS barges
            FROM ldud_barge_lines
            WHERE commenced_loading IS NOT NULL
              AND completed_loading IS NULL
              AND ldud_id = ANY(%s)
            GROUP BY ldud_id
        """, (ldud_ids,))
        under_loading = {r['ldud_id']: r['barges'] for r in cur.fetchall()}

        cur.execute("""
            SELECT ldud_id,
                   STRING_AGG(TRIM(barge_name), ', ' ORDER BY barge_name) AS barges
            FROM ldud_barge_lines
            WHERE cast_off_mv IS NOT NULL
              AND (along_side_berth IS NULL OR TRIM(COALESCE(along_side_berth,'')) = '')
              AND ldud_id = ANY(%s)
            GROUP BY ldud_id
        """, (ldud_ids,))
        at_gull_loaded = {r['ldud_id']: r['barges'] for r in cur.fetchall()}

        cur.execute("""
            SELECT ldud_id,
                STRING_AGG(TRIM(barge_name), ', ' ORDER BY barge_name) AS barges
            FROM ldud_barge_lines
            WHERE anchored_gull_island IS NOT NULL
            AND cast_off_port IS NULL
            AND ldud_id = ANY(%s)
            GROUP BY ldud_id
        """, (ldud_ids,))
        eta_to_dharamtar = {r['ldud_id']: r['barges'] for r in cur.fetchall()}

    cur.execute("""
        SELECT
            h.mbc_name,
            p.departure_gull_island,
            h.cargo_name,
            COALESCE(h.bl_quantity, 0) AS bl_qty
        FROM mbc_header h
        JOIN mbc_discharge_port_lines p ON p.mbc_id = h.id
        WHERE p.departure_gull_island IS NOT NULL
          AND TRIM(COALESCE(p.departure_gull_island, '')) <> ''
          AND (
                p.vessel_arrival_port IS NULL
                OR TRIM(COALESCE(p.vessel_arrival_port, '')) = ''
              )
          AND NULLIF(TRIM(p.departure_gull_island), '')::timestamp <= %s
        ORDER BY p.departure_gull_island
    """, (to_datetime,))
 
    mbc_eta_rows = cur.fetchall()
 
    # Format: "MBC_NAME (cargo) ETA: departure_time"
    mbc_eta_list = []
    for r in mbc_eta_rows:
        dep_dt = _parse_dt(r['departure_gull_island'])
        dep_str = dep_dt.strftime('%d-%m %H:%M') if dep_dt else ''
        entry = f"{r['mbc_name']} ({r['cargo_name'] or ''}) - Dep Gull: {dep_str}"
        mbc_eta_list.append(entry)

    cur.close()
    conn.close()

    for v in vessels:
        vid    = v.get('vcn_id')
        op     = v.get('operation_type', '')
        bl_qty = (bl_export.get(vid, 0) if op == 'Export' else bl_import.get(vid, 0))

        v['stevedore_group']  = vcn_meta.get(vid, '')
        v['bl_qty']           = bl_qty
        v['ops_24h']          = ops_24h.get(v['id'], 0)
        v['ops_till']         = ops_till.get(v['id'], 0)
        v['balance']          = round(bl_qty - ops_till.get(v['id'], 0), 2)
        v['under_loading']    = under_loading.get(v['id'], '')
        v['eta_to_dharamtar'] = eta_to_dharamtar.get(v['id'], '')
        v['wt_r19']           = ''
        v['at_gull_loaded']   = at_gull_loaded.get(v['id'], '')
        v['mbc_eta']          = ', '.join(mbc_eta_list)

    return vessels



def _fetch_tide_data(from_datetime, to_datetime):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT
            tide_datetime,
            tide_meters
        FROM tide_master
        WHERE
            tide_datetime IS NOT NULL
            AND TRIM(tide_datetime) <> ''
            AND NULLIF(TRIM(tide_datetime), '')::timestamp >= %s
        ORDER BY NULLIF(TRIM(tide_datetime), '')::timestamp
        LIMIT 6
    """, (from_datetime,))

    rows = [dict(r) for r in cur.fetchall()]

    cur.close()
    conn.close()

    types = _classify_tide_types(rows)

    tide_data = []

    for row, tide_type in zip(rows, types):
        dt = _parse_dt(row["tide_datetime"])

        tide_data.append({
            "type": tide_type,
            "time": dt.strftime("%d/%m %H:%M") if dt else "",
            "height": row["tide_meters"],
        })

    return tide_data




def get_shift_code(dt):
    if not dt:
        return None
    hour = dt.hour
    if 6 <= hour < 14:
        return "A"
    elif 14 <= hour < 22:
        return "B"
    else:
        return "C"


def _fetch_all_barges(selected_date=None, selected_shift="ALL"):

    conn = get_db()
    cur  = get_cursor(conn)

    barges = []
    occupied_berth_set = set()

    selected_dt = None
    if selected_date:
        try:
            selected_dt = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except Exception:
            pass

    cur.execute("""
        WITH discharge_sums AS (
            SELECT
                TRIM(UPPER(ll.barge_name)) AS barge_name,
                ll.source_id,
                SUM(COALESCE(ll.quantity,0)) AS discharged_qty
            FROM lueu_lines ll
            WHERE ll.is_deleted IS NOT TRUE
            AND ll.source_type = 'VCN'
            GROUP BY TRIM(UPPER(ll.barge_name)), ll.source_id
        )
        SELECT
            l.id,
            l.barge_name,
            l.trip_number,
            l.cargo_name,
            l.cast_off_port,
            l.along_side_berth,
            l.commence_discharge_berth,
            l.completed_discharge_berth,
            COALESCE(l.discharge_quantity,0) AS discharge_qty,
            (
                COALESCE(l.discharge_quantity,0)
                - COALESCE(ds.discharged_qty,0)
            ) AS balance_qty
        FROM ldud_barge_lines l
        LEFT JOIN ldud_header h
            ON h.id = l.ldud_id
        LEFT JOIN discharge_sums ds
            ON ds.barge_name = TRIM(
                UPPER(
                    CONCAT(
                        l.barge_name,
                        ' / ',
                        COALESCE(l.trip_number::text,'1')
                    )
                )
            )
            AND ds.source_id = h.vcn_id
        WHERE COALESCE(TRIM(l.barge_name),'') <> ''
    """)
    for row in cur.fetchall():
        row       = dict(row)

        
        cutoff_date = date(2026, 5, 1)

        if selected_dt and selected_dt < cutoff_date:
            continue

        # Skip completed
        if row.get("cast_off_port") and str(row.get("cast_off_port")).strip():
            continue

        if row.get("completed_discharge_berth") and str(row.get("completed_discharge_berth")).strip():
            continue

        # Waiting — alongside berth set, discharge not started
        if (
            row.get("along_side_berth")
            and str(row.get("along_side_berth")).strip()
            and (
                row.get("commence_discharge_berth") is None
                or str(row.get("commence_discharge_berth")).strip() == ""
            )
        ):
            status = "Waiting"

        # Under Discharge — discharge started, not completed, balance > 0
        elif (
            row.get("commence_discharge_berth")
            and str(row.get("commence_discharge_berth")).strip()
            and (
                row.get("completed_discharge_berth") is None
                or str(row.get("completed_discharge_berth")).strip() == ""
            )
            and float(row.get("balance_qty", 0) or 0) > 0
        ):
            status = "Under Discharge"

        else:
            continue

        # ← NO date filter, NO shift filter for barges

        completed_date = _fmt_dt(row.get("cast_off_port")) if row.get("cast_off_port") else None
        berth = (row.get("commence_discharge_berth") or row.get("along_side_berth") or "")

        barges.append({
            "id":             row["id"],
            "type":           "BARGE",
            "completed_date": completed_date,
            "barge_name":     row["barge_name"],
            "name":           row["barge_name"],
            "cargo": row.get("cargo_name") or row.get("cargo_type") or "",
            "qty":            row["discharge_qty"],
            "discharge_qty":  float(row["discharge_qty"]),
            "total_qty":      float(row["discharge_qty"]),
            "balance_qty": float(row.get("balance_qty", 0) or 0),
            "berth":          berth,
            "status":         status,
            "commence_discharge_berth": str(row.get("commence_discharge_berth") or "").strip(),
            "unloading_commenced":      "",   # barges use commence_discharge_berth
        })

    cur.execute("""
        WITH latest_mbc AS (
            SELECT
                h.id,
                h.mbc_name,
                h.cargo_name,
                COALESCE(h.bl_quantity,0) AS bl_qty,
                p.vessel_unloading_berth AS berth,
                p.vessel_arrival_port AS arrival_port,
                p.unloading_commenced,
                p.unloading_completed,
                p.vessel_cast_off AS mbc_cast_off,
                ROW_NUMBER() OVER (
                    PARTITION BY h.mbc_name
                    ORDER BY p.id DESC
                ) rn
            FROM mbc_header h
            JOIN mbc_discharge_port_lines p
                ON p.mbc_id = h.id
        )
        SELECT
            m.*,
            COALESCE(l.actual_qty,0) AS actual_qty
        FROM latest_mbc m
        LEFT JOIN (
            SELECT
                source_id,
                SUM(COALESCE(quantity,0)) AS actual_qty
            FROM lueu_lines
            WHERE source_type = 'MBC'
            AND is_deleted IS NOT TRUE
            GROUP BY source_id
        ) l ON l.source_id = m.id
        WHERE m.rn = 1
        AND m.arrival_port IS NOT NULL
        AND TRIM(COALESCE(m.arrival_port,'')) <> ''
        AND (
                m.unloading_completed IS NULL
                OR TRIM(COALESCE(m.unloading_completed,'')) = ''
            )
        AND (
                m.mbc_cast_off IS NULL
                OR TRIM(COALESCE(m.mbc_cast_off,'')) = ''
            )
        ORDER BY m.mbc_name
    """)
    for row in cur.fetchall():
        row = dict(row)

        cutoff_date = date(2026, 5, 1)

        if selected_dt and selected_dt < cutoff_date:
            continue

        berth = (row.get("berth") or "").strip()

        # Waiting
        if (
            row.get("arrival_port")
            and str(row.get("arrival_port")).strip()
            and (
                row.get("unloading_commenced") is None
                or str(row.get("unloading_commenced")).strip() == ""
            )
        ):
            status = "Waiting"

        # Discharging
        elif (
            row.get("unloading_commenced")
            and str(row.get("unloading_commenced")).strip()
        ):
            status = "Discharging"

        else:
            continue

        bl_qty = float(row["bl_qty"] or 0)
        actual_qty = float(row["actual_qty"] or 0)
        balance_qty = max(bl_qty - actual_qty, 0)
        # Hide fully discharged MBC
        if balance_qty <= 0:
            continue

        barges.append({
            "id": row["id"],
            "type": "MBC",
            "completed_date": _fmt_dt(row.get("unloading_completed")),
            "barge_name": row["mbc_name"],
            "name": row["mbc_name"],
            "cargo": row.get("cargo_name") or "",
            "qty": bl_qty,
            "discharge_qty": bl_qty,
            "total_qty": bl_qty,
            "balance_qty": balance_qty,
            "berth": "",
            "status": status,
            "unloading_commenced":      str(row.get("unloading_commenced") or "").strip(),
            "commence_discharge_berth": "",   # MBCs use unloading_commenced
        })

        if status == "Discharging" and berth:
            occupied_berth_set.add(berth)

    cur.close()
    conn.close()
    return barges, occupied_berth_set


# ── ROUTES — each defined exactly ONCE ───────────────────────────────────────

@bp.route('/module/RP01/Barge-Position-Report/')
@login_required
def barge_position_dashboard():

    barges, occupied_berth_set = _fetch_all_barges()

    waiting = [
    b for b in barges
    if b["status"] in ["Waiting", "Under Discharge"]
    ]

    discharging = [
        b for b in barges
        if b["status"] == "Discharging"
    ]
    occupied_berths = len(occupied_berth_set)

    today         = date.today()
    today_str     = today.strftime('%Y-%m-%d')

    from_date_str = request.args.get('from_date', today_str)
    from_time_str = request.args.get('from_time', '00:00')
    to_date_str   = request.args.get('to_date',   today_str)
    to_time_str   = request.args.get('to_time',   '23:59')

    to_datetime   = datetime.strptime(f"{to_date_str} {to_time_str}", '%Y-%m-%d %H:%M')

    # Widen window: start from previous day 00:00 so vessels that
    # completed early on the selected date are still included
    from_datetime = datetime.strptime(from_date_str, '%Y-%m-%d') - timedelta(days=1)
    from_datetime = from_datetime.replace(hour=0, minute=0, second=0)

    mother_vessels_raw = _fetch_mother_vessels(from_datetime, to_datetime)

    mother_vessels = [{
        'vessel_name':         v.get('vessel_name') or '',
        'stevedore_group':     v.get('stevedore_group') or '',
        'bl_qty':              v.get('bl_qty') or 0,
        'ops_24h':             v.get('ops_24h') or 0,
        'ops_till':            v.get('ops_till') or 0,
        'balance':             v.get('balance') or 0,
        'under_loading':       v.get('under_loading') or '',
        'eta_to_dharamtar':    v.get('eta_to_dharamtar') or '',
        'wt_r19':              v.get('wt_r19') or '',
        'at_gull_loaded':      v.get('at_gull_loaded') or '',
        'mbc_eta':             v.get('mbc_eta') or '',
        'nor_tendered':        _fmt_dt(v.get('nor_tendered')),
        'discharge_commenced': _fmt_dt(v.get('discharge_commenced')),
        'discharge_completed': _fmt_dt(v.get('discharge_completed')),
        'unloaded_till_date':  '',
        'disch_commenced':     '',
    } for v in mother_vessels_raw]
    
    conn = get_db()
    cur = get_cursor(conn)



    cur.execute("""
    SELECT berth_name
    FROM port_berth_master
    """)

    berths = [r["berth_name"].upper() for r in cur.fetchall()]

    old_berths = [
        "BERTH 1",
        "BERTH 2",
        "BERTH 3",
        "BERTH 4",
        "BERTH 5",
        "BERTH 5A",
    ]

    new_berths = [
        "BERTH 6",
        "BERTH 7",
        "BERTH 8",
        "BERTH 8A",
        "BERTH 9",
        "BERTH 10",
        "BERTH 11",
        "BERTH 12",
    ]

    cur.close()
    conn.close()

    tide_data = _fetch_tide_data(from_datetime, to_datetime)

    return render_template(
        "Barge_Position_Report/barge_dashboard.html",
        waiting=waiting,
        discharging=discharging,
        all_barges=barges,
        mother_vessels=mother_vessels,
        tide_data=tide_data,
        old_berths=old_berths,
        new_berths=new_berths,
        from_date=from_date_str,
        from_time=from_time_str,
        to_date=to_date_str,
        to_time=to_time_str,
        total_barges=len(barges),
        waiting_count=len(waiting),
        discharging_count=len(discharging),
        occupied_berths=occupied_berths,
        available_berths=max(0, 14 - occupied_berths),
        
    )


@bp.route('/api/module/RP01/shift-details')
@login_required
def get_shift_details():
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("SELECT name FROM port_shift_incharge ORDER BY name")
    shift_incharge_list = [r["name"] for r in cur.fetchall()]
    cur.execute("SELECT name FROM port_shift_operators ORDER BY name")
    crane_operator_list = [r["name"] for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({
        "shift_incharge_list": shift_incharge_list,
        "crane_operator_list": crane_operator_list,
    })


@bp.route('/api/module/RP01/mother-vessel-data')
@login_required
def mother_vessel_data():

    from_datetime = datetime.fromisoformat(request.args.get('from_datetime'))
    to_datetime   = datetime.fromisoformat(request.args.get('to_datetime'))

    # Widen window: previous day 00:00 so completed vessels are included
    window_start = datetime(
        from_datetime.year,
        from_datetime.month,
        from_datetime.day,
        0, 0, 0
    ) - timedelta(days=1)

    vessels = _fetch_mother_vessels(window_start, to_datetime)
    
    
    

    return jsonify([{
        "vessel_name":         v.get("vessel_name"),
        "nor_tendered":        _fmt_dt(v.get("nor_tendered")),
        "discharge_commenced": _fmt_dt(v.get("discharge_commenced")),
        "discharge_completed": _fmt_dt(v.get("discharge_completed")),
        "under_loading":       v.get("under_loading", ""),
        "eta_to_dharamtar":    v.get("eta_to_dharamtar", ""),
        "wt_r19":              "",
        "at_gull_loaded":      v.get("at_gull_loaded", ""),
        "mbc_eta":             v.get("mbc_eta", ""),
    } for v in vessels])


@bp.route('/api/module/RP01/tide-data')
@login_required
def tide_data_api():
    from_datetime = datetime.fromisoformat(request.args.get('from_datetime'))
    to_datetime   = datetime.fromisoformat(request.args.get('to_datetime'))
    return jsonify(_fetch_tide_data(from_datetime, to_datetime))


@bp.route('/api/module/RP01/berth-occupancy')
@login_required
def berth_occupancy():
    completed = request.args.get("completed") == "1"
    selected_date  = request.args.get('date')
    selected_shift = request.args.get('shift')

    items, _ = _fetch_all_barges(selected_date, selected_shift)
    

    # If popup requests completed data
    if completed:

        completed = []

        conn = get_db()
        cur = get_cursor(conn)

        # ---------------- COMPLETED BARGES ----------------
        cur.execute("""
            SELECT
                l.id,
                l.barge_name,
                l.cargo_name,
                COALESCE(l.discharge_quantity,0) AS qty,
                l.cast_off_port,
                l.completed_discharge_berth
            FROM ldud_barge_lines l
            WHERE
                l.cast_off_port IS NOT NULL
                OR l.completed_discharge_berth IS NOT NULL
        """)

        for row in cur.fetchall():
            row = dict(row)

            completed_dt = _parse_dt(
                row["cast_off_port"] or row["completed_discharge_berth"]
            )

            if completed_dt and completed_dt.date() == datetime.strptime(
                selected_date, "%Y-%m-%d"
            ).date():

                completed.append({
                    "type": "BARGE",
                    "name": row["barge_name"],
                    "cargo": row["cargo_name"] or "",
                    "qty": float(row["qty"] or 0),
                    "status": "Completed",
                    "completed_date": _fmt_dt(completed_dt)
                })

        # ---------------- COMPLETED MBC ----------------
        cur.execute("""
            SELECT
                h.id,
                h.mbc_name,
                h.cargo_name,
                COALESCE(h.bl_quantity,0) AS qty,
                p.unloading_completed,
                p.vessel_cast_off
            FROM mbc_header h
            JOIN mbc_discharge_port_lines p
                ON p.mbc_id=h.id
            WHERE
                p.unloading_completed IS NOT NULL
                OR p.vessel_cast_off IS NOT NULL
        """)

        for row in cur.fetchall():
            row = dict(row)

            completed_dt = (
                _parse_dt(row["vessel_cast_off"])
                or _parse_dt(row["unloading_completed"])
            )

            if completed_dt and completed_dt.date() == datetime.strptime(
                selected_date, "%Y-%m-%d"
            ).date():

                completed.append({
                    "type": "MBC",
                    "name": row["mbc_name"],
                    "cargo": row["cargo_name"] or "",
                    "qty": float(row["qty"] or 0),
                    "status": "Completed",
                    "completed_date": _fmt_dt(completed_dt)
                })

        cur.close()
        conn.close()

        return jsonify(completed)

    # Existing berth/waiting response remains unchanged
    return jsonify(items)
@bp.route('/api/module/RP01/berths')
@login_required
def get_berths():

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT
            id,
            berth_id,
            berth_name,
            berth_sequence
        FROM port_berth_master
        ORDER BY
            COALESCE(berth_sequence,999),
            berth_name
    """)

    data = [dict(r) for r in cur.fetchall()]

    cur.close()
    conn.close()

    return jsonify(data)

@bp.route('/api/module/RP01/shift-wise-discharge')
@login_required
def shift_wise_discharge():
    import traceback
    try:
        return _shift_wise_discharge_inner()
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

def _shift_wise_discharge_inner():
    selected_date = request.args.get('date', '')
    shift = request.args.get('shift', 'ALL')
    if not selected_date:
        return jsonify({'error': 'date is required'}), 400

    conn = get_db()
    cur = get_cursor(conn)
    is_all = shift.upper() == 'ALL'

    # ── 1. JETTY DISCHARGE ───────────────────────────────────────────────────
    if is_all:
        cur.execute("""
            SELECT cargo_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE entry_date = %s
              AND quantity > 0
              AND cargo_name IS NOT NULL AND cargo_name != ''
              AND is_deleted IS NOT TRUE
            GROUP BY cargo_name ORDER BY cargo_name
        """, (selected_date,))
    else:
        cur.execute("""
            SELECT cargo_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE entry_date = %s AND shift = %s
              AND quantity > 0
              AND cargo_name IS NOT NULL AND cargo_name != ''
              AND is_deleted IS NOT TRUE
            GROUP BY cargo_name ORDER BY cargo_name
        """, (selected_date, shift))
    jetty_rows = [dict(r) for r in cur.fetchall()]

    # ── 2. CLEANING DELAYS ───────────────────────────────────────────────────
    if is_all:
        cur.execute("""
            SELECT source_id, source_type, delay_name, barge_name
            FROM lueu_lines
            WHERE entry_date = %s
              AND is_deleted IS NOT TRUE
              AND delay_name IS NOT NULL AND delay_name != ''
              AND (
                  LOWER(delay_name) LIKE '%%payloader%%'
                  OR LOWER(delay_name) LIKE '%%Labor Cleaning%%'
                
              )
        """, (selected_date,))
    else:
        cur.execute("""
            SELECT source_id, source_type, delay_name,barge_name
            FROM lueu_lines
            WHERE entry_date = %s AND shift = %s
              AND is_deleted IS NOT TRUE
              AND delay_name IS NOT NULL AND delay_name != ''
              AND (
                  LOWER(delay_name) LIKE '%%payloader%%'
                  OR LOWER(delay_name) LIKE '%%Labor Cleaning%%'
                  
              )
        """, (selected_date, shift))

    delay_map = {}

    for row in cur.fetchall():
        key = (row['source_id'], row['source_type'])
        name = (row['delay_name'] or '').lower()

        if key not in delay_map:
            delay_map[key] = {
                'payloader': False,
                'labour': False
            }

        if 'payloader' in name:
            delay_map[key]['payloader'] = True
        else:
            delay_map[key]['labour'] = True


    # ── 3. BARGE DISCHARGE ──
    if is_all:
            cur.execute("""
                WITH actual AS (
                    SELECT
                        TRIM(UPPER(barge_name)) AS barge_key,
                        source_id,
                        SUM(COALESCE(quantity,0)) AS actual_qty
                    FROM lueu_lines
                    WHERE is_deleted IS NOT TRUE AND source_type = 'VCN'
                    AND entry_date = %s
                    GROUP BY 1, 2
                    HAVING SUM(COALESCE(quantity,0)) > 0
                )
                SELECT bl.id, bl.barge_name, bl.trip_number, bl.cargo_name,
                    COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                    COALESCE(a.actual_qty, 0) AS actual_discharge,
                    bl.along_side_berth, bl.commence_discharge_berth,
                    bl.completed_discharge_berth, bl.cast_off_port, h.vcn_id
                FROM ldud_barge_lines bl
                JOIN ldud_header h ON h.id = bl.ldud_id
                LEFT JOIN actual a
                    ON a.barge_key = TRIM(
                        UPPER(
                            CONCAT(
                                bl.barge_name,
                                ' / ',
                                COALESCE(bl.trip_number::text,'1')
                            )
                        )
                    )
                AND a.source_id = h.vcn_id
                WHERE COALESCE(TRIM(bl.barge_name),'') <> ''
                AND COALESCE(a.actual_qty,0) > 0
                ORDER BY bl.barge_name
            """, (selected_date,))
    else:
        cur.execute("""
                WITH actual AS (
                    SELECT
                        TRIM(UPPER(barge_name)) AS barge_key,
                        source_id,
                        SUM(COALESCE(quantity,0)) AS actual_qty
                    FROM lueu_lines
                    WHERE is_deleted IS NOT TRUE AND source_type = 'VCN'
                    AND entry_date = %s AND shift = %s
                    GROUP BY 1, 2
                )
                SELECT bl.id, bl.barge_name,bl.trip_number, bl.cargo_name,
                    COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                    COALESCE(a.actual_qty, 0) AS actual_discharge,
                    bl.along_side_berth, bl.commence_discharge_berth,
                    bl.completed_discharge_berth, bl.cast_off_port, h.vcn_id
                FROM ldud_barge_lines bl
                JOIN ldud_header h ON h.id = bl.ldud_id
                INNER JOIN actual a
                    ON a.barge_key = TRIM(
                        UPPER(
                            CONCAT(
                                bl.barge_name,
                                ' / ',
                                COALESCE(bl.trip_number::text,'1')
                            )
                        )
                    )
                AND a.source_id = h.vcn_id
                WHERE COALESCE(TRIM(bl.barge_name),'') <> ''
                AND COALESCE(a.actual_qty,0) > 0
                ORDER BY bl.barge_name
            """, (selected_date, shift))

    from datetime import date as date_cls
    cutoff = date_cls(2026, 5, 1)
    sel_date_obj = datetime.strptime(selected_date, '%Y-%m-%d').date()

    barge_discharge = []
    for row in cur.fetchall():
        row = dict(row)
        if sel_date_obj < cutoff:
            continue
        if row.get('cast_off_port') and str(row['cast_off_port']).strip():
            status = 'Completed'
        elif row.get('completed_discharge_berth') and str(row['completed_discharge_berth']).strip():
            status = 'Completed'
        elif row.get('commence_discharge_berth') and str(row['commence_discharge_berth']).strip():
            status = 'Under Discharge'
        else:
            status = 'Waiting'

        delays = delay_map.get(
            (row['vcn_id'], 'VCN'),
            {'payloader': False, 'labour': False}
        )
        barge_discharge.append({
            'type': 'BARGE',
            'name': row['barge_name'],
            'cargo': row.get('cargo_name') or '',
            'bl_qty': float(row['bl_qty'] or 0),
            'actual_discharge': float(row['actual_discharge'] or 0),
            'status': status,

            'payloader_cl': row['barge_name'] if delays['payloader'] else '',
            'labour_cleaned': row['barge_name'] if delays['labour'] else '',
        })

        # ── 4. MBC DISCHARGE ─────────────────────────────────────────────────────

        if is_all:
            cur.execute("""
                SELECT
                    l.source_id AS id,
                    l.barge_name AS mbc_name,
                    l.cargo_name,
                    COALESCE(h.bl_quantity,0) AS bl_qty,
                    SUM(COALESCE(l.quantity,0)) AS actual_discharge
                FROM lueu_lines l
                LEFT JOIN mbc_header h
                    ON h.id = l.source_id
                WHERE l.source_type = 'MBC'
                AND l.is_deleted IS NOT TRUE
                AND l.entry_date = %s
                AND TRIM(COALESCE(l.barge_name,'')) <> ''
                GROUP BY
                    l.source_id,
                    l.barge_name,
                    l.cargo_name,
                    h.bl_quantity
                HAVING SUM(COALESCE(l.quantity,0)) > 0
                ORDER BY l.barge_name
            """, (selected_date,))
        else:
            cur.execute("""
                SELECT
                    l.source_id AS id,
                    l.barge_name AS mbc_name,
                    l.cargo_name,
                    COALESCE(h.bl_quantity,0) AS bl_qty,
                    SUM(COALESCE(l.quantity,0)) AS actual_discharge
                FROM lueu_lines l
                LEFT JOIN mbc_header h
                    ON h.id = l.source_id
                WHERE l.source_type = 'MBC'
                AND l.is_deleted IS NOT TRUE
                AND l.entry_date = %s
                AND l.shift = %s
                AND TRIM(COALESCE(l.barge_name,'')) <> ''
                GROUP BY
                    l.source_id,
                    l.barge_name,
                    l.cargo_name,
                    h.bl_quantity
                HAVING SUM(COALESCE(l.quantity,0)) > 0
                ORDER BY l.barge_name
            """, (selected_date, shift))

    mbc_discharge = []

    for row in cur.fetchall():
        row = dict(row)

        delays = delay_map.get(
            (row['id'], 'MBC'),
            {'payloader': False, 'labour': False}
        )

        mbc_discharge.append({
            'type': 'MBC',
            'name': row['mbc_name'],
            'cargo': row.get('cargo_name') or '',
            'bl_qty': float(row['bl_qty'] or 0),
            'actual_discharge': float(row['actual_discharge'] or 0),
            'status': 'Discharging',
            'payloader_cl': row['mbc_name'] if delays['payloader'] else '',
            'labour_cleaned': row['mbc_name'] if delays['labour'] else '',
        })

    cur.close()
    conn.close()

    return jsonify({
        'jetty_discharge': jetty_rows,
        'barge_discharge': barge_discharge,
        'mbc_discharge': mbc_discharge,
    })
    
    
    


    
@bp.route('/api/module/RP01/download-barge-position-excel')
@login_required
def download_barge_position_excel():

    selected_date  = request.args.get('date', '')
    selected_shift = request.args.get('shift', 'ALL')
    shift_incharge = request.args.get('shift_incharge', '')
    bpo            = request.args.get('bpo', '')
    operator       = request.args.get('operator', '')

    barges, occupied_berth_set = _fetch_all_barges(selected_date, selected_shift)

    waiting     = [b for b in barges if b["status"] in ["Waiting", "Under Discharge"]]
    discharging = [b for b in barges if b["status"] == "Discharging"]

    from_datetime = datetime.combine(
        datetime.strptime(selected_date, "%Y-%m-%d").date(),
        datetime.min.time()
    ) - timedelta(days=1)

    to_datetime = datetime.combine(
        datetime.strptime(selected_date, "%Y-%m-%d").date(),
        datetime.max.time()
    )

    mother_vessels = _fetch_mother_vessels(from_datetime, to_datetime)
    tide_data      = _fetch_tide_data(from_datetime, to_datetime)

    # ── Fetch berth occupancy for board ──────────────────────────────────────
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("SELECT berth_name FROM port_berth_master ORDER BY berth_sequence, id")
    berths = [r["berth_name"].upper() for r in cur.fetchall()]
    cur.close()
    conn.close()

    old_berths = berths[:6]
    new_berths = berths[6:]
    positions  = ['A/S', 'D/B', 'T/B', 'F/B', 'S/B']

    # Build berth→position→item map
    berth_map = {}
    for item in barges:
        b = (item.get("berth") or "").strip().upper()
        p = (item.get("position") or "A/S").upper()
        if b:
            berth_map[(b, p)] = item

    # ── Styles ────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Barge Position"

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'),  bottom=Side(style='medium')
    )

    BLUE_HDR   = PatternFill("solid", fgColor="4D8CCD")   # header blue
    YELLOW_HDR = PatternFill("solid", fgColor="FFFF00")   # title yellow
    GREEN_CELL = PatternFill("solid", fgColor="C6EFCE")   # discharging
    BLUE_CELL  = PatternFill("solid", fgColor="DBEAFE")   # waiting-discharge
    GREY_CELL  = PatternFill("solid", fgColor="E5E7EB")   # completed
    WAIT_CELL  = PatternFill("solid", fgColor="FEF08A")   # waiting area

    white_bold   = Font(bold=True, color="FFFFFF", size=10)
    black_bold   = Font(bold=True, color="000000", size=10)
    normal_font  = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def style_cell(cell, fill=None, font=None, alignment=None, border=None):
        if fill:      cell.fill      = fill
        if font:      cell.font      = font
        if alignment: cell.alignment = alignment
        if border:    cell.border    = border

    def set_row_height(ws, row, height):
        ws.row_dimensions[row].height = height

    # ── ROW 1: Title bar (Image 1 top row) ───────────────────────────────────
    r = 1
    # A1: "date"
    ws.cell(r, 1, "date")
    style_cell(ws.cell(r,1), border=thin, alignment=center_align, font=black_bold)

    # B1: date value  (merged B1:C1)
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(r, 2, selected_date)
    style_cell(ws.cell(r,2), border=thin, alignment=center_align, font=black_bold)

    # D1: title (merged D1:H1)
    ws.merge_cells(f"D{r}:H{r}")
    ws.cell(r, 4, "DPPL OPRATION SHIFT WISE  REPORT")
    style_cell(ws.cell(r,4), border=thin, alignment=center_align, font=Font(bold=True, size=11))

    # I1: "SHIFT"
    ws.cell(r, 9, "SHIFT")
    style_cell(ws.cell(r,9), border=thin, alignment=center_align, font=black_bold)

    # J1: shift value  (merged J1:K1)
    ws.merge_cells(f"J{r}:K{r}")
    shift_label = {"A": "A- SHIFT", "B": "B- SHIFT", "C": "C- SHIFT"}.get(selected_shift, "ALL SHIFT")
    ws.cell(r, 10, shift_label)
    style_cell(ws.cell(r,10), border=thin, alignment=center_align,
               font=Font(bold=True, size=10, color="006400"))

    # L1: DOC NO
    ws.cell(r, 12, "DOC NO OPE/0100/F/01")
    style_cell(ws.cell(r,12), border=thin, alignment=center_align, font=black_bold)

    # M1: REV
    ws.cell(r, 13, "REV.02")
    style_cell(ws.cell(r,13), border=thin, alignment=center_align, font=black_bold)

    # N1: ISSUE NO
    ws.merge_cells(f"N{r}:O{r}")
    ws.cell(r, 14, "ISSUE NO. isuue date:01.04.2022")
    style_cell(ws.cell(r,14), border=thin, alignment=center_align, font=black_bold)

    set_row_height(ws, r, 20)

    # ── ROW 2: Shift Incharge / BPO / Operator ───────────────────────────────
    r = 2
    ws.cell(r, 1, "shift incharge name")
    style_cell(ws.cell(r,1), border=thin, alignment=center_align, font=black_bold)

    ws.merge_cells(f"B{r}:H{r}")
    ws.cell(r, 2, shift_incharge.upper() or "PRASHANT MHATRE")
    style_cell(ws.cell(r,2), border=thin, alignment=center_align, font=black_bold)

    ws.cell(r, 9, "BPO")
    style_cell(ws.cell(r,9), border=thin, alignment=center_align, font=black_bold)

    ws.merge_cells(f"J{r}:K{r}")
    ws.cell(r, 10, bpo.upper() or "BPO NAME")
    style_cell(ws.cell(r,10), border=thin, alignment=center_align, font=black_bold)

    ws.cell(r, 12, "operator")
    style_cell(ws.cell(r,12), border=thin, alignment=center_align, font=black_bold)

    ws.merge_cells(f"N{r}:O{r}")
    ws.cell(r, 14, operator.upper() or "oprater name")
    style_cell(ws.cell(r,14), border=thin, alignment=center_align, font=black_bold)

    set_row_height(ws, r, 20)

    r = 3  # blank separator

    # ── SUMMARY METRICS ROW ───────────────────────────────────────────────────
    r = 4
    summary_items = [
        ("TOTAL MBC & BARGES IN PORT", len(barges)),
        ("WAITING",                    len(waiting)),
        ("DISCHARGING",                len(discharging)),
        ("OCCUPIED BERTHS",            len(occupied_berth_set)),
        ("AVAILABLE BERTHS",           max(0, 14 - len(occupied_berth_set))),
    ]
    col = 1
    for label, val in summary_items:
        ws.cell(r,   col, label)
        ws.cell(r+1, col, val)
        style_cell(ws.cell(r,col),   fill=BLUE_HDR, font=white_bold,
                   alignment=center_align, border=thin)
        style_cell(ws.cell(r+1,col), font=black_bold,
                   alignment=center_align, border=thin)
        col += 3
    set_row_height(ws, r,   22)
    set_row_height(ws, r+1, 20)
    r += 3

    # ── BOARD MATRIX — OLD BERTH (left) + NEW BERTHS (right) side by side ────
    r += 1
    positions  = ['A/S', 'D/B', 'T/B', 'F/B', 'S/B']

    # Column layout:
    # Col 1      = OLD BERTH label
    # Col 2-6    = OLD positions (A/S, D/B, T/B, F/B, S/B)
    # Col 7      = empty gap
    # Col 8      = NEW BERTH label
    # Col 9-13   = NEW positions (A/S, D/B, T/B, F/B, S/B)

    OLD_START = 1   # berth label col
    OLD_POS   = 2   # first position col
    GAP_COL   = 7
    NEW_START = 8
    NEW_POS   = 9

    # ── Section title row ────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, OLD_START, f"OLD BERTH  ({len(old_berths)} BERTHS)")
    style_cell(ws.cell(r, OLD_START), fill=BLUE_HDR, font=white_bold,
               alignment=center_align, border=thin)

    ws.cell(r, GAP_COL, "")  # gap

    ws.merge_cells(f"H{r}:M{r}")
    ws.cell(r, NEW_START, f"NEW BERTHS  ({len(new_berths)} BERTHS)")
    style_cell(ws.cell(r, NEW_START), fill=BLUE_HDR, font=white_bold,
               alignment=center_align, border=thin)
    set_row_height(ws, r, 20)
    r += 1

    # ── Column header row ────────────────────────────────────────────────────
    matrix_hdrs = ["BERTH"] + positions

    for ci, h in enumerate(matrix_hdrs, OLD_START):
        ws.cell(r, ci, h)
        style_cell(ws.cell(r, ci), fill=BLUE_HDR, font=white_bold,
                   alignment=center_align, border=thin)

    ws.cell(r, GAP_COL, "")  # gap

    for ci, h in enumerate(matrix_hdrs, NEW_START):
        ws.cell(r, ci, h)
        style_cell(ws.cell(r, ci), fill=BLUE_HDR, font=white_bold,
                   alignment=center_align, border=thin)
    set_row_height(ws, r, 18)
    r += 1

    # ── Data rows — OLD and NEW side by side ─────────────────────────────────
    max_rows = max(len(old_berths), len(new_berths))

    for i in range(max_rows):

        # ── OLD BERTH side ───────────────────────────────────────────────────
        if i < len(old_berths):
            berth = old_berths[i]
            ws.cell(r, OLD_START, berth)
            style_cell(ws.cell(r, OLD_START), font=black_bold,
                       alignment=left_align, border=thin)

            for ci, pos in enumerate(positions, OLD_POS):
                item = berth_map.get((berth.upper(), pos))
                if item:
                    dqty = float(item.get('discharge_qty', 0))
                    bal  = float(item.get('balance_qty',  0))
                    txt  = (f"{item['type']} \u2013 {item['name']}\n"
                            f"{item['cargo']}\n"
                            f"Disch: {dqty} MT\n"
                            f"Bal: {bal} MT")
                    fill = (GREEN_CELL if dqty > 0 and bal > 0
                            else BLUE_CELL if bal > 0
                            else GREY_CELL)
                    style_cell(ws.cell(r, ci, txt), fill=fill,
                               font=normal_font, alignment=center_align, border=thin)
                else:
                    style_cell(ws.cell(r, ci, "\u2014"),
                               alignment=center_align, border=thin, font=normal_font)
        else:
            # fill empty cols so border still shows
            for ci in range(OLD_START, OLD_START + 6):
                style_cell(ws.cell(r, ci, ""),
                           alignment=center_align, border=thin, font=normal_font)

        # ── GAP col ──────────────────────────────────────────────────────────
        ws.cell(r, GAP_COL, "")

        # ── NEW BERTH side ───────────────────────────────────────────────────
        if i < len(new_berths):
            berth = new_berths[i]
            ws.cell(r, NEW_START, berth)
            style_cell(ws.cell(r, NEW_START), font=black_bold,
                       alignment=left_align, border=thin)

            for ci, pos in enumerate(positions, NEW_POS):
                item = berth_map.get((berth.upper(), pos))
                if item:
                    dqty = float(item.get('discharge_qty', 0))
                    bal  = float(item.get('balance_qty',  0))
                    txt  = (f"{item['type']} \u2013 {item['name']}\n"
                            f"{item['cargo']}\n"
                            f"Disch: {dqty} MT\n"
                            f"Bal: {bal} MT")
                    fill = (GREEN_CELL if dqty > 0 and bal > 0
                            else BLUE_CELL if bal > 0
                            else GREY_CELL)
                    style_cell(ws.cell(r, ci, txt), fill=fill,
                               font=normal_font, alignment=center_align, border=thin)
                else:
                    style_cell(ws.cell(r, ci, "\u2014"),
                               alignment=center_align, border=thin, font=normal_font)
        else:
            for ci in range(NEW_START, NEW_START + 6):
                style_cell(ws.cell(r, ci, ""),
                           alignment=center_align, border=thin, font=normal_font)

        set_row_height(ws, r, 55)
        r += 1

    r += 1

    # ── WAITING AREA ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, "WAITING AREA")
    style_cell(ws.cell(r,1), fill=YELLOW_HDR, font=black_bold,
               alignment=center_align, border=thin)
    set_row_height(ws, r, 18)
    r += 1

    for ci, h in enumerate(["TYPE","NAME","CARGO","DISCHARGE (MT)","BALANCE (MT)","STATUS"], 1):
        ws.cell(r, ci, h)
        style_cell(ws.cell(r,ci), fill=BLUE_HDR, font=white_bold,
                   alignment=center_align, border=thin)
    r += 1

    for item in waiting:
        data = [item["type"], item["name"], item["cargo"],
                item["discharge_qty"], item["balance_qty"], item["status"]]
        for ci, v in enumerate(data, 1):
            cell = ws.cell(r, ci, v)
            style_cell(cell, fill=WAIT_CELL, font=normal_font,
                       alignment=center_align, border=thin)
        set_row_height(ws, r, 18)
        r += 1

    r += 1

    # ── MOTHER VESSEL ─────────────────────────────────────────────────────────
    mv_params = [
        ("VSL DISCH COMMNACED",    "discharge_commenced"),
        ("VSL DISCHARGE COMPLITED","discharge_completed"),
        ("UNDER LOADING",          "under_loading"),
        ("ETA TO DHARAMTAR",       "eta_to_dharamtar"),
        ("WT @ R19",               "wt_r19"),
        ("ON THE WAY TO GULL",     "at_gull_loaded"),
        ("MBC ETA",                "mbc_eta"),
    ]

    # Header row
    ws.cell(r, 1, "Parameter")
    style_cell(ws.cell(r,1), fill=BLUE_HDR, font=white_bold,
               alignment=center_align, border=thin)
    for vi, mv in enumerate(mother_vessels, 2):
        ws.cell(r, vi, f"Vessel {vi-1}\n{mv.get('vessel_name','')}")
        style_cell(ws.cell(r,vi), fill=BLUE_HDR, font=white_bold,
                   alignment=center_align, border=thin)
    set_row_height(ws, r, 30)
    r += 1

    for label, key in mv_params:
        ws.cell(r, 1, label)
        style_cell(ws.cell(r,1), font=black_bold, alignment=left_align, border=thin)
        for vi, mv in enumerate(mother_vessels, 2):
            cell = ws.cell(r, vi, mv.get(key, "") or "")
            style_cell(cell, font=normal_font, alignment=center_align, border=thin)
        set_row_height(ws, r, 18)
        r += 1

    r += 1

    # ── TIDE TABLE ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(r, 1, "TIDE TABLE")
    style_cell(ws.cell(r,1), fill=BLUE_HDR, font=white_bold,
               alignment=center_align, border=thin)
    set_row_height(ws, r, 18)
    r += 1

    for ci, h in enumerate(["TYPE","TIME","HEIGHT (m)"], 1):
        ws.cell(r, ci, h)
        style_cell(ws.cell(r,ci), fill=BLUE_HDR, font=white_bold,
                   alignment=center_align, border=thin)
    r += 1

    for t in tide_data:
        for ci, v in enumerate([t["type"], t["time"], t["height"]], 1):
            cell = ws.cell(r, ci, v)
            fill = GREEN_CELL if t["type"] == "HW" else BLUE_CELL
            style_cell(cell, fill=fill, font=normal_font,
                       alignment=center_align, border=thin)
        set_row_height(ws, r, 16)
        r += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
            1: 14,  # OLD berth label
            2: 20, 3: 18, 4: 18, 5: 18, 6: 18,   # OLD positions
            7: 3,                                   # GAP
            8: 14,  # NEW berth label
            9: 20, 10: 18, 11: 18, 12: 18, 13: 18, # NEW positions
            14: 20, 15: 10, 16: 26,
        }
    from openpyxl.utils import get_column_letter
    for cn, w in col_widths.items():
        ws.column_dimensions[get_column_letter(cn)].width = w

    # ── Send file ─────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"Daily_Barge_Position_{selected_date}_{selected_shift}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    