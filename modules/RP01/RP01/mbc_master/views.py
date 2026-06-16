from flask import render_template, request, jsonify, session, redirect, url_for, Response
from functools import wraps
from datetime import datetime, date as date_type
import io
from .. import bp
from database import get_db, get_cursor
from datetime import datetime


# ── column definitions ─────────────────────────────────────────────────────────
# (header_label, data_key, is_calculated)
COLUMNS = [
    ('Sr. No',                        'sr_no',                    False),
    ('MBC Date',                      'mbc_date',                 False),
    ('Name of MBC',                   'mbc_name',                 False),
    ('Month',                         'month',                    False),
    ('FY Year',                       'fy_year',                  False),
    ('Date',                          'date',                     False),
    ('Arrived & Anchored',            'arrived_anchored',         False),
    ('Waiting (Perb-thing)',          'waiting_perb',             True),
    ('Loading Commenced',             'loading_commenced',        False),
    ('Loading Completed',             'loading_completed',        False),
    ('Loading',                       'loading_dur',              True),
    ('Castoff',                       'castoff',                  False),
    ('Waiting for Castoff',           'waiting_castoff',          True),
    ('Cargo',                         'cargo',                    False),
    ('Qty In',                        'qty_in',                   False),
    ('At Jaigad',                     'at_jaigad',                True),
    ('Transit - Jaigad to Dharamtar', 'transit_jaigad_dharamtar', True),
    ('Jaigad to Gull',                'jaigad_to_gull',           True),
    ('Arrived at Gull',               'arrived_gull',             False),
    ('Dept. from Gull',               'dept_gull',                False),
    ('Gull Waiting',                  'gull_waiting',             True),
    ('Arrived at Dharamtar',          'arrived_dharamtar',        False),
    ('Transit - Gull to DPPL',        'transit_gull_dppl',        True),
    ('Waiting at DPPL',               'waiting_dppl',             True),
    ('Discharge Commenced',           'discharge_commenced',      False),
    ('Discharge Completed',           'discharge_completed',      False),
    ('Unloading',                     'unloading_dur',            True),
    ('Partly Disch Stop',             'partly_disch_stop',        False),
    ('Partly Disch Start',            'partly_disch_start',       False),
    ('Disch Stop Time Taken',         'disch_stop_time',          True),
    ('Sailed from Dharamtar',         'sailed_dharamtar',         False),
    ('Breakdown Start',               'breakdown_start',          False),   # intentionally blank
    ('Breakdown End',                 'breakdown_end',            False),
    ('Breakdown Time Taken',          'breakdown_time',           True),
    ('Arrived & Anchored @ Jaigad',   'arrived_jaigad',           False),
    ('Waiting to Sail from Dharamtar','waiting_sail',             True),
    ('TAT',                           'tat',                      True),
]

# Keys whose Excel cells must always be written as empty (intentionally blank)
_BLANK_KEYS = {'breakdown_start'}


# ── display helpers ────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def _parse(ts):
    if not ts:
        return None
    if isinstance(ts, datetime):
        return ts
    if isinstance(ts, date_type):
        return datetime(ts.year, ts.month, ts.day)
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(ts), fmt)
        except ValueError:
            continue
    return None


def _dur(ts_from, ts_to):
    a, b = _parse(ts_from), _parse(ts_to)
    if not a or not b:
        return ''
    delta = b - a
    if delta.total_seconds() < 0:
        return ''
    total_minutes = int(delta.total_seconds() // 60)
    h, m = divmod(total_minutes, 60)
    return f'{h}:{m:02d}'


def _fmt_dt(ts):
    dt = _parse(ts)
    return dt.strftime('%d-%m-%Y %H:%M') if dt else ''


def _fmt_date(ts):
    if not ts:
        return ''
    try:
        if hasattr(ts, 'strftime'):
            return ts.strftime('%d-%m-%Y')
        return datetime.strptime(str(ts)[:10], '%Y-%m-%d').strftime('%d-%m-%Y')
    except Exception:
        return str(ts)


def _fy(doc_date):
    try:
        dt = doc_date if hasattr(doc_date, 'month') else datetime.strptime(str(doc_date)[:10], '%Y-%m-%d')
        y, m = dt.year, dt.month
        return f'FY {str(y)[2:]}-{str(y+1)[2:]}' if m >= 4 else f'FY {str(y-1)[2:]}-{str(y)[2:]}'
    except Exception:
        return ''


def _month_label(doc_date):
    try:
        dt = doc_date if hasattr(doc_date, 'strftime') else datetime.strptime(str(doc_date)[:10], '%Y-%m-%d')
        return dt.strftime('%b-%y')
    except Exception:
        return ''


# ── TAT helpers ────────────────────────────────────────────────────────────────
def _diff_mins(ts_from, ts_to):
    """Duration in minutes as float, or None if invalid/negative."""
    a, b = _parse(ts_from), _parse(ts_to)
    if not a or not b:
        return None
    delta = (b - a).total_seconds()
    return delta / 60 if delta >= 0 else None


def _avg_dur(values):
    """Average list of float-minutes and format as H:MM; '—' if no valid data."""
    valid = [v for v in values if v is not None]
    if not valid:
        return '—'
    avg = sum(valid) / len(valid)
    h, m = divmod(int(round(avg)), 60)
    return f'{h}:{m:02d}'


def _fy_date_range(d):
    """Return (fy_start_iso, fy_end_iso) strings for the FY containing date d."""
    y, m = d.year, d.month
    if m >= 4:
        return f'{y}-04-01', f'{y + 1}-03-31'
    return f'{y - 1}-04-01', f'{y}-03-31'


def _fy_label_str(d):
    y, m = d.year, d.month
    if m >= 4:
        return f'FY {str(y)[2:]}-{str(y + 1)[2:]}'
    return f'FY {str(y - 1)[2:]}-{str(y)[2:]}'


def _fetch_raw_trips(from_date='', to_date='', month_filter='', fy_filter=''):
    """
    Fetch raw timestamp rows for TAT computation (Import only).
    Applies the same date/month/fy filters as _fetch_rows.
    Returns plain dicts (connection-independent).
    """
    conn = get_db()
    cur  = get_cursor(conn)

    where_clauses = ["h.operation_type = 'Import'"]
    params = []
    if from_date == to_date and from_date:
        where_clauses.append("DATE(h.doc_date) = %s")
        params.append(from_date)
    else:
        if from_date:
            where_clauses.append("DATE(h.doc_date) >= %s")
            params.append(from_date)

        if to_date:
            where_clauses.append("DATE(h.doc_date) <= %s")
            params.append(to_date)

    cur.execute(f"""
        SELECT
            h.doc_date,
            lp.arrived_load_port,   lp.loading_commenced,  lp.loading_completed,
            lp.cast_off_load_port,
            dp.arrival_gull_island, dp.departure_gull_island, dp.vessel_arrival_port,
            dp.unloading_commenced, dp.unloading_completed,
            dp.vessel_cast_off,     dp.sailed_out_load_port
        FROM mbc_header h
        LEFT JOIN mbc_load_port_lines      lp ON lp.mbc_id = h.id
        LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
        WHERE {' AND '.join(where_clauses)}
        ORDER BY h.doc_date ASC, h.id ASC
    """, params)

    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    # Apply month / FY filters (same logic as _fetch_rows)
    if month_filter or fy_filter:
        filtered = []
        for r in rows:
            if month_filter and _month_label(r.get('doc_date', '')) != month_filter:
                continue
            if fy_filter and _fy(r.get('doc_date', '')) != fy_filter:
                continue
            filtered.append(r)
        return filtered

    return rows


def _compute_tat_metrics(rows):
    """
    Compute average duration (in float minutes) for each TAT segment.
    Returns a dict: segment_key → average H:MM string (or '—').
    """
    buckets = {
        'preberthing':       [],
        'loading':           [],
        'wait_after_load':   [],
        'total_jaigad':      [],
        'jaigad_to_gull':    [],
        'gull_waiting':      [],
        'gull_to_dhar':      [],
        'jaigad_to_dhar':    [],
        'preberthing_dhar':  [],
        'unloading':         [],
        'wait_after_unload': [],
        'total_dharamtar':   [],
        'dhar_to_jaigad':    [],
        'tat':               [],
    }
    for r in rows:
        al = r.get('arrived_load_port')
        lc = r.get('loading_commenced')
        lp = r.get('loading_completed')
        co = r.get('cast_off_load_port')
        ag = r.get('arrival_gull_island')
        dg = r.get('departure_gull_island')
        ad = r.get('vessel_arrival_port')
        uc = r.get('unloading_commenced')
        up = r.get('unloading_completed')
        cd = r.get('vessel_cast_off')
        so = r.get('sailed_out_load_port')

        buckets['preberthing'].append(_diff_mins(al, lc))
        buckets['loading'].append(_diff_mins(lc, lp))
        buckets['wait_after_load'].append(_diff_mins(lp, co))
        buckets['total_jaigad'].append(_diff_mins(al, co))
        buckets['jaigad_to_gull'].append(_diff_mins(co, ag))
        buckets['gull_waiting'].append(_diff_mins(ag, dg))
        buckets['gull_to_dhar'].append(_diff_mins(dg, ad))
        buckets['jaigad_to_dhar'].append(_diff_mins(co, ad))
        buckets['preberthing_dhar'].append(_diff_mins(ad, uc))
        buckets['unloading'].append(_diff_mins(uc, up))
        buckets['wait_after_unload'].append(_diff_mins(up, cd))
        buckets['total_dharamtar'].append(_diff_mins(ad, cd))
        buckets['dhar_to_jaigad'].append(_diff_mins(cd, so))
        buckets['tat'].append(_diff_mins(al, so))

    return {k: _avg_dur(v) for k, v in buckets.items()}


def _write_tat_sheet(
    ws,
    period_label, mtd_label, ytd_label,
    period_count, mtd_count, ytd_count,
    period_m, mtd_m, ytd_m
    ):
    """
    Write the MBC TAT Report sheet into an openpyxl Worksheet `ws`.
    Colours match the reference image exactly.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── palette ──────────────────────────────────────────────────────
    C_GREEN  = 'CCFF99'
    C_YELLOW = 'FFFF00'
    C_ORANGE = 'FFCC00'
    C_CYAN   = '00FFFF'
    C_WHITE  = 'FFFFFF'

    thin  = Side(style='thin', color='000000')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def fill(hex_c):
        return PatternFill('solid', fgColor=hex_c)

    def fnt(bold=False, color='000000', size=11, italic=False):
        return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)

    def _cell(row, col, val, font, bg, aln):
        c = ws.cell(row, col, val)
        c.font, c.fill, c.border, c.alignment = font, fill(bg), bdr, aln
        return c

    # ── column widths ────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 62
    ws.column_dimensions['B'].width = 14   # Target
    ws.column_dimensions['C'].width = 14   # Date
    ws.column_dimensions['D'].width = 14   # MTD
    ws.column_dimensions['E'].width = 14   # YTD

    

    # ── Row 1: period super-headers ──────────────────────────────────
    ws.row_dimensions[1].height = 18
    _cell(1, 1, '', fnt(), C_GREEN, ctr)
    _cell(1, 2, '', fnt(), C_GREEN, ctr)
    _cell(1, 3, period_label, fnt(bold=True, color='CC0000'), C_GREEN, ctr)
    _cell(1, 4, 'MTD', fnt(bold=True, color='CC0000'), C_GREEN, ctr)
    _cell(1, 5, 'YTD', fnt(bold=True, color='CC6600'), C_GREEN, ctr)


    # ── Row 2: column headers ────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    for ci, val in enumerate(
     ['Activity', 'Target (Hrs)', period_label, mtd_label, ytd_label], 1
     ):
       _cell(2, ci, val, fnt(bold=True), C_GREEN, ctr)

    # ── Row 3: Trips ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    for ci, val in enumerate(
     ['Trips', '', period_count, mtd_count, ytd_count], 1
     ):
      _cell(3, ci, val, fnt(bold=True, size=12), C_YELLOW, ctr)

    # ── Activity rows definition ─────────────────────────────────────
    # (label, metric_key, row_style)
    # row_style: 'data' | 'section_total' | 'main_total' | 'tat'
    ACTIVITIES = [
    ('Jaigad Arrival - Jaigad Loading Commenced (Preberthing delay)', 'preberthing', '1:00', 'data'),
    ('Loading Commence - Loading Completion (Loading time)', 'loading', '6:00', 'data'),
    ('Loading Completed - Cast Off from Jaigad (Waiting after loading)', 'wait_after_load', '0:00', 'data'),
    ('Total time taken at Jaigad', 'total_jaigad', '7:00', 'section_total'),

    ('Jaigad Departure to Gull Arrival (Loaded Transit time)', 'jaigad_to_gull', '12:00', 'data'),
    ('Gull Arrival - Gull Departure (Waiting at Gull)', 'gull_waiting', '10:00', 'data'),
    ('Gull Departure - Dharamtar Arrival', 'gull_to_dhar', '4:00', 'data'),
    ('Jaigad Departure - Dharamtar Arrival (Jaigad to Dharamtar)', 'jaigad_to_dhar', '26:00', 'main_total'),

    ('Dharamtar Arrival to Disch Commenced (Preberthing delay)', 'preberthing_dhar', '4:00', 'data'),
    ('Disch Commended to Disch Completed (Unloading Time)', 'unloading', '6:00', 'data'),
    ('Disch Completed to Cast Off from Dharamtar (Waiting after Unloading)', 'wait_after_unload', '1:00', 'data'),
    ('Total time taken at Dharamtar', 'total_dharamtar', '11:00', 'section_total'),

    ('Dharamtar Departure to Jaigad Arrival', 'dhar_to_jaigad', '16:00', 'main_total'),
    ('TAT', 'tat', '60:00', 'tat'),
]     

    STYLE_BG   = {'data': C_WHITE,  'section_total': C_YELLOW, 'main_total': C_ORANGE, 'tat': C_CYAN}
    STYLE_BOLD = {'data': False,    'section_total': True,     'main_total': True,      'tat': True}

    for ri, (label, key, target, style) in enumerate(ACTIVITIES, start=4):
        ws.row_dimensions[ri].height = 18
        bg      = STYLE_BG[style]
        is_bold = STYLE_BOLD[style]
        fnt_row = fnt(bold=is_bold, size=11)
        vals = [
            (1, label, left),
            (2, target, ctr),
            (3, period_m.get(key, '—'), ctr),
            (4, mtd_m.get(key, '—'), ctr),
            (5, ytd_m.get(key, '—'), ctr),
        ]
        for ci, val, aln in vals:
            _cell(ri, ci, val, fnt_row, bg, aln)

    # ── Footnote rows ─────────────────────────────────────────────────
    fn_start = 4 + len(ACTIVITIES)
    footnotes = [
        '*Unloading time includes the shifting time after partial discharge of MBC & '
        'discharge kept on hold to give priority for other MBC.',
        '*Preberthing time includes the MBC waiting at Berth 10 for the unloading turn.',
    ]
    for i, fn in enumerate(footnotes):
        r = fn_start + i
        ws.row_dimensions[r].height = 28
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(r, 1, fn)
        c.font      = fnt(italic=True, size=9, color='444444')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = bdr


# ── shared data builder ────────────────────────────────────────────────────────
def _fetch_rows(from_date, to_date, month_filter, fy_filter):
    conn = get_db()
    cur  = get_cursor(conn)

    where_clauses = ["h.operation_type = 'Import'"]
    params = []
    if from_date == to_date and from_date:
        where_clauses.append("DATE(h.doc_date) = %s")
        params.append(from_date)
    else:
        if from_date:
            where_clauses.append("DATE(h.doc_date) >= %s")
            params.append(from_date)

        if to_date:
            where_clauses.append("DATE(h.doc_date) <= %s")
            params.append(to_date)

    cur.execute(f"""
        SELECT
            h.id, h.doc_date, h.mbc_name, h.cargo_name, h.bl_quantity, h.quantity_uom,
            lp.arrived_load_port, lp.loading_commenced, lp.loading_completed, lp.cast_off_load_port,
            dp.arrival_gull_island, dp.departure_gull_island, dp.vessel_arrival_port,
            dp.unloading_commenced, dp.unloading_completed,
            dp.discharge_stop_shifting, dp.discharge_start_shifting,
            dp.vessel_cast_off, dp.cleaning_commenced, dp.cleaning_completed, dp.sailed_out_load_port
        FROM mbc_header h
        LEFT JOIN mbc_load_port_lines lp ON lp.mbc_id = h.id
        LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
        WHERE {' AND '.join(where_clauses)}
        ORDER BY h.doc_date ASC, h.id ASC
    """, params)

    rows = cur.fetchall()
    conn.close()

    result = []
    sr = 1
    for r in rows:
        month = _month_label(r['doc_date'])
        fy    = _fy(r['doc_date'])
        if month_filter and month != month_filter:
            continue
        if fy_filter and fy != fy_filter:
            continue

        al = r['arrived_load_port'];   lc = r['loading_commenced']
        lp = r['loading_completed'];   co = r['cast_off_load_port']
        ag = r['arrival_gull_island']; dg = r['departure_gull_island']
        ad = r['vessel_arrival_port']; uc = r['unloading_commenced']
        up = r['unloading_completed']; ds = r['discharge_stop_shifting']
        dt = r['discharge_start_shifting']; cd = r['vessel_cast_off']
        cc = r['cleaning_commenced'];  cx = r['cleaning_completed']
        so = r['sailed_out_load_port']

        result.append({
            'sr_no':                    sr,
            'mbc_date':                 _fmt_date(r['doc_date']),
            'mbc_name':                 r['mbc_name'] or '',
            'month':                    month,
            'fy_year':                  fy,
            'date':                     _fmt_date(al),
            'arrived_anchored':         _fmt_dt(al),
            'waiting_perb':             _dur(al, lc),
            'loading_commenced':        _fmt_dt(lc),
            'loading_completed':        _fmt_dt(lp),
            'loading_dur':              _dur(lc, lp),
            'castoff':                  _fmt_dt(co),
            'waiting_castoff':          _dur(lp, co),
            'cargo':                    r['cargo_name'] or '',
            'qty_in':                   r['bl_quantity'] if r['bl_quantity'] is not None else '',
            'at_jaigad':                _dur(al, co),
            'transit_jaigad_dharamtar': _dur(co, ad),
            'jaigad_to_gull':           _dur(co, ag),
            'arrived_gull':             _fmt_dt(ag),
            'dept_gull':                _fmt_dt(dg),
            'gull_waiting':             _dur(ag, dg),
            'arrived_dharamtar':        _fmt_dt(ad),
            'transit_gull_dppl':        _dur(dg, ad),
            'waiting_dppl':             _dur(ad, uc),
            'discharge_commenced':      _fmt_dt(uc),
            'discharge_completed':      _fmt_dt(up),
            'unloading_dur':            _dur(uc, up),
            'partly_disch_stop':        _fmt_dt(ds),
            'partly_disch_start':       _fmt_dt(dt),
            'disch_stop_time':          _dur(ds, dt),
            'sailed_dharamtar':         _fmt_dt(cd),
            'breakdown_start':          '',            # intentionally blank — do not auto-fill
            'breakdown_end':            _fmt_dt(cx),
            'breakdown_time':           _dur(cc, cx),
            'arrived_jaigad':           _fmt_dt(so),
            'waiting_sail':             _dur(up, cd),
            'tat':                      _dur(al, so),
        })
        sr += 1
    return result


# ── DPPL TAT helpers ───────────────────────────────────────────────────────────
def _fetch_dppl_tat_rows(from_date='', to_date='', month_filter='', fy_filter=''):
    """
    Fetch per-trip data for the DPPL TAT pivot sheet.
    Returns plain dicts with decimal-hour and day-fraction values ready for Excel.
    """
    conn = get_db()
    cur  = get_cursor(conn)

    where_clauses = ["h.operation_type = 'Import'"]
    params = []
    if from_date == to_date and from_date:
        where_clauses.append("DATE(h.doc_date) = %s")
        params.append(from_date)
    else:
        if from_date:
            where_clauses.append("DATE(h.doc_date) >= %s")
            params.append(from_date)

        if to_date:
            where_clauses.append("DATE(h.doc_date) <= %s")
            params.append(to_date)

    cur.execute(f"""
        SELECT
            h.doc_date, h.mbc_name, h.cargo_name, h.bl_quantity,
            mm.mbc_owner_name,
            lp.arrived_load_port,
            dp.vessel_arrival_port,  dp.unloading_commenced,
            dp.unloading_completed,  dp.vessel_cast_off,
            dp.sailed_out_load_port
        FROM mbc_header h
        LEFT JOIN mbc_master            mm ON mm.mbc_name = h.mbc_name
        LEFT JOIN mbc_load_port_lines      lp ON lp.mbc_id = h.id
        LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
        WHERE {' AND '.join(where_clauses)}
        ORDER BY mm.mbc_owner_name ASC NULLS LAST, h.doc_date ASC, h.id ASC
    """, params)

    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    if month_filter or fy_filter:
        filtered = []
        for r in rows:
            if month_filter and _month_label(r.get('doc_date', '')) != month_filter:
                continue
            if fy_filter and _fy(r.get('doc_date', '')) != fy_filter:
                continue
            filtered.append(r)
        rows = filtered

    result = []
    for r in rows:
        al = r.get('arrived_load_port')
        ad = r.get('vessel_arrival_port')
        uc = r.get('unloading_commenced')
        up = r.get('unloading_completed')
        cd = r.get('vessel_cast_off')
        so = r.get('sailed_out_load_port')

        def to_hrs(mins):
            return round(mins / 60.0, 4) if mins is not None else None

        def to_day(mins):
            return mins / 1440.0 if mins is not None else None

        result.append({
            'owner':          r.get('mbc_owner_name') or '',
            'mbc_name':       r.get('mbc_name') or '',
            'mbc_date':       _fmt_date(r.get('doc_date')),
            'cargo':          r.get('cargo_name') or '',
            'qty_in':         r.get('bl_quantity'),
            'wait_dppl_hrs':  to_hrs(_diff_mins(ad, uc)),
            'unload_hrs':     to_hrs(_diff_mins(uc, up)),
            'wait_sail_hrs':  to_hrs(_diff_mins(up, cd)),
            'total_dppl_day': to_day(_diff_mins(ad, cd)),
            'tat_day':        to_day(_diff_mins(al, so)),
        })

    return result


def _write_dppl_tat_sheet(ws, rows, period_label):
    """
    Build the DPPL TAT sheet in openpyxl Worksheet `ws`.
    Replicates the spec's PivotTable layout exactly using static data.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Palette ──────────────────────────────────────────────────────
    BLUE    = '4472C4'
    LT_BLUE = 'DEEAF1'
    NAVY    = '3F3F76'
    L_GREEN = 'E2EEDA'
    WHITE   = 'FFFFFF'

    _thin = Side(style='thin', color='000000')
    bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _fill(hex_c): return PatternFill('solid', fgColor=hex_c)
    def _fnt(bold=False, color='000000', size=11):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def _w(row, col, val, bold=False, color='000000', size=11,
           bg=None, aln=None, fmt=None, brd=False):
        c = ws.cell(row, col, val)
        c.font      = _fnt(bold=bold, color=color, size=size)
        c.fill      = _fill(bg) if bg else PatternFill()
        c.alignment = aln or ctr
        if fmt:
            c.number_format = fmt
        if brd:
            c.border = bdr
        return c

    # ── Column widths ────────────────────────────────────────────────
    # A=Owner, B=Name of MBC, C=MBC Date, D=Cargo, E=Qty,
    # F=Wait DPPL, G=Unload, H=Wait Sail, I=Total DPPL, J=TAT
    for ltr, w in [('A', 18.0), ('B', 17.856), ('C', 15.0), ('D', 12.570),
                   ('E', 8.711), ('F', 10.141), ('G', 15.141), ('H', 13.711),
                   ('I', 13.141), ('J', 14.426)]:
        ws.column_dimensions[ltr].width = w
    for jj in range(11, 17):   # K–P
        ws.column_dimensions[get_column_letter(jj)].width = 14.426
    ws.column_dimensions['R'].width = 80.0
    ws.column_dimensions['S'].width = 11.570

    # ── Row 2 (spacer) ───────────────────────────────────────────────
    ws.row_dimensions[2].height = 15

    # ── Row 3: Month label + Unit of Measure ─────────────────────────
    ws.row_dimensions[3].height = 15
    _w(3, 1,  'Month',       bold=True, color=WHITE, bg=BLUE,    brd=True)  # A3
    _w(3, 2,  period_label,            color=NAVY,  bg=LT_BLUE, brd=True)  # B3
    ws.merge_cells('O3:P3')
    _w(3, 15, 'Unit of Measure-HH:MM', bold=True, aln=ctr)   # O3:P3

    # ── Row 4 (empty) ────────────────────────────────────────────────
    ws.row_dimensions[4].height = 15

    # ── Row 5: "Values" pivot column group header ─────────────────────
    ws.row_dimensions[5].height = 15
    _w(5, 5, 'Values')   # E5 (shifted right by 1)

    # ── Row 6: Column headers (74.25 pt) ──────────────────────────────
    ws.row_dimensions[6].height = 74.25

    # A6:D6 – row-label headers (blue fill, white bold)
    for ci, lbl in [(1, 'Owner'), (2, 'Name of MBC'), (3, 'MBC Date'), (4, 'Cargo')]:
        _w(6, ci, lbl, bold=True, color=WHITE, bg=BLUE, brd=True)

    # E6:H6 – data-column headers (blue fill, white bold)
    for ci, lbl in [(5, 'Sum of MBC\nQty (MT)'),
                    (6, 'Waiting at DPPL\n(AMF to Comm.)'),
                    (7, 'Unloading Time\n(Comm. to Comp.)'),
                    (8, 'Waiting to sail\nfrom DPPL\n(Comp. to Castoff)')]:
        _w(6, ci, lbl, bold=True, color=WHITE, bg=BLUE, brd=True)

    # I6 – Total Time at DPPL (light green bg, NAVY text, bold)
    _w(6, 9, 'Total Time taken\nat DPPL', bold=True, color=NAVY, bg=L_GREEN, brd=True)

    # J6 – Average of TAT (blue fill, white bold)
    _w(6, 10, 'Average of TAT', bold=True, color=WHITE, bg=BLUE, brd=True)

    # ── Data rows ─────────────────────────────────────────────────────
    FMT_QTY  = '#,##0'
    FMT_H0   = '0.00'
    FMT_H2   = '#,##0.00'
    FMT_TIME = '[h]:mm'

    ROW_H = {7: 24, 8: 17.25, 9: 24, 10: 15.75, 11: 24}

    sum_qty   = 0.0
    sum_wdppl = 0.0
    sum_unld  = 0.0
    sum_wsail = 0.0
    tdppl_vals = []
    tat_vals   = []

    for i, trip in enumerate(rows):
        r = 7 + i
        ws.row_dimensions[r].height = ROW_H.get(r, 15.75)

        # A–D: row-label data (bold, black, 11pt)
        _w(r, 1, trip['owner'],    bold=True, size=11, aln=lft, brd=True)
        _w(r, 2, trip['mbc_name'], bold=True, size=11, brd=True)
        _w(r, 3, trip['mbc_date'], bold=True, size=11, brd=True)
        _w(r, 4, trip['cargo'],    bold=True, size=11, brd=True)

        # E: MBC Qty (MT)  — SUM, #,##0
        qty = float(trip['qty_in']) if trip['qty_in'] is not None else 0.0
        _w(r, 5, qty, size=12, brd=True, fmt=FMT_QTY)
        sum_qty += qty

        # F: Waiting at DPPL — SUM, 0.00 decimal hours
        wdppl = trip['wait_dppl_hrs']
        _w(r, 6, wdppl, size=12, brd=True, fmt=FMT_H0)
        if wdppl is not None:
            sum_wdppl += wdppl

        # G: Unloading Time — SUM, #,##0.00 decimal hours
        unld = trip['unload_hrs']
        _w(r, 7, unld, size=12, brd=True, fmt=FMT_H2)
        if unld is not None:
            sum_unld += unld

        # H: Waiting to Sail — SUM, 0.00 decimal hours
        wsail = trip['wait_sail_hrs']
        _w(r, 8, wsail, size=12, brd=True, fmt=FMT_H0)
        if wsail is not None:
            sum_wsail += wsail

        # I: Total Time at DPPL — [h]:mm, light green fill, bold
        tdppl = trip['total_dppl_day']
        _w(r, 9, tdppl, bold=True, size=12, bg=L_GREEN, brd=True, fmt=FMT_TIME)
        if tdppl is not None:
            tdppl_vals.append(tdppl)

        # J: TAT — [h]:mm
        tat = trip['tat_day']
        _w(r, 10, tat, size=12, brd=True, fmt=FMT_TIME)
        if tat is not None:
            tat_vals.append(tat)

    # ── Grand Total row ────────────────────────────────────────────────
    gt = 7 + len(rows)
    ws.row_dimensions[gt].height = 24

    avg_tdppl = (sum(tdppl_vals) / len(tdppl_vals)) if tdppl_vals else None
    avg_tat   = (sum(tat_vals)   / len(tat_vals))   if tat_vals   else None

    _w(gt, 1, 'Grand Total', bold=True, size=11, aln=lft, brd=True)
    _w(gt, 2, '', brd=True)
    _w(gt, 3, '', brd=True)
    _w(gt, 4, '', brd=True)
    _w(gt, 5,  round(sum_qty,   0), size=12, brd=True, fmt=FMT_QTY)
    _w(gt, 6,  round(sum_wdppl, 4), size=12, brd=True, fmt=FMT_H0)
    _w(gt, 7,  round(sum_unld,  4), size=12, brd=True, fmt=FMT_H2)
    _w(gt, 8,  round(sum_wsail, 4), size=12, brd=True, fmt=FMT_H0)
    _w(gt, 9,  avg_tdppl, bold=True, size=12, bg=L_GREEN, brd=True, fmt=FMT_TIME)
    _w(gt, 10, avg_tat,             size=12,             brd=True, fmt=FMT_TIME)

    # ── Static cell per spec (S19 = "Dolomite") ───────────────────────
    ws.cell(19, 19, 'Dolomite')


# ── MBC Wise helpers ────────────────────────────────────────────────────────────
def _fetch_mbc_wise_rows(from_date='', to_date='', month_filter='', fy_filter=''):
    """
    Fetch raw trips, group by mbc_name, return one dict per vessel with:
    SUM of qty and AVG of each TAT segment in day-fractions ([h]:mm).
    """
    from collections import defaultdict

    conn = get_db()
    cur  = get_cursor(conn)

    where_clauses = ["h.operation_type = 'Import'"]
    params = []
    if from_date == to_date and from_date:
        where_clauses.append("DATE(h.doc_date) = %s")
        params.append(from_date)
    else:
        if from_date:
            where_clauses.append("DATE(h.doc_date) >= %s")
            params.append(from_date)

        if to_date:
            where_clauses.append("DATE(h.doc_date) <= %s")
            params.append(to_date)

    cur.execute(f"""
        SELECT
            h.doc_date, h.mbc_name, h.bl_quantity,
            lp.arrived_load_port,   lp.loading_commenced,  lp.loading_completed,
            lp.cast_off_load_port,
            dp.arrival_gull_island, dp.departure_gull_island, dp.vessel_arrival_port,
            dp.unloading_commenced, dp.unloading_completed,
            dp.vessel_cast_off,     dp.sailed_out_load_port
        FROM mbc_header h
        LEFT JOIN mbc_load_port_lines      lp ON lp.mbc_id = h.id
        LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
        WHERE {' AND '.join(where_clauses)}
        ORDER BY h.mbc_name ASC, h.doc_date ASC
    """, params)

    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    if month_filter or fy_filter:
        filtered = []
        for r in rows:
            if month_filter and _month_label(r.get('doc_date', '')) != month_filter:
                continue
            if fy_filter and _fy(r.get('doc_date', '')) != fy_filter:
                continue
            filtered.append(r)
        rows = filtered

    SEG_KEYS = [
        'preberthing', 'loading', 'wait_after_load', 'total_jaigad',
        'jaigad_to_gull', 'gull_waiting', 'gull_to_dhar', 'jaigad_to_dhar',
        'preberthing_dhar', 'unloading', 'wait_after_unload', 'total_dharamtar',
        'dhar_to_jaigad', 'tat',
    ]

    groups = defaultdict(lambda: {'qty_vals': [], **{k: [] for k in SEG_KEYS}})

    for r in rows:
        name = r.get('mbc_name') or 'Unknown'
        al = r.get('arrived_load_port')
        lc = r.get('loading_commenced')
        lp = r.get('loading_completed')
        co = r.get('cast_off_load_port')
        ag = r.get('arrival_gull_island')
        dg = r.get('departure_gull_island')
        ad = r.get('vessel_arrival_port')
        uc = r.get('unloading_commenced')
        up = r.get('unloading_completed')
        cd = r.get('vessel_cast_off')
        so = r.get('sailed_out_load_port')
        qty = float(r['bl_quantity']) if r.get('bl_quantity') is not None else 0.0

        g = groups[name]
        g['qty_vals'].append(qty)
        g['preberthing'].append(_diff_mins(al, lc))
        g['loading'].append(_diff_mins(lc, lp))
        g['wait_after_load'].append(_diff_mins(lp, co))
        g['total_jaigad'].append(_diff_mins(al, co))
        g['jaigad_to_gull'].append(_diff_mins(co, ag))
        g['gull_waiting'].append(_diff_mins(ag, dg))
        g['gull_to_dhar'].append(_diff_mins(dg, ad))
        g['jaigad_to_dhar'].append(_diff_mins(co, ad))
        g['preberthing_dhar'].append(_diff_mins(ad, uc))
        g['unloading'].append(_diff_mins(uc, up))
        g['wait_after_unload'].append(_diff_mins(up, cd))
        g['total_dharamtar'].append(_diff_mins(ad, cd))
        g['dhar_to_jaigad'].append(_diff_mins(cd, so))
        g['tat'].append(_diff_mins(al, so))

    def _avg_day(vals):
        valid = [v for v in vals if v is not None]
        return (sum(valid) / len(valid) / 1440.0) if valid else None

    result = []
    for name in sorted(groups.keys()):
        g = groups[name]
        row = {'mbc_name': name, 'qty': sum(g['qty_vals'])}
        for k in SEG_KEYS:
            row[k] = _avg_day(g[k])
        result.append(row)

    return result


def _write_mbc_wise_sheet(ws, rows, period_label):
    """
    Build the MBC Wise sheet - pixel-accurate recreation of PivotTable1 layout.
    Columns: A=blank, B=MBC Name, C=Qty, D-Q=14 TAT segments, R=Remarks.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Palette ──────────────────────────────────────────────────────
    WHITE   = 'FFFFFF'
    YELLOW  = 'FFFFCC'
    PEACH   = 'FFCC99'
    L_GREEN = 'E2EEDA'
    L_BLUE  = 'D9E2F3'
    LTBGREY = 'DEEAF6'   # PivotStyleLight16 default row bg (stand-in for "None fill")
    NAVY    = '3F3F76'
    GY_BRD  = 'B2B2B2'
    DK_BRD  = '7F7F7F'
    TXT_YEL = '7F6000'   # dark amber for YELLOW bg
    TXT_GRN = '375623'   # dark green for L_GREEN bg
    TXT_BLU = '1F3864'   # dark blue for L_BLUE bg

    # ── Borders ──────────────────────────────────────────────────────
    _th   = Side(style='thin')
    _gyb  = Side(style='thin', color=GY_BRD)
    _dkb  = Side(style='thin', color=DK_BRD)
    bdr   = Border(left=_th,  right=_th,  top=_th,  bottom=_th)
    bdr_y = Border(left=_gyb, right=_gyb, top=_th,  bottom=_th)
    bdr_p = Border(left=_dkb, right=_dkb, top=_th,  bottom=_th)

    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    bot = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
    lbt = Alignment(horizontal='left',   vertical='bottom', wrap_text=True)

    def _fill(h): return PatternFill('solid', fgColor=h)
    def _fnt(bold=False, color='000000', size=11):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def _w(row, col, val, bold=False, color='000000', size=11,
           bg=None, aln=None, fmt=None, border=None):
        c = ws.cell(row, col, val)
        c.font      = _fnt(bold=bold, color=color, size=size)
        c.fill      = _fill(bg) if bg else PatternFill()
        c.alignment = aln or ctr
        if fmt:    c.number_format = fmt
        if border: c.border = border
        return c

    # ── Page setup ───────────────────────────────────────────────────
    ws.sheet_format.defaultRowHeight = 14.4
    ws.page_setup.orientation  = 'portrait'
    ws.page_setup.paperSize    = 9   # A4
    ws.page_margins.left = ws.page_margins.right  = 0.7
    ws.page_margins.top  = ws.page_margins.bottom = 0.75
    ws.page_margins.header = ws.page_margins.footer = 0.3
    ws.page_setup.horizontalDpi = ws.page_setup.verticalDpi = 300

    # ── Column widths (exact per spec) ───────────────────────────────
    for ltr, w in [('A', 17.887), ('B', 20.109), ('C', 8.664),  ('D', 16.109),
                   ('E', 12.441), ('F', 15.0),   ('G', 15.555), ('H', 15.332),
                   ('I', 14.441), ('J', 13.555), ('K', 14.887), ('L', 15.664),
                   ('M', 16.555), ('N', 13.887), ('O', 15.887), ('P', 14.332),
                   ('Q', 14.441), ('R', 80.0)]:
        ws.column_dimensions[ltr].width = w

    # ── Row 2: Filter 1 — MBC Date ───────────────────────────────────
    ws.row_dimensions[2].height = 14.4
    _w(2, 1, 'MBC Date', bold=True, color=WHITE, bg='4472C4', border=bdr)
    _w(2, 2, period_label, color=NAVY, bg='DEEAF1', border=bdr)

    # ── Row 3: Filter 2 — Cargo ──────────────────────────────────────
    ws.row_dimensions[3].height = 14.4
    _w(3, 1, 'Cargo', bold=True, color=WHITE, bg='4472C4', border=bdr)
    _w(3, 2, '(All)', color=NAVY, bg='DEEAF1', border=bdr)

    # O3:P3 merged — Unit of Measure
    ws.merge_cells('O3:P3')
    c = ws.cell(3, 15, 'Unit of Measure-HH:MM')
    c.font      = _fnt(bold=True, color=NAVY)
    c.alignment = bot

    # ── Row 5: "Values" label ─────────────────────────────────────────
    ws.row_dimensions[5].height = 14.4
    ws.cell(5, 3, 'Values')

    # ── Row 6: Column headers (74.25 pt) ─────────────────────────────
    ws.row_dimensions[6].height = 74.25

    hdr_specs = [
        # (col, label,                             bg,      color,   bold, bdr_)
        (1,  '',                                    LTBGREY, NAVY,    False, bdr),
        (2,  'MBC Name',                            LTBGREY, NAVY,    False, bdr),
        (3,  'MBC\nQty (MT)',                       WHITE,   NAVY,    True,  bdr),
        (4,  'Preberthing\nDelay at\nJaigad',       WHITE,   NAVY,    True,  bdr),
        (5,  'Loading\nTime',                       WHITE,   NAVY,    True,  bdr),
        (6,  'Waiting\nafter\nLoading',             WHITE,   NAVY,    True,  bdr),
        (7,  'Total Time\nat Jaigad',               YELLOW,  TXT_YEL, True,  bdr_y),
        (8,  'Jaigad to\nGull',                     WHITE,   NAVY,    True,  bdr),
        (9,  'Gull\nWaiting',                       WHITE,   NAVY,    True,  bdr),
        (10, 'Gull to\nDharamtar',                  WHITE,   NAVY,    True,  bdr),
        (11, 'Jaigad to\nDharamtar',                PEACH,   NAVY,    True,  bdr_p),
        (12, 'Preberthing\nat Dharamtar',           WHITE,   NAVY,    True,  bdr),
        (13, 'Unloading\nTime',                     WHITE,   NAVY,    True,  bdr),
        (14, 'Waiting after\nUnloading',            WHITE,   NAVY,    True,  bdr),
        (15, 'Total Time\nat Dharamtar',            L_GREEN, TXT_GRN, True,  bdr),
        (16, 'Dharamtar\nto Jaigad',                L_BLUE,  TXT_BLU, True,  bdr),
        (17, 'TAT',                                 LTBGREY, NAVY,    False, bdr),
        (18, 'Remarks',                             LTBGREY, NAVY,    True,  bdr),
    ]
    for ci, lbl, bg, color, bold, b in hdr_specs:
        _w(6, ci, lbl, bold=bold, color=color, bg=bg, border=b)

    # ── Data rows ─────────────────────────────────────────────────────
    FMT_QTY  = '#,##0'
    FMT_TIME = '[h]:mm'

    ROW_H = {7: 24.0, 8: 43.2, 9: 24.0, 10: 28.8, 11: 24.0, 12: 33.0}

    SEG_KEYS = [
        'preberthing', 'loading', 'wait_after_load', 'total_jaigad',
        'jaigad_to_gull', 'gull_waiting', 'gull_to_dhar', 'jaigad_to_dhar',
        'preberthing_dhar', 'unloading', 'wait_after_unload', 'total_dharamtar',
        'dhar_to_jaigad', 'tat',
    ]

    all_qty  = 0.0
    gt_segs  = {k: [] for k in SEG_KEYS}

    for i, trip in enumerate(rows):
        r = 7 + i
        ws.row_dimensions[r].height = ROW_H.get(r, 15.6)

        _w(r, 1,  '',                 color=NAVY, size=11, bg=LTBGREY, border=bdr)
        _w(r, 2,  trip['mbc_name'],   color=NAVY, size=11, bg=LTBGREY, border=bdr)
        qty = trip['qty']
        _w(r, 3,  qty,                color=NAVY, size=12, bg=LTBGREY, border=bdr,  fmt=FMT_QTY)
        all_qty += qty

        # D-F: Jaigad load-port segments
        _w(r, 4,  trip['preberthing'],    color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        _w(r, 5,  trip['loading'],        color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        _w(r, 6,  trip['wait_after_load'],color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        # G: Total Jaigad (YELLOW)
        _w(r, 7,  trip['total_jaigad'],   bold=True, color=TXT_YEL, size=12, bg=YELLOW, border=bdr_y, fmt=FMT_TIME)
        # H-J: Transit segments
        _w(r, 8,  trip['jaigad_to_gull'], color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        _w(r, 9,  trip['gull_waiting'],   color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        _w(r, 10, trip['gull_to_dhar'],   color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        # K: Jaigad→Dharamtar (PEACH)
        _w(r, 11, trip['jaigad_to_dhar'], bold=True, color=NAVY, size=12, bg=PEACH, border=bdr_p, fmt=FMT_TIME)
        # L: Preberthing Dharamtar (WHITE fill, dark navy)
        _w(r, 12, trip['preberthing_dhar'], color=NAVY, size=12, bg=WHITE, border=bdr, fmt=FMT_TIME)
        # M-N: Discharge segments
        _w(r, 13, trip['unloading'],        color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        _w(r, 14, trip['wait_after_unload'],color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        # O: Total Dharamtar (LIGHT GREEN)
        _w(r, 15, trip['total_dharamtar'],  bold=True, color=TXT_GRN, size=12, bg=L_GREEN, border=bdr, fmt=FMT_TIME)
        # P: Dharamtar→Jaigad (LIGHT BLUE)
        _w(r, 16, trip['dhar_to_jaigad'],   bold=True, color=TXT_BLU, size=12, bg=L_BLUE, border=bdr, fmt=FMT_TIME)
        # Q: TAT
        _w(r, 17, trip['tat'],              color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
        # R: blank
        _w(r, 18, '', color=NAVY, size=11, aln=lft, border=bdr)

        # Accumulate for grand total
        for k in SEG_KEYS:
            v = trip.get(k)
            if v is not None:
                gt_segs[k].append(v)

    # ── Grand Total row ───────────────────────────────────────────────
    gt = 7 + len(rows)
    ws.row_dimensions[gt].height = 15.6

    def _gt_avg(key):
        vals = gt_segs[key]
        return (sum(vals) / len(vals)) if vals else None

    _w(gt, 1,  'Grand Total', bold=True, color=NAVY, size=12, aln=lft, bg=LTBGREY, border=bdr)
    _w(gt, 2,  '',            color=NAVY, size=12, bg=LTBGREY, border=bdr)
    _w(gt, 3,  all_qty,       bold=True, color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_QTY)
    _w(gt, 4,  _gt_avg('preberthing'),     color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 5,  _gt_avg('loading'),         color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 6,  _gt_avg('wait_after_load'), color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 7,  _gt_avg('total_jaigad'),    bold=True, color=TXT_YEL, size=12, bg=YELLOW, border=bdr_y, fmt=FMT_TIME)
    _w(gt, 8,  _gt_avg('jaigad_to_gull'),  color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 9,  _gt_avg('gull_waiting'),    color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 10, _gt_avg('gull_to_dhar'),    color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 11, _gt_avg('jaigad_to_dhar'),  bold=True, color=NAVY, size=12, bg=PEACH, border=bdr_p, fmt=FMT_TIME)
    _w(gt, 12, _gt_avg('preberthing_dhar'),color=NAVY, size=12, bg=WHITE, border=bdr, fmt=FMT_TIME)
    _w(gt, 13, _gt_avg('unloading'),       color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 14, _gt_avg('wait_after_unload'),color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)
    _w(gt, 15, _gt_avg('total_dharamtar'), bold=True, color=TXT_GRN, size=12, bg=L_GREEN, border=bdr, fmt=FMT_TIME)
    _w(gt, 16, _gt_avg('dhar_to_jaigad'),  bold=True, color=TXT_BLU, size=12, bg=L_BLUE,  border=bdr, fmt=FMT_TIME)
    _w(gt, 17, _gt_avg('tat'),             bold=True, color=NAVY, size=12, bg=LTBGREY, border=bdr, fmt=FMT_TIME)

    # ── Footnotes R20–R22 ─────────────────────────────────────────────
    for rn, txt in [
        (20, 'MBC waiting Time at DPPL & Unloading Time Increased.'),
        (21, 'MBC waiting Time at Gull & DPPL Increased.'),
        (22, 'MBC waiting Time at Jaigad Increased.'),
    ]:
        c = ws.cell(rn, 18, txt)
        c.font      = _fnt(color=NAVY, size=11)
        c.alignment = lbt
        c.border    = bdr


# ── routes ─────────────────────────────────────────────────────────────────────
@bp.route('/module/RP01/mbc-master/')
@login_required
def mbc_master_index():
    return render_template('mbc_master/mbc_master_list.html', username=session.get('username'))


@bp.route('/api/module/RP01/mbc-master/data')
@login_required
def mbc_master_data():

    selected_date = request.args.get('selected_date', '')

    rows = _fetch_rows(
        selected_date,
        selected_date,
        request.args.get('month', ''),
        request.args.get('fy', '')
    )

    return jsonify(rows)

_TAT_ACTIVITIES = [
    ('Jaigad Arrival - Jaigad Loading Commenced (Preberthing delay)', 'preberthing', '1:00', 'data'),
    ('Loading Commence - Loading Completion (Loading time)', 'loading', '6:00', 'data'),
    ('Loading Completed - Cast Off from Jaigad (Waiting after loading)', 'wait_after_load', '0:00', 'data'),
    ('Total time taken at Jaigad', 'total_jaigad', '7:00', 'section_total'),

    ('Jaigad Departure to Gull Arrival (Loaded Transit time)', 'jaigad_to_gull', '12:00', 'data'),
    ('Gull Arrival - Gull Departure (Waiting at Gull)', 'gull_waiting', '10:00', 'data'),
    ('Gull Departure - Dharamtar Arrival', 'gull_to_dhar', '4:00', 'data'),
    ('Jaigad Departure - Dharamtar Arrival (Jaigad to Dharamtar)', 'jaigad_to_dhar', '26:00', 'main_total'),

    ('Dharamtar Arrival to Disch Commenced (Preberthing delay)', 'preberthing_dhar', '4:00', 'data'),
    ('Disch Commended to Disch Completed (Unloading Time)', 'unloading', '6:00', 'data'),
    ('Disch Completed to Cast Off from Dharamtar (Waiting after Unloading)', 'wait_after_unload', '1:00', 'data'),
    ('Total time taken at Dharamtar', 'total_dharamtar', '11:00', 'section_total'),

    ('Dharamtar Departure to Jaigad Arrival', 'dhar_to_jaigad', '16:00', 'main_total'),
    ('TAT', 'tat', '60:00', 'tat'),
]


@bp.route('/api/module/RP01/mbc-master/tat-data')
@login_required
def mbc_master_tat_data():

    selected_date = request.args.get('selected_date')

    if not selected_date:
        selected_date = date_type.today().strftime('%Y-%m-%d')

    ref_dt = datetime.strptime(selected_date, '%Y-%m-%d').date()

    # =========================
    # DAILY (Selected Date)
    # =========================
    daily_raw = _fetch_raw_trips(
        selected_date,
        selected_date
    )

    daily_count = len(daily_raw)
    daily_m = _compute_tat_metrics(daily_raw)

    # =========================
    # MTD
    # =========================
    mtd_from = ref_dt.replace(day=1).strftime('%Y-%m-%d')

    mtd_raw = _fetch_raw_trips(
        mtd_from,
        selected_date
    )

    mtd_count = len(mtd_raw)
    mtd_m = _compute_tat_metrics(mtd_raw)

    # =========================
    # YTD
    # =========================
    fy_start, _ = _fy_date_range(ref_dt)

    ytd_raw = _fetch_raw_trips(
        fy_start,
        selected_date
    )

    ytd_count = len(ytd_raw)
    ytd_m = _compute_tat_metrics(ytd_raw)

    date_label = ref_dt.strftime('%d-%m-%Y')
    mtd_label = ref_dt.strftime('%b-%y')
    ytd_label = _fy_label_str(ref_dt)

    rows = [
        {
            'label': label,
            'target': target,
            'style': style,
            'date_val': daily_m.get(key, '—'),
            'mtd_val': mtd_m.get(key, '—'),
            'ytd_val': ytd_m.get(key, '—'),
        }
        for label, key, target, style in _TAT_ACTIVITIES
    ]

    return jsonify({
        'date_label': date_label,
        'mtd_label': mtd_label,
        'ytd_label': ytd_label,
        'date_trips': daily_count,
        'mtd_trips': mtd_count,
        'ytd_trips': ytd_count,
        'rows': rows
    })

@bp.route('/api/module/RP01/mbc-master/dppl-tat-data')
@login_required
def mbc_master_dppl_tat_data():
    req_from  = request.args.get('from_date', '')
    req_to    = request.args.get('to_date', '')
    req_month = request.args.get('month', '')
    req_fy    = request.args.get('fy', '')

    rows = _fetch_dppl_tat_rows(req_from, req_to, req_month, req_fy)

    def _hhmm(val_hrs):
        """Decimal hours → 'H:MM' display string."""
        if val_hrs is None:
            return None
        total_min = round(val_hrs * 60)
        h, m = divmod(total_min, 60)
        return f'{h}:{m:02d}'

    def _hhmm_day(val_day):
        """Day fraction → 'H:MM' display string."""
        if val_day is None:
            return None
        return _hhmm(val_day * 24)

    out = []
    sum_qty   = 0.0
    sum_wdppl = 0.0
    sum_unld  = 0.0
    sum_wsail = 0.0
    tdppl_vals = []
    tat_vals   = []

    for r in rows:
        qty   = float(r['qty_in']) if r['qty_in'] is not None else 0.0
        wdppl = r['wait_dppl_hrs']
        unld  = r['unload_hrs']
        wsail = r['wait_sail_hrs']
        tdppl = r['total_dppl_day']
        tat   = r['tat_day']

        sum_qty += qty
        if wdppl  is not None: sum_wdppl += wdppl
        if unld   is not None: sum_unld  += unld
        if wsail  is not None: sum_wsail += wsail
        if tdppl  is not None: tdppl_vals.append(tdppl)
        if tat    is not None: tat_vals.append(tat)

        out.append({
            'owner':      r['owner'],
            'mbc_name':   r['mbc_name'],
            'mbc_date':   r['mbc_date'],
            'cargo':      r['cargo'],
            'qty':        f'{qty:,.0f}' if qty else '—',
            'wait_dppl':  _hhmm(wdppl),
            'unload':     _hhmm(unld),
            'wait_sail':  _hhmm(wsail),
            'total_dppl': _hhmm_day(tdppl),
            'tat':        _hhmm_day(tat),
        })

    avg_tdppl = (sum(tdppl_vals) / len(tdppl_vals)) if tdppl_vals else None
    avg_tat   = (sum(tat_vals)   / len(tat_vals))   if tat_vals   else None

    totals = {
        'qty':        f'{sum_qty:,.0f}',
        'wait_dppl':  _hhmm(sum_wdppl),
        'unload':     _hhmm(sum_unld),
        'wait_sail':  _hhmm(sum_wsail),
        'total_dppl': _hhmm_day(avg_tdppl),
        'tat':        _hhmm_day(avg_tat),
    }

    return jsonify({'rows': out, 'totals': totals})


@bp.route('/api/module/RP01/mbc-master/download')
@login_required
def mbc_master_download():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    req_from  = request.args.get('from_date', '')
    req_to    = request.args.get('to_date', '')
    req_month = request.args.get('month', '')
    req_fy    = request.args.get('fy', '')

    # ── Sheet 1: Master Data rows (display-formatted) ─────────────────
    rows = _fetch_rows(req_from, req_to, req_month, req_fy)

    # ── Reference date for MTD / YTD ──────────────────────────────────
    ref_date_str = req_to or date_type.today().strftime('%Y-%m-%d')
    try:
        ref_dt = datetime.strptime(ref_date_str, '%Y-%m-%d').date()
    except ValueError:
        ref_dt = date_type.today()

    # ── Sheet 2 period data: same filter as Sheet 1 ────────────────────
    period_raw   = _fetch_raw_trips(req_from, req_to, req_month, req_fy)
    period_count = len(period_raw)
    period_m     = _compute_tat_metrics(period_raw)

    # Period label
    if req_month:
        period_label = req_month
    elif req_fy:
        period_label = req_fy
    elif req_from and req_to:
        period_label = datetime.strptime(
            req_to, "%Y-%m-%d"
        ).strftime("%d-%m-%Y")
    else:
        period_label = 'All Periods'

    # ── Sheet 2 MTD data ───────────────────────────────────────────────
    mtd_from_str = ref_dt.replace(day=1).strftime('%Y-%m-%d')
    mtd_raw      = _fetch_raw_trips(mtd_from_str, ref_date_str)
    mtd_count    = len(mtd_raw)
    mtd_m        = _compute_tat_metrics(mtd_raw)
    mtd_label    = ref_dt.strftime('%b-%y')

    # ── Sheet 2 YTD data ───────────────────────────────────────────────
    fy_start_str, _ = _fy_date_range(ref_dt)
    ytd_raw         = _fetch_raw_trips(fy_start_str, ref_date_str)
    ytd_count       = len(ytd_raw)
    ytd_m           = _compute_tat_metrics(ytd_raw)
    ytd_label       = _fy_label_str(ref_dt)

    # ── Build workbook ─────────────────────────────────────────────────
    wb = Workbook()

    # ══ Sheet 1: MBC Wise (default first sheet) ════════════════════════
    ws0 = wb.active
    ws0.title = 'MBC Wise'
    mbc_wise_rows = _fetch_mbc_wise_rows(req_from, req_to, req_month, req_fy)
    _write_mbc_wise_sheet(ws0, mbc_wise_rows, period_label)

    # ══ Sheet 2: Master Data ═══════════════════════════════════════════
    ws1 = wb.create_sheet('Master Data')

    # Per-column header fills (1-indexed column position → hex, None = no fill)
    _HDR_FILLS = {
        1:  'A5A5A5',  # A  – Sr. No             Grey
        2:  'FFFFCC',  # B  – MBC Date            Light Yellow
        3:  'A5A5A5',  # C  – Name of MBC         Grey
        # D (4) = Month: No Fill
        5:  'FFFFCC',  6:  'FFFFCC',  7:  'FFFFCC',  8:  'FFFFCC',
        9:  'FFFFCC',  10: 'FFFFCC',  11: 'FFFFCC',  12: 'FFFFCC',
        13: 'FFFFCC',  14: 'FFFFCC',  15: 'FFFFCC',  16: 'FFFFCC',
        # E–P (5–16) = Light Yellow
        17: '70AD47',  18: '70AD47',  19: '70AD47',  20: '70AD47',
        21: '70AD47',  22: '70AD47',  23: '70AD47',  24: '70AD47',
        25: '70AD47',  26: '70AD47',  27: '70AD47',  28: '70AD47',
        # Q–AB (17–28) = Green
        29: '4472C4',  30: '4472C4',  31: '4472C4',  # AC–AE = Blue
        32: '70AD47',                                  # AF    = Green
        33: '4472C4',  34: '4472C4',  35: '4472C4',  # AG–AI = Blue
        36: '70AD47',                                  # AJ    = Green
        37: '5B9BD5',                                  # AK    = Light Blue (TAT)
    }
    cell_font       = Font(name='Calibri', size=10)
    calc_font       = Font(name='Calibri', bold=True, size=10, color='1A5FA8')
    calc_fill       = PatternFill('solid', fgColor='FFFDE7')
    alt_fill        = PatternFill('solid', fgColor='F5F8FC')
    border_side     = Side(style='thin', color='D0D0D0')
    thin_border     = Border(left=border_side, right=border_side,
                             top=border_side,  bottom=border_side)
    center_align    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align      = Alignment(horizontal='left',   vertical='center')
    right_align     = Alignment(horizontal='right',  vertical='center')

    # Header row — 60pt height, per-column fill, black font throughout
    ws1.row_dimensions[1].height = 60
    hdr_font = Font(name='Calibri', bold=True, size=10, color='000000')
    for col_idx, (label, key, is_calc) in enumerate(COLUMNS, start=1):
        fill_hex = _HDR_FILLS.get(col_idx)
        cell = ws1.cell(row=1, column=col_idx, value=label)
        cell.font      = hdr_font
        cell.fill      = PatternFill('solid', fgColor=fill_hex) if fill_hex else PatternFill()
        cell.alignment = center_align
        cell.border    = thin_border

    # Data rows
    for row_idx, record in enumerate(rows, start=2):
        is_even = (row_idx % 2 == 0)
        ws1.row_dimensions[row_idx].height = 18
        for col_idx, (label, key, is_calc) in enumerate(COLUMNS, start=1):
            # Intentionally blank columns
            val = '' if key in _BLANK_KEYS else record.get(key, '')
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if is_calc:
                cell.font      = calc_font
                cell.fill      = calc_fill
                cell.alignment = center_align
            else:
                cell.font = cell_font
                cell.fill = alt_fill if is_even else PatternFill()
                if key in ('qty_in', 'sr_no'):
                    cell.alignment = right_align
                elif key in ('mbc_name', 'cargo', 'mbc_date', 'date',
                             'arrived_anchored', 'loading_commenced', 'loading_completed',
                             'castoff', 'arrived_gull', 'dept_gull', 'arrived_dharamtar',
                             'discharge_commenced', 'discharge_completed', 'partly_disch_stop',
                             'partly_disch_start', 'sailed_dharamtar', 'breakdown_start',
                             'breakdown_end', 'arrived_jaigad'):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

    # Column widths
    width_map = {
        'sr_no': 6,   'mbc_date': 12,  'mbc_name': 20,  'month': 9,   'fy_year': 10,
        'date': 12,   'arrived_anchored': 18, 'waiting_perb': 16,
        'loading_commenced': 18, 'loading_completed': 18, 'loading_dur': 13,
        'castoff': 18, 'waiting_castoff': 16, 'cargo': 16, 'qty_in': 9,
        'at_jaigad': 13, 'transit_jaigad_dharamtar': 24, 'jaigad_to_gull': 15,
        'arrived_gull': 18, 'dept_gull': 18, 'gull_waiting': 13,
        'arrived_dharamtar': 18, 'transit_gull_dppl': 18, 'waiting_dppl': 15,
        'discharge_commenced': 18, 'discharge_completed': 18, 'unloading_dur': 13,
        'partly_disch_stop': 18, 'partly_disch_start': 18, 'disch_stop_time': 18,
        'sailed_dharamtar': 18, 'breakdown_start': 18, 'breakdown_end': 18,
        'breakdown_time': 18, 'arrived_jaigad': 18, 'waiting_sail': 26, 'tat': 8,
    }
    for col_idx, (label, key, is_calc) in enumerate(COLUMNS, start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width_map.get(key, 10)

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # ══ Sheet 2: MBC TAT Report ════════════════════════════════════════
    ws2 = wb.create_sheet('MBC TAT Report')
    _write_tat_sheet(
        ws2,
        period_label, mtd_label, ytd_label,
        period_count, mtd_count, ytd_count,
        period_m, mtd_m, ytd_m
    )

    # ══ Sheet 3: DPPL TAT ══════════════════════════════════════════════
    ws3 = wb.create_sheet('DPPL TAT')
    dppl_rows = _fetch_dppl_tat_rows(req_from, req_to, req_month, req_fy)
    _write_dppl_tat_sheet(ws3, dppl_rows, period_label)

    # ── Stream response ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname_parts = ['MBC_TAT_Report']
    if req_from:
        fname_parts.append(req_from)
    if req_to:
        fname_parts.append('to_' + req_to)
    filename = '_'.join(fname_parts) + '.xlsx'

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@bp.route('/api/module/RP01/mbc-master/mbc-wise-data')
@login_required
def mbc_master_mbc_wise_data():
    req_from  = request.args.get('from_date', '')
    req_to    = request.args.get('to_date', '')
    req_month = request.args.get('month', '')
    req_fy    = request.args.get('fy', '')

    rows = _fetch_mbc_wise_rows(req_from, req_to, req_month, req_fy)

    def _day_to_hhmm(val_day):
        if val_day is None:
            return None
        total_min = round(val_day * 1440)
        h, m = divmod(total_min, 60)
        return f'{h}:{m:02d}'

    SEG_KEYS = [
        'preberthing', 'loading', 'wait_after_load', 'total_jaigad',
        'jaigad_to_gull', 'gull_waiting', 'gull_to_dhar', 'jaigad_to_dhar',
        'preberthing_dhar', 'unloading', 'wait_after_unload', 'total_dharamtar',
        'dhar_to_jaigad', 'tat',
    ]

    out = []
    for r in rows:
        entry = {
            'mbc_name': r['mbc_name'],
            'qty':      round(r['qty']) if r['qty'] else 0,
        }
        for k in SEG_KEYS:
            entry[k] = _day_to_hhmm(r.get(k))
        out.append(entry)

    return jsonify(out)
