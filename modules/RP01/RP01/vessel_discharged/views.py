from flask import (
    render_template,
    request,
    jsonify,
    session,
    redirect,
    url_for,
    Response,
)
from functools import wraps
from datetime import date, datetime
from collections import defaultdict
import io
import re

from .. import bp
from database import get_db, get_cursor

# ── Excel colour / style constants ─────────────────────────────────────────

XL_GREY = "C0C0C0"
XL_LAVEND = "CCCCFF"
XL_CYAN = "CCFFFF"
XL_WHITE = "FFFFFF"
XL_TITLE_SZ = 14
XL_NORM_SZ = 10
XL_SMALL_SZ = 9

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_thin = Side(style="thin", color="000000")
_med = Side(style="medium", color="000000")
_bdr = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_bdr_ml = Border(left=_med, right=_thin, top=_thin, bottom=_thin)
_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
_right = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=XL_NORM_SZ, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _fmt_modu(total_mins):
    """Format minutes as decimal hours"""

    if not total_mins:
        return "0.00"

    decimal_hours = float(total_mins) / 60

    return f"{decimal_hours:.2f}"


def _fmt_qty(value):
    if value is None or value == "":
        return ""
    try:
        return "{:,}".format(int(round(float(value))))
    except (TypeError, ValueError):
        return str(value)


def _fmt_number(value, decimals=2):
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):,.{decimals}f}"
    except (TypeError, ValueError):
        return str(value)


def _parse_number(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        text = str(val).replace(",", "").strip()
        if text == "":
            return None
        if "." in text:
            return float(text)
        return int(text)
    except (TypeError, ValueError):
        return val


def _parse_dt(val):

    if not val:
        return None

    if isinstance(val, datetime):
        return val

    val = str(val).strip()

    formats = [
        "%d-%m-%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]

    for fmt in formats:

        try:
            return datetime.strptime(val, fmt)

        except Exception:
            pass

    try:
        return datetime.fromisoformat(val)

    except Exception:
        return None


def _fmt_dt(val):
    dt = _parse_dt(val)
    return dt.strftime("%d-%m-%Y %H:%M") if dt else ""


def _day_key(val):
    dt = _parse_dt(val)
    return dt.date() if dt else None


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated


# ── Data fetch ──────────────────────────────────────────────────────────────

_DATE_FIELDS = {
    "discharge_commenced": "pla_commenced.pla_discharge_started",
    "discharge_completed": "a.discharge_commenced",
    "nor_tendered": "h.nor_tendered",
}
_DATE_FIELD_DEFAULT = "discharge_commenced"


def _fetch_list(from_date, to_date, date_field=None):

    date_col = _DATE_FIELDS.get(
        date_field or _DATE_FIELD_DEFAULT,
        _DATE_FIELDS[_DATE_FIELD_DEFAULT]
    )

    conn = get_db()

    cur = get_cursor(conn)

    query = """
    SELECT
        h.id,
        h.doc_num,
        h.vcn_doc_num,
        h.vessel_name,

        pla.anchored AS pla_date,

        pla_commenced.pla_discharge_started
            AS discharge_commenced,

        a.discharge_commenced
            AS discharge_completed,

        h.nor_tendered,

        h.doc_status,

        v.vessel_agent_name,

        v.operation_type,

        COALESCE(SUM(cd.bl_quantity), 0)
            AS bl_qty,

        STRING_AGG(
            DISTINCT cd.cargo_name,
            ', '
        ) AS cargo_names

    FROM ldud_header h

    LEFT JOIN vcn_header v
        ON v.id = h.vcn_id

    LEFT JOIN vcn_cargo_declaration cd
        ON cd.vcn_id = h.vcn_id

    -- =====================================
    -- FINAL COMPLETED DATE
    -- =====================================

    LEFT JOIN LATERAL (

        SELECT
            aa.discharge_started,
            aa.discharge_commenced

        FROM ldud_anchorage aa

        WHERE aa.ldud_id = h.id

        ORDER BY
            aa.discharge_commenced DESC NULLS LAST

        LIMIT 1

    ) a ON TRUE

    -- =====================================
    -- PLA COMMENCED DATE
    -- =====================================

    LEFT JOIN LATERAL (

        SELECT
            aa.discharge_started
                AS pla_discharge_started

        FROM ldud_anchorage aa

        WHERE aa.ldud_id = h.id
        AND UPPER(aa.anchorage_name)
            LIKE '%%PLA%%'

        ORDER BY aa.anchored ASC

        LIMIT 1

    ) pla_commenced ON TRUE

    -- =====================================
    -- PLA ARRIVAL DATE
    -- =====================================

    LEFT JOIN LATERAL (

        SELECT
            aa.anchored

        FROM ldud_anchorage aa

        WHERE aa.ldud_id = h.id
        AND UPPER(aa.anchorage_name)
            LIKE '%%PLA%%'

        ORDER BY aa.anchored ASC

        LIMIT 1

    ) pla ON TRUE

    WHERE LOWER(h.operation_type) = 'import'

    AND DATE(""" + date_col + """) BETWEEN %s AND %s
"""

    # =========================================
    # ONLY ACTIVE COMMENCED VESSELS
    # =========================================

    if date_field == "discharge_commenced":

        query += """
            AND a.discharge_started IS NOT NULL
            AND a.discharge_commenced IS NULL
        """

    # =========================================
    # FINAL GROUP / ORDER
    # =========================================

    query += """
        GROUP BY
            h.id,
            h.doc_num,
            h.vcn_doc_num,
            h.vessel_name,
            h.nor_tendered,
            h.doc_status,
            v.vessel_agent_name,
            v.operation_type,
            pla.anchored,
            pla_commenced.pla_discharge_started,
            a.discharge_started,
            a.discharge_commenced

        ORDER BY
            a.discharge_started DESC NULLS LAST
    """

    cur.execute(
        query,
        (from_date, to_date),
    )

    rows = [dict(r) for r in cur.fetchall()]

    conn.close()

    return rows

def _fetch_vessel_data(ldud_id):
    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("SELECT * FROM ldud_header WHERE id = %s", (ldud_id,))
    header_row = cur.fetchone()
    if not header_row:
        conn.close()
        return {
            "header": {},
            "vcn": {},
            "cargo_list": [],
            "delays": [],
            "vessel_ops": [],
            "barge_lines": [],
            "anchorages": [],
        }
    header = dict(header_row)

    vcn_id = header.get("vcn_id")
    vcn = {}
    cargo_list = []
    if vcn_id:
        cur.execute("SELECT * FROM vcn_header WHERE id = %s", (vcn_id,))
        row = cur.fetchone()

        if row:
            vcn = dict(row)
        cur.execute(
            "SELECT cargo_name, bl_quantity, quantity_uom, bl_no FROM vcn_cargo_declaration WHERE vcn_id = %s",
            (vcn_id,),
        )
        cargo_list = [dict(r) for r in cur.fetchall()]

        cur.execute(
            """
        SELECT DISTINCT
            d.*,
            vdt.name AS delay_name,
            vdt.type AS delay_type_name
        FROM ldud_delays d
        LEFT JOIN vessel_delay_types vdt
            ON d.delay_name = vdt.name
        WHERE d.ldud_id = %s
        ORDER BY d.start_datetime
     """,
            (ldud_id,),
        )

    delays = [dict(r) for r in cur.fetchall()]

    cur.execute(
        "SELECT * FROM ldud_vessel_operations WHERE ldud_id = %s ORDER BY start_time",
        (ldud_id,),
    )
    vessel_ops = [dict(r) for r in cur.fetchall()]

    cur.execute(
        "SELECT * FROM ldud_barge_lines WHERE ldud_id = %s ORDER BY along_side_vessel",
        (ldud_id,),
    )
    barge_lines = [dict(r) for r in cur.fetchall()]

    cur.execute(
        "SELECT * FROM ldud_anchorage WHERE ldud_id = %s ORDER BY id", (ldud_id,)
    )
    anchorages = [dict(r) for r in cur.fetchall()]

    conn.close()
    return {
        "header": header,
        "vcn": vcn,
        "cargo_list": cargo_list,
        "delays": delays,
        "vessel_ops": vessel_ops,
        "barge_lines": barge_lines,
        "anchorages": anchorages,
    }


# ── Excel builder ───────────────────────────────────────────────────────────


def _write_vessel_sheet(ws, data):
    """Write one vessel's discharged report onto an existing worksheet."""
    header = data["header"]
    vcn = data["vcn"]
    cargo_list = data["cargo_list"]
    delays = data["delays"]
    vessel_ops = data["vessel_ops"]
    barge_lines = data["barge_lines"]

    vessel_name = header.get("vessel_name", "Vessel")
    doc_num = header.get("doc_num", "")
    bl_qty = sum(float(c.get("bl_quantity") or 0) for c in cargo_list)
    cargo_nm = (
        ", ".join(c["cargo_name"] for c in cargo_list if c.get("cargo_name")) or ""
    )
    type_of_disc = vcn.get("type_of_discharge", "")
    vessel_agent = vcn.get("vessel_agent_name", "")
    operation_type = header.get("operation_type", "")

    anchorages = data.get("anchorages", [])

    mfl_date = None
    pla_date = None

    for a in anchorages:

        anch_name = (a.get("anchorage_name") or "").upper()

        anchored = a.get("anchored")

        anchored_dt = _parse_dt(anchored)

        if not anchored_dt:
            continue

        # PLA
        if "PLA" in anch_name:

            if pla_date is None or anchored_dt < pla_date:
                pla_date = anchored_dt

        # MFL (X3, J3 etc.)
        else:

            if mfl_date is None or anchored_dt < mfl_date:
                mfl_date = anchored_dt

    arrived_pla = pla_date.strftime("%d-%m-%Y %H:%M") if pla_date else "-"

    arrived_mbpt = mfl_date.strftime("%d-%m-%Y %H:%M") if mfl_date else "-"

    disc_start = None
    disc_end = None

    started_list = [
        _parse_dt(a.get("discharge_started"))
        for a in anchorages
        if "PLA" in (a.get("anchorage_name") or "").upper()
        and a.get("discharge_started")
    ]

    completed_list = [
        _parse_dt(a.get("discharge_commenced"))
        for a in anchorages
        if a.get("discharge_commenced")
    ]

    started_list = [x for x in started_list if x]
    completed_list = [x for x in completed_list if x]

    if started_list:
        disc_start = min(started_list)

    if completed_list:
        disc_end = max(completed_list)

    disc_start_str = _fmt_dt(disc_start) or "-"

    disc_end_str = _fmt_dt(disc_end) or "-"

    date_field = data.get("date_field", _DATE_FIELD_DEFAULT)

    date_field = (date_field or "").strip().lower()

    # SAME LOGIC AS PREVIEW

    if date_field == "discharge_commenced":

        disc_end_str = ""
    # Format custom clearance date to match preview
    custom_cleared_fmt = _fmt_dt(header.get("custom_clearance")) or "N/A"

    # 7 columns: A=20, B=8, C=8, D=14, E=8, F=18, G=22
    for i, w in enumerate([28, 16, 16, 18, 16, 24, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    NC = 7
    _nw = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _mbdr(row, c1, c2, fill=XL_WHITE):
        """Apply perimeter thin borders to every cell in a horizontal merged range."""
        for ci in range(c1, c2 + 1):
            b = Border(
                left=_thin if ci == c1 else None,
                right=_thin if ci == c2 else None,
                top=_thin,
                bottom=_thin,
            )
            try:
                ws.cell(row, ci).border = b
                ws.cell(row, ci).fill = _fill(fill)
            except AttributeError:
                pass  # MergedCell in older openpyxl

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    c = ws.cell(1, 1, "MOTHER VESSSEL DISCHARGED REPORT")
    c.font = _font(bold=True, size=XL_TITLE_SZ)
    c.fill = _fill(XL_WHITE)
    c.alignment = _ctr
    c.border = _bdr
    _mbdr(1, 1, NC)

    # ── Row 2: Blank ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 10
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
    c = ws.cell(2, 1, "")
    c.fill = _fill(XL_WHITE)
    c.border = _bdr
    _mbdr(2, 1, NC)

    # ── Rows 3-10: Header block ───────────────────────────────────────────────
    # A:B merged = left label (bold) | C:E merged = left value | F = right label | G = right value
    def _hdr2(row, ll, lv, rl, rv, height=16):
        ws.row_dimensions[row].height = height
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row, 1, ll)
        c.font = _font(bold=True)
        c.fill = _fill(XL_WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _bdr
        _mbdr(row, 1, 2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        lv_value = _parse_number(lv)
        c = ws.cell(row, 3, lv_value if lv_value is not None else "")
        c.font = _font()
        c.fill = _fill(XL_WHITE)
        c.alignment = _left
        c.border = _bdr
        if isinstance(lv_value, (int, float)):
            c.number_format = (
                "#,##0"
                if isinstance(lv_value, int) or float(lv_value).is_integer()
                else "#,##0.00"
            )
        _mbdr(row, 3, 5)
        c = ws.cell(row, 6, rl)
        c.font = _font(bold=True)
        c.fill = _fill(XL_WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        c.border = _bdr
        rv_value = _parse_number(rv)
        c = ws.cell(row, 7, rv_value if rv_value is not None else "")
        c.font = _font()
        c.fill = _fill(XL_WHITE)
        c.alignment = _left
        c.border = _bdr
        if isinstance(rv_value, (int, float)):
            c.number_format = (
                "#,##0"
                if isinstance(rv_value, int) or float(rv_value).is_integer()
                else "#,##0.00"
            )

    # =====================================================
    # HEADER SECTION SAME AS PREVIEW
    # =====================================================

    _hdr2(3, "Mother Vessel", vessel_name, "Mother Vessel Sr. No", doc_num, height=24)

    _hdr2(
        4,
        "Cargo Handler",
        type_of_disc or "Full Discharge",
        "Stevedores",
        vessel_agent or "SAIPRO SHIPPING SERVICES",
        height=24,
    )

    _hdr2(5, "Cargo Type", cargo_nm, "Quantity as B/L", _fmt_qty(bl_qty), height=24)

    _hdr2(
        6,
        "Charter Type",
        operation_type or "Import",
        "Custom Cleared",
        custom_cleared_fmt,
        height=24,
    )

    _hdr2(7, "Arrived at PLA", arrived_pla, "Arrived at MBPT", arrived_mbpt, height=24)

    _hdr2(
        8,
        "Discharge Commenced",
        disc_start_str,
        "Discharge Completed",
        disc_end_str,
        height=24,
    )

    # _hdr2(
    #     9,
    #     "Committed Discharge Rate as per Charter Party Agreement",
    #     "-",
    #     "Demurrage Rate",
    #     "-",
    #     height=24,
    # )

    # _hdr2(
    #     10,
    #     "Committed Discharge Rate as per Barge Owner Agreement",
    #     "-",
    #     "Despatch Rate",
    #     "-",
    #     height=24,
    # )

    # ── Row 11: Blank separator ───────────────────────────────────────────────
    ws.row_dimensions[11].height = 6
    for ci in range(1, NC + 1):
        c = ws.cell(11, ci, "")
        c.fill = _fill(XL_WHITE)
        c.border = _bdr

    # ── Row 12: Day-wise section headers ──────────────────────────────────────
    ws.row_dimensions[12].height = 18
    c = ws.cell(12, 1, "Day Wise Discharge : ")
    c.font = _font(bold=True)
    c.fill = _fill(XL_GREY)
    c.alignment = _nw
    c.border = _bdr
    ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=3)
    c = ws.cell(12, 2, "M. V. Discharge")
    c.font = _font(bold=True)
    c.alignment = _ctr
    _mbdr(12, 2, 3, XL_GREY)
    c = ws.cell(12, 4, "REMARKS")
    c.font = _font(bold=True)
    c.fill = _fill(XL_GREY)
    c.alignment = _ctr
    c.border = _bdr
    c = ws.cell(12, 5, "")
    c.fill = _fill(XL_WHITE)
    c.border = _bdr
    c = ws.cell(12, 6, "Day Wise Barge Discharge : ")
    c.font = _font(bold=True)
    c.fill = _fill(XL_GREY)
    c.alignment = _nw
    c.border = _bdr
    c = ws.cell(12, 7, "Jetty Discharge")
    c.font = _font(bold=True)
    c.fill = _fill(XL_GREY)
    c.alignment = _ctr
    c.border = _bdr

    # ── Day-wise data ─────────────────────────────────────────────────────────
    mv_by_day = defaultdict(float)
    for op in vessel_ops:
        k = _day_key(op.get("start_time"))
        if k:
            mv_by_day[k] += float(op.get("quantity") or 0)

    bg_by_day = defaultdict(float)
    for bl in barge_lines:
        k = _day_key(bl.get("completed_discharge_berth") or bl.get("along_side_vessel"))
        if k:
            bg_by_day[k] += float(bl.get("discharge_quantity") or 0)

    mv_dates = sorted(mv_by_day.keys())
    bg_dates = sorted(bg_by_day.keys())
    max_rows = max(len(mv_dates), len(bg_dates), 1)
    mv_total = bg_total = 0.0

    r = 13
    for i in range(max_rows):
        ws.row_dimensions[r].height = 15
        if i < len(mv_dates):
            dk = mv_dates[i]
            qty = mv_by_day[dk]
            mv_total += qty
            c = ws.cell(r, 1, dk.strftime("%d-%b-%y").upper())
            c.font = _font()
            c.fill = _fill(XL_WHITE)
            c.alignment = _ctr
            c.border = _bdr
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            c = ws.cell(r, 2, int(round(qty)))
            c.font = _font()
            c.alignment = _ctr
            c.number_format = "#,##0"
            _mbdr(r, 2, 3)
        else:
            c = ws.cell(r, 1, "")
            c.fill = _fill(XL_WHITE)
            c.border = _bdr
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            c = ws.cell(r, 2, "")
            _mbdr(r, 2, 3)
        c = ws.cell(r, 4, "")
        c.fill = _fill(XL_WHITE)
        c.border = _bdr
        c = ws.cell(r, 5, "")
        c.fill = _fill(XL_WHITE)
        c.border = _bdr
        if i < len(bg_dates):
            dk = bg_dates[i]
            qty = bg_by_day[dk]
            bg_total += qty
            c = ws.cell(r, 6, dk.strftime("%d-%b-%y").upper())
            c.font = _font()
            c.fill = _fill(XL_WHITE)
            c.alignment = _ctr
            c.border = _bdr
            c = ws.cell(r, 7, int(round(qty)))
            c.font = _font()
            c.fill = _fill(XL_WHITE)
            c.alignment = _ctr
            c.number_format = "#,##0"
            c.border = _bdr
        else:
            c = ws.cell(r, 6, "")
            c.fill = _fill(XL_WHITE)
            c.border = _bdr
            c = ws.cell(r, 7, "")
            c.fill = _fill(XL_WHITE)
            c.border = _bdr
        r += 1

    # ── Total row (numbers only, no "TOTAL" label) ────────────────────────────
    ws.row_dimensions[r].height = 16
    c = ws.cell(r, 1, "")
    c.fill = _fill(XL_WHITE)
    c.border = _bdr
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c = ws.cell(r, 2, int(round(mv_total)))
    c.font = _font(bold=True)
    c.alignment = _ctr
    c.number_format = "#,##0"
    _mbdr(r, 2, 3)
    c = ws.cell(r, 4, "")
    c.fill = _fill(XL_WHITE)
    c.border = _bdr
    c = ws.cell(r, 5, "")
    c.fill = _fill(XL_WHITE)
    c.border = _bdr
    c = ws.cell(r, 6, "")
    c.fill = _fill(XL_WHITE)
    c.border = _bdr
    c = ws.cell(r, 7, int(round(bg_total)))
    c.font = _font(bold=True)
    c.fill = _fill(XL_WHITE)
    c.alignment = _ctr
    c.number_format = "#,##0"
    c.border = _bdr
    r += 1

    # ── Blank row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 8
    for ci in range(1, NC + 1):
        c = ws.cell(r, ci, "")
        c.fill = _fill(XL_WHITE)
        c.border = _bdr
    r += 1

    # ── Delay log ─────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, " Average Mother Vessel Delays : ")
    c.font = _font(bold=True)
    c.alignment = _left
    _mbdr(r, 1, NC, XL_GREY)
    r += 1

    ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, "DELAY TYPE NAME")
    c.font = _font(bold=True)
    c.alignment = _ctr
    _mbdr(r, 1, 2, XL_GREY)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(r, 3, "DELAY DESCRIPTION")
    c.font = _font(bold=True)
    c.alignment = _ctr
    _mbdr(r, 3, 5, XL_GREY)
    c = ws.cell(r, 6, "MODU")
    c.font = _font(bold=True)
    c.fill = _fill(XL_GREY)
    c.alignment = _ctr
    c.border = _bdr
    c = ws.cell(r, 7, "CALCULATED HRS")
    c.font = _font(bold=True)
    c.fill = _fill(XL_GREY)
    c.alignment = _ctr
    c.border = _bdr
    r += 1

    total_delay_mins = 0.0
    total_calculated_mins = 0.0
    for d in delays:
        ws.row_dimensions[r].height = 15
        mins = float(d.get("total_time_mins") or 0)
        total_delay_mins += mins
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(r, 1, d.get("delay_type_name", ""))
        c.font = _font(size=XL_SMALL_SZ)
        c.alignment = _left
        _mbdr(r, 1, 2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

        delay_name = d.get("delay_name", "") or ""
        crane_no = d.get("crane_number", "") or ""

        crane_list = [x.strip() for x in crane_no.split(",") if x.strip()]

        if len(crane_list) >= 4:

            delay_desc = f" All Crane  {delay_name}"

        elif crane_no:

            delay_desc = f" Crane No {crane_no}  {delay_name}"

        else:

            delay_desc = delay_name

        c = ws.cell(r, 3, delay_desc)

        c.font = _font(size=XL_SMALL_SZ)
        c.alignment = _left

        _mbdr(r, 3, 5)
        # MODU
        c = ws.cell(r, 6, _fmt_modu(mins))
        c.font = _font(size=XL_SMALL_SZ)
        c.fill = _fill(XL_WHITE)
        c.alignment = _ctr
        c.border = _bdr

        # CALCULATED HOURS
        crane_no = d.get("crane_number", "") or ""

        crane_list = [x.strip() for x in crane_no.split(",") if x.strip()]
        crane_count = len(crane_list)

        calc_hours = 0

        if crane_count == 1:
            calc_hours = mins / 4

        elif crane_count == 2:
            calc_hours = mins / 2

        elif crane_count == 3:
            calc_hours = (mins / 4) * 3

        elif crane_count >= 4:
            calc_hours = mins

        else:
            calc_hours = mins

        # ALWAYS add total
        total_calculated_mins += calc_hours

        c = ws.cell(r, 7, _fmt_modu(calc_hours))
        c.font = _font(size=XL_SMALL_SZ)
        c.fill = _fill(XL_WHITE)
        c.alignment = _ctr
        c.border = _bdr
        r += 1

    # Total delays: A:E right-aligned, F = total MODU
    ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, "Total Delays")
    c.font = _font(bold=True)
    c.alignment = _right
    _mbdr(r, 1, 5)
    c = ws.cell(r, 6, _fmt_modu(total_delay_mins))
    c.font = _font(bold=True)
    c.fill = _fill(XL_LAVEND)
    c.alignment = _ctr
    c.border = _bdr

    c = ws.cell(r, 7, _fmt_modu(total_calculated_mins))
    c.font = _font(bold=True)
    c.fill = _fill(XL_LAVEND)
    c.alignment = _ctr
    c.border = _bdr
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, "DELAY TYPE NAME")
    c.font = _font(bold=True)
    c.alignment = _ctr
    _mbdr(r, 1, 2, XL_CYAN)

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(r, 3, "DELAY DESCRIPTION")
    c.font = _font(bold=True)
    c.alignment = _ctr
    _mbdr(r, 3, 5, XL_CYAN)

    c = ws.cell(r, 6, "Total Hrs")
    c.font = _font(bold=True)
    c.fill = _fill(XL_CYAN)
    c.alignment = _ctr
    c.border = _bdr

    c = ws.cell(r, 7, "Calculated Hrs")
    c.font = _font(bold=True)
    c.fill = _fill(XL_CYAN)
    c.alignment = _ctr
    c.border = _bdr

    r += 1

    summary = {}

    for d in delays:

        delay_type = d.get("delay_type_name", "") or ""
        delay_name = d.get("delay_name", "") or ""

        mins = float(d.get("total_time_mins") or 0)

        crane_no = d.get("crane_number", "") or ""
        crane_count = len([x for x in crane_no.split(",") if x.strip()])

        calc_mins = mins

        if crane_count == 1:
            calc_mins = mins / 4

        elif crane_count == 2:
            calc_mins = mins / 2

        elif crane_count == 3:
            calc_mins = (mins / 4) * 3

        key = (delay_type, delay_name)

        if key not in summary:
            summary[key] = {"mins": 0, "calc": 0}

        summary[key]["mins"] += mins
        summary[key]["calc"] += calc_mins

    total_calc = 0
    total_hrs = 0

    for (delay_type, delay_name), vals in summary.items():

        # ONLY summary rows here

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        c = ws.cell(r, 1, delay_type)

        _mbdr(r, 1, 2)

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

        c = ws.cell(r, 3, delay_name)

        _mbdr(r, 3, 5)

        c = ws.cell(r, 6, _fmt_modu(vals["mins"]))
        c.alignment = _ctr
        c.border = _bdr

        c = ws.cell(r, 7, _fmt_modu(vals["calc"]))
        c.alignment = _ctr
        c.border = _bdr

        total_calc += vals["calc"]

        total_hrs += vals["mins"]

        r += 1

    # =========================================
    # ONLY ONE TOTAL ROW AFTER LOOP
    # =========================================

    ws.row_dimensions[r].height = 16

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    c = ws.cell(r, 1, "Total Delays")

    c.font = _font(bold=True)

    c.alignment = _right

    _mbdr(r, 1, 5, XL_LAVEND)

    c = ws.cell(r, 6, _fmt_modu(total_hrs))

    c.font = _font(bold=True)

    c.fill = _fill(XL_LAVEND)

    c.alignment = _ctr

    c.border = _bdr

    c = ws.cell(r, 7, _fmt_modu(total_calc))

    c.font = _font(bold=True)

    c.fill = _fill(XL_LAVEND)

    c.alignment = _ctr

    c.border = _bdr

    r += 1

    # PERFORMANCE CALCULATIONS

    gross_days = 0
    net_days = 0
    gross_rate = 0
    net_rate = 0

    if (
        date_field in ["discharge_completed", "nor_tendered"]
        and disc_start
        and disc_end
    ):

        total_hours = (disc_end - disc_start).total_seconds() / 3600

        gross_days = round(total_hours / 24, 2)

        # SAME LOGIC AS PREVIEW
        delay_display = _fmt_modu(total_calc)

        parts = str(delay_display).split(".")

        hrs = int(parts[0])

        mins = int(parts[1]) if len(parts) > 1 else 0

        total_delay_hours = hrs + (mins / 60)

        delay_days = round(total_delay_hours / 24, 2)

        net_days = max(round(gross_days - delay_days, 2), 0)

        if gross_days > 0:

            gross_rate = round(float(bl_qty) / gross_days, 2)

        if net_days > 0:

            net_rate = round(float(bl_qty) / net_days, 2)
    # Purple separator row

    ws.row_dimensions[r].height = 10

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    c = ws.cell(r, 1, "")

    _mbdr(r, 1, 7, XL_LAVEND)

    r += 1

    # Gross row

    ws.row_dimensions[r].height = 20

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    c = ws.cell(r, 1, "Gross Discharge Days")

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _left

    _mbdr(r, 1, 2)

    c = ws.cell(r, 3, float(gross_days))
    c.number_format = "0.00"

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _ctr

    c.border = _bdr

    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    c = ws.cell(r, 4, "Gross Discharge Rate Achieved (A/G)")

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _left

    _mbdr(r, 4, 6)

    c = ws.cell(r, 7, float(gross_rate))
    c.number_format = "#,##0.00"

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _ctr

    c.number_format = "#,##0.00"

    c.border = _bdr

    r += 1

    # Net row

    ws.row_dimensions[r].height = 20

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    c = ws.cell(r, 1, "Net Discharge Days")

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _left

    _mbdr(r, 1, 2)

    c = ws.cell(r, 3, float(net_days))
    c.number_format = "0.00"

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _ctr

    c.border = _bdr

    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    c = ws.cell(r, 4, "Net Discharge Rate Achieved (A/G)")

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _left

    _mbdr(r, 4, 6)

    c = ws.cell(r, 7, float(net_rate))
    c.number_format = "#,##0.00"

    c.font = _font(size=XL_SMALL_SZ)

    c.alignment = _ctr

    c.number_format = "#,##0.00"

    c.border = _bdr

    r += 1

    # Remarks header

    ws.row_dimensions[r].height = 18

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    c = ws.cell(r, 1, "Remarks")

    c.font = _font(bold=True)

    c.alignment = _left

    _mbdr(r, 1, 7, XL_LAVEND)

    r += 1

    # Empty remarks rows

    for _ in range(2):

        ws.row_dimensions[r].height = 18

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

        c = ws.cell(r, 1, "")

        _mbdr(r, 1, 7)

        r += 1


def _build_excel(data):
    """Build a single-vessel workbook."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    vessel_name = data["header"].get("vessel_name", "Vessel")
    safe_title = re.sub(r"[\\/*?\[\]:]", "_", vessel_name)[:31]
    ws.title = safe_title or "Report"
    _write_vessel_sheet(ws, data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_all_excel(vessels_data):
    """Build a multi-sheet workbook — one sheet per vessel."""
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for data in vessels_data:
        vessel_name = data["header"].get("vessel_name", "Vessel")
        safe_title = re.sub(r"[\\/*?\[\]:]", "_", vessel_name)[:31] or "Sheet"
        if first:
            ws = wb.active
            ws.title = safe_title
            first = False
        else:
            ws = wb.create_sheet(safe_title)
        _write_vessel_sheet(ws, data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ──────────────────────────────────────────────────────────────────


@bp.route("/module/RP01/vessel-discharged/")
@login_required
def vessel_discharged_index():
    return render_template(
        "vessel_discharged/vessel_discharged_list.html",
        username=session.get("username"),
    )


@bp.route("/api/module/RP01/vessel-discharged/data")
@login_required
def vessel_discharged_data():

    from_date = request.args.get(
        "from_date",
        date.today().replace(day=1).strftime("%Y-%m-%d")
    )

    to_date = request.args.get(
        "to_date",
        date.today().strftime("%Y-%m-%d")
    )

    date_field = request.args.get(
        "date_field",
        _DATE_FIELD_DEFAULT
    )

    date_field = (date_field or "").strip().lower()

    rows = _fetch_list(from_date, to_date, date_field)

    for row in rows:

        # convert datetime
        for k, v in row.items():

            if hasattr(v, "isoformat"):
                row[k] = v.isoformat()


    return jsonify(rows)

@bp.route("/api/module/RP01/vessel-discharged/download-all")
@login_required
def vessel_discharged_download_all():
    from_date = request.args.get(
        "from_date", date.today().replace(day=1).strftime("%Y-%m-%d")
    )
    to_date = request.args.get("to_date", date.today().strftime("%Y-%m-%d"))
    date_field = request.args.get("date_field", _DATE_FIELD_DEFAULT)
    rows = _fetch_list(from_date, to_date, date_field)
    if not rows:
        return Response("No records in selected range", status=404)
    vessels_data = []

    for r in rows:

        d = _fetch_vessel_data(r["id"])

        d["date_field"] = date_field

        vessels_data.append(d)
    buf = _build_all_excel(vessels_data)
    fname = f"MVDischarged_{from_date}_to_{to_date}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@bp.route("/api/module/RP01/vessel-discharged/<int:ldud_id>/download")
@login_required
def vessel_discharged_download(ldud_id):

    date_field = request.args.get("date_field", _DATE_FIELD_DEFAULT)

    date_field = (date_field or "").strip().lower()

    data = _fetch_vessel_data(ldud_id)

    data["date_field"] = date_field

    if not data["header"]:
        return jsonify({"error": "Record not found"}), 404

    buf = _build_excel(data)

    vessel = re.sub(
        r"[^A-Za-z0-9_\-]", "_", data["header"].get("vessel_name", "vessel")
    )

    fname = f"MVDischarged_{vessel}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype=(
            "application/" "vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@bp.route("/api/module/RP01/vessel-discharged/<int:ldud_id>/preview")
@login_required
def vessel_discharged_preview(ldud_id):

    date_field = request.args.get("date_field", _DATE_FIELD_DEFAULT)

    date_field = (date_field or "").strip().lower()

    conn = get_db()

    cur = get_cursor(conn)

    # =========================
    # HEADER DATA
    # =========================

    cur.execute(
        """
    SELECT DISTINCT
        h.doc_num,
        h.vessel_name,
        h.custom_clearance,

        STRING_AGG(
            DISTINCT cd.cargo_name,
            ', '
        ) AS cargo,

        COALESCE(
            SUM(DISTINCT cd.bl_quantity),
            0
        ) AS bl_qty,

        h.arrived_mfl,
        h.arrived_mbpt,

       MIN(a.discharge_started)
        AS commence_discharge_berth,

        MAX(a.discharge_commenced)
        AS completed_discharge_berth

    FROM ldud_header h

    LEFT JOIN vcn_cargo_declaration cd
        ON cd.vcn_id = h.vcn_id

    LEFT JOIN ldud_anchorage a
    ON a.ldud_id = h.id

    WHERE h.id = %s

    GROUP BY
        h.doc_num,
        h.vessel_name,
        h.custom_clearance,
        h.arrived_mfl,
        h.arrived_mbpt

""",
        (ldud_id,),
    )
    header = cur.fetchone()

    if not header:
        conn.close()
        return jsonify({"error": "Record not found"}), 404

    # =========================
    # ANCHORAGE DATA
    # =========================

    cur.execute(
        """

        SELECT
            anchorage_name,
            anchored

        FROM ldud_anchorage

        WHERE ldud_id = %s

        ORDER BY anchored

    """,
        (ldud_id,),
    )

    anchorages = cur.fetchall()

    mfl_date = None
    pla_date = None

    for a in anchorages:

        anch_name = (a["anchorage_name"] or "").upper()

        anchored = _parse_dt(a["anchored"])

        if not anchored:
            continue

        # PLA
        if "PLA" in anch_name:

            if pla_date is None or anchored < pla_date:
                pla_date = anchored

        # MFL (X3, J3 etc.)
        else:

            if mfl_date is None or anchored < mfl_date:
                mfl_date = anchored

    arrived_pla = pla_date.strftime("%d-%m-%Y %H:%M") if pla_date else "-"

    arrived_mbpt = mfl_date.strftime("%d-%m-%Y %H:%M") if mfl_date else "-"

    # =========================
    # DAYWISE MV DISCHARGE
    # =========================

    cur.execute(
        """
        SELECT
            DATE(start_time) as d,
            COALESCE(SUM(quantity), 0) as qty

        FROM ldud_vessel_operations

        WHERE ldud_id = %s

        GROUP BY DATE(start_time)

        ORDER BY DATE(start_time)
    """,
        (ldud_id,),
    )

    mv_daywise = cur.fetchall()

    # =========================
    # JETTY DISCHARGE
    # =========================

    cur.execute(
        """
        SELECT
            DATE(trip_start::timestamp) as d,
            COALESCE(SUM(discharge_quantity), 0) as qty

        FROM ldud_barge_lines

        WHERE ldud_id = %s

        GROUP BY DATE(trip_start::timestamp)

        ORDER BY DATE(trip_start::timestamp)
    """,
        (ldud_id,),
    )

    jetty_daywise = cur.fetchall()

    daywise_data = []
    mv_total = 0
    barge_total = 0
    max_rows = max(len(mv_daywise), len(jetty_daywise), 1)
    for i in range(max_rows):
        row = {}
        if i < len(mv_daywise):
            d = mv_daywise[i]["d"]
            qty = int(round(mv_daywise[i]["qty"] or 0))
            row["mv_date"] = d.strftime("%d-%b-%y").upper() if d else ""
            row["mv_qty"] = _fmt_qty(qty)
            mv_total += qty
        else:
            row["mv_date"] = ""
            row["mv_qty"] = ""
        if i < len(jetty_daywise):
            d = jetty_daywise[i]["d"]
            qty = int(round(jetty_daywise[i]["qty"] or 0))
            row["barge_date"] = d.strftime("%d-%b-%y").upper() if d else ""
            row["barge_qty"] = _fmt_qty(qty)
            barge_total += qty
        else:
            row["barge_date"] = ""
            row["barge_qty"] = ""
        daywise_data.append(row)

    mv_total = _fmt_qty(mv_total)
    barge_total = _fmt_qty(barge_total)

    # =========================
    # DELAYS
    # =========================

    cur.execute(
        """
        SELECT
            d.*,
            vdt.type AS delay_type_name

        FROM ldud_delays d
        LEFT JOIN vessel_delay_types vdt
            ON d.delay_name = vdt.name
        WHERE d.ldud_id = %s
        ORDER BY d.start_datetime
    """,
        (ldud_id,),
    )

    delay_rows = [dict(r) for r in cur.fetchall()]

    delays = []
    seen_delays = set()
    summary_map = {}
    total_delay = 0
    total_calc = 0

    for d in delay_rows:
        mins = float(d.get("total_time_mins") or 0)
        crane_no = d.get("crane_number", "") or ""
        crane_list = [x.strip() for x in crane_no.split(",") if x.strip()]

        if len(crane_list) >= 4:
            delay_desc = f" All Crane {d.get('delay_name', '')}"
        elif crane_no:
            delay_desc = f" Crane No {crane_no} {d.get('delay_name', '')}"
        else:
            delay_desc = d.get("delay_name", "")

        calc = mins
        if len(crane_list) == 1:
            calc = mins / 4
        elif len(crane_list) == 2:
            calc = mins / 2
        elif len(crane_list) == 3:
            calc = (mins / 4) * 3
        delay_key = (d.get("delay_type_name", ""), delay_desc, _fmt_modu(mins))

        if delay_key in seen_delays:
            continue

        seen_delays.add(delay_key)

        delays.append(
            {
                "delay_type": d.get("delay_type_name", "") or "",
                "delay_description": delay_desc,
                "modu": _fmt_modu(mins),
                "calc": _fmt_modu(calc),
            }
        )

        delay_type = d.get("delay_type_name", "") or ""
        delay_name = d.get("delay_name", "") or ""

        key = (delay_type, delay_name)
        if key not in summary_map:
            summary_map[key] = {"total": 0, "calc": 0}

        summary_map[key]["total"] += mins
        summary_map[key]["calc"] += calc
        total_delay += mins
        total_calc += calc

    summary = []

    for (delay_type, delay_name), vals in summary_map.items():

        summary.append(
            {
                "delay_type": delay_type,
                # same as Excel
                "delay_description": delay_name,
                "total": _fmt_modu(vals["total"]),
                "calc": _fmt_modu(vals["calc"]),
            }
        )

    totals = {
        "total_delay": _fmt_modu(total_delay),
        "total_calc": _fmt_modu(total_calc),
    }
    # =========================
    # GROSS / NET DAYS
    # =========================

    gross_days = "0"
    net_days = "0"
    gross_rate = "0"
    net_rate = "0"

    if (
        date_field in ["discharge_completed", "nor_tendered"]
        and header["commence_discharge_berth"]
        and header["completed_discharge_berth"]
        ):

        try:

            start = _parse_dt(header["commence_discharge_berth"])

            end = _parse_dt(header["completed_discharge_berth"])

            if start and end:

                total_hours = (end - start).total_seconds() / 3600

                gross_days = round(total_hours / 24, 2)

                delay_display = _fmt_modu(total_calc)

                parts = str(delay_display).split(".")

                hrs = int(parts[0])

                mins = int(parts[1]) if len(parts) > 1 else 0

                delay_hours = hrs + (mins / 60)

                delay_days = round(delay_hours / 24, 2)

                net_days = max(round(gross_days - delay_days, 2), 0)

                bl_qty_val = float(header.get("bl_qty") or 0)

                if gross_days > 0:

                    gross_rate = round(bl_qty_val / gross_days, 2)

                if net_days > 0:

                    net_rate = round(bl_qty_val / net_days, 2)

        except Exception as e:

            print(e)

    conn.close()

    custom_cleared_fmt = _fmt_dt(header.get("custom_clearance")) or "N/A"

    discharge_commenced_fmt = _fmt_dt(header["commence_discharge_berth"])

    discharge_completed_fmt = _fmt_dt(header["completed_discharge_berth"])

    # FILTER DISPLAY LOGIC

    if date_field == "discharge_commenced":

        # hide completed date
        discharge_completed_fmt = ""

        # reset calculations
        gross_days = "0"
        net_days = "0"
        gross_rate_fmt = "0"
        net_rate_fmt = "0"


    elif date_field == "discharge_completed":

        # show all
        pass

    bl_qty_fmt = _fmt_qty(header.get("bl_qty") or 0)

    if gross_rate != "":

        gross_rate_fmt = _fmt_number(gross_rate)

    else:

        gross_rate_fmt = ""

    if net_rate != "":

        net_rate_fmt = _fmt_number(net_rate)

    else:

        net_rate_fmt = ""

    return render_template(
        "vessel_discharged/vessel_discharged_preview.html",
        vessel_name=header["vessel_name"],
        doc_num=header["doc_num"],
        cargo_handler="Full Discharge",
        stevedores="SAIPRO SHIPPING SERVICES",
        cargo_type=header.get("cargo") or "-",
        bl_qty=bl_qty_fmt,
        charter_type="Import",
        custom_cleared=custom_cleared_fmt,
        arrived_pla=arrived_pla,
        arrived_mbpt=arrived_mbpt,
        discharge_commenced=discharge_commenced_fmt,
        discharge_completed=discharge_completed_fmt,
        charter_party_rate="-",
        demurrage_rate="-",
        barge_owner_rate="-",
        despatch_rate="-",
        daywise_data=daywise_data,
        mv_total=mv_total,
        barge_total=barge_total,
        delays=delays,
        summary=summary,
        total_delay=totals["total_delay"],
        total_calc=totals["total_calc"],
        gross_days=gross_days,
        net_days=net_days,
        gross_rate=gross_rate_fmt,
        net_rate=net_rate_fmt,
    )
