from flask import render_template, request, jsonify, session, redirect, url_for, Response
from functools import wraps
from datetime import date, datetime, timedelta
from collections import defaultdict
import io
import re
from openpyxl import Workbook
from .. import bp
from database import get_db, get_cursor

# ── Excel colour / style constants ─────────────────────────────────────────
XL_GREY     = 'C0C0C0'
XL_LAVEND   = 'CCCCFF'
XL_CYAN     = 'CCFFFF'
XL_WHITE    = 'FFFFFF'
XL_NAVY     = '1F4E78'
XL_TITLE_SZ = 14
XL_NORM_SZ  = 10
XL_SMALL_SZ = 9

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_thin   = Side(style='thin',   color='000000')
_med    = Side(style='medium', color='000000')
_bdr    = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
_bdr_ml = Border(left=_med,   right=_thin,  top=_thin,  bottom=_thin)
_ctr    = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_right  = Alignment(horizontal='right',  vertical='center', wrap_text=True)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=XL_NORM_SZ, color='000000'):
    return Font(name='Calibri', bold=bold, size=size, color=color)


def _fmt_dt(val):

    if not val:
        return ''

    try:
        dt = _parse_tat_datetime(val)
        return dt.strftime('%d-%m-%Y %H:%M')

    except Exception:
        return ''

def _safe_float(val):
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return 0.0

def _parse_tat_datetime(val):

    if not val:
        return None

    val = str(val)

    if "T24:" in val:

        date_part, time_part = val.split("T")

        dt = datetime.strptime(
            date_part,
            "%Y-%m-%d"
        ) + timedelta(days=1)

        hh, mm = time_part.split(":")

        return dt.replace(
            hour=0,
            minute=int(mm)
        )

    return datetime.fromisoformat(val)


def _calc_tat(start_val, end_val):

    try:


        start_dt = _parse_tat_datetime(start_val)
        end_dt = _parse_tat_datetime(end_val)



        diff = end_dt - start_dt

        return f"{round(diff.total_seconds()/3600,2)} Hrs"

    except Exception as e:

       

        return ''
    
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── Data fetch 




def _fetch_list(from_date, to_date):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
    WITH discharge_sums AS (
        -- Pre-aggregate ALL lueu_lines once — no per-row subquery
        SELECT
            ll.barge_name,
            ll.source_id,
            SUM(ll.quantity) AS discharge_done_qty
        FROM lueu_lines ll
        WHERE
            ll.is_deleted IS NOT TRUE
            AND ll.source_type = 'VCN'
            AND TO_DATE(ll.entry_date, 'YYYY-MM-DD') <= %s::date
        GROUP BY
            ll.barge_name,
            ll.source_id
    ),

    shift_latest AS (
        -- Pre-aggregate latest shift per vcn_id
        SELECT DISTINCT ON (source_id)
            source_id,
            shift
        FROM lueu_lines
        WHERE
            source_type = 'VCN'
            AND is_deleted IS NOT TRUE
        ORDER BY source_id, id DESC
    )

    SELECT
        l.id,
        l.barge_name AS original_barge_name,
        CONCAT(l.barge_name, '/', COALESCE(l.trip_number::text, '1')) AS barge_name,
        l.trip_number,
        h.vessel_name AS mother_vessel_name,
        h.vcn_id,
        l.cargo_name AS cargo_type,
        l.bpt_bfl AS mbpt_pla,
        COALESCE(l.discharge_quantity, 0) AS qty_mt,
        COALESCE(l.discharge_quantity, 0)
            - COALESCE(ds.discharge_done_qty, 0) AS qty_balance,
        l.trip_start,
        l.anchored_gull_island,
        l.aweigh_gull_island,
        l.along_side_vessel,
        l.commenced_loading,
        l.completed_loading,
        l.cast_off_mv,
        l.amf_at_port,
        l.along_side_berth,
        l.commence_discharge_berth,
        l.completed_discharge_berth,
        l.anchored_gull_island_empty,
        l.aweigh_gull_island_empty,
        l.cast_off_berth AS cast_off_berth_nt,
        l.cast_off_port,
        l.port_crane AS unloaded_by,
        sl.shift

    FROM ldud_barge_lines l

    LEFT JOIN ldud_header h
        ON l.ldud_id = h.id

    LEFT JOIN discharge_sums ds
        ON ds.barge_name = CONCAT(
            l.barge_name, ' / ', COALESCE(l.trip_number::text, '1')
        )
        AND ds.source_id = h.vcn_id

    LEFT JOIN shift_latest sl
        ON sl.source_id = h.vcn_id

    WHERE
        l.barge_name IS NOT NULL
        AND TRIM(l.barge_name) <> ''
        AND l.trip_start IS NOT NULL
        AND TRIM(l.trip_start) <> ''
        AND l.trip_start::timestamp <= %s::timestamp
        AND (
            l.completed_discharge_berth IS NULL
            OR TRIM(l.completed_discharge_berth) = ''
            OR l.completed_discharge_berth::timestamp <= %s::timestamp
        )

    ORDER BY l.trip_start, l.id
    """,
    (to_date, to_date, to_date))

    raw_rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    return raw_rows

def safe_dt(value):

    if not value:
        return None

    if isinstance(value, datetime):
        return value

    value = str(value).strip()

    formats = [
        "%d-%m-%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M"
    ]

    for fmt in formats:

        try:
            return datetime.strptime(
                value,
                fmt
            )
        except:
            pass

    return None

def get_barge_status(row, from_dt, to_dt):

    trip_start          = safe_dt(row.get('trip_start'))
    completed_discharge = safe_dt(row.get('completed_discharge_berth'))
    discharge_start     = safe_dt(row.get('commence_discharge_berth'))
    loading_end         = safe_dt(row.get('completed_loading'))
    loading_start       = safe_dt(row.get('commenced_loading'))
    vessel_side         = safe_dt(row.get('along_side_vessel'))
    qty_balance         = float(row.get('qty_balance') or 0)

    if not trip_start:
        return None

  
    if trip_start > to_dt:
        return None

    
    if completed_discharge and completed_discharge < from_dt:
        return None

   
    if qty_balance <= 0.001 and completed_discharge and completed_discharge < from_dt:
        return None

    qty_mt = float(row.get('qty_mt') or 0)
    qty_balance = float(row.get('qty_balance') or 0)

    # COMPLETED DISCHARGE
    if qty_mt > 0 and qty_balance <= 0:
        return 'completed_discharge'

    # UNDER DISCHARGE
    if discharge_start:
        return 'under_discharge'

    # WAITING FOR DISCHARGE
    if (
        row.get('along_side_berth')
        and not row.get('commence_discharge_berth')
    ):
        return 'waiting_discharge'

    # CURRENTLY LOADING
    if loading_start and not loading_end:
        return 'currently_loading'

    # LOADED & TRANSIT
    if (
        row.get('cast_off_mv')
        and not row.get('along_side_berth')
    ):
        return 'loaded_transit'
    
    
    
    return None

def _fetch_barge_data(barge_line_id):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute(
        """
    SELECT
    l.*,

    h.vessel_name,

    l.port_crane AS unloaded_by,

    COALESCE(
        v.bl_quantity,
        0
    ) AS bl_quantity

    FROM ldud_barge_lines l

    LEFT JOIN ldud_header h
        ON l.ldud_id = h.id

    LEFT JOIN (

        SELECT
            vcn_id,
            cargo_name,
            SUM(bl_quantity) AS bl_quantity

        FROM vcn_cargo_declaration

        GROUP BY
            vcn_id,
            cargo_name

    ) v
        ON h.vcn_id = v.vcn_id
        AND l.cargo_name = v.cargo_name

    WHERE l.id = %s
        """,
        (barge_line_id,)
    )

    row = cur.fetchone()

    if not row:
        conn.close()
        return {}

    data = dict(row)

    # Mother vessel name
    ldud_id = data.get('ldud_id')

    if ldud_id:

        cur.execute(
            """
            SELECT vessel_name
            FROM ldud_header
            WHERE id = %s
            """,
            (ldud_id,)
        )

        hrow = cur.fetchone()

        data['mother_vessel_name'] = (
            hrow['vessel_name']
            if hrow else ''
        )

    else:

        data['mother_vessel_name'] = ''

    # Final values
    data['cargo_type'] = data.get('cargo_name', '')

    data['quantity_uom'] = 'MT'

    conn.close()

    return data

# ── Excel helpers ───────────────────────────────────────────────────────────

def _mbdr(ws, row, c1, c2, fill=XL_WHITE):
    """Apply perimeter thin borders + fill to every cell in a merged range."""
    for ci in range(c1, c2 + 1):
        b = Border(
            left   = _thin if ci == c1 else None,
            right  = _thin if ci == c2 else None,
            top    = _thin,
            bottom = _thin,
        )
        try:
            ws.cell(row, ci).border = b
            ws.cell(row, ci).fill   = _fill(fill)
        except AttributeError:
            pass


# ── Excel sheet writer ──────────────────────────────────────────────────────

def _write_barge_sheet(ws, data):
    """Write one barge line's report onto an existing worksheet."""

    barge_name         = data.get('barge_name', '')
    mother_vessel      = data.get('mother_vessel_name', '')
    cargo_type         = data.get('cargo_type', '')
    cargo_name         = data.get('cargo_name', '')
    mbpt_pla = data.get('bpt_bfl', '')
    bl_qty             = _safe_float(data.get('bl_quantity'))
    discharge_qty      = _safe_float(data.get('discharge_quantity'))
    qty_balance = max(
    bl_qty - discharge_qty,
    0
    )
    quantity_uom       = data.get('quantity_uom', 'MT')
    unloaded_by        = data.get('unloaded_by', '')

    trip_start              = _fmt_dt(data.get('trip_start'))
    anchored_gull_island = _fmt_dt(data.get('anchored_gull_island'))
    aweigh_gull_island   = _fmt_dt(data.get('aweigh_gull_island'))
    along_side_vessel       = _fmt_dt(data.get('along_side_vessel'))
    commenced_loading       = _fmt_dt(data.get('commenced_loading'))
    completed_loading       = _fmt_dt(data.get('completed_loading'))
    cast_off_mv             = _fmt_dt(data.get('cast_off_mv'))
    anchored_loaded = _fmt_dt(data.get('anchored_gull_island_empty'))
    aweigh_loaded   = _fmt_dt(data.get('aweigh_gull_island_empty'))
    amf_at_port             = _fmt_dt(data.get('amf_at_port'))
    along_side_berth        = _fmt_dt(data.get('along_side_berth'))
    commence_discharge_berth = _fmt_dt(data.get('commence_discharge_berth'))
    completed_discharge_berth = _fmt_dt(data.get('completed_discharge_berth'))
    cast_off_berth_nt       = _fmt_dt(data.get('cast_off_berth_nt'))
    cast_off_port           = _fmt_dt(data.get('cast_off_port'))

    # Column widths: A=22, B=22, C=22, D=22, E=22 (5 columns)
    NC = 5
    for i, w in enumerate([22, 22, 22, 22, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _nw = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    c = ws.cell(1, 1, 'MV BARGE DISCHARGE REPORT')
    c.font      = _font(bold=True, size=XL_TITLE_SZ, color='FFFFFF')
    c.fill      = _fill(XL_NAVY)
    c.alignment = _ctr
    c.border    = _bdr
    _mbdr(ws, 1, 1, NC, XL_NAVY)

    # ── Row 2: Blank ──────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 8
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
    c = ws.cell(2, 1, '')
    c.fill   = _fill(XL_WHITE)
    c.border = _bdr
    _mbdr(ws, 2, 1, NC)

    # ── Helper: 2-column label|value row ─────────────────────────────────────
    def _hdr2(row, label, value, height=16):
        ws.row_dimensions[row].height = height
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row, 1, label)
        c.font      = _font(bold=True)
        c.fill      = _fill(XL_WHITE)
        c.alignment = _left
        c.border    = _bdr
        _mbdr(ws, row, 1, 2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=NC)
        c = ws.cell(row, 3, value if value is not None else '')
        c.font      = _font()
        c.fill      = _fill(XL_WHITE)
        c.alignment = _left
        c.border    = _bdr
        _mbdr(ws, row, 3, NC)

    # ── Rows 3-9: Header block ────────────────────────────────────────────────
    _hdr2(3,  'Barge / MBC Name',    barge_name)
    _hdr2(4,  'Mother Vessel Name',  mother_vessel)
    _hdr2(5,  'Cargo Type',          cargo_type)
    _hdr2(6,  'Cargo',               cargo_name)
    _hdr2(7,  'MBPT / PLA',          mbpt_pla)
    _hdr2(8,  f'Qty ({quantity_uom})',  bl_qty or '')
    _hdr2(9,  f'Qty Balance ({quantity_uom})', round(qty_balance, 2) if qty_balance else '')

    # ── Row 10: Blank separator ───────────────────────────────────────────────
    ws.row_dimensions[10].height = 8
    for ci in range(1, NC + 1):
        c = ws.cell(10, ci, '')
        c.fill   = _fill(XL_WHITE)
        c.border = _bdr

    # ── Row 11: Section header ────────────────────────────────────────────────
    ws.row_dimensions[11].height = 18
    ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=NC)
    c = ws.cell(11, 1, ' Barge Movement Timeline')
    c.font      = _font(bold=True)
    c.alignment = _left
    _mbdr(ws, 11, 1, NC, XL_GREY)

    # ── Row 12: Sub-header labels ─────────────────────────────────────────────
    ws.row_dimensions[12].height = 18
    headers_12 = ['EVENT', 'DATE / TIME', 'EVENT', 'DATE / TIME', '']
    for ci, h in enumerate(headers_12, 1):
        c = ws.cell(12, ci, h)
        c.font      = _font(bold=True)
        c.fill      = _fill(XL_GREY)
        c.alignment = _ctr
        c.border    = _bdr

    # ── Helper: side-by-side event row (col1|col2 | col3|col4 | col5 empty) ──
    def _evt2(row, lbl1, val1, lbl2='', val2='', height=15):
        ws.row_dimensions[row].height = height
        c = ws.cell(row, 1, lbl1)
        c.font      = _font(bold=True, size=XL_SMALL_SZ)
        c.fill      = _fill(XL_WHITE)
        c.alignment = _left
        c.border    = _bdr
        c = ws.cell(row, 2, val1)
        c.font      = _font(size=XL_SMALL_SZ)
        c.fill      = _fill(XL_WHITE)
        c.alignment = _ctr
        c.border    = _bdr
        c = ws.cell(row, 3, lbl2)
        c.font      = _font(bold=True, size=XL_SMALL_SZ)
        c.fill      = _fill(XL_WHITE)
        c.alignment = _left
        c.border    = _bdr
        c = ws.cell(row, 4, val2)
        c.font      = _font(size=XL_SMALL_SZ)
        c.fill      = _fill(XL_WHITE)
        c.alignment = _ctr
        c.border    = _bdr
        c = ws.cell(row, 5, '')
        c.fill      = _fill(XL_WHITE)
        c.border    = _bdr

    r = 13
    timeline_rows = [
        ('Trip Start',                  trip_start,               'Anchored Gull Island',          anchored_gull_island),
        ('Aweigh Gull Island',          aweigh_gull_island,        'Alongside Vessel (MV)',          along_side_vessel),
        ('Loading Start',               commenced_loading,        'Loading End',                    completed_loading),
        ('Cast Off MV',                 cast_off_mv,              'Anch. Gull Island (Loaded)',     anchored_loaded),
        ('Aweigh Gull Island (Loaded)', aweigh_loaded,            'AMF at Port',                    amf_at_port),
        ('Alongside Berth',             along_side_berth,         'Discharge Start (Berth)',         commence_discharge_berth),
        ('Discharge End (Berth)',        completed_discharge_berth,'Cast Off Berth NT',              cast_off_berth_nt),
        ('Cast Off Port',               cast_off_port,            'Unloaded By',                    unloaded_by),
    ]
    for lbl1, val1, lbl2, val2 in timeline_rows:
        _evt2(r, lbl1, val1, lbl2, val2)
        r += 1

    # ── Blank separator ───────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 8
    for ci in range(1, NC + 1):
        c = ws.cell(r, ci, '')
        c.fill   = _fill(XL_WHITE)
        c.border = _bdr
    r += 1

    # ── Summary section header ────────────────────────────────────────────────
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, ' Quantity Summary')
    c.font      = _font(bold=True)
    c.alignment = _left
    _mbdr(ws, r, 1, NC, XL_CYAN)
    r += 1

    def _sumrow(row, label, value, height=15):
        ws.row_dimensions[row].height = height
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, label)
        c.font      = _font(bold=True, size=XL_SMALL_SZ)
        c.alignment = _left
        _mbdr(ws, row, 1, 3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NC)
        c = ws.cell(row, 4, value)
        c.font      = _font(size=XL_SMALL_SZ)
        c.alignment = _ctr
        _mbdr(ws, row, 4, NC)

    _sumrow(r, f'BL Quantity ({quantity_uom})',          bl_qty or 0)
    r += 1
    _sumrow(r, f'Discharge Quantity ({quantity_uom})',   round(discharge_qty, 2))
    r += 1
    _sumrow(r, f'Balance Quantity ({quantity_uom})',     round(qty_balance, 2))
    r += 1
    _sumrow(
    r,
    'TAT',
    _calc_tat(
        data.get('trip_start'),
        data.get('cast_off_port')
    )
    )
    r += 1

    # ── Lavender separator ────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 10
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, '')
    _mbdr(ws, r, 1, NC, XL_LAVEND)
    r += 1

    # ── Remarks header ────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, 'Remarks')
    c.font      = _font(bold=True)
    c.alignment = _left
    _mbdr(ws, r, 1, NC, XL_LAVEND)
    r += 1

    for _ in range(3):
        ws.row_dimensions[r].height = 18
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, '')
        _mbdr(ws, r, 1, NC)
        r += 1


# ── Flat summary sheet (all barges) ─────────────────────────────────────────

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def _write_summary_sheet(ws, rows):



    thin = Side(style='thin', color='D0D7E2')

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    header_fill = PatternFill(
        start_color='E6E6E6',
        end_color='E6E6E6',
        fill_type='solid'
    )

    header_font = Font(
        bold=True,
        size=10,
        color='000000'
    )

    body_font = Font(
        size=10,
        color='333333'
    )

    center = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

    left = Alignment(
        horizontal='left',
        vertical='center'
    )

    headers = [
        'Sr#',
        'Barge / MBC Name',
        'Mother Vessel Name',
        'Cargo Type',
        'MBPT / PLA',
        'Qty (MT)',
        'Qty Balance (MT)',
        'Trip Start',
        'Anch. Gull Island',
        'Aweigh Gull Island',
        'Alongside Vessel (MV)',
        'Loading Start',
        'Loading End',
        'Cast Off MV',
        'Anch. Gull Island (Loaded)',
        'Aweigh Gull Island (Loaded)',
        'AMF at Port',
        'Alongside Berth',
        'Discharge Start (Berth)',
        'Discharge End (Berth)',
        'Cast Off Berth NT',
        'Cast Off Port',
        'Unloaded By',
        'TAT'
    ]

    widths = [
        8, 28, 28, 18, 12,
        14, 16, 20, 20, 20,
        22, 18, 18, 18, 24,
        24, 18, 18, 24, 24,
        20, 18, 18, 10
    ]

    # column widths
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # header row
    ws.row_dimensions[1].height = 28

    for col_num, text in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col_num)

        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    # data rows
    for idx, row in enumerate(rows, start=2):

        values = [
            idx - 1,

            row.get('barge_name', ''),
            row.get('mother_vessel_name', ''),
            row.get('cargo_type', ''),
            row.get('mbpt_pla', ''),

            row.get('qty_mt', ''),
            row.get('qty_balance', ''),

            _fmt_dt(row.get('trip_start')),
            _fmt_dt(row.get('anchored_gull_island')),
            _fmt_dt(row.get('aweigh_gull_island')),

            _fmt_dt(row.get('along_side_vessel')),

            _fmt_dt(row.get('commenced_loading')),
            _fmt_dt(row.get('completed_loading')),

            _fmt_dt(row.get('cast_off_mv')),

            _fmt_dt(row.get('anchored_gull_island_empty')),
            _fmt_dt(row.get('aweigh_gull_island_empty')),

            _fmt_dt(row.get('amf_at_port')),
            _fmt_dt(row.get('along_side_berth')),

            _fmt_dt(row.get('commence_discharge_berth')),

            _fmt_dt(row.get('completed_discharge_berth')),

            _fmt_dt(row.get('cast_off_berth_nt')),
            _fmt_dt(row.get('cast_off_port')),

            row.get('unloaded_by', ''),

            _calc_tat(
                row.get('trip_start'),
                row.get('cast_off_port')
            )
        ]
        
        row_fill = None

        if row.get('completed_discharge_berth'):
            row_fill = PatternFill(
                start_color='C6EFCE',
                end_color='C6EFCE',
                fill_type='solid'
            )

        elif row.get('commence_discharge_berth'):
            row_fill = PatternFill(
                start_color='FFE699',
                end_color='FFE699',
                fill_type='solid'
            )

        elif row.get('completed_loading'):
            row_fill = PatternFill(
                start_color='BDD7EE',
                end_color='BDD7EE',
                fill_type='solid'
            )

        elif row.get('commenced_loading'):
            row_fill = PatternFill(
                start_color='D9D2E9',
                end_color='D9D2E9',
                fill_type='solid'
            )

        elif row.get('trip_start'):
            row_fill = PatternFill(
                start_color='FFF2CC',
                end_color='FFF2CC',
                fill_type='solid'
            )
        
        for col_num, value in enumerate(values, 1):

            cell = ws.cell(row=idx, column=col_num)

            cell.value = value
            cell.font = body_font
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            # Completed Discharge = Green
        

            if col_num in [1, 6, 7, 24]:
                cell.alignment = center
            else:
                cell.alignment = left

      # Freeze header row + first 3 data columns (Sr#, Barge Name, Mother Vessel Name, Cargo Type, MBPT/PLA)
            ws.freeze_panes = 'F2'   
    # filter
    ws.auto_filter.ref = ws.dimensions

# ── Excel builder ────────────────────────────────────────────────────────────



def _build_excel(data):

    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active

    barge_name = data.get('barge_name', 'Barge')

    safe_title = re.sub(r'[\\/*?\[\]:]', '_', barge_name)[:31]

    ws.title = safe_title or 'Report'

    # IMPORTANT
    _write_summary_sheet(ws, [data])

    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    return buf


def _build_all_excel(
    barge_rows,
    mbc_rows,
    shift_rows
):

    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Barge Report"
    _write_summary_sheet(ws1, barge_rows)

    # Sheet 2
    ws2 = wb.create_sheet("MBC Report")
    _write_mbc_sheet(ws2, mbc_rows)

    # Sheet 3
    ws3 = wb.create_sheet("Shift Report")
    _write_discharge_sheet(ws3, shift_rows)

    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    return buf


# ═══════════════════════════════════════════════════════════════════════════════
#  DROP-IN REPLACEMENTS  for  _write_mbc_sheet  and  _write_discharge_sheet
#  Paste these two functions into your views.py, replacing the originals.
#  Do NOT touch _write_summary_sheet (Barge Report) — it stays as-is.
# ═══════════════════════════════════════════════════════════════════════════════


from collections import defaultdict

# ── Row fill colours (same as Barge Report legend) ───────────────────────────
_FILL_COMPLETED = 'C6EFCE'   # green  – cast off done
_FILL_DISCHARGE = 'FFE699'   # yellow – under discharge / unloading
_FILL_WAITING   = 'F9CB9C'   # orange – waiting for discharge
_FILL_LOADED    = 'BDD7EE'   # blue   – loaded / loading completed
_FILL_LOADING   = 'D9D2E9'   # purple – currently loading
_FILL_TRANSIT   = 'FFF2CC'   # pale   – trip started only
_FILL_HEADER    = 'E6E6E6'   # grey   – header row


def _pfill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _thin_border(color='D0D7E2'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


_CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LFT = Alignment(horizontal='left',   vertical='center', wrap_text=False)
_RGT = Alignment(horizontal='right',  vertical='center', wrap_text=False)


def _row_fill_mbc(row: dict) -> str | None:
    if row.get('mbc_cast_off'):
        return _FILL_COMPLETED
    if row.get('unloading_commenced'):
        return _FILL_DISCHARGE
    if row.get('mbc_arrival_port') and not row.get('unloading_commenced'):
        return _FILL_WAITING
    if row.get('completed_loading'):
        return _FILL_LOADED
    if row.get('commenced_loading'):
        return _FILL_LOADING
    if row.get('trip_start'):
        return _FILL_TRANSIT
    return None


def _safe_float(val):
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
#  MBC REPORT SHEET — flat tabular, same style as Barge Report
# ═══════════════════════════════════════════════════════════════════════════════

def _write_mbc_sheet(ws, rows):
 
    
 
    thin      = Side(style='thin', color='D0D7E2')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    hdr_font  = Font(name='Calibri', bold=True,  size=10, color='000000')
    body_font = Font(name='Calibri', bold=False, size=10, color='333333')
    CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT = Alignment(horizontal='left',   vertical='center', wrap_text=False)
    RGT = Alignment(horizontal='right',  vertical='center', wrap_text=False)
 
    COLS = [
        ('Sr#',                         7,   CTR),
        ('MBC Name',                   28,   LFT),
        ('Cargo Type',                  18,   LFT),
        ('BL Qty (MT)',                 14,   RGT),
        ('Balance Qty (MT)',            17,   RGT),
        ('Arrived Load Port',           22,   LFT),
        ('Alongside Berth',             22,   LFT),
        ('Loading Commenced',           22,   LFT),
        ('Loading Completed',           22,   LFT),
        ('Cast Off Load Port',          22,   LFT),
        ('Arrival Gull Island',         22,   LFT),
        ('Departure Gull Island',       22,   LFT),
        ('MBC Arrival Port',            22,   LFT),
        ('MBC AMF At Unloading Berth',  26,   LFT),
        ('Unloading Commenced',         22,   LFT),
        ('Cleaning Commenced',          22,   LFT),
        ('Unloading Completed',         22,   LFT),
        ('MBC Cast Off',                22,   LFT),
        ('Sailed Out Load Port',        22,   LFT),
        ('Unloaded By',                 20,   LFT),
        ('Unloaded Berth',              18,   LFT),
        ('TAT',                         12,   CTR),
    ]
 
    # header row
    ws.row_dimensions[1].height = 28
    for i, (title, w, align) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i)
        c.value = title; c.fill = hdr_fill
        c.font  = hdr_font; c.border = border; c.alignment = CTR
 
    # colour — matches UI STATUS_COLORS exactly
    _STATUS_COLOR = {
        'completed_discharge': 'C6EFCE',
        'waiting_castoff':     'C6EFCE',
        'under_discharge':     'FFE699',
        'waiting_gull':        'FFE599',
        'on_way_dharamtar':    'F6B26B',
        'waiting_discharge':   'F9CB9C',
        'loaded_transit':      '9FC5E8',
        'currently_loading':   'D9D2E9',
    }
 
    def _row_fill(row):
        hex_c = _STATUS_COLOR.get(row.get('current_status', ''))
        if not hex_c:
            # milestone fallback for rows without current_status
            if   row.get('mbc_cast_off') or row.get('vessel_cast_off'):   hex_c = 'C6EFCE'
            elif row.get('unloading_commenced'):                           hex_c = 'FFE699'
            elif row.get('mbc_arrival_port') or row.get('vessel_arrival_port'): hex_c = 'F9CB9C'
            elif row.get('completed_loading') or row.get('loading_completed'):  hex_c = 'BDD7EE'
            elif row.get('commenced_loading') or row.get('loading_commenced'):  hex_c = 'D9D2E9'
            elif row.get('trip_start') or row.get('arrived_load_port'):    hex_c = 'FFF2CC'
        return PatternFill(start_color=hex_c, end_color=hex_c, fill_type='solid') if hex_c else None
 
    # field helper with fallbacks
    def g(row, *keys):
        for k in keys:
            v = row.get(k)
            if v: return v
        return ''
 
    # data rows
    for sr, row in enumerate(rows, 1):
        ridx = sr + 1
        fill = _row_fill(row)
        values = [
            sr,
            g(row, 'mbc_name'),
            g(row, 'cargo_type', 'cargo_name'),
            g(row, 'qty_mt'),
            g(row, 'qty_balance'),
            _fmt_dt(g(row, 'trip_start',              'arrived_load_port')),
            _fmt_dt(g(row, 'along_side_vessel',        'alongside_berth')),
            _fmt_dt(g(row, 'commenced_loading',        'loading_commenced')),
            _fmt_dt(g(row, 'completed_loading',        'loading_completed')),
            _fmt_dt(g(row, 'cast_off_mv',              'cast_off_load_port')),
            _fmt_dt(g(row, 'arrival_gull_island')),
            _fmt_dt(g(row, 'departure_gull_island')),
            _fmt_dt(g(row, 'mbc_arrival_port',         'vessel_arrival_port')),
            _fmt_dt(g(row, 'mbc_amf_unloading_berth',  'vessel_all_made_fast')),
            _fmt_dt(g(row, 'unloading_commenced')),
            _fmt_dt(g(row, 'cleaning_commenced')),
            _fmt_dt(g(row, 'unloading_completed')),
            _fmt_dt(g(row, 'mbc_cast_off',             'vessel_cast_off')),
            _fmt_dt(g(row, 'sailed_out_load_port')),
            g(row,  'vessel_unloaded_by',  'unloaded_by'),
            g(row,  'unloaded_berth',      'vessel_unloading_berth'),
            '',  # TAT — blank for MBC
        ]
        for ci, (val, (_, _, align)) in enumerate(zip(values, COLS), 1):
            c = ws.cell(row=ridx, column=ci)
            c.value = val; c.font = body_font
            c.border = border; c.alignment = align
            if fill: c.fill = fill
 
    ws.freeze_panes    = 'F2'
    ws.auto_filter.ref = ws.dimensions



def get_filtered_mbc_rows(
    from_dt,
    to_dt,
    column_filter=None,
    status_filter='all',
    selected_mbc=None,
    selected_shift=None
):
    # use same query from get_mbc_data()

    rows = fetch_mbc_rows()

    filtered_rows = []

    for row in rows:

        status = get_mbc_status(row)
        row['current_status'] = status

        if status is None:
            continue

        if status_filter != 'all' and status != status_filter:
            continue

        if selected_mbc and row.get('mbc_name') != selected_mbc:
            continue

        if selected_shift and row.get('shift') != selected_shift:
            continue

        filtered_rows.append(row)

    return filtered_rows



#  DROP-IN REPLACEMENT  for  _write_discharge_sheet  ONLY



from collections import defaultdict


def _pfill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _bdr(color='000000'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


_CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LFT = Alignment(horizontal='left',   vertical='center', wrap_text=False)
_RGT = Alignment(horizontal='right',  vertical='center', wrap_text=False)


def _safe_float(val):
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return 0.0


def _write_discharge_sheet(ws, rows):

    fill_green  = _pfill('C6EFCE')
    fill_yellow = _pfill('FFE699')
    fill_blue   = _pfill('BDD7EE')
    fill_grey   = _pfill('D9D9D9')
    fill_white  = _pfill('FFFFFF')

    border      = _bdr()
    bold        = Font(name='Calibri', bold=True,  size=10)
    bold_green  = Font(name='Calibri', bold=True,  size=10, color='006100')
    norm        = Font(name='Calibri', bold=False, size=10)

    EQUIPMENTS = [
        'Barge Unloader 1', 'Barge Unloader 2',
        'SANY 285-Exavator',
        'Sennebogen J1', 'Sennebogen J5',
        'BUL-01', 'BUL-02', 'BUL-03', 'BUL-04', 'BUL-05',
    ]
    SHIFTS     = ['A Shift', 'B Shift', 'C Shift']
    DATA_ROWS  = 5
    BLOCK_SIZE = DATA_ROWS + 1   # 5 data + 1 total = 6 rows per shift

    # grand_col is ONE column AFTER the last equipment column
    # equipment cols: 2 .. (1 + len(EQUIPMENTS)*4)
    # grand_col     : 2 + len(EQUIPMENTS)*4
    grand_col = 2 + len(EQUIPMENTS) * 4

    
    # ── column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 18
    for col in range(2, grand_col + 2):
        # Every 4th col starting at col 2 = Barge/MBC col → wider
        offset = (col - 2) % 4
        if offset == 0:   # Barge / MBC
            ws.column_dimensions[get_column_letter(col)].width = 22
        elif offset == 1: # Qty (MT)
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif offset == 2: # Cargo
            ws.column_dimensions[get_column_letter(col)].width = 20
        else:             # MV Name
            ws.column_dimensions[get_column_letter(col)].width = 20
    # grand_col (Total Discharge)
    ws.column_dimensions[get_column_letter(grand_col)].width = 16

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24

    # ── Row 1-2 col A: "Shift / Details" ──────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c = ws.cell(1, 1)           # ← write to TOP-LEFT cell of merge only
    c.value     = 'Shift / Details'
    c.font      = bold
    c.alignment = _CTR
    c.border    = border
    c.fill      = fill_grey

    # ── Equipment headers ──────────────────────────────────────────────────────
    eq_start_cols = {}
    cur_col = 2

    for eq in EQUIPMENTS:
        eq_start_cols[eq] = cur_col

        # Row 1: equipment name (merged across 4 cols)
        ws.merge_cells(start_row=1, start_column=cur_col,
                       end_row=1,   end_column=cur_col + 3)
        c = ws.cell(1, cur_col)   # TOP-LEFT of merge
        c.value     = eq
        c.font      = bold_green
        c.fill      = fill_green
        c.alignment = _CTR
        c.border    = border

        # Row 2: sub-headers (individual cells — NOT merged)
        for off, txt in enumerate(['Barge / MBC', 'Qty (MT)', 'Cargo', 'MV Name']):
            c = ws.cell(2, cur_col + off)
            c.value     = txt
            c.font      = bold
            c.fill      = fill_grey
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = border

        cur_col += 4

    # ── grand_col header: rows 1 & 2 separately (NO merge — avoids the bug) ──
    for r in [1, 2]:
        c = ws.cell(r, grand_col)
        c.value     = 'Total Discharge' if r == 1 else ''
        c.font      = bold
        c.fill      = fill_blue
        c.alignment = _CTR
        c.border    = border

    # freeze: header rows fixed, col A (shift label) not frozen so it scrolls
    ws.freeze_panes = 'B3'

    # ── Build shift → equipment → [item, …] map ───────────────────────────────
    shift_map = {s: {eq: [] for eq in EQUIPMENTS} for s in SHIFTS}

    _SHIFT_KEY = {
        'A': 'A Shift', 'B': 'B Shift', 'C': 'C Shift',
        'A SHIFT': 'A Shift', 'B SHIFT': 'B Shift', 'C SHIFT': 'C Shift',
    }

    agg = {}

    for row in rows:
        raw = (row.get('shift') or '').strip().upper()
        key = _SHIFT_KEY.get(raw)
        if not key:
            continue
        eq_val = (row.get('unloaded_by') or row.get('equipment_name') or '').strip()
        if not eq_val:
            continue
        for e in [x.strip() for x in eq_val.split(',') if x.strip()]:
            if e not in shift_map[key]:
                continue
            barge  = row.get('barge_name', '')
            cargo  = row.get('cargo_name') or row.get('cargo_type', '')
            mv     = row.get('mv_name') or row.get('mother_vessel_name', '')
            qty    = _safe_float(row.get('quantity') or row.get('qty_mt'))
            agg_key = (key, e, barge, cargo, mv)
            if agg_key in agg:
                agg[agg_key]['quantity'] += qty
            else:
                agg[agg_key] = {
                    'barge_name': barge,
                    'cargo_name': cargo,
                    'quantity':   qty,
                    'mv_name':    mv,
                }

    for (shift_key, eq, barge, cargo, mv), item in agg.items():
        shift_map[shift_key][eq].append(item)

    # ── Write shift blocks ────────────────────────────────────────────────────
    start_row = 3

    for shift in SHIFTS:
        eq_data = shift_map[shift]

        # Col A shift label — merge only the DATA rows (not the total row)
        ws.merge_cells(start_row=start_row,
                       start_column=1,
                       end_row=start_row + DATA_ROWS - 1,
                       end_column=1)
        c = ws.cell(start_row, 1)   # TOP-LEFT of merge
        c.value     = shift
        c.font      = bold
        c.fill      = fill_white
        c.alignment = _CTR
        c.border    = border

        # Data rows
        for dr in range(DATA_ROWS):
            r = start_row + dr
            ws.row_dimensions[r].height = 28

            for eq in EQUIPMENTS:
                bc    = eq_start_cols[eq]
                items = eq_data.get(eq, [])
                if dr < len(items):
                    item = items[dr]
                    qty  = _safe_float(
                        item.get('quantity') or item.get('qty_mt')
                    )
                    vals = [
                        item.get('barge_name', ''),
                        round(qty, 2) if qty else '',
                        item.get('cargo_name') or item.get('cargo_type', ''),
                        item.get('mv_name') or item.get('mother_vessel_name', ''),
                    ]
                else:
                    vals = ['', '', '', '']

                for off, v in enumerate(vals):
                    c = ws.cell(r, bc + off)
                    c.value     = v
                    c.font      = norm
                    c.border    = border
                    c.alignment = Alignment(
                        horizontal='right' if off == 1 else 'left',
                        vertical='center',
                        wrap_text=True
                    )

            # grand_col data cell (blank for data rows)
            c = ws.cell(r, grand_col)
            c.border = border

        # Shift total row
        total_row = start_row + DATA_ROWS
        ws.row_dimensions[total_row].height = 22

        c = ws.cell(total_row, 1)
        c.value     = f'{shift} Total'
        c.font      = bold
        c.fill      = fill_yellow
        c.alignment = _CTR
        c.border    = border

        shift_grand = 0.0
        for eq in EQUIPMENTS:
            bc       = eq_start_cols[eq]
            items    = eq_data.get(eq, [])
            eq_total = sum(
                _safe_float(i.get('quantity') or i.get('qty_mt'))
                for i in items
            )
            shift_grand += eq_total

            c = ws.cell(total_row, bc)
            c.value = 'Total'; c.font = bold
            c.fill = fill_yellow; c.alignment = _CTR; c.border = border

            c = ws.cell(total_row, bc + 1)
            c.value = round(eq_total, 2) if eq_total else '-'
            c.font = bold; c.fill = fill_yellow
            c.alignment = _RGT; c.border = border

            for off in [2, 3]:
                c = ws.cell(total_row, bc + off)
                c.fill = fill_yellow; c.border = border

        c = ws.cell(total_row, grand_col)
        c.value     = round(shift_grand, 2) if shift_grand else '-'
        c.font      = bold
        c.fill      = fill_yellow
        c.alignment = _RGT
        c.border    = border

        start_row += BLOCK_SIZE

    # ── All Shift Total row ───────────────────────────────────────────────────
    ws.row_dimensions[start_row].height = 24

    c = ws.cell(start_row, 1)
    c.value     = 'All Shift Total'
    c.font      = bold
    c.fill      = fill_blue
    c.alignment = _CTR
    c.border    = border

    overall_total = 0.0
    for eq in EQUIPMENTS:
        bc        = eq_start_cols[eq]
        all_items = [i for s in SHIFTS for i in shift_map[s].get(eq, [])]
        eq_total  = sum(
            _safe_float(i.get('quantity') or i.get('qty_mt'))
            for i in all_items
        )
        overall_total += eq_total

        c = ws.cell(start_row, bc)
        c.value = 'Total'; c.font = bold
        c.fill = fill_blue; c.alignment = _CTR; c.border = border

        c = ws.cell(start_row, bc + 1)
        c.value = round(eq_total, 2) if eq_total else '-'
        c.font = bold; c.fill = fill_blue
        c.alignment = _RGT; c.border = border

        for off in [2, 3]:
            c = ws.cell(start_row, bc + off)
            c.fill = fill_blue; c.border = border

    c = ws.cell(start_row, grand_col)
    c.value     = round(overall_total, 2) if overall_total else '-'
    c.font      = bold
    c.fill      = fill_blue
    c.alignment = _RGT
    c.border    = border

# ── Routes ───────────────────────────────────────────────────────────────────

@bp.route('/module/RP01/mv-barge-report/')
@login_required
def mv_barge_report_index():
    return render_template('daily_barge_report/mv_barge_report.html',
                           username=session.get('username'))


@bp.route('/api/module/RP01/mv-barge-report/data')
@login_required
def mv_barge_report_data():

    from_datetime = request.args.get('from_date', '')
    to_datetime = request.args.get('to_date', '')
    column_filter = request.args.get('column_filter', '')
    status_filter = request.args.get('status_filter', 'all')

    try:
        if from_datetime:
            from_dt = datetime.fromisoformat(from_datetime)
        else:
            from_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        if to_datetime:
            to_dt = datetime.fromisoformat(to_datetime)
        else:
            to_dt = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    except Exception as e:
        print(f"Date parsing error: {e}")
        return jsonify({'error': 'Invalid date format'}), 400

    
    rows = _fetch_list(
        from_dt.strftime('%Y-%m-%d %H:%M:%S'),
        to_dt.strftime('%Y-%m-%d %H:%M:%S')
    )

    filtered_rows = []

    for row in rows:

        status = get_barge_status(
            row,
            from_dt,
            to_dt
        )

        row['current_status'] = status

        if status is None:
            continue

        if status_filter == 'all':
            filtered_rows.append(row)
        elif status == status_filter:
            filtered_rows.append(row)

    rows = filtered_rows
    status_order = {
    'currently_loading': 1,
    'loaded_transit': 2,
    'waiting_discharge': 3,
    'under_discharge': 4,
    'completed_discharge': 5
    }

    rows.sort(
        key=lambda x: (
            status_order.get(x.get('current_status'), 999),
            x.get('barge_name', '')
        )
    )

    if column_filter:
        column_field_map = {
            'trip_start': 'trip_start',
            'anchored_gull_island': 'anchored_gull_island',
            'aweigh_gull_island': 'aweigh_gull_island',
            'along_side_vessel': 'along_side_vessel',
            'commenced_loading': 'commenced_loading',
            'completed_loading': 'completed_loading',
            'cast_off_mv': 'cast_off_mv',
            'anchored_gull_island_empty': 'anchored_gull_island_empty',
            'aweigh_gull_island_empty': 'aweigh_gull_island_empty',
            'amf_at_port': 'amf_at_port',
            'along_side_berth': 'along_side_berth',
            'commence_discharge_berth': 'commence_discharge_berth',
            'completed_discharge_berth': 'completed_discharge_berth',
            'cast_off_berth_nt': 'cast_off_berth_nt',
            'cast_off_port': 'cast_off_port'
        }

        db_field = column_field_map.get(column_filter)
        if db_field:
            rows = [r for r in rows if r.get(db_field)]
            
            

    for row in rows:
        
        row['tat'] = _calc_tat(
            row.get('trip_start'),
            row.get('cast_off_port')
        )

        date_fields = [
            'trip_start',
            'anchored_gull_island',
            'aweigh_gull_island',
            'along_side_vessel',
            'commenced_loading',
            'completed_loading',
            'cast_off_mv',
            'amf_at_port',
            'along_side_berth',
            'commence_discharge_berth',
            'completed_discharge_berth',
            'cast_off_berth_nt',
            'cast_off_port',
            'anchored_gull_island_empty',
            'aweigh_gull_island_empty'
        ]

        for fld in date_fields:
            row[fld] = _fmt_dt(row.get(fld))

    return jsonify(rows)

def get_mbc_status(row):

    # Currently Loading: loading_commenced exists, loading_completed does NOT
    if row.get('commenced_loading') and not row.get('completed_loading'):
        return 'currently_loading'

    # Loaded & Transit: loading_completed exists, arrival_gull_island does NOT
    if row.get('completed_loading') and not row.get('arrival_gull_island'):
        return 'loaded_transit'

    # Waiting at Gull: arrival_gull_island exists, departure_gull_island does NOT
    if row.get('arrival_gull_island') and not row.get('departure_gull_island'):
        return 'waiting_gull'

    # On the Way to Dharamtar: departure_gull_island exists, mbc_arrival_port does NOT
    if row.get('departure_gull_island') and not row.get('mbc_arrival_port'):
        return 'on_way_dharamtar'

    # Waiting for Discharge: mbc_arrival_port exists, unloading_commenced does NOT
    if row.get('mbc_arrival_port') and not row.get('unloading_commenced'):
        return 'waiting_discharge'

    # Under Discharge: unloading_commenced exists, unloading_completed does NOT
    if row.get('unloading_commenced') and not row.get('unloading_completed'):
        return 'under_discharge'

    # Waiting for Cast Off: unloading_completed exists, mbc_cast_off does NOT
    if row.get('unloading_completed') and not row.get('mbc_cast_off'):
        return 'waiting_castoff'

    # ✅ ADD THIS — Completed: mbc_cast_off exists (voyage fully done)
    if row.get('mbc_cast_off'):
        return 'completed_discharge'

    # ✅ ADD THIS — No data at all yet
    return None

@bp.route('/api/module/RP01/mv-barge-report/download-all')
@login_required
def mv_barge_report_download_all():

    from_datetime = request.args.get('from_date', '')
    to_datetime = request.args.get('to_date', '')
    column_filter = request.args.get('column_filter', '')
    status_filter = request.args.get('status_filter', 'all')

    try:
        if from_datetime:
            from_dt = datetime.fromisoformat(from_datetime)
        else:
            from_dt = datetime.now().replace(
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0
            )

        if to_datetime:
            to_dt = datetime.fromisoformat(to_datetime)
        else:
            to_dt = datetime.now().replace(
                hour=23,
                minute=59,
                second=59,
                microsecond=999999
            )

    except Exception:
        return jsonify(
            {'error': 'Invalid date format'}
        ), 400

    # ==================================================
    # BARGE DATA
    # ==================================================

    list_rows = _fetch_list(
        from_dt.strftime('%Y-%m-%d %H:%M:%S'),
        to_dt.strftime('%Y-%m-%d %H:%M:%S')
    )

    filtered_rows = []

    for row in list_rows:

        status = get_barge_status(
            row,
            from_dt,
            to_dt
        )

        if status is None:
            continue

        row['current_status'] = status

        if status_filter == 'all' or status == status_filter:
            filtered_rows.append(row)

    barge_rows = filtered_rows

    if column_filter:

        column_field_map = {
            'trip_start': 'trip_start',
            'anchored_gull_island': 'anchored_gull_island',
            'aweigh_gull_island': 'aweigh_gull_island',
            'along_side_vessel': 'along_side_vessel',
            'commenced_loading': 'commenced_loading',
            'completed_loading': 'completed_loading',
            'cast_off_mv': 'cast_off_mv',
            'anchored_gull_island_empty': 'anchored_gull_island_empty',
            'aweigh_gull_island_empty': 'aweigh_gull_island_empty',
            'amf_at_port': 'amf_at_port',
            'along_side_berth': 'along_side_berth',
            'commence_discharge_berth': 'commence_discharge_berth',
            'completed_discharge_berth': 'completed_discharge_berth',
            'cast_off_berth_nt': 'cast_off_berth_nt',
            'cast_off_port': 'cast_off_port'
        }

        db_field = column_field_map.get(column_filter)

        if db_field:
            barge_rows = [
                r for r in barge_rows
                if r.get(db_field)
            ]

    # ==================================================
    # MBC DATA
    # ==================================================

    conn = get_db()
    cur = get_cursor(conn)

    
    cur.execute("""
        SELECT DISTINCT ON (h.id)
            h.id                            AS mbc_id,
            h.mbc_name,
            h.cargo_name                    AS cargo_type,
            COALESCE(h.bl_quantity, 0)      AS qty_mt,
            (
                COALESCE(h.bl_quantity, 0)
                -
                COALESCE(
                    (SELECT SUM(ll.quantity)
                     FROM lueu_lines ll
                     WHERE ll.source_type = 'MBC'
                       AND ll.source_id   = h.id
                       AND ll.is_deleted IS NOT TRUE),
                    0
                )
            ) AS qty_balance,
            lp.arrived_load_port            AS trip_start,
            lp.alongside_berth              AS along_side_vessel,
            lp.loading_commenced            AS commenced_loading,
            lp.loading_completed            AS completed_loading,
            lp.cast_off_load_port           AS cast_off_mv,
            dp.arrival_gull_island          AS arrival_gull_island,
            dp.departure_gull_island        AS departure_gull_island,
            dp.vessel_arrival_port          AS mbc_arrival_port,
            dp.vessel_all_made_fast         AS mbc_amf_unloading_berth,
            dp.unloading_commenced          AS unloading_commenced,
            dp.cleaning_commenced           AS cleaning_commenced,
            dp.unloading_completed          AS unloading_completed,
            dp.vessel_cast_off              AS mbc_cast_off,
            dp.sailed_out_load_port         AS sailed_out_load_port,
            dp.vessel_unloaded_by           AS vessel_unloaded_by,
            dp.vessel_unloading_berth       AS unloaded_berth
        FROM mbc_header h
        LEFT JOIN mbc_load_port_lines lp ON lp.mbc_id = h.id
        LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
        WHERE h.mbc_name IS NOT NULL
        ORDER BY h.id
    """)
 
    all_mbc = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
 
    _mbc_order = {
        'currently_loading': 1, 'loaded_transit': 2,
        'waiting_gull': 3,      'on_way_dharamtar': 4,
        'waiting_discharge': 5, 'under_discharge': 6,
        'waiting_castoff': 7,   'completed_discharge': 8,
    }
    mbc_rows = []
    for row in all_mbc:
        t  = safe_dt(row.get('trip_start'))
        co = safe_dt(row.get('mbc_cast_off'))
        ud = safe_dt(row.get('unloading_completed'))
        if not t or t > to_dt:
            continue
        if ud and ud < from_dt:
            continue
        if co and co < from_dt:
            continue
        status = get_mbc_status(row)
        if status is None:
            continue
        row['current_status'] = status
        if status_filter == 'all' or status == status_filter:
            mbc_rows.append(row)
    mbc_rows.sort(key=lambda x: _mbc_order.get(x.get('current_status'), 999))
 
 
# ═══════════════════════════════════════════════════════════════════════════════
# CHANGE 2 of 3
# In mv_barge_report_download_all()
# Replace the ENTIRE SHIFT DATA cur.execute block + shift_rows = [...]
# ═══════════════════════════════════════════════════════════════════════════════
 
    conn = get_db()
    cur  = get_cursor(conn)
 
    entry_date_from = from_dt.date()
    entry_date_to   = to_dt.date()
    if (to_dt - from_dt).total_seconds() <= 24 * 3600:
        entry_date_to = entry_date_from
 
    cur.execute("""
        SELECT
            ll.shift,
            ll.equipment_name               AS unloaded_by,
            ll.barge_name,
            ll.cargo_name,
            COALESCE(ll.quantity, 0)        AS quantity,
            ll.source_type,
            CASE
                WHEN ll.source_type = 'VCN' THEN
                    (SELECT h.vessel_name FROM ldud_header h
                     WHERE h.vcn_id = ll.source_id LIMIT 1)
                WHEN ll.source_type = 'MBC' THEN
                    (SELECT h.mbc_name FROM mbc_header h
                     WHERE h.id = ll.source_id LIMIT 1)
                ELSE ''
            END AS mv_name
        FROM lueu_lines ll
        WHERE
            ll.is_deleted IS NOT TRUE
            AND COALESCE(ll.quantity, 0) > 0
            AND ll.entry_date IS NOT NULL
            AND TO_DATE(ll.entry_date, 'YYYY-MM-DD') BETWEEN %s AND %s
        ORDER BY ll.shift, ll.equipment_name, ll.barge_name
    """, (entry_date_from, entry_date_to))
 
    shift_rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
 

    # ==================================================
    # NO DATA
    # ==================================================

    if not barge_rows and not mbc_rows and not shift_rows:
        return Response(
            'No records found',
            status=404
        )

    # ==================================================
    # EXCEL
    # ==================================================

    buf = _build_all_excel(
        barge_rows,
        mbc_rows,
        shift_rows
    )

    fname = (
        f'MVBargeReport_'
        f'{from_dt.strftime("%Y-%m-%d")}'
        f'_to_'
        f'{to_dt.strftime("%Y-%m-%d")}.xlsx'
    )

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition':
            f'attachment; filename="{fname}"'
        }
    )
@bp.route('/api/module/RP01/mv-barge-report/<int:barge_line_id>/download')
@login_required
def mv_barge_report_download(barge_line_id):
    data = _fetch_barge_data(barge_line_id)
    if not data:
        return jsonify({'error': 'Record not found'}), 404
    buf   = _build_excel(data)
    barge = re.sub(r'[^A-Za-z0-9_\-]', '_', data.get('barge_name', 'barge'))
    fname = f'MVBargeReport_{barge}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
    
@bp.route('/api/module/RP01/mv-barge-report/mbc-data')
@login_required
def get_mbc_data():

    from_date     = request.args.get('from_date')
    to_date       = request.args.get('to_date')
    column_filter = request.args.get('column_filter')
    status_filter = request.args.get('status_filter', 'all')
    
    

    try:
        from_dt = datetime.fromisoformat(from_date) if from_date else \
                  datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        to_dt   = datetime.fromisoformat(to_date) if to_date else \
                  datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    except Exception as e:
        return jsonify({'error': 'Invalid date format'}), 400

    conn = get_db()
    cur  = get_cursor(conn)



    # ── Fetch ALL rows — no date casting in SQL to avoid corrupt data crash ──
    # Replace the MBC query with this version that pre-aggregates lueu_lines
    MBC_QUERY = """
    WITH mbc_discharge_sums AS (
        SELECT
            source_id,
            SUM(quantity) AS discharged_qty
        FROM lueu_lines
        WHERE
            source_type = 'MBC'
            AND is_deleted IS NOT TRUE
        GROUP BY source_id
    )
    SELECT DISTINCT ON (h.id)
        h.id                                AS mbc_id,
        CONCAT(h.mbc_name, ' / ', COALESCE(h.doc_num, '')) AS mbc_name,
        h.mbc_name                          AS mbc_name_raw,
        h.cargo_name                        AS cargo_type,
        COALESCE(h.bl_quantity, 0)          AS qty_mt,
        COALESCE(h.bl_quantity, 0)
            - COALESCE(ds.discharged_qty, 0) AS qty_balance,
        lp.arrived_load_port                AS trip_start,
        lp.alongside_berth                  AS along_side_vessel,
        lp.loading_commenced                AS commenced_loading,
        lp.loading_completed                AS completed_loading,
        lp.cast_off_load_port               AS cast_off_mv,
        dp.arrival_gull_island,
        dp.departure_gull_island,
        dp.vessel_arrival_port              AS mbc_arrival_port,
        dp.vessel_all_made_fast             AS mbc_amf_unloading_berth,
        dp.unloading_commenced,
        dp.cleaning_commenced,
        dp.unloading_completed,
        dp.vessel_cast_off                  AS mbc_cast_off,
        dp.sailed_out_load_port,
        dp.vessel_unloaded_by               AS vessel_unloaded_by,
        dp.vessel_unloading_berth           AS unloaded_berth,
        dp.cleaning_completed
    FROM mbc_header h
    LEFT JOIN mbc_load_port_lines lp    ON lp.mbc_id = h.id
    LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
    LEFT JOIN mbc_discharge_sums ds     ON ds.source_id = h.id
    WHERE h.mbc_name IS NOT NULL
    ORDER BY h.id
    """

    cur.execute(MBC_QUERY)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()

    # ── Apply date filter in Python (safe — no SQL casting of corrupt values) ─
    filtered_rows = []

    for row in rows:

        trip_start          = safe_dt(row.get('trip_start'))
        mbc_cast_off        = safe_dt(row.get('mbc_cast_off'))
        unloading_completed = safe_dt(row.get('unloading_completed'))

        # Must have a trip start
        if not trip_start:
            continue

        # Trip must have started on or before to_dt
        if trip_start > to_dt:
            continue

        # ── Mirror barge logic exactly ──────────────────────────────────────

        # If unloading is fully completed AND completed before from_dt → exclude
        # (same as barge: completed_discharge < from_dt → skip)
        if unloading_completed and unloading_completed < from_dt:
            continue

        # If mbc_cast_off exists AND is before from_dt → exclude
        # (extra safety for fully-done voyages with corrupt or old cast_off)
        if mbc_cast_off and mbc_cast_off < from_dt:
            continue

        # ── Optional column filter ──────────────────────────────────────────
        if column_filter:
            col_field_map = {
                'trip_start':                 'trip_start',
                'along_side_vessel':          'along_side_vessel',
                'commenced_loading':          'commenced_loading',
                'completed_loading':          'completed_loading',
                'cast_off_mv':                'cast_off_mv',
                'anchored_gull_island_empty': 'arrival_gull_island',
                'aweigh_gull_island_empty':   'departure_gull_island',
                'amf_at_port':                'mbc_arrival_port',
                'along_side_berth':           'unloaded_berth',
                'commence_discharge_berth':   'unloading_commenced',
                'completed_discharge_berth':  'unloading_completed',
                'cast_off_port':              'mbc_cast_off',
            }
            py_field = col_field_map.get(column_filter)
            if py_field:
                col_dt = safe_dt(row.get(py_field))
                if not col_dt or not (from_dt <= col_dt <= to_dt):
                    continue

        # ── Status determination ────────────────────────────────────────────
        status = get_mbc_status(row)
        row['current_status'] = status

        if status is None:
            continue
        

        if status_filter == 'all' or status == status_filter:
            filtered_rows.append(row)

        status_order = {
        'currently_loading': 1,
        'loaded_transit': 2,
        'waiting_gull': 3,
        'on_way_dharamtar': 4,
        'waiting_discharge': 5,
        'under_discharge': 6,
        'waiting_castoff': 7,
        'completed_discharge': 8
    }

    filtered_rows.sort(
        key=lambda x: status_order.get(
            x.get('current_status'),
            999
        )
    )


    return jsonify(filtered_rows)

@bp.route('/api/module/RP01/mv-barge-report/shift-data')
@login_required
def get_shift_data():

    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')

    try:
        # ── FROM date = shift day (06:00 starts)
        from_dt = datetime.fromisoformat(from_date) if from_date else datetime.now()
        to_dt   = datetime.fromisoformat(to_date)   if to_date   else datetime.now()

        # ── Entry date = FROM date का date part
        # ── Example: 2026-04-01T06:00 → entry_date = 2026-04-01
        # ── But if time < 06:00, it belongs to previous day's shift
        # ── So use FROM date's date directly
        entry_date_from = from_dt.date()
        entry_date_to   = to_dt.date()

        # ── If TO time is exactly 06:00 of next day,
        # ── then entry_date_to should be FROM date only
        # ── Because 2026-04-02T06:00 means end of 2026-04-01 shift day
        if (to_dt - from_dt).total_seconds() <= 24 * 3600:
            # Single shift day — use only FROM date
            entry_date_to = entry_date_from

    except Exception:
        return jsonify({'error': 'Invalid date format'}), 400

    conn = get_db()
    cur  = get_cursor(conn)

    cur.execute("""
        SELECT
            ll.shift,
            ll.equipment_name,
            ll.barge_name,
            ll.cargo_name,
            COALESCE(ll.quantity, 0) AS quantity,
            ll.source_type,

            CASE
                WHEN ll.source_type = 'VCN' THEN
                    (SELECT h.vessel_name
                     FROM ldud_header h
                     WHERE h.vcn_id = ll.source_id
                     LIMIT 1)
                WHEN ll.source_type = 'MBC' THEN
                    (SELECT h.mbc_name
                     FROM mbc_header h
                     WHERE h.id = ll.source_id
                     LIMIT 1)
                ELSE ''
            END AS mv_name

        FROM lueu_lines ll

        WHERE
            ll.is_deleted IS NOT TRUE
            AND COALESCE(ll.quantity, 0) > 0
            AND ll.entry_date IS NOT NULL
            AND TO_DATE(ll.entry_date, 'YYYY-MM-DD')
                BETWEEN %s AND %s

        ORDER BY
            ll.shift,
            ll.equipment_name,
            ll.barge_name
    """, (entry_date_from, entry_date_to))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    # ── Build shift map ───────────────────────────────────────
    equipments = [
        'Barge Unloader 1', 'Barge Unloader 2',
        'SANY 285-Exavator',
        'Sennebogen J1', 'Sennebogen J5',
        'BUL-01', 'BUL-02', 'BUL-03', 'BUL-04', 'BUL-05'
    ]

    shift_map = {
        'A Shift': {},
        'B Shift': {},
        'C Shift': {}
    }

    for shift in shift_map:
        for eq in equipments:
            shift_map[shift][eq] = []

    # ── First pass: aggregate by (shift, equipment, barge_name, cargo_name, mv_name) ──
    agg = {}  # key → combined item

    for row in rows:
        raw_shift = (row['shift'] or '').strip().upper()

        if raw_shift == 'A':
            shift_key = 'A Shift'
        elif raw_shift == 'B':
            shift_key = 'B Shift'
        elif raw_shift == 'C':
            shift_key = 'C Shift'
        else:
            continue

        equipment = (row['equipment_name'] or '').strip()

        if equipment not in shift_map[shift_key]:
            continue

        barge   = row['barge_name'] or ''
        cargo   = row['cargo_name'] or ''
        mv      = row['mv_name']    or ''
        source  = row['source_type'] or ''
        qty     = float(row['quantity'] or 0)

        key = (shift_key, equipment, barge, cargo, mv)

        if key in agg:
            agg[key]['quantity'] += qty
        else:
            agg[key] = {
                'barge_name':  barge,
                'cargo_name':  cargo,
                'quantity':    qty,
                'mv_name':     mv,
                'source_type': source,
            }

    # ── Second pass: populate shift_map from aggregated data ──
    for (shift_key, equipment, barge, cargo, mv), item in agg.items():
        shift_map[shift_key][equipment].append(item)

    return jsonify(shift_map)

@bp.route('/api/module/RP01/berth-details', methods=['GET', 'POST'])
@login_required
def berth_details():
    conn = get_db()
    cur = get_cursor(conn)

    if request.method == 'POST':
        data = request.get_json()

        def to_iso(val):
            """Convert DD-MM-YYYY HH:MM → YYYY-MM-DD HH:MM:SS for DB storage"""
            if not val:
                return None
            val = str(val).strip()
            # Already ISO format
            if len(val) >= 10 and val[4] == '-':
                return val
            # DD-MM-YYYY HH:MM format
            try:
                from datetime import datetime
                dt = datetime.strptime(val, '%d-%m-%Y %H:%M')
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return val  # return as-is if parsing fails

        cur.execute("""
            UPDATE ldud_barge_lines
            SET
                along_side_berth          = %s,
                commence_discharge_berth  = %s,
                completed_discharge_berth = %s,
                cast_off_berth            = %s,
                cast_off_port             = %s,
                port_crane                = %s
            WHERE id = %s
        """, (
            to_iso(data.get('along_side_berth')),
            to_iso(data.get('commence_discharge_berth')),
            to_iso(data.get('completed_discharge_berth')),
            to_iso(data.get('cast_off_berth_nt')),
            to_iso(data.get('cast_off_port')),
            data.get('unloaded_by'),
            data.get('id')
        ))

        conn.commit()
        conn.close()
        return jsonify({"success": True})
    
@bp.route('/api/module/RP01/mbc-berth-details', methods=['POST'])
@login_required
def mbc_berth_details():
    conn = get_db()
    cur = get_cursor(conn)

    data = request.get_json()

    def to_iso(val):
        if not val:
            return None
        val = str(val).strip()
        if len(val) >= 10 and val[4] == '-':
            return val
        try:
            dt = datetime.strptime(val, '%d-%m-%Y %H:%M')
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return val

    cur.execute("""
        UPDATE mbc_discharge_port_lines
        SET
            vessel_arrival_port    = %s,
            vessel_all_made_fast   = %s,
            unloading_commenced    = %s,
            cleaning_commenced     = %s,
            unloading_completed    = %s,
            vessel_cast_off        = %s,
            sailed_out_load_port   = %s,
            vessel_unloaded_by     = %s,
            vessel_unloading_berth = %s
        WHERE mbc_id = %s
    """, (
        to_iso(data.get('mbc_arrival_port')),
        to_iso(data.get('mbc_amf_unloading_berth')),
        to_iso(data.get('unloading_commenced')),
        to_iso(data.get('cleaning_commenced')),
        to_iso(data.get('unloading_completed')),
        to_iso(data.get('mbc_cast_off')),
        to_iso(data.get('sailed_out_load_port')),
        data.get('vessel_unloaded_by'),
        data.get('unloaded_berth'),
        data.get('id')
    ))

    conn.commit()
    conn.close()
    return jsonify({"success": True})