"""RP01 Historical Data — parsing, validation, reconciliation, full-replace,
template export. Pure helpers have no DB dependency; DB functions open their
own connection (mirrors the rest of the codebase)."""
import difflib
from datetime import date, datetime

from database import get_db, get_cursor

# Ordered base columns (one source of truth). Header labels in the template
# match these exact keys so upload mapping is a direct dict lookup.
COLUMNS = [
    'entry_date', 'shift', 'equipment_name', 'from_time', 'to_time',
    'source_display', 'barge_name', 'cargo_name', 'delay_name', 'system_name',
    'route_name', 'berth_name', 'shift_incharge', 'operator_name',
    'quantity', 'quantity_uom', 'remarks',
]

# Master-backed columns → (master table, master column). barge_name is special
# (matches barge OR mbc master) and handled separately in reconcile().
MASTER_MAP = {
    'equipment_name': ('equipment', 'name'),
    'cargo_name':     ('vessel_cargo', 'cargo_name'),
    'delay_name':     ('port_delay_types', 'name'),
    'route_name':     ('conveyor_routes', 'route_name'),
    'system_name':    ('port_systems', 'name'),
    'berth_name':     ('port_berth_master', 'berth_name'),
    'operator_name':  ('port_shift_operators', 'name'),
    'shift_incharge': ('port_shift_incharge', 'name'),
    'source_display': ('vessels', 'vessel_name'),
}

# Columns whose unknown values may be ADDED to a master from the preview. Only
# single-master, name-column tables (source_display/barge_name are excluded —
# they are vessel/barge OR MBC, so they must be replaced, not added). Value:
# (table, name_column, {extra_fixed_columns}).
ADDABLE_MASTERS = {
    'equipment_name': ('equipment', 'name', {}),
    'cargo_name':     ('vessel_cargo', 'cargo_name', {}),
    'delay_name':     ('port_delay_types', 'name', {}),
    'route_name':     ('conveyor_routes', 'route_name', {'is_active': 1}),
    'system_name':    ('port_systems', 'name', {}),
    'berth_name':     ('port_berth_master', 'berth_name', {}),
    'operator_name':  ('port_shift_operators', 'name', {}),
    'shift_incharge': ('port_shift_incharge', 'name', {}),
}


# ── Pure validators ──────────────────────────────────────────────────────────
def _blank(v):
    return v is None or (isinstance(v, str) and v.strip() == '')


def parse_date(v):
    """'YYYY-MM-DD' (optionally with time) → 'YYYY-MM-DD'. Blank→None. Else ValueError."""
    if _blank(v):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    raise ValueError(f"invalid date '{s}' (expected YYYY-MM-DD)")


def parse_hhmm(v):
    """'HH:MM' or 'HH:MM:SS' → 'HH:MM'. Blank→None. Else ValueError."""
    if _blank(v):
        return None
    if isinstance(v, (datetime,)):
        return v.strftime('%H:%M')
    s = str(v).strip()
    parts = s.split(':')
    if len(parts) >= 2:
        try:
            h, m = int(parts[0]), int(parts[1])
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    raise ValueError(f"invalid time '{s}' (expected HH:MM)")


def parse_number(v):
    """Numeric → float. Blank→None. Else ValueError."""
    if _blank(v):
        return None
    try:
        return float(str(v).strip().replace(',', ''))
    except (TypeError, ValueError):
        raise ValueError(f"invalid number '{v}'")


def suggest_matches(value, master_values, n=3, cutoff=0.6):
    """Closest master values to `value` (case-insensitive), up to n."""
    if not value or not master_values:
        return []
    lower_map = {}
    for m in master_values:
        lower_map.setdefault(str(m).lower(), m)
    hits = difflib.get_close_matches(str(value).lower(), list(lower_map.keys()), n=n, cutoff=cutoff)
    return [lower_map[h] for h in hits]


def parse_rows(headers, raw_rows):
    """Map raw spreadsheet rows to field dicts using `headers`.

    Returns (rows, errors). A row is skipped if every cell is blank.
    Required: entry_date (valid), equipment_name (non-empty). Format errors on
    entry_date/from_time/to_time/quantity are collected per row (1-based, +1 for
    the header row → matches the spreadsheet row number).
    """
    idx = {h: i for i, h in enumerate(headers) if h in COLUMNS}
    rows, errors = [], []

    def cell(r, key):
        i = idx.get(key)
        return r[i] if (i is not None and i < len(r)) else None

    for n, r in enumerate(raw_rows, start=2):  # +1 header, +1 to 1-base
        if all(_blank(c) for c in r):
            continue
        rec, row_errs = {}, []
        # text fields straight through (trimmed)
        for key in COLUMNS:
            if key in ('entry_date', 'from_time', 'to_time', 'quantity'):
                continue
            v = cell(r, key)
            rec[key] = (str(v).strip() if not _blank(v) else None)
        # typed fields
        try:
            rec['entry_date'] = parse_date(cell(r, 'entry_date'))
        except ValueError as e:
            row_errs.append(str(e))
        try:
            rec['from_time'] = parse_hhmm(cell(r, 'from_time'))
        except ValueError as e:
            row_errs.append(str(e))
        try:
            rec['to_time'] = parse_hhmm(cell(r, 'to_time'))
        except ValueError as e:
            row_errs.append(str(e))
        try:
            rec['quantity'] = parse_number(cell(r, 'quantity'))
        except ValueError as e:
            row_errs.append(str(e))
        # required
        if not rec.get('entry_date'):
            row_errs.append('entry_date is required')
        if _blank(rec.get('equipment_name')):
            row_errs.append('equipment_name is required')

        if row_errs:
            errors.append({'row': n, 'message': '; '.join(row_errs)})
        else:
            rows.append(rec)
    return rows, errors


# ── Master lookups ───────────────────────────────────────────────────────────
def _fetch_master(cur, table, col, extra_where=''):
    cur.execute(f"SELECT {col} AS v FROM {table} {extra_where} ORDER BY {col}")
    return [r['v'] for r in cur.fetchall() if r['v'] not in (None, '')]


def get_all_masters():
    """Return {logical_name: [values]} for every list shown in the template
    Masters sheet and used for reconciliation."""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        masters = {
            'equipment_name': _fetch_master(cur, 'equipment', 'name'),
            'cargo_name':     _fetch_master(cur, 'vessel_cargo', 'cargo_name'),
            'delay_name':     _fetch_master(cur, 'port_delay_types', 'name'),
            'route_name':     _fetch_master(cur, 'conveyor_routes', 'route_name', 'WHERE is_active = 1'),
            'system_name':    _fetch_master(cur, 'port_systems', 'name'),
            'berth_name':     _fetch_master(cur, 'port_berth_master', 'berth_name'),
            'operator_name':  _fetch_master(cur, 'port_shift_operators', 'name'),
            'shift_incharge': _fetch_master(cur, 'port_shift_incharge', 'name'),
            'barge_name':     _fetch_master(cur, 'barges', 'barge_name'),
            'mbc_name':       _fetch_master(cur, 'mbc_master', 'mbc_name'),
            'source_display': _fetch_master(cur, 'vessels', 'vessel_name'),
        }
    finally:
        conn.close()
    return masters


# Columns that may legitimately hold a vessel name OR an MBC name, so they are
# reconciled against the union of both masters:
#   - source_display: vessel name (vessel ops) OR MBC name (MBC ops)
#   - barge_name:     barge name (vessel ops) OR MBC name (MBC ops)
UNION_MASTER_KEYS = {
    'source_display': ['source_display', 'mbc_name'],
    'barge_name':     ['barge_name', 'mbc_name'],
}


def reconcile(rows, masters):
    """For each master-backed column, split the rows' distinct values into
    recognized vs unknown (with fuzzy suggestions). source_display and barge_name
    match against the vessel/barge master OR the MBC master.
    Returns {column: {recognized:[...], unknown:[{value,count,suggestions}]}}."""
    out = {}
    cols = list(MASTER_MAP.keys()) + ['barge_name']
    for col in cols:
        # distinct values + counts
        counts = {}
        for r in rows:
            v = r.get(col)
            if v:
                counts[v] = counts.get(v, 0) + 1
        keys = UNION_MASTER_KEYS.get(col, [col])
        suggest_pool = []
        for k in keys:
            suggest_pool += list(masters.get(k, []))
        valid = {str(x).lower() for x in suggest_pool}
        recognized, unknown = [], []
        for value, count in sorted(counts.items()):
            if str(value).lower() in valid:
                recognized.append(value)
            else:
                unknown.append({'value': value, 'count': count,
                                'suggestions': suggest_matches(value, suggest_pool)})
        out[col] = {'recognized': recognized, 'unknown': unknown}
    return out


def master_options(masters):
    """Return {column: [values]} the full candidate list for each reconcilable
    column's replace-picker. source_display/barge_name use the vessel/barge ∪ MBC
    union (sorted, de-duped, case-insensitive)."""
    out = {}
    for col in list(MASTER_MAP.keys()) + ['barge_name']:
        keys = UNION_MASTER_KEYS.get(col, [col])
        seen, vals = set(), []
        for k in keys:
            for v in masters.get(k, []):
                lk = str(v).lower()
                if lk not in seen:
                    seen.add(lk)
                    vals.append(v)
        out[col] = sorted(vals, key=lambda s: str(s).lower())
    return out


def apply_resolutions(rows, resolutions):
    """Apply 'replace' resolutions to parsed rows (pure). `resolutions` is
    {column: {old_value: {'action': 'replace'|'add'|'keep', 'target': new}}}.
    Only 'replace' with a non-empty target rewrites the cell; 'add'/'keep' leave
    the value as-is. Returns a new list of row dicts."""
    # Build {column: {old: new}} for replace actions only.
    repl = {}
    for col, mapping in (resolutions or {}).items():
        for old, res in (mapping or {}).items():
            if isinstance(res, dict) and res.get('action') == 'replace' and res.get('target'):
                repl.setdefault(col, {})[old] = res['target']
    if not repl:
        return [dict(r) for r in rows]
    out = []
    for r in rows:
        nr = dict(r)
        for col, m in repl.items():
            if nr.get(col) in m:
                nr[col] = m[nr[col]]
        out.append(nr)
    return out


def add_to_master(column, value):
    """Insert `value` into the master backing `column` (only single-master,
    addable columns). No-op if the value already exists. Returns True if a row
    was inserted, False if it already existed. Raises KeyError for non-addable
    columns."""
    table, namecol, extra = ADDABLE_MASTERS[column]
    cols = [namecol] + list(extra.keys())
    vals = [value] + list(extra.values())
    placeholders = ', '.join(['%s'] * len(cols))
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute(
            f"INSERT INTO {table} ({', '.join(cols)}) "
            f"SELECT {placeholders} "
            f"WHERE NOT EXISTS (SELECT 1 FROM {table} WHERE LOWER({namecol}) = LOWER(%s))",
            vals + [value],
        )
        inserted = cur.rowcount > 0
        conn.commit()
        return inserted
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ── Full replace ─────────────────────────────────────────────────────────────
def replace_all(rows, uploaded_by):
    """TRUNCATE then bulk-insert all rows, in one transaction. Returns inserted count."""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("TRUNCATE rp01_historical_lueu RESTART IDENTITY")
        insert_cols = COLUMNS + ['uploaded_by']
        placeholders = ', '.join(['%s'] * len(insert_cols))
        sql = (f"INSERT INTO rp01_historical_lueu ({', '.join(insert_cols)}) "
               f"VALUES ({placeholders})")
        for r in rows:
            cur.execute(sql, [r.get(c) for c in COLUMNS] + [uploaded_by])
        conn.commit()
        return len(rows)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_status():
    """Return {count, uploaded_by, uploaded_at} for the current dataset."""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("SELECT COUNT(*) AS c, MAX(uploaded_at) AS at, MAX(uploaded_by) AS by_id FROM rp01_historical_lueu")
        r = cur.fetchone()
        at = r['at']
        return {'count': r['c'], 'uploaded_by': r['by_id'],
                'uploaded_at': at.isoformat() if at else None}
    finally:
        conn.close()


# ── Stored-rows preview + inline edit ────────────────────────────────────────
# Columns that aren't plain text need a ::TEXT cast for ILIKE filtering.
_CAST_FILTER_COLS = {'entry_date', 'quantity'}


def get_rows(page=1, size=50, filters=None):
    """Paginated stored rows for the preview grid. Returns (rows, total).
    `filters` is a list of {field, value}; each is matched (ILIKE) against the
    FULL dataset (applied before pagination) and AND-combined. Any base column is
    filterable; `entry_date`/`quantity` are cast to text for matching."""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        allowed = set(COLUMNS)
        where_parts, params = [], []
        for f in (filters or []):
            field = f.get('field')
            val = (f.get('value') or '').strip() if f.get('value') is not None else ''
            if field in allowed and val:
                col = f"{field}::TEXT" if field in _CAST_FILTER_COLS else field
                where_parts.append(f"{col} ILIKE %s")
                params.append(f"%{val}%")
        where = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''
        cur.execute(f"SELECT COUNT(*) AS c FROM rp01_historical_lueu {where}", params)
        total = cur.fetchone()['c']
        offset = (page - 1) * size
        cur.execute(
            f"SELECT * FROM rp01_historical_lueu {where} ORDER BY entry_date, id LIMIT %s OFFSET %s",
            params + [size, offset])
        rows = []
        for r in cur.fetchall():
            d = dict(r)
            if d.get('entry_date') is not None:
                d['entry_date'] = str(d['entry_date'])
            if d.get('quantity') is not None:
                d['quantity'] = float(d['quantity'])
            if d.get('uploaded_at') is not None:
                d['uploaded_at'] = str(d['uploaded_at'])
            rows.append(d)
        return rows, total
    finally:
        conn.close()


# Free-text columns editable inline (typed columns handled via parse_* below).
_TEXT_EDIT_COLS = ['shift', 'equipment_name', 'source_display', 'barge_name',
                   'cargo_name', 'delay_name', 'system_name', 'route_name',
                   'berth_name', 'shift_incharge', 'operator_name', 'quantity_uom', 'remarks']


def update_row(row_id, data):
    """Validate + update a single stored row. Returns {'success': True} or
    {'error': msg}. Mirrors the upload validators for the typed columns."""
    clean = {}
    try:
        clean['entry_date'] = parse_date(data.get('entry_date'))
        clean['from_time'] = parse_hhmm(data.get('from_time'))
        clean['to_time'] = parse_hhmm(data.get('to_time'))
        clean['quantity'] = parse_number(data.get('quantity'))
    except ValueError as e:
        return {'error': str(e)}
    if not clean['entry_date']:
        return {'error': 'entry_date is required'}
    for k in _TEXT_EDIT_COLS:
        v = data.get(k)
        clean[k] = (str(v).strip() if v not in (None, '') else None)
    if not clean.get('equipment_name'):
        return {'error': 'equipment_name is required'}

    conn = get_db()
    cur = get_cursor(conn)
    try:
        sets = ', '.join([f"{c}=%s" for c in COLUMNS])
        cur.execute(f"UPDATE rp01_historical_lueu SET {sets} WHERE id=%s",
                    [clean.get(c) for c in COLUMNS] + [row_id])
        conn.commit()
        return {'success': True}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def delete_row(row_id):
    """Delete one stored row by id."""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("DELETE FROM rp01_historical_lueu WHERE id=%s", [row_id])
        conn.commit()
    finally:
        conn.close()


# ── Template / upload parsing ────────────────────────────────────────────────
def _example_rows():
    """Illustrative sample rows for the Example sheet. Demonstrates a vessel
    unloading + delay (barge in barge_name), an MBC unloading (MBC name in BOTH
    source_display and barge_name), and an equipment-only delay (no source)."""
    return [
        {'entry_date': '2026-04-01', 'shift': 'A', 'equipment_name': 'BARGE UNLOADER 1',
         'from_time': '06:00', 'to_time': '08:00', 'source_display': 'M.V EVA SHANGHAI',
         'barge_name': 'KINGFISHER', 'cargo_name': 'Dolomite', 'quantity': 700,
         'quantity_uom': 'MT', 'route_name': 'C-131', 'operator_name': 'Harshad Thakur',
         'shift_incharge': 'Rakesh Mhatre', 'remarks': 'Vessel op — barge unloading'},
        {'entry_date': '2026-04-01', 'shift': 'A', 'equipment_name': 'BARGE UNLOADER 1',
         'from_time': '08:00', 'to_time': '08:30', 'source_display': 'M.V EVA SHANGHAI',
         'barge_name': 'KINGFISHER', 'delay_name': 'Crane idle due to hopper full',
         'remarks': 'Vessel op — delay (no quantity)'},
        {'entry_date': '2026-04-01', 'shift': 'B', 'equipment_name': 'BARGE UNLOADER 2',
         'from_time': '14:00', 'to_time': '16:00', 'source_display': 'JSW ARJUNGAD',
         'barge_name': 'JSW ARJUNGAD', 'cargo_name': 'BRBF Fines', 'quantity': 500,
         'quantity_uom': 'MT', 'remarks': 'MBC op — MBC name in BOTH source_display and barge_name'},
        {'entry_date': '2026-04-02', 'shift': 'A', 'equipment_name': 'BU 1 & BU 2',
         'from_time': '06:00', 'to_time': '06:30', 'delay_name': 'Hopper Empty Out',
         'remarks': 'Equipment-only delay — no vessel/MBC source'},
    ]


def build_template_workbook():
    """Return an openpyxl Workbook with three sheets: 'Data' (headers +
    instructions + dropdowns to fill in), 'Example' (sample rows), and 'Masters'
    (live master values backing the dropdowns)."""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    masters = get_all_masters()
    wb = openpyxl.Workbook()

    # Masters sheet first (so we can reference its ranges)
    ms = wb.create_sheet('Masters')
    # column order on Masters sheet → maps to a Data column for dropdowns
    master_order = ['equipment_name', 'cargo_name', 'delay_name', 'route_name',
                    'system_name', 'berth_name', 'operator_name', 'shift_incharge',
                    'barge_name', 'mbc_name', 'source_display']
    master_col_letter = {}
    for ci, key in enumerate(master_order, start=1):
        col = get_column_letter(ci)
        master_col_letter[key] = col
        ms.cell(1, ci, key)
        for ri, val in enumerate(masters.get(key, []), start=2):
            ms.cell(ri, ci, val)

    note_font = Font(size=10, italic=True, color='744210')
    note_fill = PatternFill('solid', fgColor='FFFBEB')
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    wrap = Alignment(wrap_text=True, vertical='top')

    note = (
        "HOW TO FILL — one row per equipment time-slice (see the Example sheet).\n"
        "• entry_date = YYYY-MM-DD (required)   • from_time / to_time = HH:MM (24h)   "
        "• quantity = number (unloading rows only)   • equipment_name = required\n"
        "• SOURCE: enter the VESSEL name OR the MBC name in 'source_display'.\n"
        "   – Vessel source → put the barge name in 'barge_name'.\n"
        "   – MBC source → put the MBC name in BOTH 'source_display' AND 'barge_name'.\n"
        "• Leave source_display & barge_name blank for equipment-only delays.   "
        "Pick values from the dropdowns (sourced from the Masters sheet)."
    )

    def write_grid(ws, with_dropdowns, example_rows=None):
        ws.cell(1, 1, note).font = note_font
        ws.cell(1, 1).fill = note_fill
        ws.cell(1, 1).alignment = wrap
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        ws.row_dimensions[1].height = 92
        for ci, key in enumerate(COLUMNS, start=1):
            c = ws.cell(2, ci, key)
            c.font = hdr_font
            c.fill = hdr_fill
            ws.column_dimensions[get_column_letter(ci)].width = max(14, len(key) + 2)
        ws.freeze_panes = 'A3'
        if example_rows:
            for ri, rec in enumerate(example_rows, start=3):
                for ci, key in enumerate(COLUMNS, start=1):
                    ws.cell(ri, ci, rec.get(key))
        if with_dropdowns:
            dropdown_for = {
                'equipment_name': 'equipment_name', 'cargo_name': 'cargo_name',
                'delay_name': 'delay_name', 'route_name': 'route_name',
                'system_name': 'system_name', 'berth_name': 'berth_name',
                'operator_name': 'operator_name', 'shift_incharge': 'shift_incharge',
                'barge_name': 'barge_name', 'source_display': 'source_display',
            }
            for data_key, master_key in dropdown_for.items():
                if data_key not in COLUMNS:
                    continue
                mcol = master_col_letter[master_key]
                dcol = get_column_letter(COLUMNS.index(data_key) + 1)
                dv = DataValidation(type='list',
                                    formula1=f"Masters!${mcol}$2:${mcol}$2000",
                                    allow_blank=True, showErrorMessage=False)
                ws.add_data_validation(dv)
                dv.add(f"{dcol}3:{dcol}10000")

    # Data sheet (fill-in, with dropdowns)
    ds = wb.active
    ds.title = 'Data'
    write_grid(ds, with_dropdowns=True)

    # Example sheet (sample rows, no dropdowns) — placed right after Data
    ex = wb.create_sheet('Example', index=1)
    write_grid(ex, with_dropdowns=False, example_rows=_example_rows())

    return wb


def parse_upload(file_storage):
    """Parse a werkzeug FileStorage (.xlsx/.xls/.csv) → (rows, errors).
    Reads the 'Data' sheet for workbooks (falls back to the active sheet).
    Header row is the row whose first cell equals 'entry_date'."""
    import io, csv as _csv
    name = (file_storage.filename or '').lower()
    raw = file_storage.read()

    def from_matrix(matrix):
        header_idx = None
        for i, row in enumerate(matrix):
            if row and str(row[0]).strip() == 'entry_date':
                header_idx = i
                break
        if header_idx is None:
            return [], [{'row': 0, 'message': "Could not find a header row starting with 'entry_date'"}]
        headers = [str(c).strip() if c is not None else '' for c in matrix[header_idx]]
        return parse_rows(headers, matrix[header_idx + 1:])

    if name.endswith('.csv'):
        text = raw.decode('utf-8-sig', errors='replace')
        matrix = list(_csv.reader(io.StringIO(text)))
        return from_matrix(matrix)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active
        matrix = [list(r) for r in ws.iter_rows(values_only=True)]
        return from_matrix(matrix)
