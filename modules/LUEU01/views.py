from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, Response
from werkzeug.security import check_password_hash
from functools import wraps
from . import model
from database import get_user_permissions, get_db, get_cursor

bp = Blueprint('LUEU01', __name__, template_folder='.')
MODULE_CODE = 'LUEU01'
MODULE_INFO = {'code': 'LUEU01', 'name': 'Load Unload Equipment Utilization'}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def get_perms():
    if session.get('is_admin'):
        return {'can_read': 1, 'can_add': 1, 'can_edit': 1, 'can_delete': 1}
    return get_user_permissions(session.get('user_id'), MODULE_CODE)

@bp.route('/module/LUEU01/')
@login_required
def view():
    perms = get_perms()
    if not perms.get('can_read'):
        return render_template('no_access.html'), 403
    return render_template('lueu01.html', permissions=perms)

# Data endpoints
@bp.route('/api/module/LUEU01/data')
@login_required
def get_data():
    import json as _json
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)
    equipment_name = request.args.get('equipment', None)
    filters_raw = request.args.get('filters', None)
    filters = _json.loads(filters_raw) if filters_raw else []
    return jsonify(model.get_all_lines(page, size, equipment_name, filters))

@bp.route('/api/module/LUEU01/save', methods=['POST'])
@login_required
def save_data():
    perms = get_perms()
    if not perms.get('can_add') and not perms.get('can_edit'):
        return jsonify({'error': 'No permission'}), 403

    data = request.json
    data['created_by'] = session.get('username')
    result = model.save_line(data)   # {'id': ..., 'rejections': [...]}
    return jsonify(result)

@bp.route('/api/module/LUEU01/split', methods=['POST'])
@login_required
def split_line():
    perms = get_perms()
    if not perms.get('can_edit'):
        return jsonify({'error': 'No permission'}), 403
    data = request.json
    line_id = data.get('line_id')
    split_qty = data.get('split_quantity')
    split_remark = data.get('split_remark', '')
    if not line_id or not split_qty:
        return jsonify({'error': 'Missing line_id or split_quantity'}), 400
    result = model.split_line(line_id, split_qty, split_remark, session.get('username'))
    if result:
        return jsonify({'success': True, **result})
    return jsonify({'error': 'Line not found'}), 404

@bp.route('/api/module/LUEU01/delete', methods=['POST'])
@login_required
def delete_data():
    perms = get_perms()
    if not perms.get('can_delete'):
        return jsonify({'error': 'No permission to delete'}), 403

    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400

    password = request.json.get('password', '')
    username = session.get('username')

    # Verify the current user's password before allowing deletion
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT id, password FROM users WHERE username = %s', (username,))
    user = cur.fetchone()
    conn.close()
    if not user or not check_password_hash(user['password'], password):
        return jsonify({'error': 'Incorrect password. Deletion not authorised.'}), 403

    model.soft_delete_lines(ids, username=username)

    return jsonify({
        'success': True,
        'deleted_count': len(ids)
    })

# Dropdown data endpoints
@bp.route('/api/module/LUEU01/vcn-options')
@login_required
def get_vcn_options():
    options = model.get_vcn_options()
    result = []
    for opt in options:
        anchored = opt.get('anchorage_arrival', '')
        if anchored:
            anchored = anchored[:16].replace('T', ' ')
        display = f"{opt['vcn_doc_num']} / {opt['vessel_name']}"
        result.append({
            'value': display,
            'label': display,
            'type': 'VCN',
            'id': opt['id']
        })
    return jsonify(result)

@bp.route('/api/module/LUEU01/mbc-options')
@login_required
def get_mbc_options():
    options = model.get_mbc_options()
    result = []
    for opt in options:
        display = f"{opt['doc_num']} / {opt['mbc_name']}"
        result.append({
            'value': display,
            'label': display,
            'type': 'MBC',
            'id': opt['id'],
            'cargo_name': opt.get('cargo_name') or '',
            'completed': bool(opt.get('completed'))
        })
    return jsonify(result)

@bp.route('/api/module/LUEU01/equipment')
@login_required
def get_equipment():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT name FROM equipment ORDER BY name')
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@bp.route('/api/module/LUEU01/delays')
@login_required
def get_delays():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT name FROM port_delay_types ORDER BY name')
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@bp.route('/api/module/LUEU01/cargo')
@login_required
def get_cargo():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT cargo_name FROM vessel_cargo ORDER BY cargo_name')
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['cargo_name'] for r in rows])

@bp.route('/api/module/LUEU01/operation-types')
@login_required
def get_operation_types():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT name FROM vessel_operation_types ORDER BY name')
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@bp.route('/api/module/LUEU01/uom')
@login_required
def get_uom():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT name, is_default FROM quantity_uom ORDER BY name')
    rows = cur.fetchall()
    conn.close()
    names = [r['name'] for r in rows]
    default_uom = next((r['name'] for r in rows if r['is_default']), '')
    return jsonify({'names': names, 'default': default_uom})

@bp.route('/api/module/LUEU01/barges/<int:vcn_id>')
@login_required
def get_barges_for_vcn(vcn_id):
    barges = model.get_vcn_barges(vcn_id)
    return jsonify(barges)

@bp.route('/api/module/LUEU01/mbc-names')
@login_required
def get_mbc_names():
    names = model.get_mbc_names()
    return jsonify(names)

@bp.route('/api/module/LUEU01/routes')
@login_required
def get_routes():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT route_name FROM conveyor_routes WHERE is_active = 1 ORDER BY route_name')
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['route_name'] for r in rows])

@bp.route('/api/module/LUEU01/systems')
@login_required
def get_systems():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute("SELECT name FROM port_systems WHERE name IS NOT NULL AND name != '' ORDER BY name")
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@bp.route('/api/module/LUEU01/berths')
@login_required
def get_berths():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('SELECT berth_name FROM port_berth_master ORDER BY berth_name')
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['berth_name'] for r in rows])

@bp.route('/api/module/LUEU01/shift-incharge')
@login_required
def get_shift_incharge():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute("SELECT name FROM port_shift_incharge WHERE name IS NOT NULL AND name != '' ORDER BY name")
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@bp.route('/api/module/LUEU01/shift-operators')
@login_required
def get_shift_operators():
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute("SELECT name FROM port_shift_operators WHERE name IS NOT NULL AND name != '' ORDER BY name")
    rows = cur.fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@bp.route('/api/module/LUEU01/bl-progress/<source_type>/<int:source_id>')
@login_required
def get_bl_progress(source_type, source_id):
    if source_type not in ('VCN', 'MBC'):
        return jsonify({'error': 'Invalid source type'}), 400
    return jsonify(model.get_bl_progress(source_type, source_id))


@bp.route('/api/module/LUEU01/export')
@login_required
def export_excel():
    """Dump the entire lueu_lines table to .xlsx, with a trailing Issues column
    flagging 'possible over quantity' / 'possible overlap' per row."""
    if not get_perms().get('can_read'):
        return jsonify({'error': 'No permission'}), 403

    import io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = model.get_export_rows()

    columns = [
        ('Equipment', 'equipment_name'),
        ('Date', 'entry_date'),
        ('Shift', 'shift'),
        ('From', 'from_time'),
        ('To', 'to_time'),
        ('Diff (Hrs)', 'diff_hrs'),
        ('VCN / MBC', 'source_display'),
        ('Barge / MBC Name', 'barge_name'),
        ('Cargo', 'cargo_name'),
        ('Delay', 'delay_name'),
        ('System', 'system_name'),
        ('Receiving Route', 'route_name'),
        ('Berth', 'berth_name'),
        ('Shift Incharge', 'shift_incharge'),
        ('Operator', 'operator_name'),
        ('Quantity', 'quantity'),
        ('Qty UOM', 'quantity_uom'),
        ('Remarks', 'remarks'),
        ('Deleted', '_deleted_flag'),
        ('Deleted By', 'deleted_by'),
        ('Deleted Date', 'deleted_date'),
        ('Issues', 'issues'),
    ]

    thin = Side(style='thin', color='000000')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)
    issue_fill = PatternFill('solid', fgColor='FDE2E1')
    issue_font = Font(name='Calibri', size=10, bold=True, color='B91C1C')
    del_font = Font(name='Calibri', size=10, italic=True, color='9CA3AF')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    rgt = Alignment(horizontal='right', vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LUEU Lines'
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    for col_idx, (label, _key) in enumerate(columns, 1):
        c = ws.cell(1, col_idx, label)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bdr; c.alignment = ctr
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(label) + 2)

    for row_idx, r in enumerate(rows, 2):
        is_del = bool(r.get('is_deleted'))
        for col_idx, (label, key) in enumerate(columns, 1):
            if key == '_deleted_flag':
                val = 'Yes' if is_del else ''
            else:
                val = r.get(key)
            if val is None:
                val = ''
            elif hasattr(val, 'isoformat'):
                val = str(val)
            c = ws.cell(row_idx, col_idx, val)
            c.border = bdr
            c.alignment = rgt if key in ('quantity', 'diff_hrs') else lft
            if key == 'issues' and val:
                c.fill = issue_fill; c.font = issue_font
            elif is_del:
                c.font = del_font
            else:
                c.font = cell_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"LUEU01_export_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@bp.route('/module/LUEU01/dashboard')
@login_required
def dashboard():
    return render_template('lueu01_dashboard.html')

@bp.route('/api/module/LUEU01/dashboard-data')
@login_required
def dashboard_data():
    return jsonify(model.get_dashboard_data())

@bp.route('/api/module/LUEU01/barge-cargos/<int:vcn_id>')
@login_required
def get_barge_cargos(vcn_id):
    barge_name = request.args.get('barge', '')
    cargos = model.get_barge_cargos(vcn_id, barge_name)
    return jsonify(cargos)
